#    MetExtract II
#    Copyright (C) 2015
#
#    This program is free software; you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation; either version 2 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with this program; if not, write to the Free Software
#    Foundation, Inc., 51 Franklin St, Fifth Floor, Boston, MA  02110-1301  USA


from __future__ import absolute_import, division, print_function
import asyncio
import os
import shutil
import subprocess
import sys
import threading
import time
import psutil
import logging
import pprint
import traceback
import base64
import functools
import json
import re
from collections import defaultdict
from copy import deepcopy
from math import log10
from multiprocessing import Manager, Pool, cpu_count, freeze_support, set_start_method, Lock
from operator import itemgetter
from optparse import OptionParser
from pickle import dumps, loads
from xml.parsers.expat import ExpatError
import polars as pl
import matplotlib
from PySide6 import QtCore, QtGui, QtWidgets
from PySide6.QtGui import QColor
from PySide6.QtWidgets import QCheckBox, QComboBox, QHBoxLayout, QPushButton, QTableWidgetItem, QWidget
import matplotlib.patches as patches
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
from matplotlib.ticker import ScalarFormatter
from .utils import get_app_folder, get_main_dir, getDBFormat, getDBSuffix
from . import LoggingSetup
from .utilities.AboutDialog import AboutDialog
from .mePyGuis.adductsEdit import adductsEdit, ConfiguredAdduct, ConfiguredElement
from .mePyGuis.calcIsoEnrichmentDialog import calcIsoEnrichmentDialog
from .mePyGuis.groupEdit import groupEdit
from .mePyGuis.heteroAtomEdit import heteroAtomEdit, ConfiguredHeteroAtom
from .mePyGuis.PeakPickingSettingsDialog import PeakPickingSettingsDialog
from .mePyGuis.ProgressWrapper import ProgressWrapper
from .mePyGuis.RegExTestDialog import RegExTestDialog
from .MetExtractII_Main import MetExtractVersion
from .utils import (
    Bunch,
    CallBackMethod,
    ChromPeakPair,
    FuncProcess,
    SampleGroup,
    SQLInsert,
    getNormRatio,
    mean,
    natSort,
    sd,
)
from .PolarsDB import PolarsDB
from .Chromatogram import Chromatogram
from . import HCA_general, annotateResultMatrix, pyperclip
from .annotateResultMatrix import addGroup as grpAdd
from .annotateResultMatrix import addStatsColumnToResults
from .annotateResultMatrix import performGroupOmit as grpOmit
from .bracketResults import bracketResults, calculateMetaboliteGroups, compute_sample_stats
from .mePyGuis.mainWindow import Ui_MainWindow
from .mePyGuis.TracerEdit import ConfiguredTracer, tracerEdit
from .MSMS import optimizeMSMSTargets
from .reIntegration import reIntegrateResultsFile
from .exportToInclusionList import writeIQXInclusionList, writeQExactiveInclusionList
from .resultsPostProcessing import searchDatabases as searchDatabases
from .formulaTools import formulaTools, getElementOfIsotope, getIsotopeMass

try:
    from matchms import Spectrum as MatchmsSpectrum
    from matchms.similarity import CosineGreedy as MatchmsCosineGreedy
    from matchms.similarity import CosineHungarian as MatchmsCosineHungarian
    from matchms.similarity import ModifiedCosineGreedy as MatchmsModifiedCosineGreedy
    from matchms.similarity import ModifiedCosineHungarian as MatchmsModifiedCosineHungarian
    from matchms.similarity import NeutralLossesCosine as MatchmsNeutralLossesCosine
    from matchms.filtering import normalize_intensities as matchms_normalize_intensities
    from matchms.filtering import select_by_relative_intensity as matchms_select_by_relative_intensity

    MATCHMS_AVAILABLE = True
except Exception:
    MATCHMS_AVAILABLE = False

# Registry of matchms similarity algorithms selectable in the "Show options" panel.
# The values are the matchms classes themselves; instances are created on demand with
# the user-configured m/z tolerance (see _get_msms_similarity_algorithm).
MSMS_SIMILARITY_ALGORITHMS = {}
if MATCHMS_AVAILABLE:
    MSMS_SIMILARITY_ALGORITHMS = {
        "ModifiedCosineHungarian": MatchmsModifiedCosineHungarian,
        "ModifiedCosineGreedy": MatchmsModifiedCosineGreedy,
        "CosineHungarian": MatchmsCosineHungarian,
        "CosineGreedy": MatchmsCosineGreedy,
        "NeutralLossesCosine": MatchmsNeutralLossesCosine,
    }

app = None

# Set local folder for MetExtract II
local_folder = get_app_folder()
OBO_DOWNLOAD_URL = "https://bioportal.bioontology.org/ontologies/MS"
if __name__ == "__main__":
    print(f"Using local folder '{local_folder}'")

LoggingSetup.LoggingSetup.Instance().initLogging(location=local_folder)

sys.displayhook = pprint.pprint


TRACER = object()
METABOLOME = object()

FILENAME_SAFE_PATTERN = r"[^A-Za-z0-9_.-]+"

# Boxplot layout constants for abundance-profile group comparison plots
ABUNDANCE_BOXPLOT_CLUSTER_WIDTH = 0.75
ABUNDANCE_BOXPLOT_SLOT_FILL_RATIO = 0.8


class _NumericDateSortItem(QTableWidgetItem):
    """QTableWidgetItem that sorts numerically or by ISO-timestamp when possible."""

    def __lt__(self, other):
        t_self = self.text()
        t_other = other.text()
        # Numeric comparison
        try:
            return float(t_self) < float(t_other)
        except (ValueError, TypeError):
            pass
        # ISO timestamp comparison  (e.g. "2026-04-02T14:51:17.405324")
        try:
            from datetime import datetime as _dt

            return _dt.fromisoformat(t_self) < _dt.fromisoformat(t_other)
        except (ValueError, TypeError):
            pass
        return t_self < t_other


# Helper function to safely load pickled data with error handling for old cached data
def safe_pickle_loads(data, default_value=None, operation_name="loading data"):
    """
    Safely load pickled data with error handling for compatibility issues.

    Args:
        data: The base64 encoded data to unpickle
        default_value: Value to return if unpickling fails
        operation_name: Description of the operation for logging

    Returns:
        Unpickled data or default_value if unpickling fails
    """
    try:
        return loads(base64.b64decode(data.encode("utf-8")))
    except (ImportError, ModuleNotFoundError, SyntaxError) as e:
        logging.warning(f"Could not load cached data when {operation_name}: {e}")
        logging.warning("This may be due to cached data from an older version. Using default value.")
        return default_value
    except Exception as e:
        logging.error(f"Unexpected error when {operation_name}: {e}")
        return default_value


def sanitize_filename(text):
    return re.sub(FILENAME_SAFE_PATTERN, "_", str(text))


# </editor-fold>
# <editor-fold desc="### PyQT 4 Imports">
# </editor-fold>
# <editor-fold desc="### MatPlotLib imports and setup">

matplotlib.use("Qt5Agg")  # Use Qt5Agg backend for PySide6 compatibility

# from mpldatacursor import datacursor, HighlightingDataCursor

matplotlib.rcParams["savefig.dpi"] = 300
font = {"size": 16}
matplotlib.rc("font", **font)
globAlpha = 0.05
predefinedColors = [
    "FireBrick",
    "YellowGreen",
    "SteelBlue",
    "DarkOrange",
    "Teal",
    "DarkSlateGrey",
    "CornflowerBlue",
    "DarkOliveGreen",
    "SlateGrey",
    "CadetBlue",
    "DarkCyan",
    "Black",
    "DarkSeaGreen",
    "DimGray",
    "GoldenRod",
    "LightBlue",
    "MediumTurquoise",
    "RoyalBlue",
]


# required for Plotting table
def convertXaToX(x, a):
    return ((1 - a) * 1) + (a * x)


# taken from http://www.scipy.org/Cookbook/Matplotlib/Transformations
# New enough versions have offset_copy (by Eric Firing):
if "offset_copy" in dir(matplotlib.transforms):
    from matplotlib.transforms import offset_copy

    def offset(ax, x, y):
        return offset_copy(ax.transData, x=x, y=y, units="dots")
else:  # Without offset_copy we have to do some black transform magic
    from matplotlib.transforms import blend_xy_sep_transform, identity_transform

    def offset(ax, x, y):
        # This trick makes a shallow copy of ax.transData (but fails for polar plots):
        trans = blend_xy_sep_transform(ax.transData, ax.transData)
        # Now we set the offset in pixels
        trans.set_offset((x, y), identity_transform())
        return trans


# show only bottom and left axes in a matplotlib graph
def simpleaxis(ax):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.get_xaxis().tick_bottom()
    ax.get_yaxis().tick_left()


# remove all axes in a matplotlib graph
def noaxis(ax):
    ax.spines["bottom"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.axis("off")


# from mpltools import style
# style.use('ggplot')
# </editor-fold>

# <editor-fold desc="### MZXML">

# </editor-fold>
# <editor-fold desc="### FindIsoPairs Import">
try:
    from .findIsoPairs import FindIsoPairs
except Exception as err:
    if __name__ in ["__main__", "src.MExtract"]:
        logging.error("Identification/Processing of new files is not available: %s" % (str(err)))
# </editor-fold>
# <editor-fold desc="### Group Results Import">

# </editor-fold>
# <editor-fold desc="### UI Window Imports">

# </editor-fold>
# <editor-fold desc="### Re-integration Import">

# </editor-fold>
# <editor-fold desc="### Various Imports">


def memory_usage_psutil():
    """Return memory usage in MB (cross-platform, Python 3)."""
    try:
        process = psutil.Process(os.getpid())
        # rss is in bytes
        return process.memory_info().rss / float(2**20)
    except Exception as exc:
        # Fallback for Unix-like systems without psutil
        logging.info(f"Exception: {exc}")
        return -1


# </editor-fold>

# <editor-fold desc="### debug imports">

pp = pprint.PrettyPrinter(indent=1)
# </editor-fold>


# helper method for findIsoPairs and multiprocessing (multi core support)
def findSILFeatures(rI):
    rI.findIsoPairs()


peakAbundanceUseSignals = 5
peakAbundanceUseSignalsSides = int((peakAbundanceUseSignals - 1) / 2)


# <editor-fold desc="Re-Integrate Function definitions">
# Re-integration functions have been moved to src/reIntegration.py
# This module now uses PolarsDB for re-integration with threading support
# </editor-fold>


def closeCallBack(selfObj):
    if (
        QtWidgets.QMessageBox.question(
            selfObj,
            "MetExtract",
            "Are you sure you want to cancel?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
        )
        == QtWidgets.QMessageBox.Yes
    ):
        selfObj.terminateJobs = True
        logging.info("Processing stopped by user")

        return True
    else:
        return False  # don't close progresswrapper and continue processing files


# </editor-fold>


def interruptIndividualFilesProcessing(selfObj, pool):
    if (
        QtWidgets.QMessageBox.question(
            selfObj,
            "MetExtract",
            "Are you sure you want to cancel?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
        )
        == QtWidgets.QMessageBox.Yes
    ):
        pool.close()
        pool.terminate()
        pool.join()

        selfObj.terminateJobs = True
        logging.info("Processing stopped by user")

        return True
    else:
        return False  # don't close progresswrapper and continue processing files


def interruptBracketingOfFeaturePairs(selfObj, funcProc):
    if (
        QtWidgets.QMessageBox.question(
            selfObj,
            "MetExtract",
            "Are you sure you want to cancel?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
        )
        == QtWidgets.QMessageBox.Yes
    ):
        funcProc.terminate()
        funcProc.join()

        selfObj.terminateJobs = True
        logging.info("Processing stopped by user")
    else:
        return False  # don't close progresswrapper and continue processing files


def interruptConvolutingOfFeaturePairs(selfObj, funcProc):
    if (
        QtWidgets.QMessageBox.question(
            selfObj,
            "MetExtract",
            "Are you sure you want to cancel?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
        )
        == QtWidgets.QMessageBox.Yes
    ):
        funcProc.terminate()
        funcProc.join()

        selfObj.terminateJobs = True
        logging.info("Processing stopped by user")
    else:
        return False  # don't close progresswrapper and continue processing files


def loadMZXMLFile(params):
    # {"File":fi, "Group":group.name, "IntensityThreshold":intensityThrehold}

    mzFilter = None
    if params["selectedMZs"] is not None:
        mzFilter = []
        for selMZ in params["selectedMZs"]:
            mzFilter.append(
                (
                    selMZ * (1.0 - params["ppm"] / 1.0e6),
                    selMZ * (1.0 + params["ppm"] / 1.0e6),
                )
            )

    mzXML = Chromatogram()
    mzXML.parse_file(params["File"], intensityCutoff=params["IntensityThreshold"], mzFilter=mzFilter)

    ret = {
        "File": params["File"],
        "Group": params["Group"],
        "IntensityThreshold": params["IntensityThreshold"],
    }
    ret["mzXMLFile"] = mzXML

    return ret


# Custom data role used to store the relative intensity ratio (float 0.0–1.0) on tree items.
_RELATIVE_BAR_ROLE = QtCore.Qt.ItemDataRole.UserRole + 50


class _RelativeBarDelegate(QtWidgets.QStyledItemDelegate):
    """Paints a faint green bar filling the left portion of the cell proportional to the stored ratio."""

    _COLOR_FILL = QtGui.QColor(80, 180, 80, 120)  # green bar
    _COLOR_BG = QtGui.QColor(200, 240, 200, 50)  # very faint green background tint

    def paint(self, painter, option, index):
        super().paint(painter, option, index)

        ratio = index.data(_RELATIVE_BAR_ROLE)
        try:
            ratio = float(ratio)
        except (TypeError, ValueError):
            ratio = 0.0

        if ratio > 0.0:
            ratio = min(ratio, 1.0)
            painter.save()
            rect = option.rect
            # Draw after default painting so it also remains visible on selected rows.
            painter.fillRect(rect, self._COLOR_BG)
            bar = QtCore.QRect(rect.x(), rect.y(), max(1, int(rect.width() * ratio)), rect.height())
            painter.fillRect(bar, self._COLOR_FILL)
            painter.restore()


class _MSMSTableDelegate(QtWidgets.QStyledItemDelegate):
    """Delegate for msms_SpectraList_exp: retains row background color on selection
    (with lighter transparency) and makes selected row text bold."""

    def paint(self, painter, option, index):
        # Get the stored background color from the item
        bg_color = index.data(QtCore.Qt.BackgroundRole)
        is_selected = option.state & QtWidgets.QStyle.State_Selected

        # Build a modified option so the default selection highlight is suppressed
        opt = QtWidgets.QStyleOptionViewItem(option)
        if is_selected and bg_color is not None:
            # Remove the selected state flag so the default blue selection is not drawn
            opt.state = opt.state & ~QtWidgets.QStyle.State_Selected
            # Draw standard item without selection highlight
            self.initStyleOption(opt, index)
            opt.state = opt.state & ~QtWidgets.QStyle.State_Selected

        super().paint(painter, opt, index)

        if is_selected:
            painter.save()
            # Determine lighter overlay color from existing background
            if bg_color is not None:
                if hasattr(bg_color, "color"):
                    base = bg_color.color()
                else:
                    base = bg_color
                overlay = QtGui.QColor(base.red(), base.green(), base.blue(), 160)
            else:
                overlay = QtGui.QColor(100, 160, 255, 120)
            painter.fillRect(option.rect, overlay)
            # Draw bold text
            font = painter.font()
            font.setBold(True)
            painter.setFont(font)
            text_rect = option.rect.adjusted(4, 0, -4, 0)
            text = index.data(QtCore.Qt.DisplayRole)
            if text:
                palette = option.palette
                painter.setPen(palette.color(QtGui.QPalette.Text))
                painter.drawText(text_rect, QtCore.Qt.AlignVCenter | QtCore.Qt.AlignLeft, str(text))
            painter.restore()


class _DBTestWorker(QtCore.QThread):
    """Background worker that test-imports database files without blocking the UI."""

    finished = QtCore.Signal(list)

    def __init__(self, dbFiles, parent=None):
        super().__init__(parent)
        self.dbFiles = dbFiles

    def run(self):
        results = annotateResultMatrix.testDatabaseImports(self.dbFiles)
        self.finished.emit(results)


class mainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    # <editor-fold desc="### check group functions">
    def forceUpdateFL(self):
        self.updateLCMSSampleSettings()

    # check, if a specified mzxml file can be loaded successfully
    def checkFileImport(self, file):
        # fhash="%s_%s"%(file, sha256(open(file, 'rb').read()).hexdigest())  #self.ckeckedLCMSFiles[fhash]=Bunch(parsed=parsed, fls=tm.getFilterLinesPerPolarity(), pols=tm.getPolarities(), tics=tics)
        fhash = "%s_NOHash" % (file)  ## ignore hash, files are not likely to change

        cache_file = local_folder + "/fileImport.pqts.cache"

        # Load existing cache if it exists
        cache_df = None
        if os.path.exists(cache_file):
            try:
                cache_df = pl.read_parquet(cache_file)
            except Exception as e:
                logging.warning(f"Could not load cache file: {e}. Creating new cache.")
                cache_df = None

        if fhash not in self.checkedLCMSFiles.keys():
            fetched = []
            if cache_df is not None:
                filtered = cache_df.filter(pl.col("filePath") == fhash)
                for row_dict in filtered.to_dicts():
                    fetched.append(str(row_dict["parsingInfo"]))

            if len(fetched) > 0:
                try:
                    b = loads(base64.b64decode(fetched[0].encode("utf-8")))
                    self.checkedLCMSFiles[fhash] = b
                except (ImportError, ModuleNotFoundError, SyntaxError) as e:
                    logging.warning(f"Could not load cached file import data: {e}")
                    logging.warning("This may be due to cached data from an older version. Re-parsing file.")
                    # Continue with normal file parsing by not setting the cache
            else:
                f = file.replace("\\", "/")
                f = f[f.rfind("/") + 1 : f.rfind(".")]

                if not (re.match("^[a-zA-Z0-9_]*$", f)):
                    self.checkedLCMSFiles[fhash] = Bunch(parsed="Invalid characters in file name")
                else:
                    parsed = False
                    try:
                        tm = Chromatogram()
                        tm.parse_file(file, ignoreCharacterData=True)

                        parsed = True
                        pols = tm.getPolarities()

                        fls = tm.getFilterLinesPerPolarity()

                        tics = {}
                        if "+" in pols:
                            tic, times, scanIDs = tm.getTIC(filterLine=fls["+"].pop())
                            tics["+"] = Bunch(tic=tic, times=times)
                        if "-" in pols:
                            tic, times, scanIDs = tm.getTIC(filterLine=fls["-"].pop())
                            tics["-"] = Bunch(tic=tic, times=times)

                        self.checkedLCMSFiles[fhash] = Bunch(
                            parsed=parsed,
                            fls=tm.getFilterLinesPerPolarity(),
                            pols=tm.getPolarities(),
                            tics=tics,
                        )

                        # Add to cache
                        new_row = pl.DataFrame({"filePath": [fhash], "parsingInfo": [base64.b64encode(dumps(self.checkedLCMSFiles[fhash])).decode("utf-8")]})

                        if cache_df is not None:
                            # Remove old entry if it exists and append new one
                            cache_df = cache_df.filter(pl.col("filePath") != fhash)
                            cache_df = pl.concat([cache_df, new_row])
                        else:
                            cache_df = new_row

                        # Write cache back to file
                        cache_df.write_parquet(cache_file)

                    except ExpatError as ex:
                        self.checkedLCMSFiles[fhash] = Bunch(parsed="Parsing error " + str(ex))
                        print(ex)
                    except Exception as ex:
                        self.checkedLCMSFiles[fhash] = Bunch(parsed="General error " + str(ex))

                        traceback.print_exc()

        return self.checkedLCMSFiles[fhash].parsed

    def updateLCMSSampleSettings(self):
        # fetch LC-HRMS data (polarity, filter-lines)
        self.ui.processedFilesComboBox.clear()
        self.ui.processedFilesComboBox.addItem("--", Bunch(file=None))
        self.ui.res_ExtractedData.clear()
        self.ui.chromPeakName.setText("")

        filterLines = set()
        polarities = set()

        # find out, how many files are to be inspected
        all = 0
        done = []
        indGroups = {}
        filesDict = {}

        for group in self.getAllSampleGroups():
            indGroups[group.name] = natSort(group.files)
            for file in natSort(group.files):
                if file not in done:
                    all += 1
                    done.append(file)

                if file not in filesDict.keys():
                    filesDict[file] = set()
                filesDict[file].add(group.name)

        multipleFiles = {}
        for file in filesDict.keys():
            if len(filesDict[file]) > 1:
                multipleFiles[file] = filesDict[file]

        if len(multipleFiles) > 0:
            fc = []
            for f in multipleFiles.keys():
                fc.append(f)
                for group in multipleFiles[f]:
                    fc.append("  - group %s" % group)

            QtWidgets.QMessageBox.warning(
                self,
                "MetExtract II",
                "Some files were used more than once. Please check if all groups have been imported / created correctly\n\nThe files are:\n%s" % ("\n".join(fc)),
                QtWidgets.QMessageBox.Ok,
            )

        pw = ProgressWrapper(1, parent=self, showIndProgress=True, indGroups=indGroups)
        pw.show()
        pw.getCallingFunction()("max")(all)

        i = 0
        done = []
        commonFilterLines = {}
        color = False
        failed = defaultdict(list)
        self.clearPlot(self.ui.pl_tic)

        # try to import each LC-HRMS file and get its polarities and filter lines (MS only)
        # draw the TICs of the individual filter lines
        for group in self.getAllSampleGroups():
            self.ui.processedFilesComboBox.addItem("%s" % group.name, userData=Bunch(file=None))

            for file in natSort(group.files):
                if file not in done:
                    pw.getCallingFunction()("text")("Importing group " + group.name + "\n" + file)
                    pw.getCallingFunction()("statuscolor")(file, "Orange")
                    pw.getCallingFunction()("statustext")(file, text="File: %s\nStatus: %s" % (file, "importing"))

                    f = file.replace("\\", "/")
                    f = f[f.rfind("/") + 1 : f.rfind(".")]

                    if not (re.match("^[a-zA-Z0-9_]*$", f)):
                        failed["Invalid characters"].append(file)
                        pw.getCallingFunction()("statuscolor")(file, "firebrick")
                        continue

                    if not (os.path.isfile(file)):
                        failed["File not found"].append(file)
                        pw.getCallingFunction()("statuscolor")(file, "firebrick")
                        continue

                    try:
                        if self.checkFileImport(file):
                            # fhash="%s_%s"%(file, sha256(open(file, 'rb').read()).hexdigest())
                            fhash = "%s_NOHash" % (file)  ## ignore hash, files are not likely to change
                            b = self.checkedLCMSFiles[fhash]

                            pols = b.pols
                            fls = b.fls

                            if "+" in pols:
                                self.drawPlot(
                                    self.ui.pl_tic,
                                    0,
                                    [t / 60.0 for t in b.tics["+"].times],
                                    b.tics["+"].tic,
                                    useCol=group.color,
                                    label=file,
                                    gid=file,
                                    addDC=True,
                                )
                            if "-" in pols:
                                mult = 1
                                if "+" in pols and "-" in pols:
                                    mult = -1
                                self.drawPlot(
                                    self.ui.pl_tic,
                                    0,
                                    [t / 60.0 for t in b.tics["-"].times],
                                    [e * mult for e in b.tics["-"].tic],
                                    useCol=group.color,
                                    label=file,
                                    gid=file,
                                    addDC=True,
                                )

                            for pol in pols:
                                if len(fls[pol]) > 0:
                                    if pol not in commonFilterLines:
                                        commonFilterLines[pol] = fls[pol]
                                    else:
                                        commonFilterLines[pol] = commonFilterLines[pol].intersection(fls[pol])

                            for pol in pols:
                                polarities.add(pol)

                            # if file has been processed successfully (FileName.meii DB exists) add it to the successfully processed list
                            if os.path.exists(file + getDBSuffix()) and os.path.isfile(file + getDBSuffix()):
                                fname = fname = file[(file.rfind("/") + 1) :]
                                qpixmap = QtGui.QPixmap(10, 10)
                                qpixmap.fill(QtGui.QColor(group.color))
                                icon = QtGui.QIcon(qpixmap)
                                self.ui.processedFilesComboBox.addItem(
                                    icon,
                                    "%s - %s" % (group.name, fname.replace(".mzXML", "")),
                                    userData=Bunch(file=file),
                                )

                            done.append(file)

                            pw.getCallingFunction()("statuscolor")(file, "olivedrab")
                            pw.getCallingFunction()("statustext")(file, text="File: %s\nStatus: %s" % (file, "imported"))
                        else:
                            pw.getCallingFunction()("statuscolor")(file, "firebrick")
                            pw.getCallingFunction()("statustext")(file, text="File: %s\nStatus: %s" % (file, "failed"))
                            failed["General error"].append(file)

                    except ExpatError:
                        pw.getCallingFunction()("statuscolor")(file, "firebrick")
                        pw.getCallingFunction()("statustext")(file, text="File: %s\nStatus: %s" % (file, "failed"))
                        failed["Parsing error"].append(file)
                    except Exception:
                        pw.getCallingFunction()("statuscolor")(file, "firebrick")
                        pw.getCallingFunction()("statustext")(file, text="File: %s\nStatus: %s" % (file, "failed"))
                        failed["General error"].append(file)

                    i += 1
                    pw.getCallingFunction()("value")(i)
            color = not color

            self.ui.processedFilesComboBox.addItem(" ", userData=Bunch(file=None))

        # update the TIC graph
        pw.hide()
        self.drawCanvas(self.ui.pl_tic, showLegendOverwrite=False)

        if len(failed) > 0:
            t = []
            for q in failed.keys():
                t.append(q)
                t.append("\n")
                for w in failed[q]:
                    t.append("  - ")
                    t.append(w)
                    t.append("\n")
            t = "".join(t)
            QtWidgets.QMessageBox.warning(
                self,
                "MetExtract",
                "%d failed to import correctly. Please resolve the following problems before continuing.\n\nFailed files:\n%s" % (len(failed), t),
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            )
            return False, failed

        if len(commonFilterLines) == 0:
            QtWidgets.QMessageBox.information(
                self,
                "MetExtract",
                "No common scan events among files. Please use files with at least one identical measurement method",
                QtWidgets.QMessageBox.Ok,
            )
            return False, []

        # Create list with positive mode scan events
        qslpos = []
        if "+" in commonFilterLines.keys():
            for fl in commonFilterLines["+"]:
                filterLines.add(fl)
                qslpos.append(fl)

        # Create list with negative mode scan events
        qslneg = []
        if "-" in commonFilterLines.keys():
            for fl in commonFilterLines["-"]:
                filterLines.add(fl)
                qslneg.append(fl)

        # Add empty scan event to both ionisation modes if positive and negative ionisation scans are available
        if (len(qslneg) == 0 and len(qslpos) == 0) or (len(qslneg) >= 1 and len(qslpos) >= 1):
            qslneg.insert(0, "None")
            qslpos.insert(0, "None")

        # Add empty scan event if only one ionisation mode was used to the opposite ionisation mode
        if len(qslneg) == 0 and len(qslpos) >= 1:
            qslneg.insert(0, "None")
        if len(qslpos) == 0 and len(qslneg) >= 1:
            qslpos.insert(0, "None")

        # Set scan event in user GUI
        self.ui.positiveScanEvent.clear()
        self.ui.positiveScanEvent.addItems(qslpos)
        self.ui.negativeScanEvent.clear()
        self.ui.negativeScanEvent.addItems(qslneg)

        # Set selected scan event. If both ionisation modes have scan events, select the first of both
        # If only one ionisation mode was used, set the first there and set the other ionisation mode to empty
        self.ui.positiveScanEvent.setCurrentIndex(0)
        self.ui.negativeScanEvent.setCurrentIndex(0)
        if len(qslpos) > 1 and len(qslneg) > 1:
            self.ui.positiveScanEvent.setCurrentIndex(1)
            self.ui.negativeScanEvent.setCurrentIndex(1)

        # Allow user to process files and view the results
        self.ui.pickingTab.setEnabled(len(filterLines) > 0)
        self.ui.resultsTab.setEnabled(self.ui.processedFilesComboBox.count() > 0)

        return True

    def smoothingWindowChanged(self):
        e = self.ui.eicSmoothingWindow.currentText() != "None"
        self.ui.eicSmoothingWindowSize.setVisible(e)
        self.ui.smoothingWindowSizeLabel.setVisible(e)

        e = self.ui.eicSmoothingWindow.currentText() == "SavitzkyGolay"
        self.ui.smoothingWindowPolynomLabel.setVisible(e)
        self.ui.smoothingPolynom_spinner.setVisible(e)

    def isoSearchChanged(self):
        self.ui.label_73.setVisible(self.ui.isoAbundance.checkState() == QtCore.Qt.Checked)
        self.ui.intensityThresholdIsotopologs.setVisible(self.ui.isoAbundance.checkState() == QtCore.Qt.Checked)
        # if self.ui.isoAbundance.checkState()==QtCore.Qt.Checked:
        #    self.ui.intensityThresholdIsotopologs.setValue(self.ui.intensityThreshold.value())

    # check if all imported LC-HRMS data files were processed with the same MetExtract settings
    def checkIndFilesSettings(self):
        done = []
        inValidfiles = ""
        unprocessed = 0
        i = 0
        settings = {}

        # check each imported LC-HRMS file individually
        for group in self.getAllSampleGroups():
            for file in natSort(group.files):
                if file not in done:
                    done.append(file)
                    if os.path.exists(file + getDBSuffix()) and os.path.isfile(file + getDBSuffix()):
                        db_con = None
                        try:
                            db_con = PolarsDB(file + getDBSuffix(), format=getDBFormat())

                            # fetch processing configuration (table config)
                            if "config" in db_con.tables:
                                config_df = db_con.tables["config"]
                                # create an entry for each processed setting
                                for row_dict in config_df.to_dicts():
                                    key = row_dict["key"]
                                    value = row_dict["value"]
                                    if key not in [
                                        "experimentOperator",
                                        "experimentID",
                                        "experimentComments",
                                        "experimentName",
                                        "processingUUID_ext",
                                    ]:
                                        if key not in settings:
                                            settings[key] = {}
                                        if value not in settings[key]:
                                            settings[key][value] = 0
                                        settings[key][value] += 1
                            if db_con is not None:
                                db_con.close()
                            db_con = None
                        except Exception:
                            try:
                                if db_con is not None:
                                    db_con.close()
                            except Exception:
                                logging.warning("Warning: Could not close intermediate (parquet) file")
                    else:
                        if len(inValidfiles) > 0:
                            inValidfiles += "\n"
                        inValidfiles = inValidfiles + file
                        unprocessed += 1
                    i += 1
        # if only one setting key-value pair was used for each LC-HRMS file, the processing settings of all LC-HRMS
        # data files are identical
        corruptParameters = []
        for sett in settings.keys():
            if len(settings[sett]) != 1:
                if sett not in [
                    "heteroAtoms",
                    "adducts",
                    "elements",
                    "tracerConfiguration",
                    "heteroElements",
                    "adducts",
                    "elementsForNL",
                ]:
                    corruptParameters.append("  * " + sett + ": " + ", ".join(settings[sett]))

        if not self.silent:
            if len(inValidfiles) > 0 and unprocessed != i:
                QtWidgets.QMessageBox.warning(
                    self,
                    "MetExtract",
                    "Not all files were processed\nUnprocessed files:\n%s" % inValidfiles,
                    QtWidgets.QMessageBox.Ok,
                )
        if not self.silent:
            if len(corruptParameters) > 0:
                QtWidgets.QMessageBox.warning(
                    self,
                    "MetExtract",
                    "Not all files were processed using the same parameters.\nIndividual files should be reprocessed\n\n%s" % ("\n".join(sorted(corruptParameters))),
                    QtWidgets.QMessageBox.Ok,
                )
            # self.updateIndividualFileProcessing = False
            # self.ui.processIndividualFiles.setChecked(True)
            # self.updateIndividualFileProcessing = True

    # </editor-fold>

    # <editor-fold desc="### add/modify/remove group functions">

    def getAllSampleGroups(self):
        """Return a natSort-sorted list of all SampleGroup objects from the groups table."""
        groups = []
        for row in range(self.ui.groupsList.rowCount()):
            grp = self._getGroupFromRow(row)
            if grp is not None:
                groups.append(grp)
        return natSort(groups, key=lambda x: str(x.name))

    def _getGroupFromRow(self, row):
        """Extract a SampleGroup from a table row."""
        tbl = self.ui.groupsList
        nameItem = tbl.item(row, 0)
        if nameItem is None:
            return None
        name = nameItem.text()
        filesItem = tbl.item(row, 1)
        files = filesItem.data(QtCore.Qt.UserRole) if filesItem else []

        colorWidget = tbl.cellWidget(row, 2)
        if colorWidget is not None:
            btn = colorWidget.findChild(QPushButton)
            color = btn.property("colorName") if btn else predefinedColors[0]
        else:
            color = predefinedColors[0]

        minFound = 1
        minFoundItem = tbl.item(row, 3)
        if minFoundItem:
            try:
                minFound = int(minFoundItem.text())
            except ValueError:
                minFound = 1

        omitWidget = tbl.cellWidget(row, 4)
        omitFeatures = omitWidget.findChild(QCheckBox).isChecked() if omitWidget else True

        metGrpWidget = tbl.cellWidget(row, 5)
        useForMetaboliteGrouping = metGrpWidget.findChild(QCheckBox).isChecked() if metGrpWidget else True

        fpWidget = tbl.cellWidget(row, 6)
        removeAsFalsePositive = fpWidget.findChild(QCheckBox).isChecked() if fpWidget else False

        # MSMS target selection is intentionally disabled in the UI.
        useAsMSMSTarget = False

        return SampleGroup(name, files, minFound, omitFeatures, useForMetaboliteGrouping, removeAsFalsePositive, color, useAsMSMSTarget)

    def _makeCenteredCheckbox(self, checked=False):
        """Return a centred checkbox widget for insertion in the table."""
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setAlignment(QtCore.Qt.AlignCenter)
        cb = QCheckBox()
        cb.setChecked(checked)
        layout.addWidget(cb)
        return container, cb

    def _propagateCheckboxChange(self, sourceCb, col):
        """When a checkbox changes, apply the same value to all other selected rows in the same column."""
        tbl = self.ui.groupsList
        sourceRow = -1
        for row in range(tbl.rowCount()):
            w = tbl.cellWidget(row, col)
            if w is not None:
                cb = w.findChild(QCheckBox)
                if cb is sourceCb:
                    sourceRow = row
                    break
        if sourceRow == -1:
            return
        selectedRows = {idx.row() for idx in tbl.selectedIndexes()}
        if sourceRow not in selectedRows or len(selectedRows) <= 1:
            return
        for row in selectedRows:
            if row == sourceRow:
                continue
            w = tbl.cellWidget(row, col)
            if w is not None:
                cb = w.findChild(QCheckBox)
                if cb is not None:
                    cb.blockSignals(True)
                    cb.setChecked(sourceCb.isChecked())
                    cb.blockSignals(False)
        self.grpFileEdited = True

    def _propagateComboChange(self, sourceCombo, col):
        """When a combobox changes, apply the same value to all other selected rows in the same column."""
        tbl = self.ui.groupsList
        sourceRow = -1
        for row in range(tbl.rowCount()):
            w = tbl.cellWidget(row, col)
            if w is sourceCombo:
                sourceRow = row
                break
        if sourceRow == -1:
            return
        selectedRows = {idx.row() for idx in tbl.selectedIndexes()}
        if sourceRow not in selectedRows or len(selectedRows) <= 1:
            return
        for row in selectedRows:
            if row == sourceRow:
                continue
            w = tbl.cellWidget(row, col)
            if w is not None and isinstance(w, QComboBox):
                w.blockSignals(True)
                idx = w.findText(sourceCombo.currentText())
                if idx >= 0:
                    w.setCurrentIndex(idx)
                w.blockSignals(False)
        self.grpFileEdited = True

    def _addGroupRow(self, row, name, files, minGrpFound, omitFeatures, useForMetaboliteGrouping, removeAsFalsePositive, color, useAsMSMSTarget):
        """Populate a single row in the groups table."""
        tbl = self.ui.groupsList
        tbl.blockSignals(True)

        # Col 0 – name (editable)
        nameItem = QTableWidgetItem(str(name))
        nameItem.setFlags(nameItem.flags() | QtCore.Qt.ItemIsEditable)
        tbl.setItem(row, 0, nameItem)

        # Col 1 – files  (non-editable display + UserRole data)
        filesItem = QTableWidgetItem("%d file(s)" % len(files))
        filesItem.setFlags(filesItem.flags() & ~QtCore.Qt.ItemIsEditable)
        filesItem.setData(QtCore.Qt.UserRole, list(files))
        filesItem.setToolTip("\n".join(files))
        tbl.setItem(row, 1, filesItem)

        # Col 2 – color picker button
        colorBtn = QPushButton()
        colorBtn.setProperty("colorName", color)
        self._styleColorButton(colorBtn, color)
        colorBtn.setToolTip("Click to choose a color")
        colorBtn.clicked.connect(lambda checked, btn=colorBtn, r=row: self._onColorButtonClicked(btn, r))
        container2 = QWidget()
        layout2 = QtWidgets.QHBoxLayout(container2)
        layout2.setContentsMargins(2, 0, 2, 0)
        layout2.addWidget(colorBtn)
        tbl.setCellWidget(row, 2, container2)

        # Col 3 – minFound (editable int)
        minFoundItem = QTableWidgetItem(str(minGrpFound))
        minFoundItem.setFlags(minFoundItem.flags() | QtCore.Qt.ItemIsEditable)
        tbl.setItem(row, 3, minFoundItem)

        # Col 4 – omitFeatures checkbox
        container4, cb4 = self._makeCenteredCheckbox(omitFeatures)
        cb4.stateChanged.connect(lambda state, cb=cb4: self._propagateCheckboxChange(cb, 4))
        tbl.setCellWidget(row, 4, container4)

        # Col 5 – useForMetaboliteGrouping
        container5, cb5 = self._makeCenteredCheckbox(useForMetaboliteGrouping)
        cb5.stateChanged.connect(lambda state, cb=cb5: self._propagateCheckboxChange(cb, 5))
        tbl.setCellWidget(row, 5, container5)

        # Col 6 – removeAsFalsePositive
        container6, cb6 = self._makeCenteredCheckbox(removeAsFalsePositive)
        cb6.stateChanged.connect(lambda state, cb=cb6: self._propagateCheckboxChange(cb, 6))
        tbl.setCellWidget(row, 6, container6)

        # Col 7 – useAsMSMSTarget (kept for backward compatibility, hidden in UI)
        container7, cb7 = self._makeCenteredCheckbox(False)
        cb7.stateChanged.connect(lambda state, cb=cb7: self._propagateCheckboxChange(cb, 7))
        tbl.setCellWidget(row, 7, container7)

        tbl.blockSignals(False)

    def _styleColorButton(self, btn, color_name):
        """Style a QPushButton as a solid color swatch."""
        qc = QtGui.QColor(color_name)
        if not qc.isValid():
            qc = QtGui.QColor("gray")
        luminance = 0.299 * qc.red() + 0.587 * qc.green() + 0.114 * qc.blue()
        fg = "black" if luminance > 128 else "white"
        btn.setText(color_name)
        btn.setStyleSheet("background-color: %s; color: %s; border: 1px solid gray; padding: 2px 8px;" % (qc.name(), fg))

    def _onColorButtonClicked(self, btn, row):
        """Open QColorDialog and apply the chosen colour to the button and all selected rows."""
        current = QtGui.QColor(btn.property("colorName"))
        chosen = QtWidgets.QColorDialog.getColor(current, self, "Choose group color")
        if not chosen.isValid():
            return
        color_name = chosen.name()
        btn.setProperty("colorName", color_name)
        self._styleColorButton(btn, color_name)
        # Also apply to all selected rows
        tbl = self.ui.groupsList
        selectedRows = {idx.row() for idx in tbl.selectedIndexes()}
        selectedRows.add(row)
        for r in selectedRows:
            w = tbl.cellWidget(r, 2)
            if w is not None:
                b = w.findChild(QPushButton)
                if b is not None and b is not btn:
                    b.setProperty("colorName", color_name)
                    self._styleColorButton(b, color_name)
        self.grpFileEdited = True

    def addGroup(
        self,
        name,
        files,
        minGrpFound,
        omitFeatures,
        useForMetaboliteGrouping,
        removeAsFalsePositive,
        color,
        atPos=None,
        useAsMSMSTarget=False,
    ):
        useAsMSMSTarget = False
        self.loadedMZXMLs = None

        failed = defaultdict(list)

        pw = ProgressWrapper(1, parent=self, showIndProgress=True, indGroups={"files": files})
        pw.show()
        pw.getCallingFunction()("max")(len(files))
        pw.getCallingFunction()("text")("Checking group %s" % (name))
        i = 0

        for f in files:
            pw.getCallingFunction()("text")("Checking group %s\nFile: %s" % (name, f))
            x = self.checkFileImport(f)
            if x:
                pw.getCallingFunction()("statuscolor")(f, "olivedrab")
                pw.getCallingFunction()("statustext")(f, text="File: %s\nStatus: %s" % (f, "imported"))
            else:
                failed[x].append(f)
                pw.getCallingFunction()("statuscolor")(f, "Firebrick")
                pw.getCallingFunction()("statustext")(f, text="File: %s\nStatus: %s" % (f, x))
            i += 1
            pw.getCallingFunction()("value")(i)
        pw.hide()

        if len(failed) == 0:
            tbl = self.ui.groupsList
            if atPos is None:
                atPos = tbl.rowCount()
            tbl.insertRow(atPos)
            self._addGroupRow(atPos, name, files, minGrpFound, omitFeatures, useForMetaboliteGrouping, removeAsFalsePositive, color, useAsMSMSTarget)
        else:
            t = []
            for q in failed.keys():
                t.append(q)
                t.append("\n")
                for w in failed[q]:
                    t.append("  - ")
                    t.append(w)
                    t.append("\n")
            t = "".join(t)
            QtWidgets.QMessageBox.warning(
                self,
                "MetExtract",
                "%d files failed to import correctly. Group '%s' will not be created.\nPlease resolve the following issues (see documentation for further help)\n\n%s" % (len(failed), name, t),
                QtWidgets.QMessageBox.Ok,
            )

    def showAddGroupDialogClicked(self):
        self.showAddGroupDialog()

    def showFileStatsPopup(self):
        """Collect all sample files from the groups table, compute stats, and show them in a popup dialog."""

        # Build ordered list of (filepath, group_name, group_color) from all groups
        all_files_with_groups = []
        for grp in self.getAllSampleGroups():
            for f in natSort(grp.files):
                all_files_with_groups.append((str(f), str(grp.name), str(grp.color)))

        if not all_files_with_groups:
            QtWidgets.QMessageBox.information(self, "File Stats", "No sample files configured in the groups table.")
            return

        pos_se = str(self.ui.positiveScanEvent.currentText()) if self.ui.positiveScanEvent.count() > 0 else None
        neg_se = str(self.ui.negativeScanEvent.currentText()) if self.ui.negativeScanEvent.count() > 0 else None

        pw = QtWidgets.QProgressDialog("Computing file stats...", "Cancel", 0, len(all_files_with_groups), self)
        pw.setWindowTitle("File Stats")
        pw.setWindowModality(QtCore.Qt.WindowModal)
        # Show "current / total  (percent%)" inside the bar
        bar = pw.findChild(QtWidgets.QProgressBar)
        if bar is not None:
            bar.setFormat("%v / %m  (%p%)")
        pw.setValue(0)
        pw.show()

        stats_rows = []
        for i, (f, group_name, group_color) in enumerate(all_files_with_groups):
            pw.setLabelText(f"Processing: {os.path.basename(f)}")
            pw.setValue(i)
            QtWidgets.QApplication.processEvents()
            if pw.wasCanceled():
                return
            rows = compute_sample_stats([f], pos_se, neg_se)
            for row in rows:
                row["_group_name"] = group_name
                row["_group_color"] = group_color
            stats_rows.extend(rows)
        pw.setValue(len(all_files_with_groups))
        pw.close()

        if not stats_rows:
            return

        # Columns to display (exclude internal _ keys)
        data_columns = [k for k in stats_rows[0].keys() if not k.startswith("_")]
        column_labels = {
            "file": "File",
            "startTimeStamp": "Start Timestamp",
            "ms1_pos": "MS1 Pos",
            "ms1_neg": "MS1 Neg",
            "ms2_pos": "MS2 Pos",
            "ms2_neg": "MS2 Neg",
            "last_rt_min": "Last RT (min)",
            "ms1_dt_min": "MS1 dT Min (s)",
            "ms1_dt_p10": "MS1 dT 10%",
            "ms1_dt_p25": "MS1 dT 25%",
            "ms1_dt_median": "MS1 dT Median",
            "ms1_dt_mean": "MS1 dT Mean",
            "ms1_dt_p75": "MS1 dT 75%",
            "ms1_dt_p90": "MS1 dT 90%",
            "ms1_dt_max": "MS1 dT Max",
            "ms1_dt_sd": "MS1 dT SD",
            "ms2_dt_min": "MS2 dT Min (s)",
            "ms2_dt_p10": "MS2 dT 10%",
            "ms2_dt_p25": "MS2 dT 25%",
            "ms2_dt_median": "MS2 dT Median",
            "ms2_dt_mean": "MS2 dT Mean",
            "ms2_dt_p75": "MS2 dT 75%",
            "ms2_dt_p90": "MS2 dT 90%",
            "ms2_dt_max": "MS2 dT Max",
            "ms2_dt_sd": "MS2 dT SD",
            "ms1_signalInt_pos_min": "MS1 log10(int) Pos Min",
            "ms1_signalInt_pos_p10": "MS1 log10(int) Pos 10%",
            "ms1_signalInt_pos_p25": "MS1 log10(int) Pos 25%",
            "ms1_signalInt_pos_median": "MS1 log10(int) Pos Median",
            "ms1_signalInt_pos_p75": "MS1 log10(int) Pos 75%",
            "ms1_signalInt_pos_p90": "MS1 log10(int) Pos 90%",
            "ms1_signalInt_pos_p91": "MS1 log10(int) Pos 91%",
            "ms1_signalInt_pos_p92": "MS1 log10(int) Pos 92%",
            "ms1_signalInt_pos_p93": "MS1 log10(int) Pos 93%",
            "ms1_signalInt_pos_p94": "MS1 log10(int) Pos 94%",
            "ms1_signalInt_pos_p95": "MS1 log10(int) Pos 95%",
            "ms1_signalInt_pos_p96": "MS1 log10(int) Pos 96%",
            "ms1_signalInt_pos_p97": "MS1 log10(int) Pos 97%",
            "ms1_signalInt_pos_p98": "MS1 log10(int) Pos 98%",
            "ms1_signalInt_pos_p99": "MS1 log10(int) Pos 99%",
            "ms1_signalInt_pos_max": "MS1 log10(int) Pos Max",
            "ms1_signalInt_neg_min": "MS1 log10(int) Neg Min",
            "ms1_signalInt_neg_p10": "MS1 log10(int) Neg 10%",
            "ms1_signalInt_neg_p25": "MS1 log10(int) Neg 25%",
            "ms1_signalInt_neg_median": "MS1 log10(int) Neg Median",
            "ms1_signalInt_neg_p75": "MS1 log10(int) Neg 75%",
            "ms1_signalInt_neg_p90": "MS1 log10(int) Neg 90%",
            "ms1_signalInt_neg_p91": "MS1 log10(int) Neg 91%",
            "ms1_signalInt_neg_p92": "MS1 log10(int) Neg 92%",
            "ms1_signalInt_neg_p93": "MS1 log10(int) Neg 93%",
            "ms1_signalInt_neg_p94": "MS1 log10(int) Neg 94%",
            "ms1_signalInt_neg_p95": "MS1 log10(int) Neg 95%",
            "ms1_signalInt_neg_p96": "MS1 log10(int) Neg 96%",
            "ms1_signalInt_neg_p97": "MS1 log10(int) Neg 97%",
            "ms1_signalInt_neg_p98": "MS1 log10(int) Neg 98%",
            "ms1_signalInt_neg_p99": "MS1 log10(int) Neg 99%",
            "ms1_signalInt_neg_max": "MS1 log10(int) Neg Max",
            "MS:1000073": "MS:1000073",
            "MS:1000079": "MS:1000079",
        }

        # "Group" is the first column; the rest follow
        all_columns = ["_group"] + data_columns
        all_headers = ["Group"] + [column_labels.get(c, c) for c in data_columns]

        # Compute per-column means for numeric columns (for deviation colour-coding)
        numeric_cols = set()
        col_values = {c: [] for c in data_columns}
        for row in stats_rows:
            for c in data_columns:
                v = row.get(c)
                if v is not None:
                    try:
                        col_values[c].append(float(v))
                        numeric_cols.add(c)
                    except (ValueError, TypeError):
                        pass
        col_means = {}
        for c in numeric_cols:
            vals = col_values[c]
            if vals:
                col_means[c] = sum(vals) / len(vals)

        def _deviation_color(col, val_str):
            """Return a QColor based on fractional deviation from the column mean, or None."""
            if col not in numeric_cols or col not in col_means:
                return None
            try:
                val = float(val_str)
            except (ValueError, TypeError):
                return None
            mean_v = col_means[col]
            if mean_v == 0:
                return None
            dev = abs(val - mean_v) / abs(mean_v)
            if dev >= 0.10:
                c = QtGui.QColor(178, 34, 34)  # firebrick
                c.setAlpha(200)
                return c
            # Gradient: alpha 0 → 200 linearly from dev=0 to dev=0.10
            alpha = int(dev / 0.10 * 200)
            c = QtGui.QColor(178, 34, 34)
            c.setAlpha(alpha)
            return c

        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("File Stats")
        dialog.resize(1400, 600)
        layout = QtWidgets.QVBoxLayout(dialog)

        table = QtWidgets.QTableWidget(len(stats_rows), len(all_columns))
        table.setHorizontalHeaderLabels(all_headers)
        table.horizontalHeader().setStretchLastSection(False)
        table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        table.setAlternatingRowColors(False)
        table.setSortingEnabled(True)
        table.setSortingEnabled(False)  # disable while populating

        for r, row_data in enumerate(stats_rows):
            gname = row_data.get("_group_name", "")
            gcolor_str = row_data.get("_group_color", "")
            gcolor = QtGui.QColor(gcolor_str)
            if gcolor.isValid():
                gcolor.setAlpha(80)
            else:
                gcolor = None

            # Group column
            grp_item = _NumericDateSortItem(gname)
            if gcolor:
                grp_item.setBackground(gcolor)
            table.setItem(r, 0, grp_item)

            # Data columns
            for c_idx, col in enumerate(data_columns, start=1):
                val = row_data.get(col)
                if val is None:
                    text = ""
                elif isinstance(val, float):
                    text = f"{val:.4f}"
                else:
                    text = str(val)
                item = _NumericDateSortItem(text)
                dev_color = _deviation_color(col, text)
                if dev_color is not None:
                    item.setBackground(dev_color)
                table.setItem(r, c_idx, item)

        table.setSortingEnabled(True)
        table.horizontalHeader().setSectionsMovable(True)
        table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        table.resizeColumnsToContents()
        layout.addWidget(table)

        btn_close = QtWidgets.QPushButton("Close")
        btn_close.clicked.connect(dialog.accept)
        layout.addWidget(btn_close)
        dialog.exec()

    # show an import group dialog to the user
    def showAddGroupDialog(self, initWithFiles=[], initWithGroupName=""):
        tdiag = groupEdit(self, self.lastOpenDir, colors=predefinedColors)
        if (
            tdiag.executeDialog(
                groupName=initWithGroupName,
                groupfiles=initWithFiles,
                activeColor=predefinedColors[self.ui.groupsList.rowCount() % len(predefinedColors)],
            )
            == QtWidgets.QDialog.Accepted
        ):
            self.lastOpenDir = tdiag.getOpenDir()

            failed = defaultdict(list)
            pw = ProgressWrapper(
                1,
                parent=self,
                showIndProgress=True,
                indGroups={tdiag.getGroupName(): tdiag.getGroupFiles()},
            )
            pw.show()
            pw.getCallingFunction()("max")(len(tdiag.getGroupFiles()))
            pw.getCallingFunction()("value")(0)

            i = 0
            for f in tdiag.getGroupFiles():
                pw.getCallingFunction()("value")(i)
                pw.getCallingFunction()("text")(f)
                pw.getCallingFunction()("statuscolor")(f, "orange")
                pw.getCallingFunction()("statustext")(f, text="File: %s\nStatus: %s" % (f, "importing"))

                i += 1

                x = self.checkFileImport(f)

                if x:
                    pw.getCallingFunction()("statuscolor")(f, "olivedrab")
                    pw.getCallingFunction()("statustext")(f, text="File: %s\nStatus: %s" % (f, "imported"))
                else:
                    failed[x].append(f)
                    pw.getCallingFunction()("statuscolor")(f, "firebrick")
                    pw.getCallingFunction()("statustext")(f, text="File: %s\nStatus: %s" % (f, "failed to import"))

            pw.hide()

            if len(failed) > 0:
                t = []
                for q in failed.keys():
                    t.append(q)
                    t.append("\n")
                    for w in failed[q]:
                        t.append("  - ")
                        t.append(w)
                        t.append("\n")
                t = "".join(t)
                QtWidgets.QMessageBox.warning(
                    self,
                    "MetExtract",
                    "%d files failed to import correctly. Group will not be created.\nPlease resolve the following issues (see documentation for further help)\n\n%s" % (len(failed), t),
                    QtWidgets.QMessageBox.Ok,
                )
                return None

            self.addGroup(
                name=tdiag.getGroupName(),
                files=tdiag.getGroupFiles(),
                minGrpFound=tdiag.getMinimumGroupFound(),
                omitFeatures=tdiag.getOmitFeatures(),
                useForMetaboliteGrouping=tdiag.getUseForMetaboliteGrouping(),
                removeAsFalsePositive=tdiag.getRemoveAsFalsePositive(),
                color=str(tdiag.getGroupColor()),
            )

            if self.updateLCMSSampleSettings():
                self.grpFileEdited = True

    # remove selected groups from the input
    def remGrp(self):
        self.loadedMZXMLs = None

        selectedRows = sorted({idx.row() for idx in self.ui.groupsList.selectedIndexes()}, reverse=True)
        for row in selectedRows:
            self.ui.groupsList.removeRow(row)

        self.updateLCMSSampleSettings()
        self.grpFileEdited = True

    # Double-click on files column opens group-edit dialog; name/minFound remain inline-editable
    def editGroup(self, index):
        if index.column() != 1:
            return
        self.loadedMZXMLs = None
        row = index.row()
        grp = self._getGroupFromRow(row)
        if grp is None:
            return
        t = groupEdit(colors=predefinedColors)
        if (
            t.executeDialog(
                groupName=grp.name,
                groupfiles=grp.files,
                minimumGroupFound=grp.minFound,
                omitFeatures=grp.omitFeatures,
                useForMetaboliteGrouping=grp.useForMetaboliteGrouping,
                removeAsFalsePositive=grp.removeAsFalsePositive,
                activeColor=grp.color,
                useAsMSMSTarget=grp.useAsMSMSTarget,
            )
            == QtWidgets.QDialog.Accepted
        ):
            self.ui.groupsList.removeRow(row)
            self.addGroup(
                name=t.getGroupName(),
                files=t.getGroupFiles(),
                minGrpFound=t.getMinimumGroupFound(),
                omitFeatures=t.getOmitFeatures(),
                useForMetaboliteGrouping=t.getUseForMetaboliteGrouping(),
                removeAsFalsePositive=t.getRemoveAsFalsePositive(),
                color=str(t.getGroupColor()),
                useAsMSMSTarget=False,
                atPos=row,
            )
            self.updateLCMSSampleSettings()
            self.grpFileEdited = True

    def _onGroupTableItemChanged(self, item):
        """Propagate text-cell edits (name, minFound) to all other selected rows in the same column."""
        col = item.column()
        if col not in (0, 3):
            return
        tbl = self.ui.groupsList
        row = item.row()
        selectedRows = {idx.row() for idx in tbl.selectedIndexes()}
        if row not in selectedRows or len(selectedRows) <= 1:
            self.grpFileEdited = True
            return
        tbl.blockSignals(True)
        for r in selectedRows:
            if r == row:
                continue
            existingItem = tbl.item(r, col)
            if existingItem is not None:
                existingItem.setText(item.text())
        tbl.blockSignals(False)
        self.grpFileEdited = True

    # </editor-fold>

    # <editor-fold desc="### load/save group compilation">
    def saveGroupsClicked(self):
        groupFile = QtWidgets.QFileDialog.getSaveFileName(
            caption="Select group file",
            dir=self.lastOpenDir,
            filter="Group file (*.grp);;All files(*.*)",
        )[0]
        groupFile = str(groupFile).replace("\\", "/")

        doAsk = True
        text = ""
        addWarning = False

        while doAsk:
            text, ok = QtWidgets.QInputDialog.getText(
                self,
                "Custom evaluation name",
                "%sDo you want to specify a custom evaluation name for this experiment?<br>"
                "This lets you test different parameter settings without overwriting the results of other processings using the same input files<br>"
                "For each evaluation with a provided name, a subfolder will be created where all results will be saved (input files will be duplicated and specified groups will be changed)<br>"
                "If you want to create or use a custom evaluation name enter its name in the text box below<br>"
                "If you don't want to use a custom evaluation name leave the line empty or cancel this dialog<br>"
                "Attention: Existing files will be overwritten and the results groups and data matrix files will be changed to this directory!"
                % ("<b>Error: Specified experiment evaluation (%s) may already be in use. Verify or use a different evaluation name<br><br></b>" % (text) if addWarning else ""),
            )
            if ok and text != "":
                text = "EVAL_" + str(text)

                ## verify directory is empty
                if os.path.exists(str(groupFile[: groupFile.rfind("/")] + "/" + text)):
                    addWarning = True
                else:
                    ## make directory
                    os.mkdir(str(groupFile[: groupFile.rfind("/")] + "/" + text))
                    os.mkdir(str(groupFile[: groupFile.rfind("/")] + "/" + text + "/data"))

                    # groupFile=groupFile[:groupFile.rfind("/")]+"/"+text+groupFile[groupFile.rfind("/"):]
                    ## copy files
                    totalCopyFiles = 0
                    for group in self.getAllSampleGroups():
                        for i in range(len(group.files)):
                            totalCopyFiles += 1

                    pw = ProgressWrapper()
                    pw.setMax(totalCopyFiles)
                    pw.setValue(0)
                    pw.show()

                    done = 0
                    for group in self.getAllSampleGroups():
                        os.mkdir(str(groupFile[: groupFile.rfind("/")] + "/" + text + "/data/" + group.name))
                        for i in range(len(group.files)):
                            pw.setTextu("Copying " + group.files[i][group.files[i].rfind("/") :])
                            shutil.copy(
                                str(group.files[i]),
                                str(groupFile[: groupFile.rfind("/")] + "/" + text + "/data/" + group.name + group.files[i][group.files[i].rfind("/") :]),
                            )
                            group.files[i] = str(groupFile[: groupFile.rfind("/")] + "/" + text + "/data/" + group.name + group.files[i][group.files[i].rfind("/") :])
                            done = done + 1
                            pw.setValueu(done)

                    try:
                        pw.setTextu("Zipping current MetExtract II application for documentation")
                        import zipfile

                        def zipdir(path, zip):
                            for root, dirs, files in os.walk(path):
                                for file in files:
                                    zip.write(os.path.join(root, file))

                        zipF = zipfile.ZipFile(
                            str(groupFile[: groupFile.rfind("/")] + "/" + text + "/MetExtractII_application.zip"),
                            "w",
                        )
                        zipdir(get_main_dir(), zipF)

                        zipF.close()

                    except Exception as ex:
                        print(ex)
                        logging.warning("Could not zip MetExtract II application for documentation")

                    pw.close()

                    ## rename settings
                    groupFile = str(groupFile[: groupFile.rfind("/")] + "/" + text + groupFile[groupFile.rfind("/") :])
                    self.ui.groupsSave.setText(str(groupFile[: groupFile.rfind("/")] + "/results.xlsx"))
                    self.ui.msms_fileLocation.setText(str(groupFile[: groupFile.rfind("/")] + "/MSMStargets.tsv"))

                    doAsk = False
                    logging.info("New experiment evaluation. Name: %s" % text)
            else:
                doAsk = False

        if groupFile is not None and len(groupFile) > 0:
            self.lastOpenDir = str(groupFile).replace("\\", "/")
            self.lastOpenDir = self.lastOpenDir[: self.lastOpenDir.rfind("/")]

            self.saveGroups(groupFile, saveSettings=True)

    # save group compilation and settings
    def saveGroups(self, groupFile=None, saveSettings=False):
        if groupFile is not None and len(groupFile) > 0:
            print(f"saving groups and settings to {groupFile}")
            grps = QtCore.QSettings(groupFile, QtCore.QSettings.IniFormat)
            grps.clear()

            grps.beginGroup("ExperimentDescription")
            grps.setValue("ExperimentName", self.ui.exExperimentName_LineEdit.text())
            grps.setValue("Operator", self.ui.exOperator_LineEdit.text())
            grps.setValue("ExperimentID", self.ui.exExperimentID_LineEdit.text())
            grps.setValue("ExperimentComments", self.ui.exComments_TextEdit.toPlainText())
            grps.endGroup()

            for group in self.getAllSampleGroups():
                grps.beginGroup(group.name)
                # for i in range(len(group.files)):
                #    try:
                #        relFilePath = "./" + str(os.path.relpath(group.files[i], os.path.split(str(groupFile))[0]).replace("\\", "/"))
                #    except ValueError:
                #        # Files are on different drives, use absolute path
                #        relFilePath = str(group.files[i]).replace("\\", "/")
                #    grps.setValue(group.name + "__" + str(i), relFilePath)
                grps.setValue("files", "§".join(group.files))
                grps.setValue("Min_Peaks_Found", group.minFound)
                grps.setValue("OmitFeatures", group.omitFeatures)
                grps.setValue("RemoveAsFalsePositive", group.removeAsFalsePositive)
                grps.setValue("Color", group.color)
                grps.setValue("UseForMetaboliteGrouping", group.useForMetaboliteGrouping)
                grps.setValue("useAsMSMSTarget", group.useAsMSMSTarget)
                grps.endGroup()

            grps.beginGroup("ExperimentResults")

            resFile = str(self.ui.groupsSave.text().replace("\\", "/"))
            if resFile.startswith("./"):
                pat = os.path.split(str(groupFile))[0] + "/" + resFile
                grps.setValue("GroupSaveFile", resFile)
                self.ui.groupsSave.setText(pat)
            else:
                grps.setValue("GroupSaveFile", resFile)
                self.ui.groupsSave.setText(resFile)

            grps.setValue(
                "msms_fileLocation",
                str(self.ui.msms_fileLocation.text()).replace("\\", "/"),
            )

            grps.endGroup()

            # ask user, if currently loaded settings shall also be saved to this compilation
            if (
                saveSettings
                or QtWidgets.QMessageBox.question(
                    self,
                    "MetExtract",
                    "Do you want to save the current settings with this group?",
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                )
                == QtWidgets.QMessageBox.Yes
            ):
                self.saveSettingsFile(groupFile, clear=False)

    # Helper methods for PySide6 compatibility with QSettings values
    def to_bool(self, value):
        """Convert QSettings value to boolean (PySide6 compatible)"""
        if value is None:
            return False
        if hasattr(value, "toBool"):
            return value.toBool()
        elif isinstance(value, str):
            return value.lower() in ("true", "1", "yes", "on")
        else:
            return bool(value)

    def to_int(self, value):
        """Convert QSettings value to int (PySide6 compatible)"""
        if value is None:
            return 0
        if hasattr(value, "toInt"):
            return value.toInt()[0]
        else:
            return int(float(value))

    def to_double(self, value):
        """Convert QSettings value to float (PySide6 compatible)"""
        if value is None:
            return 0.0
        if hasattr(value, "toDouble"):
            return value.toDouble()[0]
        else:
            return float(value)

    def to_str(self, value):
        """Convert QSettings value to str (PySide6 compatible)"""
        if value is None:
            return ""
        if hasattr(value, "toString"):
            return value.toString()
        else:
            return str(value)

    def loadGroupsClicked(self):
        groupFile = QtWidgets.QFileDialog.getOpenFileName(
            caption="Select group file",
            dir=self.lastOpenDir,
            filter="Group file (*.grp);;All files(*.*)",
        )

        if groupFile is not None and len(groupFile) > 0:
            self.lastOpenDir = str(groupFile).replace("\\", "/")
            self.lastOpenDir = self.lastOpenDir[: self.lastOpenDir.rfind("/")]

            self.loadGroups(groupFile)

    def loadGroups(self, groupFile=None, forceLoadSettings=False, askLoadSettings=True):
        if groupFile is not None and len(groupFile) > 0:
            if len(groupFile) > 0:
                self.lastOpenDir = str(groupFile).replace("\\", "/")
                self.lastOpenDir = self.lastOpenDir[: self.lastOpenDir.rfind("/")]

            grps = QtCore.QSettings(groupFile, QtCore.QSettings.IniFormat)

            grps.beginGroup("ExperimentDescription")

            if grps.contains("ExperimentName"):
                self.ui.exExperimentName_LineEdit.setText(str(grps.value("ExperimentName")))
            if grps.contains("Operator"):
                self.ui.exOperator_LineEdit.setText(str(grps.value("Operator")))
            if grps.contains("ExperimentID"):
                self.ui.exExperimentID_LineEdit.setText(str(grps.value("ExperimentID")))
            if grps.contains("ExperimentComments"):
                self.ui.exComments_TextEdit.setPlainText(str(grps.value("ExperimentComments")))

            grps.endGroup()

            procGrps = []
            for grp in grps.childGroups():
                if str(grp) != "Settings" and str(grp) != "ExperimentDescription" and str(grp) != "ExperimentResults" and str(grp) != "MetExtract" and str(grp) != "WorkingDirectory":
                    procGrps.append(grp)

            grpi = 0
            groupsToAdd = []
            for grp in natSort(procGrps):
                grps.beginGroup(grp)
                kids = []
                minFound = 1
                color = predefinedColors[grpi % len(predefinedColors)]
                omitFeatures = True
                useForMetaboliteGrouping = True
                removeAsFalsePositive = False
                useAsMSMSTarget = False
                for kid in grps.childKeys():
                    if str(kid) == "Min_Peaks_Found":
                        minFound = self.to_int(grps.value(kid))
                    elif str(kid) == "OmitFeatures":
                        omitFeatures = self.to_bool(grps.value(kid))
                    elif str(kid) == "Color":
                        color = str(grps.value(kid))
                    elif str(kid) == "UseForMetaboliteGrouping":
                        useForMetaboliteGrouping = self.to_bool(grps.value(kid))
                    elif str(kid) == "RemoveAsFalsePositive":
                        removeAsFalsePositive = self.to_bool(grps.value(kid))
                    elif str(kid) == "useAsMSMSTarget":
                        useAsMSMSTarget = self.to_bool(grps.value(kid))
                    elif str(kid) == "files":
                        files = self.to_str(grps.value(kid))
                        files = files.strip().split("§")
                        files = [f.replace("\\", "/") for f in files]
                        for file_i in range(len(files)):
                            file = files[file_i]
                            if os.path.isabs(file):
                                kids.append(file)
                            else:
                                kids.append(os.path.split(str(groupFile))[0].replace("\\", "/") + "/" + file)
                    elif str(kid).startswith(grp):
                        if os.path.isabs(str(grps.value(kid)).replace("\\", "/")):
                            kids.append(str(grps.value(kid)).replace("\\", "/"))
                        else:
                            kids.append(os.path.split(str(groupFile))[0].replace("\\", "/") + "/" + str(grps.value(kid)).replace("\\", "/"))

                kids = natSort(kids)
                groupsToAdd.append(
                    Bunch(
                        name=grp,
                        files=kids,
                        minGrpFound=minFound,
                        omitFeatures=omitFeatures,
                        useForMetaboliteGrouping=useForMetaboliteGrouping,
                        removeAsFalsePositive=removeAsFalsePositive,
                        color=color,
                        useAsMSMSTarget=useAsMSMSTarget,
                    )
                )

                grps.endGroup()
                grpi += 1

            for grpToAdd in natSort(groupsToAdd, key=lambda x: x.name):
                self.addGroup(
                    name=grpToAdd.name,
                    files=grpToAdd.files,
                    minGrpFound=grpToAdd.minGrpFound,
                    omitFeatures=grpToAdd.omitFeatures,
                    useForMetaboliteGrouping=grpToAdd.useForMetaboliteGrouping,
                    removeAsFalsePositive=grpToAdd.removeAsFalsePositive,
                    color=grpToAdd.color,
                    useAsMSMSTarget=grpToAdd.useAsMSMSTarget,
                )

            grps.beginGroup("ExperimentResults")
            if grps.contains("GroupSaveFile"):
                resFile = str(grps.value("GroupSaveFile"))

                if resFile.startswith("./"):
                    resFile = os.path.split(str(groupFile))[0] + "/" + resFile

                self.ui.groupsSave.setText(resFile)
                self.ui.msms_fileLocation.setText(grps.value("msms_fileLocation"))
                self.loadGroupsResultsFile(resFile)

            grps.endGroup()

            self.updateLCMSSampleSettings()

            if forceLoadSettings or (
                askLoadSettings
                and QtWidgets.QMessageBox.question(
                    self,
                    "MetExtract",
                    "Do you want to load the associated settings with this group?",
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                )
                == QtWidgets.QMessageBox.Yes
            ):
                self.loadSettingsFile(str(groupFile))
            self.grpFile = str(groupFile)
            self.grpFileEdited = False

    # </editor-fold>

    # <editor-fold desc="### group results visualisation functions">
    # EXPERIMENTAL
    def loadGroupsResultsFile(self, groupsResFile):
        experimentalGroups = self.getAllSampleGroups()

        try:
            self.ui.resultsExperiment_TreeWidget.clear()
            self.ui.resultsExperiment_TreeWidget.setHeaderLabels(["OGroup", "MZ", "RT", "Xn", "Z", "IonMode", "MS2 N/L"])

            if os.path.exists(groupsResFile) and os.path.isfile(groupsResFile):
                widths = [160, 80, 60, 30, 30, 30, 60]
                for i in range(len(widths)):
                    self.ui.resultsExperiment_TreeWidget.setColumnWidth(i, widths[i])

                self.experimentResults = Bunch(db_con=None, file=None)

                self.experimentResults.db_con = PolarsDB(groupsResFile, format="xlsx")

                # show a dialog with a drop-down list asking the user to specify the table to load
                options = self.experimentResults.db_con.list_tables()
                options = [opt for opt in options if opt not in ["Parameters", "__dTypes__", "2_StatColumns_FalsePositives", "2_StatColumns_Omitted", "4_Convoluted_doublePeaks", "5_Annotated_Compounds", "5_Annotated_SumFormulas", "0_sampleStats", "DB_info"]][::-1]

                mgsBox = QtWidgets.QMessageBox(self)
                mgsBox.setWindowTitle("Select results to load")
                mgsBox.setText("Multiple result tables found in the selected file. Please select the table to load:")
                combo = QtWidgets.QComboBox(mgsBox)
                combo.addItems(options)
                mgsBox.layout().addWidget(combo)
                mgsBox.setStandardButtons(QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel)
                if mgsBox.exec() == QtWidgets.QMessageBox.Ok:
                    selected_table = combo.currentText()
                else:
                    return

                self.experimentResults.selected_table = selected_table

                metaboliteGroupTreeItems = {}
                # Get distinct OGroup values from GroupResults, ordered by rt
                group_results_df = self.experimentResults.db_con.tables[selected_table]
                distinct_groups = group_results_df.group_by("OGroup").agg(RT=pl.col("RT").mean()).sort("RT")

                for row_dict in distinct_groups.to_dicts():
                    metaboliteGroup = Bunch(type="metaboliteGroup", metaboliteGroupID=row_dict["OGroup"])
                    metaboliteGroupTreeItem = QtWidgets.QTreeWidgetItem(["%s" % metaboliteGroup.metaboliteGroupID])
                    metaboliteGroupTreeItem.bunchData = metaboliteGroup
                    self.ui.resultsExperiment_TreeWidget.addTopLevelItem(metaboliteGroupTreeItem)
                    metaboliteGroupTreeItems[metaboliteGroup.metaboliteGroupID] = metaboliteGroupTreeItem
                kids = []

                if False:
                    fileMappingData = {}
                    # Get file results from FoundFeaturePairs, ordered by areaN DESC
                    found_fps_df = self.experimentResults.db_con.tables["FoundFeaturePairs"].sort("areaN", descending=True)

                    for row_dict in found_fps_df.to_dicts():
                        fileRes = Bunch(type="fileResult", resID=row_dict["resID"], file=row_dict["file"], areaN=row_dict["areaN"], areaL=row_dict["areaL"])
                        if fileRes.resID not in fileMappingData.keys():
                            fileMappingData[fileRes.resID] = []

                        fileMappingData[fileRes.resID].append(fileRes)

                    # Get feature pairs with count of found files
                    found_counts = self.experimentResults.db_con.tables["FoundFeaturePairs"].group_by("resID").agg(pl.count("resID").alias("FOUNDINCOUNT"))
                    fp_with_counts = group_results_df.join(found_counts, left_on="id", right_on="resID", how="left")
                    fp_with_counts = fp_with_counts.sort("mz")

                # Build max-normalized abundance ratios per group: ratio = Average_peakarea / group max.
                exp_ratio_by_num = {}
                if "Average_peakarea" in group_results_df.columns and "OGroup" in group_results_df.columns and "Num" in group_results_df.columns:
                    _grp_max = {row["OGroup"]: float(row["_max"]) for row in group_results_df.group_by("OGroup").agg(pl.col("Average_peakarea").max().alias("_max")).to_dicts() if row.get("_max") is not None and float(row["_max"]) > 0.0}
                    for row in group_results_df.select(["Num", "OGroup", "Average_peakarea"]).to_dicts():
                        gmax = _grp_max.get(row["OGroup"])
                        avg_area = row.get("Average_peakarea")
                        if gmax is not None and avg_area is not None:
                            exp_ratio_by_num[row["Num"]] = float(avg_area) / gmax
                elif "Relative_peakarea_in_group" in group_results_df.columns and "Num" in group_results_df.columns:
                    # Fallback only when average peak area is unavailable.
                    exp_ratio_by_num = {row["Num"]: float(row["Relative_peakarea_in_group"]) for row in group_results_df.select(["Num", "Relative_peakarea_in_group"]).to_dicts() if row.get("Relative_peakarea_in_group") is not None}

                for row_dict in group_results_df.to_dicts():
                    # Count N_found_Samples from per-file _Found columns or use pre-computed value
                    n_found_samples = row_dict.get("N_found_Samples")
                    if n_found_samples is None:
                        n_found_samples = 0
                        for col_name in row_dict.keys():
                            if col_name.endswith("_Found") and row_dict[col_name] is not None:
                                found_val = str(row_dict[col_name])
                                if "Direct" in found_val or "Reintegrated" in found_val:
                                    n_found_samples += 1

                    fp = Bunch(
                        type="featurePair",
                        id=row_dict["Num"],
                        metaboliteGroupID=row_dict["OGroup"],
                        mz=row_dict["MZ"],
                        lmz=row_dict.get("L_MZ"),
                        dmz=row_dict.get("D_MZ"),
                        rt=row_dict["RT"] * 60.0,
                        xn=row_dict["Xn"],
                        charge=row_dict["Charge"],
                        scanEvent=row_dict.get("ScanEvent"),
                        ionisationMode=row_dict.get("Ionisation_Mode"),
                        tracer=row_dict.get("Tracer"),
                        N_found_Samples=n_found_samples,
                    )

                    title = "%s" % (str(fp.id))
                    try:
                        title = "%s / %d rep." % (title, int(fp.N_found_Samples))
                    except Exception:
                        pass
                    try:
                        rel_ratio = exp_ratio_by_num.get(fp.id)
                        if rel_ratio is not None:
                            title = "%s / %.1f%%" % (title, rel_ratio * 100.0)
                    except Exception:
                        pass
                    try:
                        title = "%s / %.4g" % (title, row_dict.get("Average_peakarea", -1))
                    except Exception:
                        pass
                    try:
                        title = "%s / %s" % (title, row_dict.get("Ion", ""))
                    except Exception:
                        pass

                    featurePair = QtWidgets.QTreeWidgetItem(
                        [
                            title,
                            "%.4f" % fp.mz,
                            "%.2f" % (fp.rt / 60.0),
                            str(fp.xn),
                            str(fp.charge),
                            str(fp.ionisationMode),
                            "",
                        ]
                    )
                    featurePair.bunchData = fp

                    ## TODO
                    if False:
                        abundances = []
                        for fileRes in fileMappingData[fp.id]:
                            fileResNode = QtWidgets.QTreeWidgetItem(
                                [
                                    str(fileRes.file),
                                    "%.2g" % (fileRes.areaN),
                                    "%.2g" % (fileRes.areaL),
                                ]
                            )
                            fileResNode.bunchData = fileRes

                            color = None
                            for sampleGroup in experimentalGroups:
                                for file in sampleGroup.files:
                                    if str(fileRes.file) in file:
                                        color = sampleGroup.color

                            if color is not None:
                                fileResNode.setBackground(0, QColor(color))

                            featurePair.addChild(fileResNode)

                            abundances.append(fileRes.areaN)

                    kids.append((featurePair, -1, fp.metaboliteGroupID))

                # Build a lookup {Num -> relative_ratio} for the bar delegate.
                # Reuse the same max-normalized lookup for delegate bars.
                _exp_bar_ratio = exp_ratio_by_num

                for fg in set([k[2] for k in kids]):
                    ckids = sorted(
                        [k for k in kids if k[2] == fg],
                        key=lambda x: x[1],
                        reverse=True,
                    )
                    for kid in ckids:
                        ratio = _exp_bar_ratio.get(kid[0].bunchData.id)
                        if ratio is not None:
                            kid[0].setData(0, _RELATIVE_BAR_ROLE, float(ratio))
                        metaboliteGroupTreeItems[kid[2]].addChild(kid[0])

                for grpID in metaboliteGroupTreeItems.keys():
                    kids = []
                    for i in range(metaboliteGroupTreeItems[grpID].childCount()):
                        kids.append(metaboliteGroupTreeItems[grpID].child(i))
                    meanRT = mean([float(kid.text(2)) for kid in kids])
                    metaboliteGroupTreeItems[grpID].setText(1, "%d" % len(kids))
                    metaboliteGroupTreeItems[grpID].setText(2, "%.2f" % meanRT)

                # Load data into Statistics tab
                self._loadStatisticsData(from_sheet=selected_table)

        except Exception as e:
            traceback.print_exc()
            logging.error(str(traceback))

            logging.error("Multiple file results could not be fetched correctly: " + str(e))
            pass

    def closeLoadedGroupsResultsFile(self):
        if hasattr(self, "experimentResults"):
            self.ui.resultsExperiment_TreeWidget.clear()
            self.experimentResults.db_con = None
            delattr(self, "experimentResults")
        # Clear feature map when results are closed
        self._featureMapData = []
        if hasattr(self.ui, "expFeatureMap_plot"):
            self.ui.expFeatureMap_plot.axes.clear()
            self.ui.expFeatureMap_plot.canvas.draw()

    # ── Feature Map ──────────────────────────────────────────────────────────

    # Colors for positive / negative polarity
    _FM_COLOR_POS = "#DD9D48"
    _FM_COLOR_NEG = "#8D86B8"

    def _toggleFeatureMap(self, checked: bool):
        """Show or hide the feature map panel; rebuild when shown."""
        self.ui.expFeatureMapContainer.setVisible(checked)
        if checked:
            self._buildFeatureMap()

    def _buildFeatureMap(self):
        """Render all features as a RT-vs-MZ scatter plot grouped by OGroup."""
        if not hasattr(self, "experimentResults") or self.experimentResults is None:
            return

        try:
            df = self.experimentResults.db_con.tables[self.experimentResults.selected_table]
        except Exception:
            return

        ax = self.ui.expFeatureMap_plot.axes
        ax.clear()
        self._featureMapData = []
        self._featureMapAnnotation = None

        has_mode = "Ionisation_Mode" in df.columns
        has_avg = "Average_peakarea" in df.columns
        has_found = "N_found_Samples" in df.columns

        # Determine which metric drives dot size from the combo box
        size_metric = self.ui.expFeatureMapSizeCombo.currentText()

        # Try to derive MSMS spectra counts from per-file columns (columns ending in _MSMS or similar)
        # We count non-null/non-zero entries in native+labeled MSMS columns if available,
        # otherwise fall back to counting _Found columns as a proxy.
        def _msms_count(r: dict) -> int:
            """Count MSMS spectra recorded for this feature (native + labeled)."""
            total = 0
            for col, val in r.items():
                if val and (col.endswith("_N_MSMS") or col.endswith("_L_MSMS") or col.endswith("_MSMS_N") or col.endswith("_MSMS_L")):
                    try:
                        total += int(val)
                    except (ValueError, TypeError):
                        pass
            return total

        def _found_count(r: dict) -> int:
            if has_found:
                try:
                    return int(r.get("N_found_Samples") or 0)
                except (ValueError, TypeError):
                    pass
            # fallback: count _Found columns with detected values
            count = 0
            for col, val in r.items():
                if col.endswith("_Found") and val and ("Direct" in str(val) or "Reintegrated" in str(val)):
                    count += 1
            return count

        rows = df.to_dicts()

        # Restrict to features currently visible in the tree (respects active filters)
        visible_nums = self._getVisibleFeatureNums()
        if visible_nums is not None:
            rows = [r for r in rows if r.get("Num") in visible_nums]

        # Collect raw metric values for normalisation
        if size_metric == "Average peak area":
            raw_vals = [float(r.get("Average_peakarea") or 0.0) for r in rows] if has_avg else [0.0] * len(rows)
        elif size_metric == "Number of MSMS spectra":
            raw_vals = [float(_msms_count(r)) for r in rows]
        else:  # Found in n samples
            raw_vals = [float(_found_count(r)) for r in rows]

        max_val = max(raw_vals) if raw_vals else 1.0
        if max_val <= 0:
            max_val = 1.0

        # Collect per-OGroup data for connecting lines
        groups_for_lines: dict[str, list[dict]] = {}
        for r in rows:
            groups_for_lines.setdefault(r["OGroup"], []).append(r)

        # Draw connecting lines (sorted by mz) using the polarity color of the first member
        for og, grp_rows in groups_for_lines.items():
            if len(grp_rows) < 2:
                continue
            rep_mode = str(grp_rows[0].get("Ionisation_Mode", "+") or "+")
            line_color = self._FM_COLOR_POS if "+" in rep_mode else self._FM_COLOR_NEG
            sorted_rows = sorted(grp_rows, key=lambda r: r["MZ"])
            ax.plot(
                [r["RT"] for r in sorted_rows],
                [r["MZ"] for r in sorted_rows],
                color=line_color,
                linewidth=0.8,
                alpha=0.45,
                zorder=1,
            )

        # Accumulate scatter data separated by polarity
        pos_rts, pos_mzs, pos_sizes = [], [], []
        neg_rts, neg_mzs, neg_sizes = [], [], []

        for i, r in enumerate(rows):
            rt = float(r["RT"])
            mz = float(r["MZ"])
            ion_mode = str(r.get("Ionisation_Mode", "+") or "+") if has_mode else "+"
            area = float(r.get("Average_peakarea") or 0.0) if has_avg else 0.0
            n_msms = _msms_count(r)
            n_found = _found_count(r)

            raw = raw_vals[i]
            size = 20.0 + (raw / max_val) * 230.0

            self._featureMapData.append(
                {
                    "rt": rt,
                    "mz": mz,
                    "num": r.get("Num"),
                    "ogroup": r["OGroup"],
                    "charge": r.get("Charge"),
                    "polarity": ion_mode,
                    "xn": r.get("Xn"),
                    "avg_peakarea": area,
                    "n_msms": n_msms,
                    "n_found": n_found,
                }
            )

            if "+" in ion_mode:
                pos_rts.append(rt)
                pos_mzs.append(mz)
                pos_sizes.append(size)
            else:
                neg_rts.append(rt)
                neg_mzs.append(mz)
                neg_sizes.append(size)

        if pos_rts:
            ax.scatter(
                pos_rts,
                pos_mzs,
                s=pos_sizes,
                c=self._FM_COLOR_POS,
                marker="o",
                zorder=2,
                alpha=0.85,
                linewidths=0,
                label="positive",
            )
        if neg_rts:
            ax.scatter(
                neg_rts,
                neg_mzs,
                s=neg_sizes,
                c=self._FM_COLOR_NEG,
                marker="o",
                zorder=2,
                alpha=0.85,
                linewidths=0,
                label="negative",
            )

        if pos_rts or neg_rts:
            ax.legend(fontsize=8, markerscale=0.8, framealpha=0.7)

        ax.set_xlabel("RT (min)", fontsize=10)
        ax.set_ylabel("m/z", fontsize=10)
        ax.set_title(f"Feature Map  ·  size = {size_metric}", fontsize=11)
        ax.tick_params(labelsize=9)

        self._featureMapAnnotation = ax.annotate(
            "",
            xy=(0, 0),
            xytext=(15, 15),
            textcoords="offset points",
            bbox={"boxstyle": "round,pad=0.4", "fc": "lightyellow", "ec": "gray", "lw": 0.8},
            arrowprops={"arrowstyle": "->", "color": "gray"},
            fontsize=8,
            visible=False,
        )

        self.ui.expFeatureMap_plot.fig.tight_layout()
        self.ui.expFeatureMap_plot.canvas.draw()

    def _getVisibleFeatureNums(self) -> set | None:
        """Return the set of Num values currently visible in the tree.
        Returns None if all items are visible (no filter active)."""
        tree = self.ui.resultsExperiment_TreeWidget
        visible: set = set()
        any_hidden = False
        for i in range(tree.topLevelItemCount()):
            top = tree.topLevelItem(i)
            for c in range(top.childCount()):
                child = top.child(c)
                if child.isHidden():
                    any_hidden = True
                else:
                    bd = getattr(child, "bunchData", None)
                    if bd is not None:
                        visible.add(getattr(bd, "id", None))
        return visible if any_hidden else None

    def _featureMapFindNearest(self, event) -> dict | None:
        """Return the nearest feature within ~10 pixels of the mouse event, or None."""
        if event.xdata is None or event.ydata is None:
            return None
        if not self._featureMapData:
            return None

        ax = self.ui.expFeatureMap_plot.axes
        # Transform data coords to display coords for pixel distance
        x_disp, y_disp = ax.transData.transform((event.xdata, event.ydata))

        best = None
        best_dist = float("inf")
        for feat in self._featureMapData:
            fx, fy = ax.transData.transform((feat["rt"], feat["mz"]))
            dist = ((fx - x_disp) ** 2 + (fy - y_disp) ** 2) ** 0.5
            if dist < best_dist:
                best_dist = dist
                best = feat

        return best if best_dist <= 10.0 else None

    @staticmethod
    def _mathbf(value) -> str:
        """Return the value as a bold matplotlib mathtext string, escaping special characters."""
        s = str(value)
        for ch in ["\\", "{", "}", "$", "_", "^", "#", "%", "&", "~"]:
            s = s.replace(ch, "\\" + ch)
        s = s.replace(" ", "\\ ")
        return r"$\mathbf{" + s + r"}$"

    def _onFeatureMapHover(self, event):
        """Show hover annotation near the cursor when close to a feature dot."""
        if not hasattr(self, "_featureMapAnnotation") or self._featureMapAnnotation is None:
            return
        if event.inaxes != self.ui.expFeatureMap_plot.axes:
            if self._featureMapAnnotation.get_visible():
                self._featureMapAnnotation.set_visible(False)
                self.ui.expFeatureMap_plot.canvas.draw_idle()
            return

        feat = self._featureMapFindNearest(event)
        ann = self._featureMapAnnotation

        if feat is None:
            if ann.get_visible():
                ann.set_visible(False)
                self.ui.expFeatureMap_plot.canvas.draw_idle()
            return

        # Build tooltip text: one key-value pair per line, value in bold
        lines = [
            f"Num: {self._mathbf(feat['num'])}",
            f"OGroup: {self._mathbf(feat['ogroup'])}",
            f"m/z: {self._mathbf('%.4f' % feat['mz'])}",
            f"RT: {self._mathbf('%.3f' % feat['rt'])} min",
            f"Charge: {self._mathbf(feat['charge'])}",
            f"Polarity: {self._mathbf(feat['polarity'])}",
            f"Xn: {self._mathbf(feat['xn'])}",
        ]
        avg = feat.get("avg_peakarea")
        if avg is not None and avg > 0:
            lines.append(f"Avg peak area: {self._mathbf('%.3g' % avg)}")
        n_msms = feat.get("n_msms")
        if n_msms is not None:
            lines.append(f"MSMS spectra: {self._mathbf(n_msms)}")
        n_found = feat.get("n_found")
        if n_found is not None:
            lines.append(f"Found in samples: {self._mathbf(n_found)}")
        ann.set_text("\n".join(lines))
        ann.xy = (feat["rt"], feat["mz"])
        ann.set_visible(True)
        self.ui.expFeatureMap_plot.canvas.draw_idle()

    def _onFeatureMapClick(self, event):
        """On left-click near a feature dot, select it in the tree widget."""
        if event.button != 1:
            return
        if event.inaxes != self.ui.expFeatureMap_plot.axes:
            return
        # Don't trigger selection when the pan or zoom tool consumed the drag
        toolbar_mode = str(self.ui.expFeatureMap_plot.mpl_toolbar.mode)
        if toolbar_mode in ("pan/zoom", "zoom rect"):
            return

        feat = self._featureMapFindNearest(event)
        if feat is None:
            return

        num = feat.get("num")
        if num is None:
            return

        # Switch to experiment results tab if needed
        for i in range(self.ui.tabWidget.count()):
            if self.ui.tabWidget.widget(i) == self.ui.bracketedResultsTab:
                self.ui.tabWidget.setCurrentIndex(i)
                break

        # Find and select the matching feature in the tree
        tree = self.ui.resultsExperiment_TreeWidget
        for top_idx in range(tree.topLevelItemCount()):
            top_item = tree.topLevelItem(top_idx)
            # Check top-level item itself
            if hasattr(top_item, "bunchData") and getattr(top_item.bunchData, "id", None) == num:
                tree.setCurrentItem(top_item)
                tree.scrollToItem(top_item)
                return
            # Check children
            for child_idx in range(top_item.childCount()):
                child = top_item.child(child_idx)
                if hasattr(child, "bunchData") and getattr(child.bunchData, "id", None) == num:
                    top_item.setExpanded(True)
                    tree.setCurrentItem(child)
                    tree.scrollToItem(child)
                    return

    # ── Feature-field copy context menu (shared by both result tree views) ────
    def _addFeatureCopyMenu(self, menu, values: dict) -> dict:
        """Add a 'Copy' submenu to *menu* offering the available feature fields.

        *values* maps the keys ogroup/num/polarity/rt/mz/lmz to their string values
        (any of which may be None to omit it). Returns {QAction: text-to-copy}.
        """
        field_order = [
            ("OGroup", "ogroup"),
            ("Feature-Num", "num"),
            ("Polarity", "polarity"),
            ("Mean RT (min)", "rt"),
            ("Mean m/z (native)", "mz"),
            ("Mean m/z (labeled)", "lmz"),
        ]
        present = [(label, key) for label, key in field_order if values.get(key) is not None]
        if not present:
            return {}

        copyMenu = menu.addMenu("Copy")
        actions: dict = {}
        for label, key in present:
            act = copyMenu.addAction(label)
            actions[act] = str(values[key])

        copyMenu.addSeparator()
        tab_str = "\t".join(str(values[key]) for _, key in present)
        semi_str = ";".join(str(values[key]) for _, key in present)
        actions[copyMenu.addAction("All (tab-separated)")] = tab_str
        actions[copyMenu.addAction("All (semicolon-separated)")] = semi_str
        return actions

    def _featureValuesFromBunch(self, bd) -> dict:
        """Build the copy-value dict for an experiment-results tree item's bunchData."""
        values = {"ogroup": None, "num": None, "polarity": None, "rt": None, "mz": None, "lmz": None}
        if bd is None:
            return values
        bd_type = getattr(bd, "type", None)
        if bd_type == "featurePair":
            values["ogroup"] = getattr(bd, "metaboliteGroupID", None)
            values["num"] = getattr(bd, "id", None)
            values["polarity"] = getattr(bd, "ionisationMode", None)
            rt = getattr(bd, "rt", None)
            if rt is not None:
                values["rt"] = "%.3f" % (float(rt) / 60.0)
            mz = getattr(bd, "mz", None)
            if mz is not None:
                values["mz"] = "%.4f" % float(mz)
            lmz = getattr(bd, "lmz", None)
            if lmz is not None:
                values["lmz"] = "%.4f" % float(lmz)
        elif bd_type == "metaboliteGroup":
            values["ogroup"] = getattr(bd, "metaboliteGroupID", None)
        return values

    def _showExperimentTreeContextMenu(self, position):
        tree = self.ui.resultsExperiment_TreeWidget
        item = tree.itemAt(position)
        if item is None:
            return
        bd = getattr(item, "bunchData", None)
        values = self._featureValuesFromBunch(bd)
        menu = QtWidgets.QMenu()
        actions = self._addFeatureCopyMenu(menu, values)
        if not actions:
            return
        chosen = menu.exec_(tree.mapToGlobal(position))
        if chosen in actions:
            pyperclip.copy(actions[chosen])

    def _showFeatureInExperimentResults(self, feature_index: int):
        """Navigate to the experiment results tab and select the specified feature."""
        # Switch to the experiment results tab (bracketedResultsTab)
        for i in range(self.ui.tabWidget.count()):
            if self.ui.tabWidget.widget(i) == self.ui.bracketedResultsTab:
                self.ui.tabWidget.setCurrentIndex(i)
                break

        # Try to find and select the feature in the tree widget
        if hasattr(self, "experimentResults") and self.experimentResults is not None:
            # Iterate through all top-level items (metabolite groups and standalone features)
            for i in range(self.ui.resultsExperiment_TreeWidget.topLevelItemCount()):
                item = self.ui.resultsExperiment_TreeWidget.topLevelItem(i)

                # Check if this is a feature (standalone or metabolite group)
                if hasattr(item, "bunchData") and hasattr(item.bunchData, "id"):
                    if item.bunchData.id == feature_index:
                        self.ui.resultsExperiment_TreeWidget.setCurrentItem(item)
                        self.ui.resultsExperiment_TreeWidget.scrollToItem(item)
                        self.ui.resultsExperiment_TreeWidget.expandItem(item.parent() if item.parent() else item)
                        return

                # If this is a metabolite group, check its children
                for child_idx in range(item.childCount()):
                    child = item.child(child_idx)
                    if hasattr(child, "bunchData") and hasattr(child.bunchData, "id"):
                        if child.bunchData.id == feature_index:
                            self.ui.resultsExperiment_TreeWidget.setCurrentItem(child)
                            self.ui.resultsExperiment_TreeWidget.scrollToItem(child)
                            self.ui.resultsExperiment_TreeWidget.expandItem(item)
                            return

    def _loadStatisticsData(self, from_sheet="GroupResults"):
        """Load experiment data into the Statistics tab."""
        if not hasattr(self.ui, "statisticsWidget"):
            return

        if not hasattr(self, "experimentResults") or self.experimentResults is None:
            return

        try:
            # Collect feature data and group info
            experiment_data = {"features": {}, "groups": {}, "metadata": {}}

            # Get group info
            for grp in self.getAllSampleGroups():
                group_name = str(grp.name)
                basenames = [os.path.splitext(os.path.basename(str(f)))[0] for f in grp.files]
                experiment_data["groups"][group_name] = basenames

            # Load feature data from the single results table
            if hasattr(self, "experimentResults") and self.experimentResults.db_con is not None:
                table_df = self.experimentResults.db_con.tables[from_sheet]
                columns = table_df.columns

                # Discover sample names from columns ending with _Area_N or _Area_L
                suffix_n = "_Area_N"
                suffix_l = "_Area_L"
                samples_n = {col[: -len(suffix_n)] for col in columns if col.endswith(suffix_n)}
                samples_l = {col[: -len(suffix_l)] for col in columns if col.endswith(suffix_l)}
                all_samples = samples_n | samples_l

                features_N = {}
                features_L = {}
                features_total = {}
                metadata = {"mz": {}, "rt": {}, "num": {}, "ogroup": {}, "featurePairID": {}, "featureGroupID": {}}

                for row in table_df.to_dicts():
                    idx = row["Num"]
                    metadata["mz"][idx] = row["MZ"]
                    metadata["rt"][idx] = row["RT"]
                    metadata["num"][idx] = idx
                    metadata["ogroup"][idx] = row["OGroup"]
                    metadata["featurePairID"][idx] = row.get("featurePairID", 0) or 0
                    metadata["featureGroupID"][idx] = row.get("featureGroupID", 0) or 0

                    row_n = {}
                    row_l = {}
                    row_total = {}
                    for sample in all_samples:
                        area_n = row.get(sample + suffix_n) or 0.0
                        area_l = row.get(sample + suffix_l) or 0.0
                        if area_n is None or area_n == "":
                            area_n = 0.0
                        if area_l is None or area_l == "":
                            area_l = 0.0
                        if type(area_n) is str:
                            area_n = float(area_n)
                        if type(area_l) is str:
                            area_l = float(area_l)
                        row_n[sample] = area_n
                        row_l[sample] = area_l
                        row_total[sample] = area_n + area_l

                    features_N[idx] = row_n
                    features_L[idx] = row_l
                    features_total[idx] = row_total

                experiment_data["features"] = features_total  # Default to total
                experiment_data["features_N"] = features_N
                experiment_data["features_L"] = features_L
                experiment_data["metadata"] = metadata

                logging.info(f"Statistics: {len(all_samples)} samples detected from columns")
                logging.info(f"Statistics: Group names: {list(experiment_data.get('groups', {}).keys())}")
                for group_name, samples in experiment_data.get("groups", {}).items():
                    logging.info(f"Statistics: Group '{group_name}' samples (first 3): {samples[:3]}")

            self.ui.statisticsWidget.load_experiment_data(experiment_data)
            self.ui.statisticsTab.setEnabled(True)
            logging.info(f"Statistics data loaded: {len(experiment_data.get('features', {}))} features")

        except Exception as e:
            logging.error(f"Error loading statistics data: {e}")
            import traceback

            logging.error(traceback.format_exc())

    def resultsExperimentChanged(self):
        self.clearPlot(self.ui.resultsExperiment_plot)
        self.clearPlot(self.ui.resultsExperimentSeparatedPeaks_plot)
        self.clearPlot(self.ui.resultsExperimentMSScanPeaks_plot)
        self.clearPlot(self.ui.resultsExperimentAbundance_plot)

        if not hasattr(self, "experimentResults") or self.experimentResults is None:
            self.drawCanvas(self.ui.resultsExperiment_plot)
            self.drawCanvas(self.ui.resultsExperimentSeparatedPeaks_plot)
            self.drawCanvas(self.ui.resultsExperimentMSScanPeaks_plot)
            self.drawCanvas(self.ui.resultsExperimentAbundance_plot, showLegendOverwrite=False)
            self.updateSamplePeaksTab([])
            return

        if len(self.ui.resultsExperiment_TreeWidget.selectedItems()) == 0:
            self.drawCanvas(self.ui.resultsExperiment_plot)
            self.drawCanvas(self.ui.resultsExperimentSeparatedPeaks_plot)
            self.drawCanvas(self.ui.resultsExperimentMSScanPeaks_plot)
            self.drawCanvas(self.ui.resultsExperimentAbundance_plot, showLegendOverwrite=False)
            self.updateSamplePeaksTab([])
            return

        plotItems = self._getSelectedExperimentPlotItems()
        self.updateExperimentAbundancePlot(plotItems)

        # Load raw mzXML files if not already loaded
        if not hasattr(self, "loadedMZXMLs") or self.loadedMZXMLs is None:
            selectedMZs = None
            if (
                QtWidgets.QMessageBox.question(
                    self,
                    "MetExtract",
                    "Raw data needs to be loaded to display EICs.\nLoad entire chromatograms (Yes) or just detected m/z values (No)?\n\nIf you have limited RAM, choose No.\nNote: loading only detected m/z values disables 'Show custom feature'.",
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                )
                == QtWidgets.QMessageBox.No
            ):
                selectedMZs = []
                for itemI in range(self.ui.resultsExperiment_TreeWidget.topLevelItemCount()):
                    item = self.ui.resultsExperiment_TreeWidget.topLevelItem(itemI)
                    if item.bunchData.type == "metaboliteGroup":
                        for i in range(item.childCount()):
                            child = item.child(i)
                            if child.bunchData.type == "featurePair":
                                selectedMZs.append(child.bunchData.mz)
                                if child.bunchData.lmz is not None:
                                    selectedMZs.append(child.bunchData.lmz)
                    if item.bunchData.type == "featurePair":
                        selectedMZs.append(item.bunchData.mz)
                        if item.bunchData.lmz is not None:
                            selectedMZs.append(item.bunchData.lmz)

            self.loadAllSamples(selectedMZs=selectedMZs, ppm=25.0)

        if not hasattr(self, "loadedMZXMLs") or self.loadedMZXMLs is None:
            self.drawCanvas(self.ui.resultsExperiment_plot)
            self.drawCanvas(self.ui.resultsExperimentSeparatedPeaks_plot)
            self.drawCanvas(self.ui.resultsExperimentMSScanPeaks_plot)
            return

        definedGroups = self.getAllSampleGroups()

        if not plotItems:
            self.drawCanvas(self.ui.resultsExperiment_plot)
            self.drawCanvas(self.ui.resultsExperimentSeparatedPeaks_plot)
            self.drawCanvas(self.ui.resultsExperimentMSScanPeaks_plot)
            return

        ppm = self.ui.doubleSpinBox_resultsExperiment_EICppm.value()
        borderOffset = self.ui.doubleSpinBox_resultsExperiment_PeakWidth.value()
        shiftMinutes = self.ui.doubleSpinBox_separatePeaksShift.value()
        separateBy = self.ui.comboBox_separatePeaks.currentText()  # "Group" or "Sample"
        meanRT = []
        intlim = [0, 0]

        # Per-feature row lookup to determine which files detected the feature
        # pair (i.e. have a feature ID in their "<sample>_FID" column).
        rowsByNum = {}
        selected_table = getattr(self.experimentResults, "selected_table", None)
        if selected_table is not None and selected_table in self.experimentResults.db_con.tables:
            table_df = self.experimentResults.db_con.tables[selected_table]
            feature_ids = [pi.id for pi in plotItems if getattr(pi, "id", None) is not None]
            if feature_ids:
                rowsByNum = {row["Num"]: row for row in table_df.filter(pl.col("Num").is_in(feature_ids)).to_dicts()}

        # Solid line for detected feature pairs, densely dashdotted otherwise.
        solidStyle = "-"
        denselyDashDotted = (0, (3, 1, 1, 1))

        all_files = sum(len(g.files) for g in definedGroups)

        pw = ProgressWrapper(1, parent=self, showIndProgress=False)
        pw.show()
        pw.getCallingFunction()("max")(len(plotItems) * all_files)
        pw.getCallingFunction()("value")(0)

        done = 0
        for h, pi in enumerate(plotItems):
            meanRT.append(pi.rt)

            rtBorderMin = pi.rt / 60.0 - borderOffset
            rtBorderMax = pi.rt / 60.0 + borderOffset

            singleOffset = 0
            offsetOrder = []
            for grpInd, group in enumerate(definedGroups):
                if separateBy == "Group":
                    offsetOrder.append((group.name, group.color if group.color else "gray"))
                for i in range(len(group.files)):
                    fi = str(group.files[i]).replace("\\", "/")
                    a = fi[fi.rfind("/") + 1 :]
                    if ".mzXML" in a:
                        a = a[: a.find(".mzXML")]
                    elif ".mzxml" in a.lower():
                        a = a[: a.lower().find(".mzxml")]
                    elif ".mzML" in a:
                        a = a[: a.find(".mzML")]

                    pw.getCallingFunction()("text")("MZ: %.5f\nFile: '%s'" % (pi.mz, a))
                    pw.getCallingFunction()("value")(done)
                    done = done + 1

                    if fi not in self.loadedMZXMLs:
                        continue

                    scanEvent = pi.scanEvent if hasattr(pi, "scanEvent") and pi.scanEvent else None
                    if scanEvent is None:
                        # Try to find a valid filter line
                        filterLines = self.loadedMZXMLs[fi].getFilterLines(
                            includeMS1=True,
                            includeMS2=False,
                            includePosPolarity=True,
                            includeNegPolarity=True,
                        )
                        if filterLines:
                            # Pick the filter line matching the ionisation mode if available
                            ionMode = getattr(pi, "ionisationMode", None)
                            if ionMode and "+" in str(ionMode):
                                scanEvent = next((fl for fl in filterLines if "+" in fl), filterLines[0])
                            elif ionMode and "-" in str(ionMode):
                                scanEvent = next((fl for fl in filterLines if "-" in fl), filterLines[0])
                            else:
                                scanEvent = filterLines[0]

                    if scanEvent is None:
                        continue

                    try:
                        availableFilterLines = self.loadedMZXMLs[fi].getFilterLines(
                            includeMS1=True,
                            includeMS2=False,
                            includePosPolarity=True,
                            includeNegPolarity=True,
                        )
                        if scanEvent not in availableFilterLines:
                            continue

                        eic, times, scanIds, mzs = self.loadedMZXMLs[fi].getEIC(pi.mz, ppm=ppm, filterLine=scanEvent)
                        lmz = pi.lmz if pi.lmz is not None else pi.mz
                        eicL, timesL, scanIdsL, mzsL = self.loadedMZXMLs[fi].getEIC(lmz, ppm=ppm, filterLine=scanEvent)

                        groupColor = group.color if group.color else "gray"

                        # Solid line if this file detected the feature pair (has a
                        # feature ID), densely dashdotted otherwise.
                        row_data = rowsByNum.get(pi.id)
                        fid_val = row_data.get(a + "_FID") if row_data is not None else None
                        detected = fid_val is not None and str(fid_val).strip() not in ("", "None")
                        eicStyle = solidStyle if detected else denselyDashDotted

                        maxN = 1
                        maxL = 1

                        if self.ui.resultsExperimentNormaliseXICs_checkBox.isChecked() or self.ui.resultsExperimentNormaliseXICsSeparately_checkBox.isChecked():
                            m = 0
                            ml = 0
                            for j in range(len(eic)):
                                if abs(pi.rt / 60 - (times[j] / 60.0)) <= 0.2:
                                    m = max(m, eic[j])
                                    ml = max(ml, eicL[j])
                            if m != 0:
                                maxN = m
                            if ml != 0:
                                maxL = ml

                        if self.ui.resultsExperimentNormaliseXICs_checkBox.isChecked():
                            maxN = maxL

                        intlim[0] = min(
                            intlim[0],
                            min([1] + [-eicL[j] / maxL for j in range(len(eic)) if rtBorderMin <= times[j] / 60.0 <= rtBorderMax]),
                        )
                        intlim[1] = max(
                            intlim[1],
                            max([1] + [eic[j] / maxN for j in range(len(eic)) if rtBorderMin <= times[j] / 60.0 <= rtBorderMax]),
                        )

                        # --- MS1 scan per file ---
                        scan = self.loadedMZXMLs[fi].getClosestMS1Scan(pi.rt / 60.0, filterLine=scanEvent)
                        if scan is not None:
                            ms1_mzs = scan.mz_list
                            ms1_ints = scan.intensity_list
                            use_inds = []
                            for ind_i in range(len(ms1_mzs)):
                                if ms1_mzs[ind_i] >= pi.mz - 5 and ms1_mzs[ind_i] <= lmz + 5:
                                    use_inds.append(ind_i)
                            if len(use_inds) > 0:
                                plot_mzs = [ms1_mzs[ii] for ii in use_inds]
                                plot_ints = [ms1_ints[ii] for ii in use_inds]

                                max_plotted_int = max(plot_ints) if plot_ints else 1
                                corrFact = 1.0 / max_plotted_int * 9 if max_plotted_int > 0 else 1

                                self.ui.resultsExperimentMSScanPeaks_plot.axes.vlines(
                                    x=plot_mzs,
                                    ymin=done * 10,
                                    ymax=[done * 10 + ii * corrFact for ii in plot_ints],
                                    color=groupColor,
                                    alpha=0.3,
                                )
                                # Highlight M and M' isotope peaks
                                for target_mz in [pi.mz + k * 1.00335484 for k in [0, 1, 2, 3]] + [lmz + k * 1.00335484 for k in [0, 1, 2, 3]] + [pi.mz - k * 1.00335484 for k in [0, 1, 2, 3]] + [lmz - k * 1.00335484 for k in [0, 1, 2, 3]]:
                                    peakID = scan.findMZ(target_mz, ppm=ppm)
                                    if peakID[0] != -1:
                                        self.ui.resultsExperimentMSScanPeaks_plot.axes.vlines(
                                            x=scan.mz_list[peakID[0]],
                                            ymin=done * 10,
                                            ymax=done * 10 + scan.intensity_list[peakID[0]] * corrFact,
                                            color=groupColor,
                                        )
                                for target_mz in [pi.mz, lmz]:
                                    peakID = scan.findMZ(target_mz, ppm=ppm)
                                    if peakID[0] != -1:
                                        self.ui.resultsExperimentMSScanPeaks_plot.axes.vlines(
                                            x=scan.mz_list[peakID[0]],
                                            ymin=done * 10,
                                            ymax=done * 10 + scan.intensity_list[peakID[0]] * corrFact,
                                            color=groupColor,
                                            linewidth=2.0,
                                        )

                                self.ui.resultsExperimentMSScanPeaks_plot.axes.hlines(
                                    y=done * 10,
                                    xmin=pi.mz - 5,
                                    xmax=lmz + 5,
                                    color=groupColor,
                                    alpha=0.33,
                                )
                                self.ui.resultsExperimentMSScanPeaks_plot.axes.text(
                                    y=done * 10,
                                    x=lmz + 5.5,
                                    color=groupColor,
                                    s="%s, max.int. %.3g" % (a, max_plotted_int),
                                )

                        # --- Overlaid EICs ---
                        self.ui.resultsExperiment_plot.axes.plot(
                            [t / 60.0 for t in times],
                            [e / maxN for e in eic],
                            color=groupColor,
                            linestyle=eicStyle,
                            label="M: %s" % (a),
                        )
                        self.ui.resultsExperiment_plot.axes.plot(
                            [t / 60.0 for t in times],
                            [-e / maxL for e in eicL],
                            color=groupColor,
                            linestyle=eicStyle,
                            label="M': %s" % (a),
                        )

                        # --- Separated EICs with artificial RT shift ---
                        if separateBy == "Group":
                            offset = grpInd * shiftMinutes
                        else:
                            offset = singleOffset * shiftMinutes
                            offsetOrder.append((a, groupColor))

                        self.ui.resultsExperimentSeparatedPeaks_plot.axes.plot(
                            [t / 60.0 + offset for t in times if rtBorderMin <= t / 60.0 <= rtBorderMax],
                            [eic[j] / maxN for j in range(len(eic)) if rtBorderMin <= times[j] / 60.0 <= rtBorderMax],
                            color=groupColor,
                            linestyle=eicStyle,
                            label="M: %s" % (a),
                        )
                        self.ui.resultsExperimentSeparatedPeaks_plot.axes.plot(
                            [t / 60.0 + offset for t in times if rtBorderMin <= t / 60.0 <= rtBorderMax],
                            [-eicL[j] / maxL for j in range(len(eicL)) if rtBorderMin <= times[j] / 60.0 <= rtBorderMax],
                            color=groupColor,
                            linestyle=eicStyle,
                            label="M': %s" % (a),
                        )

                        singleOffset += 1

                    except Exception as e:
                        logging.warning("Could not compute EIC for file '%s': %s" % (fi, str(e)))

        pw.hide()
        if done > 0:
            pi = plotItems[0]
            for oi, o in enumerate(offsetOrder):
                self.ui.resultsExperimentSeparatedPeaks_plot.axes.axvline(x=oi * shiftMinutes + pi.rt / 60.0, color=o[1])
                label = o[0]
                if separateBy == "Group":
                    found, rsd = self._groupSampleStats(o[0], plotItems, definedGroups, rowsByNum)
                    label = "%s\n%s\n%s" % (o[0], found, rsd)
                self.ui.resultsExperimentSeparatedPeaks_plot.axes.text(
                    x=oi * shiftMinutes + pi.rt / 60.0,
                    y=intlim[1] * 1.05,
                    s=label,
                    rotation=0,
                    horizontalalignment="center",
                    verticalalignment="bottom",
                    color=o[1],
                    backgroundcolor="white",
                    weight="bold",
                )
            intlim[1] = intlim[1] * 1.5

            self.ui.resultsExperiment_plot.axes.set_title("Overlaid EICs of selected feature pairs or groups")
            self.ui.resultsExperiment_plot.axes.set_xlabel("Retention time (min)")
            self.ui.resultsExperiment_plot.axes.set_ylabel("Intensity")
            sepLabel = "experimental group" if separateBy == "Group" else "sample"
            self.ui.resultsExperimentSeparatedPeaks_plot.axes.set_title("Overlaid EICs (separated artificially by %s, shift=%.2f min)" % (sepLabel, shiftMinutes))
            if len(plotItems) == 1:
                self.ui.resultsExperimentSeparatedPeaks_plot.axes.set_title(
                    "EICs of %.5f (%.5f), %.2f min, %s\n(separated by %s, shift=%.2f min)"
                    % (
                        plotItems[0].mz,
                        plotItems[0].lmz if plotItems[0].lmz else 0,
                        plotItems[0].rt / 60.0,
                        plotItems[0].scanEvent if plotItems[0].scanEvent else "",
                        sepLabel,
                        shiftMinutes,
                    )
                )
            self.ui.resultsExperimentSeparatedPeaks_plot.axes.set_xlabel("Retention time (min) + %s-index × %.2f min" % (sepLabel, shiftMinutes))
            self.ui.resultsExperimentSeparatedPeaks_plot.axes.set_ylabel("Intensity")
            self.ui.resultsExperimentMSScanPeaks_plot.axes.set_xlabel("M/Z")
            self.ui.resultsExperimentMSScanPeaks_plot.axes.set_ylabel("Normalized Intensity\nSeparated by sample")

            rtlim = [
                mean(meanRT) / 60.0 - borderOffset,
                mean(meanRT) / 60.0 + borderOffset,
            ]
            intlim = [intlim[0] * 1.1, intlim[1] * 1.1]
            self.drawCanvas(self.ui.resultsExperiment_plot, xlim=rtlim, ylim=intlim, showLegendOverwrite=False)
            self.drawCanvas(self.ui.resultsExperimentSeparatedPeaks_plot, showLegendOverwrite=self.ui.showLegend_experiment.isChecked())
            lmz_val = plotItems[0].lmz if plotItems[0].lmz else plotItems[0].mz
            self.drawCanvas(
                self.ui.resultsExperimentMSScanPeaks_plot,
                xlim=[plotItems[0].mz - 5, lmz_val + 10],
                ylim=[0, (done + 1) * 10],
            )
        else:
            self.drawCanvas(self.ui.resultsExperiment_plot)
            self.drawCanvas(self.ui.resultsExperimentSeparatedPeaks_plot)
            self.drawCanvas(self.ui.resultsExperimentMSScanPeaks_plot)

        # Update MSMS spectra list for selected features
        selectedItems = self.ui.resultsExperiment_TreeWidget.selectedItems()
        self.updateMSMSList_exp(selectedItems)

        # Update peak details tab
        self.updatePeakDetailsTab(plotItems)

        # Update sample peaks tab
        self.updateSamplePeaksTab(plotItems)

    def _getSelectedExperimentPlotItems(self):
        plotItems = []
        for item in self.ui.resultsExperiment_TreeWidget.selectedItems():
            if item.bunchData.type == "metaboliteGroup":
                for i in range(item.childCount()):
                    child = item.child(i)
                    if child.bunchData.type == "featurePair":
                        plotItems.append(child.bunchData)
            if item.bunchData.type == "featurePair":
                plotItems.append(item.bunchData)
        return plotItems

    def _groupSampleStats(self, groupName, plotItems, definedGroups, rowsByNum):
        """Compute label strings for a group in the Separated peaks plot.

        Returns a tuple ("found x / y / z", "RSD % a / b") where:
          x = samples in which the feature pair was detected (has a feature ID),
          y = samples re-integrated as the native form (Area_N not empty),
          z = samples re-integrated as the labeled form (Area_L not empty),
          a/b = relative standard deviation (%) of native/labeled peak areas.
        """
        group = next((g for g in definedGroups if g.name == groupName), None)
        if group is None:
            return ("found 0 / 0 / 0", "RSD % n/a / n/a")

        sampleNames = []
        for f in group.files:
            fi = str(f).replace("\\", "/")
            a = fi[fi.rfind("/") + 1 :]
            for ext in (".mzXML", ".mzxml", ".mzML", ".mzml"):
                if a.lower().endswith(ext.lower()):
                    a = a[: -len(ext)]
                    break
            sampleNames.append(a)

        def parseArea(val):
            if val is None:
                return None
            try:
                return float(str(val).split(";")[0])
            except (TypeError, ValueError):
                return None

        nFound = nNative = nLabeled = 0
        areasN = []
        areasL = []
        for a in sampleNames:
            sampleFound = sampleNative = sampleLabeled = False
            for pi in plotItems:
                row = rowsByNum.get(pi.id)
                if row is None:
                    continue
                fid = row.get(a + "_FID")
                if fid is not None and str(fid).strip() not in ("", "None"):
                    sampleFound = True
                aN = parseArea(row.get(a + "_Area_N"))
                if aN is not None:
                    sampleNative = True
                    areasN.append(aN)
                aL = parseArea(row.get(a + "_Area_L"))
                if aL is not None:
                    sampleLabeled = True
                    areasL.append(aL)
            nFound += sampleFound
            nNative += sampleNative
            nLabeled += sampleLabeled

        def rsd(vals):
            if len(vals) < 2:
                return "n/a"
            m = sum(vals) / len(vals)
            if m == 0:
                return "n/a"
            sd = (sum((v - m) ** 2 for v in vals) / (len(vals) - 1)) ** 0.5
            return "%.1f" % (sd / m * 100.0)

        return ("found %d / %d / %d" % (nFound, nNative, nLabeled), "RSD %% %s / %s" % (rsd(areasN), rsd(areasL)))

    def updateExperimentAbundancePlot(self, plotItems):
        self.clearPlot(self.ui.resultsExperimentAbundance_plot)
        if not plotItems or not hasattr(self, "experimentResults") or self.experimentResults is None:
            self.drawCanvas(self.ui.resultsExperimentAbundance_plot, showLegendOverwrite=False)
            return

        selected_table = getattr(self.experimentResults, "selected_table", None)
        if selected_table is None or selected_table not in self.experimentResults.db_con.tables:
            self.drawCanvas(self.ui.resultsExperimentAbundance_plot, showLegendOverwrite=False)
            return

        table_df = self.experimentResults.db_con.tables[selected_table]
        selected_ids = [pi.id for pi in plotItems if getattr(pi, "id", None) is not None]
        if len(selected_ids) == 0:
            self.drawCanvas(self.ui.resultsExperimentAbundance_plot, showLegendOverwrite=False)
            return

        selected_df = table_df.filter(pl.col("Num").is_in(selected_ids))
        rows_by_num = {row["Num"]: row for row in selected_df.to_dicts()}

        sample_entries = []
        for group in self.getAllSampleGroups():
            group_name = str(group.name)
            for fi in group.files:
                fi = str(fi)
                sample_name = fi[max(fi.rfind("/") + 1, fi.rfind("\\") + 1) :]
                sample_name = re.sub(r"\.(mzxml|mzml)$", "", sample_name, flags=re.IGNORECASE)
                sample_entries.append((group_name, sample_name))

        sample_entries = sorted(set(sample_entries), key=lambda x: (x[0].lower(), x[1].lower()))
        value_type_map = {
            "Native area (_Area_N)": "_Area_N",
            "Labeled area (_Area_L)": "_Area_L",
            "Native intensity (_Abundance_N)": "_Abundance_N",
            "Labeled intensity (_Abundance_L)": "_Abundance_L",
        }
        try:
            value_type_text = self.ui.comboBox_abundanceValueType.currentText()
        except Exception:
            value_type_text = "Native area (_Area_N)"

        both_mode = "both" in value_type_text.lower()
        if both_mode:
            if "intensity" in value_type_text.lower():
                blocks = [("Native", "_Abundance_N"), ("Labeled", "_Abundance_L")]
                metric_label = "intensity"
            else:
                blocks = [("Native", "_Area_N"), ("Labeled", "_Area_L")]
                metric_label = "area"
        else:
            abundance_suffix = value_type_map.get(value_type_text, "_Area_N")
            # Fall back to any available value type if the requested one is missing in the result table
            if not any(f"{entry[1]}{abundance_suffix}" in table_df.columns for entry in sample_entries):
                for fallback in ["_Area_N", "_Abundance_N", "_Area_L", "_Abundance_L"]:
                    if any(f"{entry[1]}{fallback}" in table_df.columns for entry in sample_entries):
                        abundance_suffix = fallback
                        break
            blocks = [("", abundance_suffix)]

        # Build group-name -> defined color map for boxes and sample dots
        group_color_map = {}
        for grp in self.getAllSampleGroups():
            group_color_map[str(grp.name)] = str(grp.color)

        plot_type = self.ui.comboBox_abundancePlotType.currentText()
        log_scale = self.ui.comboBox_abundanceScale.currentText() == "Logarithmic"
        scaling_mode = self.ui.comboBox_abundanceScalingMode.currentText()

        group_names = sorted({entry[0] for entry in sample_entries}, key=lambda x: x.lower())
        samples_by_group = {group_name: [sample for grp, sample in sample_entries if grp == group_name] for group_name in group_names}

        # Collect (and optionally scale) values per isotopolog block:
        # {suffix: {feature_id: {(group, sample): value}}}
        block_feature_values = {}
        for _block_label, block_suffix in blocks:
            feature_sample_values = {}
            for feature_id in sorted(set(selected_ids)):
                row = rows_by_num.get(feature_id)
                if row is None:
                    continue
                values = {}
                for group_name, sample_name in sample_entries:
                    col = sample_name + block_suffix
                    if col not in table_df.columns:
                        continue
                    val = row.get(col)
                    if val in [None, ""]:
                        continue
                    try:
                        values[(group_name, sample_name)] = float(val)
                    except Exception:
                        continue
                if values:
                    feature_sample_values[feature_id] = values

            if scaling_mode != "None":
                for feature_id, values in feature_sample_values.items():
                    if scaling_mode == "Scale to max sample":
                        ref = max(values.values()) if values else 0.0
                    else:
                        group_means = []
                        for group_name in group_names:
                            group_vals = [values[(group_name, sample)] for sample in samples_by_group.get(group_name, []) if (group_name, sample) in values]
                            if group_vals:
                                group_means.append(mean(group_vals))
                        ref = max(group_means) if group_means else 0.0
                    if ref > 0:
                        for key in list(values.keys()):
                            values[key] = values[key] / ref

            block_feature_values[block_suffix] = feature_sample_values

        ax = self.ui.resultsExperimentAbundance_plot.axes
        if both_mode:
            _value_label = metric_label
        else:
            _value_label = {
                "_Area_N": "native area",
                "_Area_L": "labeled area",
                "_Abundance_N": "native intensity",
                "_Abundance_L": "labeled intensity",
            }.get(blocks[0][1], "abundance")
        ax.set_title("Abundance profiles of selected features")
        ax.set_ylabel(f"{_value_label.capitalize()} ({'log' if log_scale else 'linear'} scale)")
        if scaling_mode != "None":
            ax.set_ylabel(f"Relative {_value_label} ({'log' if log_scale else 'linear'} scale)")

        if "Boxplot" in plot_type:
            box_data = []
            positions = []
            box_colors = []
            scatter_x = []
            scatter_y = []
            scatter_colors = []
            legend_handles = []
            block_label_positions = []  # (center_x, label)
            xtick_positions = []
            xtick_labels = []
            box_width = ABUNDANCE_BOXPLOT_CLUSTER_WIDTH

            # union of feature ids present in any block, stable order
            feature_ids = sorted({fid for fsv in block_feature_values.values() for fid in fsv.keys()})
            block_gap = 1.5

            if len(feature_ids) > 0 and len(group_names) > 0:
                # keep each group's feature boxes tightly clustered while still visibly separated
                slot_width = ABUNDANCE_BOXPLOT_CLUSTER_WIDTH / max(1, len(feature_ids))
                box_width = slot_width * ABUNDANCE_BOXPLOT_SLOT_FILL_RATIO

                # Legend: one entry per group
                for group_index, group_name in enumerate(group_names):
                    gcolor = group_color_map.get(group_name, f"C{group_index % 10}")
                    legend_handles.append(patches.Patch(facecolor=gcolor, alpha=0.35, label=group_name))

                base_cursor = 0.0
                for block_label, block_suffix in blocks:
                    feature_sample_values = block_feature_values.get(block_suffix, {})
                    block_group_centers = []
                    for group_index, group_name in enumerate(group_names):
                        group_base = base_cursor + group_index
                        xtick_positions.append(group_base)
                        xtick_labels.append(group_name)
                        block_group_centers.append(group_base)
                        gcolor = group_color_map.get(group_name, f"C{group_index % 10}")
                        for feature_index, feature_id in enumerate(feature_ids):
                            if feature_id not in feature_sample_values:
                                continue
                            offset = (-ABUNDANCE_BOXPLOT_CLUSTER_WIDTH / 2.0) + (feature_index + 0.5) * slot_width
                            vals = []
                            for sample_name in samples_by_group.get(group_name, []):
                                key = (group_name, sample_name)
                                if key not in feature_sample_values[feature_id]:
                                    continue
                                value = feature_sample_values[feature_id][key]
                                if log_scale and value <= 0:
                                    continue
                                vals.append(value)
                            if vals:
                                pos = group_base + offset
                                box_data.append(vals)
                                positions.append(pos)
                                box_colors.append(gcolor)
                                # Overlay individual samples as evenly jittered dots
                                n_vals = len(vals)
                                for j, value in enumerate(vals):
                                    if n_vals > 1:
                                        jitter = (j / (n_vals - 1) - 0.5) * box_width * 0.6
                                    else:
                                        jitter = 0.0
                                    scatter_x.append(pos + jitter)
                                    scatter_y.append(value)
                                    scatter_colors.append(gcolor)
                    if block_label and block_group_centers:
                        block_label_positions.append((sum(block_group_centers) / len(block_group_centers), block_label))
                    base_cursor += len(group_names) + block_gap

            if box_data:
                bp = ax.boxplot(box_data, positions=positions, widths=box_width, patch_artist=True, showfliers=False)
                for patch, c in zip(bp["boxes"], box_colors):
                    patch.set_facecolor(c)
                    patch.set_alpha(0.35)
                if scatter_x:
                    ax.scatter(scatter_x, scatter_y, c=scatter_colors, s=18, edgecolors="black", linewidths=0.4, alpha=0.9, zorder=3)
                ax.set_xticks(xtick_positions)
                ax.set_xticklabels(xtick_labels, rotation=25, ha="right")
                ax.set_xlabel("Experimental group")
                # Isotopolog block labels above the plot and separators between blocks
                if block_label_positions:
                    xaxis_transform = ax.get_xaxis_transform()
                    for center_x, label in block_label_positions:
                        ax.text(center_x, 1.0, label, transform=xaxis_transform, ha="center", va="bottom", fontsize=9, fontweight="bold")
                    for sep_block in range(1, len(blocks)):
                        sep_x = sep_block * (len(group_names) + block_gap) - block_gap / 2.0 - 0.5
                        ax.axvline(sep_x, color="gray", linestyle="--", linewidth=0.8)
                if legend_handles:
                    ax.legend(handles=legend_handles, loc="best", title="Sample group")
            else:
                ax.text(0.5, 0.5, "No abundance values available", transform=ax.transAxes, ha="center", va="center")
        else:
            n_samples = len(sample_entries)
            sample_label = [f"{grp}|{sample}" for grp, sample in sample_entries]

            ax.set_xlabel("Sample")
            cmap = matplotlib.cm.get_cmap("tab10")
            feature_color = {fid: cmap(i % 10) for i, fid in enumerate(sorted(set(selected_ids)))}

            block_gap = 2.0
            xtick_positions = []
            xtick_labels = []
            block_label_positions = []
            block_boundaries = []
            base_cursor = 0

            for block_index, (block_label, block_suffix) in enumerate(blocks):
                feature_sample_values = block_feature_values.get(block_suffix, {})
                line_style = "--" if block_label == "Labeled" else "-"
                x = [base_cursor + i for i in range(n_samples)]

                for i in range(n_samples):
                    xtick_positions.append(base_cursor + i)
                    xtick_labels.append(sample_label[i])

                # Group-change separators within this block
                last_group = None
                for i, (grp, _) in enumerate(sample_entries):
                    if grp != last_group and i > 0:
                        ax.axvline(base_cursor + i - 0.5, color="lightgray", linestyle="--", linewidth=0.8)
                    last_group = grp

                for feature_id in sorted(feature_sample_values.keys()):
                    y_vals = []
                    for grp, sample in sample_entries:
                        key = (grp, sample)
                        if key not in feature_sample_values[feature_id]:
                            y_vals.append(float("nan"))
                            continue
                        val = feature_sample_values[feature_id][key]
                        if log_scale and val <= 0:
                            y_vals.append(float("nan"))
                        else:
                            y_vals.append(val)
                    label = f"Feature {feature_id}" + (f" ({block_label})" if block_label else "")
                    ax.plot(x, y_vals, marker="o", linewidth=1.2, linestyle=line_style, color=feature_color.get(feature_id), label=label)

                if block_label:
                    block_label_positions.append((base_cursor + (n_samples - 1) / 2.0, block_label))
                if block_index < len(blocks) - 1:
                    block_boundaries.append(base_cursor + n_samples - 0.5 + block_gap / 2.0)

                base_cursor += n_samples + block_gap

            for sep_x in block_boundaries:
                ax.axvline(sep_x, color="gray", linestyle="--", linewidth=0.8)

            ax.set_xticks(xtick_positions)
            ax.set_xticklabels(xtick_labels, rotation=45, ha="right")
            if block_label_positions:
                xaxis_transform = ax.get_xaxis_transform()
                for center_x, label in block_label_positions:
                    ax.text(center_x, 1.0, label, transform=xaxis_transform, ha="center", va="bottom", fontsize=9, fontweight="bold")
            ax.legend()

        if log_scale:
            ax.set_yscale("log")
        self.drawCanvas(self.ui.resultsExperimentAbundance_plot, showLegendOverwrite=False)

    def _refreshExperimentEICs(self, *args):
        """Re-draw experiment EICs when separation/normalisation controls change, but only if raw data is already loaded."""
        if hasattr(self, "loadedMZXMLs") and self.loadedMZXMLs is not None:
            self.resultsExperimentChanged()

    def _refreshExperimentMSMS(self, *args):
        """Refresh the MSMS spectra list when RT window or precursor intensity threshold changes."""
        if hasattr(self, "loadedMZXMLs") and self.loadedMZXMLs is not None:
            selectedItems = self.ui.resultsExperiment_TreeWidget.selectedItems()
            self.updateMSMSList_exp(selectedItems)

    def _refreshExperimentMSMS(self, *args):
        """Refresh the MSMS spectra list when RT window or precursor intensity threshold changes."""
        if hasattr(self, "loadedMZXMLs") and self.loadedMZXMLs is not None:
            selectedItems = self.ui.resultsExperiment_TreeWidget.selectedItems()
            self.updateMSMSList_exp(selectedItems)

    def _build_msms_feature_forms(self):
        """Build a map of feature_num -> set of MS2 forms (native/labeled)."""
        counts = self._build_msms_feature_counts()
        forms = {}
        for num, (n_native, n_labeled) in counts.items():
            s = set()
            if n_native > 0:
                s.add("native")
            if n_labeled > 0:
                s.add("labeled")
            if s:
                forms[num] = s
        return forms

    def _build_msms_feature_counts(self):
        """Build a map of feature_num -> [native_count, labeled_count] counting the
        matching MS2 spectra across all loaded files."""
        msms_feature_counts = {}

        if not hasattr(self, "loadedMZXMLs") or self.loadedMZXMLs is None:
            return msms_feature_counts

        # Show progress dialog
        pw = ProgressWrapper(1, parent=self, showIndProgress=False)
        pw.show()

        try:
            ppm = self.ui.doubleSpinBox_resultsExperiment_EICppm.value()
        except Exception:
            ppm = 5.0

        try:
            msms_rt_window = self.ui.doubleSpinBox_resultsExperiment_MSMSRTWindow.value()
        except Exception:
            msms_rt_window = 0.5

        try:
            prec_intens_percent = self.ui.doubleSpinBox_resultsExperiment_MSMSPrecIntensPercent.value() / 100.0
        except Exception:
            prec_intens_percent = 0.5

        try:
            abs_intens_threshold = self.ui.doubleSpinBox_resultsExperiment_MSMSAbsIntensThreshold.value()
        except Exception:
            abs_intens_threshold = 0.0

        file_keys = [k for k in self.loadedMZXMLs if k.lower().endswith(".mzxml") or k.lower().endswith(".mzml")]

        # Collect all features from the tree
        all_features = {}
        tree = self.ui.resultsExperiment_TreeWidget
        for i in range(tree.topLevelItemCount()):
            top = tree.topLevelItem(i)
            for c in range(top.childCount()):
                child = top.child(c)
                bd = getattr(child, "bunchData", None)
                if bd and hasattr(bd, "id"):
                    all_features[bd.id] = bd

        if not all_features:
            pw.close()
            return msms_feature_counts

        # Get feature data from DB for per-file RT bounds
        rows_by_num = {}
        if hasattr(self, "experimentResults") and self.experimentResults is not None and self.experimentResults.db_con is not None:
            for sheet in ["4_Convoluted", "3_Reintegrated", "1_Bracketed"]:
                try:
                    tdf = self.experimentResults.db_con.get_table(sheet)
                    if tdf is not None and not tdf.is_empty():
                        rows_by_num = {r["Num"]: r for r in tdf.to_dicts()}
                        break
                except Exception:
                    pass

        pw.getCallingFunction()("max")(sum(len(mzxml.MS2_list) if hasattr(mzxml, "MS2_list") else 0 for mzxml in self.loadedMZXMLs.values()))
        pw.getCallingFunction()("value")(0)

        total_checked = 0

        # For each file and MS2 scan, determine if it matches a feature
        for file_key in file_keys:
            mzxml_file = self.loadedMZXMLs[file_key]
            if not (hasattr(mzxml_file, "MS2_list") and len(mzxml_file.MS2_list) > 0):
                continue

            for ms2_scan in mzxml_file.MS2_list:
                total_checked += 1
                if total_checked % 50 == 0:
                    pw.getCallingFunction()("value")(total_checked)
                    QtWidgets.QApplication.processEvents()

                # Try to match this scan to any feature
                for feature_num, bd in all_features.items():
                    # Calculate m/z windows
                    native_mz_min = bd.mz * (1 - ppm / 1000000.0)
                    native_mz_max = bd.mz * (1 + ppm / 1000000.0)
                    labeled_mz_min = bd.lmz * (1 - ppm / 1000000.0) if bd.lmz is not None else None
                    labeled_mz_max = bd.lmz * (1 + ppm / 1000000.0) if bd.lmz is not None else None

                    # RT window around feature
                    rt_min_s = bd.rt - msms_rt_window * 60.0
                    rt_max_s = bd.rt + msms_rt_window * 60.0

                    # Check RT window
                    if not (rt_min_s <= ms2_scan.retention_time <= rt_max_s):
                        continue

                    # Check polarity matches the feature's ionisation mode
                    ion_mode = getattr(bd, "ionisationMode", None)
                    if ion_mode and ms2_scan.polarity and ms2_scan.polarity != ion_mode:
                        continue

                    # Check precursor m/z and determine form
                    form = None
                    if native_mz_min <= ms2_scan.precursor_mz <= native_mz_max:
                        form = "native"
                    elif labeled_mz_min is not None and labeled_mz_min <= ms2_scan.precursor_mz <= labeled_mz_max:
                        form = "labeled"

                    if form is None:
                        continue

                    # Check percent threshold if applicable
                    if prec_intens_percent > 0.0:
                        row_data = rows_by_num.get(feature_num)
                        if row_data is not None:
                            fname = os.path.basename(file_key)
                            for ext in [".mzxml", ".mzml"]:
                                if fname.lower().endswith(ext):
                                    fname = fname[: -len(ext)]
                                    break
                            try:
                                peak_rt_min = min(
                                    float(str(row_data.get(f"{fname}_N_startRT", "")).split(";")[0]) * 60.0 if row_data.get(f"{fname}_N_startRT") else float("inf"),
                                    float(str(row_data.get(f"{fname}_L_startRT", "")).split(";")[0]) * 60.0 if row_data.get(f"{fname}_L_startRT") else float("inf"),
                                )
                                peak_rt_max = max(
                                    float(str(row_data.get(f"{fname}_N_endRT", "")).split(";")[0]) * 60.0 if row_data.get(f"{fname}_N_endRT") else float("-inf"),
                                    float(str(row_data.get(f"{fname}_L_endRT", "")).split(";")[0]) * 60.0 if row_data.get(f"{fname}_L_endRT") else float("-inf"),
                                )
                                if peak_rt_min != float("inf") and peak_rt_max != float("-inf"):
                                    try:
                                        eic, times, _, _ = mzxml_file.getEIC(bd.mz, ppm=ppm, filterLine=None)
                                        peak_max = 0.0
                                        for intensity, t in zip(eic, times):
                                            if peak_rt_min <= t <= peak_rt_max:
                                                if intensity > peak_max:
                                                    peak_max = intensity
                                        if peak_max > 0.0 and ms2_scan.precursor_intensity < prec_intens_percent * peak_max:
                                            continue
                                    except Exception:
                                        pass
                            except (TypeError, ValueError):
                                pass

                    # Check absolute threshold
                    if abs_intens_threshold > 0.0 and ms2_scan.precursor_intensity < abs_intens_threshold:
                        continue

                    # This scan matches the feature
                    if feature_num not in msms_feature_counts:
                        msms_feature_counts[feature_num] = [0, 0]
                    if form == "native":
                        msms_feature_counts[feature_num][0] += 1
                    else:
                        msms_feature_counts[feature_num][1] += 1
                    break

        pw.close()
        return msms_feature_counts

    def _updateExperimentMSMSCountsColumn(self):
        """Fill the 'MS2 N/L' column of the experiment feature tree with the number of
        native/labeled MS2 spectra per feature and prefix the feature id with '*' when
        at least one MS2 spectrum exists."""
        tree = self.ui.resultsExperiment_TreeWidget
        if tree.columnCount() < 7:
            return
        if not hasattr(self, "loadedMZXMLs") or self.loadedMZXMLs is None:
            return

        counts = self._build_msms_feature_counts()

        for i in range(tree.topLevelItemCount()):
            top = tree.topLevelItem(i)
            for c in range(top.childCount()):
                child = top.child(c)
                bd = getattr(child, "bunchData", None)
                if bd is None or not hasattr(bd, "id"):
                    continue
                n_native, n_labeled = counts.get(bd.id, (0, 0))
                has_msms = (n_native > 0) or (n_labeled > 0)
                n_txt = str(n_native) if n_native > 0 else ""
                l_txt = str(n_labeled) if n_labeled > 0 else ""
                child.setText(6, f"{n_txt}/{l_txt}" if has_msms else "")

                title = child.text(0)
                if has_msms and not title.startswith("*"):
                    child.setText(0, "*" + title)
                elif not has_msms and title.startswith("*"):
                    child.setText(0, title[1:])

    def _updateMSMSTreeClicked(self):
        """Handler for the 'Update Tree' button in the Show Options MS/MS group: re-evaluates
        the current MS/MS filters (RT window, intensity thresholds, filter-string regex,
        polarity) against all loaded raw data and refreshes the tree's MS/MS indicator."""
        if not hasattr(self, "loadedMZXMLs") or self.loadedMZXMLs is None:
            QtWidgets.QMessageBox.information(self, "Update Tree", "No raw MS/MS data loaded.")
            return
        QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.WaitCursor)
        try:
            self._updateExperimentMSMSCountsColumn()
        except Exception:
            logging.exception("Could not update MS2 spectra counts column")
        finally:
            QtWidgets.QApplication.restoreOverrideCursor()

    def _refreshExperimentAbundancePlot(self, *args):
        self.updateExperimentAbundancePlot(self._getSelectedExperimentPlotItems())

    def updateSamplePeaksTab(self, plotItems):
        """Prepare data for per-sample peak plots and render the first page."""
        # Cache data for pagination
        self._samplePeaks_plotItems = list(plotItems) if plotItems else []
        self._samplePeaks_page = 0

        # Pre-build sorted sample list once
        self._samplePeaks_entries = []
        if plotItems and hasattr(self, "loadedMZXMLs") and self.loadedMZXMLs is not None:
            all_groups = self.getAllSampleGroups()
            group_order = {str(grp.name): i for i, grp in enumerate(all_groups)}
            seen = set()
            raw = []
            for grp in all_groups:
                group_name = str(grp.name)
                for fi in grp.files:
                    fi_str = str(fi).replace("\\", "/")
                    a = fi_str[fi_str.rfind("/") + 1 :]
                    for ext in (".mzXML", ".mzxml", ".mzML", ".mzml"):
                        if a.lower().endswith(ext.lower()):
                            a = a[: -len(ext)]
                            break
                    key = (group_name, a)
                    if key not in seen:
                        seen.add(key)
                        raw.append((group_name, a, fi_str))
            self._samplePeaks_entries = sorted(raw, key=lambda x: (group_order.get(x[0], 0), x[1].lower()))

        self._renderSamplePeaksPage()

    def _renderSamplePeaksPage(self):
        """Render the current page of per-sample peak subplots."""
        try:
            _PAGE_ROWS = int(self.ui.spinBox_samplePeaksRows.value())
        except Exception:
            _PAGE_ROWS = 4
        try:
            _PAGE_COLS = int(self.ui.spinBox_samplePeaksCols.value())
        except Exception:
            _PAGE_COLS = 4
        _PAGE_ROWS = max(1, _PAGE_ROWS)
        _PAGE_COLS = max(1, _PAGE_COLS)
        _PAGE_SIZE = _PAGE_ROWS * _PAGE_COLS

        try:
            norm_mode = self.ui.comboBox_samplePeaksNorm.currentText()
        except Exception:
            norm_mode = "Normalize to peak apex"

        try:
            labeled_negative = self.ui.checkBox_samplePeaksLabeledNegative.isChecked()
        except Exception:
            labeled_negative = False
        try:
            labeled_norm_separately = self.ui.checkBox_samplePeaksLabeledNormSeparately.isChecked()
        except Exception:
            labeled_norm_separately = False

        fig = self.ui.resultsExperimentSamplePeaks_plot.fig
        fig.clear()
        canvas = self.ui.resultsExperimentSamplePeaks_plot.canvas

        sample_entries = getattr(self, "_samplePeaks_entries", [])
        plotItems = getattr(self, "_samplePeaks_plotItems", [])
        page = getattr(self, "_samplePeaks_page", 0)

        n_total = len(sample_entries)
        n_pages = max(1, (n_total + _PAGE_SIZE - 1) // _PAGE_SIZE)
        page = max(0, min(page, n_pages - 1))
        self._samplePeaks_page = page

        # Update navigation buttons / label
        prev_btn = self.ui.pushButton_samplePeaksPrev
        next_btn = self.ui.pushButton_samplePeaksNext
        page_lbl = self.ui.label_samplePeaksPage
        prev_btn.setEnabled(page > 0)
        next_btn.setEnabled(page < n_pages - 1)
        if n_total > 0:
            page_lbl.setText(f"Page {page + 1} / {n_pages}  ({n_total} samples)")
        else:
            page_lbl.setText("")

        if not plotItems or n_total == 0:
            canvas.draw()
            return

        if not hasattr(self, "experimentResults") or self.experimentResults is None:
            canvas.draw()
            return

        selected_table = getattr(self.experimentResults, "selected_table", None)
        if selected_table is None or selected_table not in self.experimentResults.db_con.tables:
            canvas.draw()
            return

        table_df = self.experimentResults.db_con.tables[selected_table]
        feature_ids = [pi.id for pi in plotItems if getattr(pi, "id", None) is not None]
        rows_by_num = {}
        if feature_ids:
            filtered = table_df.filter(pl.col("Num").is_in(feature_ids))
            rows_by_num = {row["Num"]: row for row in filtered.to_dicts()}

        ppm = self.ui.doubleSpinBox_resultsExperiment_EICppm.value()

        # Slice the current page
        page_start = page * _PAGE_SIZE
        page_entries = sample_entries[page_start : page_start + _PAGE_SIZE]
        n_on_page = len(page_entries)

        n_cols = min(_PAGE_COLS, n_on_page)
        n_rows = (n_on_page + n_cols - 1) // n_cols

        fig_width = n_cols * 4.5
        fig_height = n_rows * 3.8
        fig.set_size_inches(fig_width, fig_height)

        # Reserve a fixed absolute band (~0.55 in) at the top for the legend so the
        # gap stays small regardless of how many rows are shown.
        legend_fraction = min(0.4, 0.55 / fig_height)

        axes = fig.subplots(n_rows, n_cols, squeeze=False)

        # Build colour map per feature
        cmap = matplotlib.cm.get_cmap("tab10")
        feature_colors = {pi.id: cmap(i % 10) for i, pi in enumerate(plotItems)}

        # Legend handles
        import matplotlib.lines as mlines

        legend_handles = [mlines.Line2D([], [], color=feature_colors[pi.id], linewidth=1.5, label=f"Feature {pi.id} ({pi.mz:.4f})") for pi in plotItems]
        # Native (solid) vs labeled (dashed) line-style legend
        legend_handles.append(mlines.Line2D([], [], color="gray", linewidth=1.5, linestyle="-", label="Native"))
        legend_handles.append(mlines.Line2D([], [], color="gray", linewidth=1.5, linestyle="--", label="Labeled"))

        for s_idx, (group_name, sample_name, fi_str) in enumerate(page_entries):
            row = s_idx // n_cols
            col = s_idx % n_cols
            ax = axes[row][col]
            ax.set_title(f"{group_name}\n{sample_name}", fontsize=8, pad=3)
            ax.tick_params(labelsize=7)
            ax.set_xlabel("RT (min)", fontsize=7)
            ax.set_ylabel("Intensity" if norm_mode == "None" else "Scaled intensity", fontsize=7)
            if labeled_negative:
                ax.axhline(0.0, color="gray", linewidth=0.6, alpha=0.7)

            if not hasattr(self, "loadedMZXMLs") or self.loadedMZXMLs is None or fi_str not in self.loadedMZXMLs:
                ax.text(0.5, 0.5, "Not loaded", ha="center", va="center", transform=ax.transAxes, fontsize=7, color="gray")
                continue

            mzxml_obj = self.loadedMZXMLs[fi_str]

            for pi in plotItems:
                color = feature_colors[pi.id]
                row_data = rows_by_num.get(pi.id, {})

                # Determine scan event (shared by native and labeled forms)
                scanEvent = pi.scanEvent if hasattr(pi, "scanEvent") and pi.scanEvent else None
                if scanEvent is None:
                    filter_lines = mzxml_obj.getFilterLines(includeMS1=True, includeMS2=False, includePosPolarity=True, includeNegPolarity=True)
                    if filter_lines:
                        ion_mode = getattr(pi, "ionisationMode", None)
                        if ion_mode and "+" in str(ion_mode):
                            scanEvent = next((fl for fl in filter_lines if "+" in fl), filter_lines[0])
                        elif ion_mode and "-" in str(ion_mode):
                            scanEvent = next((fl for fl in filter_lines if "-" in fl), filter_lines[0])
                        else:
                            scanEvent = filter_lines[0]

                if scanEvent is None:
                    continue

                available = mzxml_obj.getFilterLines(includeMS1=True, includeMS2=False, includePosPolarity=True, includeNegPolarity=True)
                if scanEvent not in available:
                    continue

                # Native and labeled forms: (DB tag, m/z, line style)
                forms = [("N", pi.mz, "-")]
                if getattr(pi, "lmz", None):
                    forms.append(("L", pi.lmz, "--"))

                native_scale = None
                for form_tag, form_mz, form_ls in forms:
                    if form_mz is None:
                        continue

                    start_rt_min = None
                    apex_rt_min = None
                    end_rt_min = None

                    if row_data:
                        sv = row_data.get(f"{sample_name}_{form_tag}_startRT")
                        av = row_data.get(f"{sample_name}_{form_tag}_apexRT")
                        ev = row_data.get(f"{sample_name}_{form_tag}_endRT")
                        if sv is not None and av is not None and ev is not None:
                            try:
                                # DB columns store RT already in minutes (seconds / 60 at write time)
                                # Use first value when multiple peaks are stored as semicolon list
                                start_rt_min = float(str(sv).split(";")[0])
                                apex_rt_min = float(str(av).split(";")[0])
                                end_rt_min = float(str(ev).split(";")[0])
                            except (TypeError, ValueError):
                                start_rt_min = None

                    if start_rt_min is None or end_rt_min is None:
                        global_rt_min = pi.rt / 60.0
                        start_rt_min = global_rt_min - 0.3
                        apex_rt_min = global_rt_min
                        end_rt_min = global_rt_min + 0.3

                    peak_width = max(end_rt_min - start_rt_min, 1e-6)
                    context = 0.75 * peak_width
                    window_start = start_rt_min - context
                    window_end = end_rt_min + context

                    try:
                        eic, times, _scan_ids, _mzs = mzxml_obj.getEIC(form_mz, ppm=ppm, filterLine=scanEvent)
                    except Exception:
                        continue

                    if len(times) == 0:
                        continue

                    times_min = [t / 60.0 for t in times]
                    eic_list = list(eic)

                    # Crop to view window
                    window_indices = [j for j, t in enumerate(times_min) if window_start <= t <= window_end]
                    if not window_indices:
                        continue

                    cropped_times = [times_min[j] for j in window_indices]
                    cropped_eic = [eic_list[j] for j in window_indices]

                    # Normalisation according to the selected mode
                    if norm_mode == "None":
                        scale = 1.0
                    elif norm_mode == "Normalize to maximum":
                        local_max = max(cropped_eic) if cropped_eic else 0.0
                        scale = (1.0 / local_max) if local_max > 0 else 1.0
                    else:  # "Normalize to peak apex"
                        apex_idx = min(range(len(times_min)), key=lambda j: abs(times_min[j] - apex_rt_min))
                        apex_int = eic_list[apex_idx]
                        scale = (1.0 / apex_int) if apex_int > 0 else 1.0

                    if form_tag == "N":
                        native_scale = scale
                    elif form_tag == "L" and not labeled_norm_separately and native_scale is not None:
                        # Keep labeled on the native scale so the abundance ratio is
                        # preserved unless the user requests separate normalisation.
                        scale = native_scale

                    # Optionally mirror the labeled trace below the x-axis
                    sign = -1.0 if (form_tag == "L" and labeled_negative) else 1.0
                    scaled_eic = [sign * v * scale for v in cropped_eic]
                    ax.plot(cropped_times, scaled_eic, color=color, linewidth=1.0, alpha=0.85, linestyle=form_ls)

                    # Shade the peak region
                    boundary = [(t, v) for t, v in zip(cropped_times, scaled_eic) if start_rt_min <= t <= end_rt_min]
                    if boundary:
                        ax.fill_between([x[0] for x in boundary], [x[1] for x in boundary], alpha=0.12, color=color)

        # Hide unused subplots in last row
        for s_idx in range(n_on_page, n_rows * n_cols):
            axes[s_idx // n_cols][s_idx % n_cols].set_visible(False)

        if legend_handles:
            fig.legend(
                handles=legend_handles,
                loc="upper center",
                bbox_to_anchor=(0.5, 1.0),
                ncol=max(1, min(len(legend_handles), n_cols)),
                fontsize=8,
                frameon=True,
                title="Features",
                title_fontsize=8,
            )

        fig.tight_layout(rect=[0, 0, 1.0, 1.0 - legend_fraction])

        dpi = self.ui.resultsExperimentSamplePeaks_plot.dpi
        w_px = int(fig_width * dpi)
        h_px = int(fig_height * dpi)
        canvas.setMinimumSize(w_px, h_px)
        self.ui.resultsExperimentSamplePeaks_widget.setMinimumSize(w_px, h_px + 40)

        canvas.draw()

    def _samplePeaksPrevPage(self):
        self._samplePeaks_page = max(0, getattr(self, "_samplePeaks_page", 0) - 1)
        self._renderSamplePeaksPage()

    def _samplePeaksNextPage(self):
        self._samplePeaks_page = getattr(self, "_samplePeaks_page", 0) + 1
        self._renderSamplePeaksPage()

    # </editor-fold>

    # <editor-fold desc="### load/save settings functions">
    def loadSettings(self):
        settingsFile = QtWidgets.QFileDialog.getOpenFileName(
            caption="Select settings file",
            dir=self.lastOpenDir,
            filter="Setting file (*.ini);;Group file with settings (*.grp);;All files(*.*)",
        )
        if len(settingsFile) > 0:
            self.lastOpenDir = str(settingsFile).replace("\\", "/")
            self.lastOpenDir = self.lastOpenDir[: self.lastOpenDir.rfind("/")]
        self.loadSettingsFile(settingsFile)

    # load settings from key-value pair file
    def loadSettingsFile(self, settingsFile, checkExperimentType=True, silent=False):
        if len(settingsFile) > 0:
            sett = QtCore.QSettings(settingsFile, QtCore.QSettings.IniFormat)

            sett.beginGroup("MetExtract")

            if sett.contains("Version"):
                version = str(sett.value("Version"))
                if (
                    not (silent)
                    and version != MetExtractVersion
                    and QtWidgets.QMessageBox.question(
                        self,
                        "MetExtract",
                        "Settings were created from a different MetExtract II version. Some settings may not be imported correctly.\nDo you want to continue?",
                        QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    )
                    == QtWidgets.QMessageBox.No
                ):
                    return

            sett.endGroup()

            sett.beginGroup("WorkingDirectory")

            if sett.contains("workingDirectory"):
                try:
                    os.chdir(str(sett.value("workingDirectory")))
                    LoggingSetup.LoggingSetup.Instance().initLogging(location=str(sett.value("workingDirectory")))
                    logging.info("Working directory changed to '%s'" % str(sett.value("workingDirectory")))
                    if not (silent):
                        QtWidgets.QMessageBox.information(
                            self,
                            "MetExtract",
                            "The current working directory was changed to\n'%s'" % str(sett.value("workingDirectory")),
                            QtWidgets.QMessageBox.Ok,
                        )
                except WindowsError:
                    QtWidgets.QMessageBox.information(
                        self,
                        "MetExtract",
                        "The working directory cannot be changed. The specified path does not exists.\nPlease set the directory manually",
                        QtWidgets.QMessageBox.Ok,
                    )

            sett.endGroup()

            sett.beginGroup("Settings")

            if sett.contains("processIndividualFiles"):
                self.ui.processIndividualFiles.setChecked(self.to_bool(sett.value("processIndividualFiles")))

            if checkExperimentType:
                if sett.contains("TracerExperiment"):
                    if self.to_bool(sett.value("TracerExperiment")) and self.labellingExperiment == METABOLOME:
                        QtWidgets.QMessageBox.warning(
                            self,
                            "MetExtract",
                            "Warning: You are trying to open a TracExtract experiment with AllExtract. Labelling parameters will not be loaded. Please switch to TracExtract",
                            QtWidgets.QMessageBox.Ok,
                        )
                    elif not (self.to_bool(sett.value("TracerExperiment"))) and self.labellingExperiment == TRACER:
                        QtWidgets.QMessageBox.warning(
                            self,
                            "MetExtract",
                            "Warning: You are trying to open an AllExtract experiment with TracExtract. Labelling parameters will not be loaded. Please switch to AllExtract",
                            QtWidgets.QMessageBox.Ok,
                        )
                else:
                    QtWidgets.QMessageBox.warning(
                        self,
                        "MetExtract",
                        "Warning: Experiment type is not stored in the settings. Please ensure that the correct module has been loaded for your experiment",
                        QtWidgets.QMessageBox.Ok,
                    )

            if sett.contains("tracerConfiguration") and self.labellingExperiment == TRACER:
                try:
                    self.configuredTracer = loads(base64.b64decode(sett.value("tracerConfiguration").encode("utf-8")))[0]
                except (ImportError, ModuleNotFoundError, SyntaxError) as e:
                    logging.warning(f"Could not load tracer configuration from settings file: {e}")
                    logging.warning("This may be due to cached data from an older version. Configuration will be skipped.")
                    self.configuredTracer = None

            if sett.contains("LabellingElementA") and self.labellingExperiment == METABOLOME:
                self.ui.isotopeAText.setText(str(sett.value("LabellingElementA")))
            if sett.contains("IsotopicAbundanceA") and self.labellingExperiment == METABOLOME:
                self.ui.isotopicAbundanceA.setValue(self.to_double(sett.value("IsotopicAbundanceA")) * 100.0)
            if sett.contains("LabellingElementB") and self.labellingExperiment == METABOLOME:
                self.ui.isotopeBText.setText(str(sett.value("LabellingElementB")))
            if sett.contains("IsotopicAbundanceB") and self.labellingExperiment == METABOLOME:
                self.ui.isotopicAbundanceB.setValue(self.to_double(sett.value("IsotopicAbundanceB")) * 100.0)
            if sett.contains("useCValidation"):
                self.ui.useCValidation.setCheckState(QtCore.Qt.CheckState(self.to_int(sett.value("useCValidation"))))
            if sett.contains("useRatio") and self.labellingExperiment == METABOLOME:
                self.ui.useRatio.setChecked(self.to_bool(sett.value("useRatio")))
            if sett.contains("minRatio") and self.labellingExperiment == METABOLOME:
                self.ui.minRatio.setValue(self.to_double(sett.value("minRatio")))
            if sett.contains("maxRatio") and self.labellingExperiment == METABOLOME:
                self.ui.maxRatio.setValue(self.to_double(sett.value("maxRatio")))

            if sett.contains("IntensityThreshold"):
                self.ui.intensityThreshold.setValue(self.to_int(sett.value("IntensityThreshold")))
            if sett.contains("IntensityCutoff"):
                self.ui.intensityCutoff.setValue(self.to_int(sett.value("IntensityCutoff")))
            if sett.contains("intensityThresholdIsotopologs"):
                self.ui.intensityThresholdIsotopologs.setValue(self.to_int(sett.value("intensityThresholdIsotopologs")))

            if sett.contains("ScanStart"):
                self.ui.scanStartTime.setValue(self.to_double(sett.value("ScanStart")))
            if sett.contains("ScanEnd"):
                self.ui.scanEndTime.setValue(self.to_double(sett.value("ScanEnd")))

            if sett.contains("xCounts"):
                self.ui.xCountSearch.setText(str(sett.value("xCounts")))
            else:
                if self.labellingExperiment == TRACER:
                    self.ui.xCountSearch.setText("10-15, 30, 45")
                elif self.labellingExperiment == METABOLOME:
                    self.ui.xCountSearch.setText("3-60")

            if sett.contains("MaxLoading"):
                self.ui.maxLoading.setValue(self.to_int(sett.value("MaxLoading")))
            if sett.contains("MaxMassDeviation"):
                self.ui.ppmRangeIdentification.setValue(self.to_double(sett.value("MaxMassDeviation")))
            if sett.contains("IsotopicPatternCountA"):
                self.ui.isotopePatternCountA.setValue(self.to_int(sett.value("IsotopicPatternCountA")))
            if sett.contains("IsotopicPatternCountB"):
                self.ui.isotopePatternCountB.setValue(self.to_int(sett.value("IsotopicPatternCountB")))
            if sett.contains("ScanIndexOffset"):
                self.ui.scanIndexOffset.setValue(self.to_int(sett.value("ScanIndexOffset")))
            if sett.contains("lowAbundanceIsotopeCutoff"):
                self.ui.isoAbundance.setChecked(self.to_bool(sett.value("lowAbundanceIsotopeCutoff")))
            if sett.contains("IntensityAbundanceErrorA"):
                self.ui.baseRange.setValue(self.to_double(sett.value("IntensityAbundanceErrorA")) * 100.0)
            if sett.contains("IntensityAbundanceErrorB"):
                self.ui.isotopeRange.setValue(self.to_double(sett.value("IntensityAbundanceErrorB")) * 100.0)

            if sett.contains("ClustPPM"):
                self.ui.clustPPM.setValue(self.to_double(sett.value("ClustPPM")))
            if sett.contains("minSpectraCount"):
                self.ui.minSpectraCount.setValue(self.to_int(sett.value("minSpectraCount")))
            if sett.contains("correctCCount"):
                self.ui.correctcCount.setChecked(self.to_bool(sett.value("correctcCount")))

            if sett.contains("EICppm"):
                self.ui.wavelet_EICppm.setValue(self.to_double(sett.value("EICppm")))
            if sett.contains("EICSmoothingWindow"):
                eicSmoothingWindow = str(sett.value("EICSmoothingWindow"))
                pos = self.ui.eicSmoothingWindow.findText(eicSmoothingWindow)
                if pos >= 0:
                    self.ui.eicSmoothingWindow.setCurrentIndex(pos)
                else:
                    QtWidgets.QMessageBox.information(
                        self,
                        "MetExtract",
                        "Invalid EIC smoothing option",
                        QtWidgets.QMessageBox.Ok,
                    )
                    self.ui.eicSmoothingWindow.setCurrentIndex(0)
            if sett.contains("EICSmoothingWindowSize"):
                self.ui.eicSmoothingWindowSize.setValue(self.to_int(sett.value("EICSmoothingWindowSize")))
            if sett.contains("smoothingPolynom_spinner"):
                self.ui.smoothingPolynom_spinner.setValue(self.to_int(sett.value("smoothingPolynom_spinner")))
            if sett.contains("artificialMPshift_start"):
                self.ui.spinBox_artificialMPshift_start.setValue(self.to_int(sett.value("artificialMPshift_start")))
            if sett.contains("artificialMPshift_stop"):
                self.ui.spinBox_artificialMPshift_stop.setValue(self.to_int(sett.value("artificialMPshift_stop")))

            if sett.contains("Wavelet_MinScale"):
                self.ui.wavelet_minScale.setValue(self.to_int(sett.value("Wavelet_MinScale")))
            if sett.contains("Wavelet_MaxScale"):
                self.ui.wavelet_maxScale.setValue(self.to_int(sett.value("Wavelet_MaxScale")))
            if sett.contains("Wavelet_SNRTh"):
                self.ui.wavelet_SNRThreshold.setValue(self.to_int(sett.value("Wavelet_SNRTh")))
            if sett.contains("Peak_CenterError"):
                self.ui.peak_centerError.setValue(self.to_int(sett.value("Peak_CenterError")))
            if sett.contains("Peak_WidthError"):
                self.ui.peak_scaleError.setValue(self.to_int(sett.value("Peak_WidthError")))
            if sett.contains("Peak_minPeakCorr"):
                self.ui.minPeakCorr.setValue(self.to_double(sett.value("Peak_minPeakCorr")) * 100.0)
            if sett.contains("checkBox_checkPeakRatio"):
                self.ui.checkBox_checkPeakRatio.setChecked(self.to_bool(sett.value("checkBox_checkPeakRatio")))
            if sett.contains("doubleSpinBox_minPeakRatio"):
                self.ui.doubleSpinBox_minPeakRatio.setValue(self.to_double(sett.value("doubleSpinBox_minPeakRatio")))
            if sett.contains("doubleSpinBox_maxPeakRatio"):
                self.ui.doubleSpinBox_maxPeakRatio.setValue(self.to_double(sett.value("doubleSpinBox_maxPeakRatio")))

            # Peak picking settings (algorithm + post-processing filters)
            if sett.contains("peakPickingSettings"):
                try:
                    self._peakPickingSettings = json.loads(sett.value("peakPickingSettings"))
                except Exception:
                    pass

            if sett.contains("peakPickingEicRows"):
                try:
                    raw = json.loads(sett.value("peakPickingEicRows"))
                    # Each entry is [file_path, filter_text, mz, ppm]
                    self._peakPickingEicRows = [tuple(row) for row in raw]
                except Exception:
                    pass

            if sett.contains("calcIsoRatioNative"):
                self.ui.calcIsoRatioNative_spinBox.setValue(self.to_int(sett.value("calcIsoRatioNative")))
            if sett.contains("calcIsoRatioLabelled"):
                self.ui.calcIsoRatioLabelled_spinBox.setValue(self.to_int(sett.value("calcIsoRatioLabelled")))
            if sett.contains("calcIsoRatioMoiety"):
                self.ui.calcIsoRatioMoiety_spinBox.setValue(self.to_int(sett.value("calcIsoRatioMoiety")))

            if sett.contains("hAIntensityError"):
                self.ui.hAIntensityError.setValue(self.to_double(sett.value("hAIntensityError")) * 100.0)
            if sett.contains("hAMinScans"):
                self.ui.hAMinScans.setValue(self.to_int(sett.value("hAMinScans")))
            if sett.contains("heteroAtoms") or sett.contains("heteroElements"):
                try:
                    if sett.contains("heteroAtoms"):
                        self.heteroElements = loads(base64.b64decode(sett.value("heteroAtoms").encode("utf-8")))
                    elif sett.contains("heteroElements"):
                        self.heteroElements = loads(base64.b64decode(sett.value("heteroElements").encode("utf-8")))
                except (ImportError, ModuleNotFoundError, SyntaxError) as e:
                    logging.warning(f"Could not load hetero elements from settings: {e}")
                    logging.warning("This may be due to cached data from an older version. Using default elements.")
                    self.heteroElements = []
                if any([not (isinstance(d, ConfiguredHeteroAtom)) for d in self.heteroElements]):
                    QtWidgets.QMessageBox.warning(
                        self,
                        "MetExtract",
                        "Could not parse hetero elements. Predefined elements will be used instead. Please adapt the list accordingly and save the compilation again.",
                    )
                    self.heteroElements = self.preConfigured_heteroElements

            if sett.contains("saveTSV"):
                self.ui.saveCSV.setChecked(self.to_bool(sett.value("saveTSV")))
            if sett.contains("saveFeatureML"):
                self.ui.saveFeatureML.setChecked(self.to_bool(sett.value("saveFeatureML")))
            if sett.contains("saveMZXML"):
                self.ui.saveMZXML.setChecked(self.to_bool(sett.value("saveMZXML")))
            if sett.contains("writeMZXMLOptions"):
                writeMZXMLOptions = self.to_int(sett.value("writeMZXMLOptions"))
                self.ui.wm_ia.setChecked(writeMZXMLOptions & 1)
                self.ui.wm_iap.setChecked(writeMZXMLOptions & 2)
                self.ui.wm_imb.setChecked(writeMZXMLOptions & 4)
                self.ui.wm_ib.setChecked(writeMZXMLOptions & 8)
            if sett.contains("minCorrelation"):
                self.ui.minCorrelation.setValue(self.to_double(sett.value("minCorrelation")) * 100.0)
            if sett.contains("minCorrelationConnections"):
                self.ui.minCorrelationConnections.setValue(self.to_double(sett.value("minCorrelationConnections")) * 100.0)
            if sett.contains("adducts"):
                try:
                    self.adducts = loads(base64.b64decode(sett.value("adducts").encode("utf-8")))
                except (ImportError, ModuleNotFoundError, SyntaxError) as e:
                    logging.warning(f"Could not load adducts from settings: {e}")
                    logging.warning("This may be due to cached data from an older version. Using default adducts.")
                    self.adducts = []
                if any([not (isinstance(d, ConfiguredAdduct)) for d in self.adducts]):
                    QtWidgets.QMessageBox.warning(
                        self,
                        "MetExtract",
                        "Could not parse adducts for ion annotation. Predefined adducts will be used instead. Please adapt the list accordingly and save the compilation again.",
                    )
                    self.adducts = self.preConfigured_adducts

            if sett.contains("elements") or sett.contains("elementsForNL"):
                try:
                    if sett.contains("elements"):
                        self.elementsForNL = loads(base64.b64decode(sett.value("elements").encode("utf-8")))
                    elif sett.contains("elementsForNL"):
                        self.elementsForNL = loads(base64.b64decode(sett.value("elementsForNL").encode("utf-8")))
                except (ImportError, ModuleNotFoundError, SyntaxError) as e:
                    logging.warning(f"Could not load elements for neutral losses from settings: {e}")
                    logging.warning("This may be due to cached data from an older version. Using default elements.")
                    self.elementsForNL = []
                if any([not (isinstance(d, ConfiguredElement)) for d in self.elementsForNL]):
                    QtWidgets.QMessageBox.warning(
                        self,
                        "MetExtract",
                        "Could not parse elements for neutral loss calculation. Predefined elements will be used instead. Please adapt the list accordingly and save the compilation again.",
                    )
                    self.elementsForNL = self.preConfigured_elementsForNL
            if sett.contains("simplifyInSourceFragments"):
                self.ui.checkBox_simplifyInSourceFragments.setChecked(self.to_bool(sett.value("simplifyInSourceFragments")))

            if sett.contains("processMultipleFiles"):
                self.ui.processMultipleFiles.setChecked(self.to_bool(sett.value("processMultipleFiles")))

            if sett.contains("GroupFeatures"):
                self.ui.groupResults.setChecked(self.to_bool(sett.value("GroupFeatures")))
            if sett.contains("GroupPPM"):
                self.ui.groupPpm.setValue(self.to_double(sett.value("GroupPPM")))
            if sett.contains("GroupAlign"):
                self.ui.alignChromatograms.setChecked(self.to_bool(sett.value("GroupAlign")))
            if sett.contains("GroupNPolynom"):
                self.ui.polynomValue.setValue(self.to_int(sett.value("GroupNPolynom")))
            if sett.contains("GroupTimeWindow"):
                self.ui.groupingRT.setValue(self.to_double(sett.value("GroupTimeWindow")))

            if sett.contains("ConvoluteMetaboliteGroups"):
                self.ui.convoluteResults.setChecked(self.to_bool(sett.value("ConvoluteMetaboliteGroups")))
            if sett.contains("maxAnnotationTimeWindow"):
                self.ui.maxAnnotationTimeWindow.setValue(self.to_double(sett.value("maxAnnotationTimeWindow")))
            if sett.contains("MetaboliteClusterMinConnections"):
                self.ui.metaboliteClusterMinConnections.setValue(self.to_int(sett.value("MetaboliteClusterMinConnections")))
            if sett.contains("useSILRatioForConvolution"):
                self.ui.useSILRatioForConvolution.setChecked(self.to_bool(sett.value("useSILRatioForConvolution")))
            if sett.contains("minConnectionRate"):
                self.ui.minConnectionRate.setValue(self.to_double(sett.value("minConnectionRate")) * 100.0)
            if sett.contains("useAbundanceSimilarityForConvolution"):
                self.ui.useAbundanceSimilarityForConvolution.setChecked(self.to_bool(sett.value("useAbundanceSimilarityForConvolution")))
            if sett.contains("abundanceSimilarityThreshold"):
                self.ui.abundanceSimilarityThreshold.setValue(self.to_double(sett.value("abundanceSimilarityThreshold")) * 100.0)

            if sett.contains("GroupIntegrateMissedPeaks"):
                self.ui.integratedMissedPeaks.setChecked(self.to_bool(sett.value("GroupIntegrateMissedPeaks")))
            if sett.contains("integrationMaxTimeDifference"):
                self.ui.integrationMaxTimeDifference.setValue(self.to_double(sett.value("integrationMaxTimeDifference")))
            if sett.contains("reintegrateIntensityCutoff"):
                self.ui.reintegrateIntensityCutoff.setValue(self.to_double(sett.value("reintegrateIntensityCutoff")))

            if sett.contains("annotateMetabolites"):
                self.ui.annotateMetabolites_CheckBox.setChecked(self.to_bool(sett.value("annotateMetabolites")))
            if sett.contains("annotateMetabolites_generateSumFormulas"):
                self.ui.generateSumFormulas_CheckBox.setChecked(self.to_bool(sett.value("annotateMetabolites_generateSumFormulas")))
            if sett.contains("annotateMetabolites_sumFormulasMinimumElements"):
                self.ui.sumFormulasMinimumElements_lineEdit.setText(str(sett.value("annotateMetabolites_sumFormulasMinimumElements")))
            if sett.contains("annotateMetabolites_sumFormulasMaximumElements"):
                self.ui.sumFormulasMaximumElements_lineEdit.setText(str(sett.value("annotateMetabolites_sumFormulasMaximumElements")))
            if sett.contains("annotateMetabolites_sumFormulasUseExactXn"):
                te = str(sett.value("annotateMetabolites_sumFormulasUseExactXn"))
                opts = {"Exact": 0, "Don't use": 1, "Minimum": 2, "PlusMinus": 3}
                self.ui.sumFormulasUseExactXn_ComboBox.setCurrentIndex(opts[te])
            if sett.contains("annotateMetabolites_plusMinus"):
                self.ui.sumFormulasPlusMinus_spinBox.setValue(self.to_int(sett.value("annotateMetabolites_plusMinus")))
            if sett.contains("annotateMetabolites_checkRT"):
                self.ui.checkRTInHits_checkBox.setChecked(self.to_bool(sett.value("annotateMetabolites_checkRT")))
            if sett.contains("annotateMetabolites_maxRTError"):
                self.ui.maxRTErrorInHits_spinnerBox.setValue(self.to_double(sett.value("annotateMetabolites_maxRTError")))
            if sett.contains("annotateMetabolites_maxPPM"):
                self.ui.annotationMaxPPM_doubleSpinBox.setValue(self.to_double(sett.value("annotateMetabolites_maxPPM")))
            if sett.contains("annotation_correctMassByPPM"):
                self.ui.annotation_correctMassByPPMposMode.setValue(self.to_double(sett.value("annotation_correctMassByPPM")))
                self.ui.annotation_correctMassByPPMnegMode.setValue(self.to_double(sett.value("annotation_correctMassByPPM")))
            if sett.contains("annotation_correctMassByPPMposMode"):
                self.ui.annotation_correctMassByPPMposMode.setValue(self.to_double(sett.value("annotation_correctMassByPPMposMode")))
            if sett.contains("annotation_correctMassByPPMnegMode"):
                self.ui.annotation_correctMassByPPMnegMode.setValue(self.to_double(sett.value("annotation_correctMassByPPMnegMode")))

            if sett.contains("annotateMetabolites_usedDatabases"):
                usedDBs = loads(base64.b64decode(sett.value("annotateMetabolites_usedDatabases").encode("utf-8")))
                for dbName, dbFile in usedDBs:
                    print("Loading DB %s from file '%s'" % (dbName, dbFile))
                    item = QtGui.QStandardItem("%s" % dbName)
                    item.setData(dbFile)
                    self.ui.dbList_listView.model().appendRow(item)

            if sett.contains("generateMSMSInfo_CheckBox"):
                self.ui.generateMSMSInfo_CheckBox.setChecked(self.to_bool(sett.value("generateMSMSInfo_CheckBox")))
            if sett.contains("msms_minCounts"):
                self.ui.msms_minCounts.setValue(self.to_double(sett.value("msms_minCounts")))
            if sett.contains("msms_rtWindow"):
                self.ui.msms_rtWindow.setValue(self.to_double(sett.value("msms_rtWindow")))
            if sett.contains("msms_maxParallelTargets"):
                self.ui.msms_maxParallelTargets.setValue(self.to_int(sett.value("msms_maxParallelTargets")))
            if sett.contains("msms_numberOfSamples"):
                self.ui.msms_numberOfSamples.setValue(self.to_int(sett.value("msms_numberOfSamples")))
            if sett.contains("msms_nOffsprings"):
                self.ui.nOffsprings.setValue(self.to_int(sett.value("msms_nOffpsrings")))
            if sett.contains("msms_nGenerations"):
                self.ui.nGenerations.setValue(self.to_int(sett.value("msms_nGenerations")))

            sett.endGroup()

            self.updateTracerInfo()

    def saveSettings(self):
        settingsFile = QtWidgets.QFileDialog.getSaveFileName(
            caption="Select settings file",
            dir=self.lastOpenDir,
            filter="Setting file (*.ini);;All files (*.*)",
        )
        if len(settingsFile) > 0:
            self.lastOpenDir = str(settingsFile).replace("\\", "/")
            self.lastOpenDir = self.lastOpenDir[: self.lastOpenDir.rfind("/")]
        if len(settingsFile) > 0:
            self.lastOpenDir = str(settingsFile).replace("\\", "/")
            self.lastOpenDir = self.lastOpenDir[: self.lastOpenDir.rfind("/")]
        self.saveSettingsFile(settingsFile)

    # save currently loaded settings to key-value pair file
    def saveSettingsFile(self, settingsFile, clear=True):
        if len(settingsFile) > 0:
            sett = QtCore.QSettings(settingsFile, QtCore.QSettings.IniFormat)
            if clear:
                sett.clear()

            sett.beginGroup("WorkingDirectory")
            sett.setValue("workingDirectory", os.getcwd().replace("\\", "/"))
            sett.endGroup()

            sett.beginGroup("Settings")

            sett.setValue(
                "processIndividualFiles",
                self.ui.processIndividualFiles.checkState() == QtCore.Qt.Checked,
            )

            sett.setValue("TracerExperiment", self.labellingExperiment == TRACER)

            if self.labellingExperiment == TRACER:
                sett.setValue(
                    "tracerConfiguration",
                    base64.b64encode(dumps([self.configuredTracer])).decode("utf-8"),
                )
            else:
                sett.setValue("LabellingElementA", self.ui.isotopeAText.text())
                sett.setValue("IsotopicAbundanceA", self.ui.isotopicAbundanceA.value() / 100.0)
                sett.setValue("LabellingElementB", self.ui.isotopeBText.text())
                sett.setValue("IsotopicAbundanceB", self.ui.isotopicAbundanceB.value() / 100.0)
                sett.setValue("useRatio", self.ui.useRatio.isChecked())
                sett.setValue("minRatio", self.ui.minRatio.value())
                sett.setValue("maxRatio", self.ui.maxRatio.value())

            sett.setValue("useCValidation", self.ui.useCValidation.checkState().value)

            sett.setValue("IntensityThreshold", self.ui.intensityThreshold.value())
            sett.setValue("IntensityCutoff", self.ui.intensityCutoff.value())
            sett.setValue("ScanStart", self.ui.scanStartTime.value())
            sett.setValue("ScanEnd", self.ui.scanEndTime.value())
            sett.setValue("xCounts", str(self.ui.xCountSearch.text()))
            sett.setValue("MaxLoading", self.ui.maxLoading.value())
            sett.setValue("MaxMassDeviation", self.ui.ppmRangeIdentification.value())
            sett.setValue("IsotopicPatternCountA", self.ui.isotopePatternCountA.value())
            sett.setValue("IsotopicPatternCountB", self.ui.isotopePatternCountB.value())
            sett.setValue("ScanIndexOffset", self.ui.scanIndexOffset.value())
            sett.setValue(
                "lowAbundanceIsotopeCutoff",
                self.ui.isoAbundance.checkState() == QtCore.Qt.Checked,
            )
            sett.setValue(
                "intensityThresholdIsotopologs",
                self.ui.intensityThresholdIsotopologs.value(),
            )
            sett.setValue("IntensityAbundanceErrorA", self.ui.baseRange.value() / 100.0)
            sett.setValue("IntensityAbundanceErrorB", self.ui.isotopeRange.value() / 100.0)

            sett.setValue("ClustPPM", self.ui.clustPPM.value())
            sett.setValue("minSpectraCount", self.ui.minSpectraCount.value())
            sett.setValue("correctCCount", self.ui.correctcCount.checkState() == QtCore.Qt.Checked)
            sett.setValue("EICppm", self.ui.wavelet_EICppm.value())
            sett.setValue("EICSmoothingWindow", str(self.ui.eicSmoothingWindow.currentText()))
            sett.setValue("EICSmoothingWindowSize", self.ui.eicSmoothingWindowSize.value())
            sett.setValue("smoothingPolynom_spinner", self.ui.smoothingPolynom_spinner.value())
            sett.setValue(
                "artificialMPshift_start",
                self.ui.spinBox_artificialMPshift_start.value(),
            )
            sett.setValue("artificialMPshift_stop", self.ui.spinBox_artificialMPshift_stop.value())

            sett.setValue("Wavelet_MinScale", self.ui.wavelet_minScale.value())
            sett.setValue("Wavelet_MaxScale", self.ui.wavelet_maxScale.value())
            sett.setValue("Wavelet_SNRTh", self.ui.wavelet_SNRThreshold.value())
            sett.setValue("Peak_CenterError", self.ui.peak_centerError.value())
            sett.setValue("Peak_WidthError", self.ui.peak_scaleError.value())
            sett.setValue("Peak_minPeakCorr", self.ui.minPeakCorr.value() / 100.0)
            sett.setValue("checkBox_checkPeakRatio", self.ui.checkBox_checkPeakRatio.isChecked())
            sett.setValue("doubleSpinBox_minPeakRatio", self.ui.doubleSpinBox_minPeakRatio.value())
            sett.setValue("doubleSpinBox_maxPeakRatio", self.ui.doubleSpinBox_maxPeakRatio.value())

            # Peak picking settings (algorithm + post-processing filters)
            sett.setValue("peakPickingSettings", json.dumps(self._peakPickingSettings))
            sett.setValue("peakPickingEicRows", json.dumps(getattr(self, "_peakPickingEicRows", [])))

            sett.setValue("calcIsoRatioNative", self.ui.calcIsoRatioNative_spinBox.value())
            sett.setValue("calcIsoRatioLabelled", self.ui.calcIsoRatioLabelled_spinBox.value())
            sett.setValue("calcIsoRatioMoiety", self.ui.calcIsoRatioMoiety_spinBox.value())

            sett.setValue("hAIntensityError", self.ui.hAIntensityError.value() / 100.0)
            sett.setValue("hAMinScans", self.ui.hAMinScans.value())
            sett.setValue(
                "heteroElements",
                base64.b64encode(dumps(self.heteroElements)).decode("utf-8"),
            )

            sett.setValue("saveTSV", self.ui.saveCSV.checkState() == QtCore.Qt.Checked)
            sett.setValue("saveFeatureML", self.ui.saveFeatureML.checkState() == QtCore.Qt.Checked)
            sett.setValue("saveMZXML", self.ui.saveMZXML.isChecked())
            writeMZXMLOptions = 0
            if self.ui.wm_ia.checkState() == QtCore.Qt.Checked:
                writeMZXMLOptions |= 1
            if self.ui.wm_iap.checkState() == QtCore.Qt.Checked:
                writeMZXMLOptions |= 2
            if self.ui.wm_imb.checkState() == QtCore.Qt.Checked:
                writeMZXMLOptions |= 4
            if self.ui.wm_ib.checkState() == QtCore.Qt.Checked:
                writeMZXMLOptions |= 8
            sett.setValue("writeMZXMLOptions", writeMZXMLOptions)

            sett.setValue("minCorrelation", self.ui.minCorrelation.value() / 100.0)
            sett.setValue(
                "minCorrelationConnections",
                self.ui.minCorrelationConnections.value() / 100.0,
            )
            sett.setValue("adducts", base64.b64encode(dumps(self.adducts)).decode("utf-8"))
            sett.setValue(
                "elementsForNL",
                base64.b64encode(dumps(self.elementsForNL)).decode("utf-8"),
            )
            sett.setValue(
                "simplifyInSourceFragments",
                self.ui.checkBox_simplifyInSourceFragments.checkState() == QtCore.Qt.Checked,
            )

            sett.setValue(
                "processMultipleFiles",
                self.ui.processMultipleFiles.checkState() == QtCore.Qt.Checked,
            )

            sett.setValue("GroupFeatures", self.ui.groupResults.isChecked())
            sett.setValue("GroupPPM", self.ui.groupPpm.value())
            sett.setValue(
                "GroupAlign",
                self.ui.alignChromatograms.checkState() == QtCore.Qt.Checked,
            )
            sett.setValue("GroupNPolynom", self.ui.polynomValue.value())
            sett.setValue("GroupTimeWindow", self.ui.groupingRT.value())

            sett.setValue("ConvoluteMetaboliteGroups", self.ui.convoluteResults.isChecked())
            sett.setValue("maxAnnotationTimeWindow", self.ui.maxAnnotationTimeWindow.value())
            sett.setValue(
                "MetaboliteClusterMinConnections",
                self.ui.metaboliteClusterMinConnections.value(),
            )
            sett.setValue(
                "useSILRatioForConvolution",
                self.ui.useSILRatioForConvolution.checkState() == QtCore.Qt.Checked,
            )
            sett.setValue("minConnectionRate", self.ui.minConnectionRate.value() / 100.0)
            sett.setValue(
                "useAbundanceSimilarityForConvolution",
                self.ui.useAbundanceSimilarityForConvolution.checkState() == QtCore.Qt.Checked,
            )
            sett.setValue("abundanceSimilarityThreshold", self.ui.abundanceSimilarityThreshold.value() / 100.0)

            sett.setValue("GroupIntegrateMissedPeaks", self.ui.integratedMissedPeaks.isChecked())
            sett.setValue(
                "integrationMaxTimeDifference",
                self.ui.integrationMaxTimeDifference.value(),
            )
            sett.setValue("reintegrateIntensityCutoff", self.ui.reintegrateIntensityCutoff.value())

            sett.setValue("annotateMetabolites", self.ui.annotateMetabolites_CheckBox.isChecked())
            sett.setValue(
                "annotateMetabolites_generateSumFormulas",
                self.ui.generateSumFormulas_CheckBox.isChecked(),
            )
            sett.setValue(
                "annotateMetabolites_sumFormulasMinimumElements",
                self.ui.sumFormulasMinimumElements_lineEdit.text(),
            )
            sett.setValue(
                "annotateMetabolites_sumFormulasMaximumElements",
                self.ui.sumFormulasMaximumElements_lineEdit.text(),
            )
            sett.setValue(
                "annotateMetabolites_sumFormulasUseExactXn",
                self.ui.sumFormulasUseExactXn_ComboBox.currentText(),
            )
            sett.setValue(
                "annotateMetabolites_plusMinus",
                self.ui.sumFormulasPlusMinus_spinBox.value(),
            )
            sett.setValue(
                "annotateMetabolites_checkRT",
                self.ui.checkRTInHits_checkBox.isChecked(),
            )
            sett.setValue(
                "annotateMetabolites_maxRTError",
                self.ui.maxRTErrorInHits_spinnerBox.value(),
            )
            sett.setValue(
                "annotateMetabolites_maxPPM",
                self.ui.annotationMaxPPM_doubleSpinBox.value(),
            )
            sett.setValue(
                "annotation_correctMassByPPMposMode",
                self.ui.annotation_correctMassByPPMposMode.value(),
            )
            sett.setValue(
                "annotation_correctMassByPPMnegMode",
                self.ui.annotation_correctMassByPPMnegMode.value(),
            )

            usedDBs = []
            for entryInd in range(self.ui.dbList_listView.model().rowCount()):
                dbFile = str(self.ui.dbList_listView.model().item(entryInd, 0).data())
                dbName = dbFile[dbFile.rfind("/") + 1 : dbFile.rfind(".")]
                usedDBs.append([dbName, dbFile])
            sett.setValue(
                "annotateMetabolites_usedDatabases",
                base64.b64encode(dumps(usedDBs)).decode("utf-8"),
            )

            sett.setValue(
                "generateMSMSInfo_CheckBox",
                self.ui.generateMSMSInfo_CheckBox.isChecked(),
            )
            sett.setValue("msms_minCounts", self.ui.msms_minCounts.value())
            sett.setValue("msms_rtWindow", self.ui.msms_rtWindow.value())
            sett.setValue("msms_maxParallelTargets", self.ui.msms_maxParallelTargets.value())
            sett.setValue("msms_numberOfSamples", self.ui.msms_numberOfSamples.value())
            sett.setValue("msms_nGenerations", self.ui.nGenerations.value())
            sett.setValue("msms_nOffsprings", self.ui.nOffsprings.value())

            sett.endGroup()

            sett.beginGroup("MetExtract")

            sett.setValue("Version", MetExtractVersion)
            import datetime
            import platform
            import uuid

            sett.setValue(
                "UUID_ext",
                "%s_%s_%s"
                % (
                    str(uuid.uuid1()),
                    str(platform.node()),
                    str(datetime.datetime.now()),
                ),
            )

            sett.endGroup()

    # </editor-fold>

    def updateGroupPPM(self):
        self.ui.groupPpm.setValue(self.ui.ppmRangeIdentification.value() * 2)

    def selectGroupsFile(self):
        grpFile = QtWidgets.QFileDialog.getSaveFileName(
            caption="Select groups file",
            dir=self.lastOpenDir,
            filter="Excel file(*.xlsx);;TSV file (*.tsv);;CSV file (*.csv);;All files (*.*)",
        )
        if len(grpFile) > 0:
            self.lastOpenDir = str(grpFile).replace("\\", "/")
            self.lastOpenDir = self.lastOpenDir[: self.lastOpenDir.rfind("/")]

        if len(grpFile) > 0:
            self.ui.groupsSave.setText(grpFile)
        else:
            self.ui.groupsSave.setText("./results.xlsx")

    def selectMSMSTargetFile(self):
        msmsTargetFile = QtWidgets.QFileDialog.getSaveFileName(
            caption="Select MSMS target file",
            dir=self.lastOpenDir,
            filter="Excel file(*.xlsx);;TSV file (*.tsv);;CSV file (*.csv);;All files (*.*)",
        )
        if len(msmsTargetFile) > 0:
            self.lastOpenDir = str(msmsTargetFile).replace("\\", "/")
            self.lastOpenDif = self.lastOpenDir[: self.lastOpenDir.rfind("/")]

        if len(msmsTargetFile) > 0:
            self.ui.msms_fileLocation.setText(msmsTargetFile)
        else:
            self.ui.msms_fileLocation.setText("./MSMStargets.tsv")

    def exitMe(self):
        QtWidgets.QApplication.exit()

    def showCalcIsoEnrichmentDialog(self):
        diag = calcIsoEnrichmentDialog()
        diag.executeDialog()

    def setWorkingDirectoryDialog(self):
        file = str(
            QtWidgets.QFileDialog.getExistingDirectory(
                self,
                "Select new working directory. The current one is '%s'" % os.getcwd(),
            )
        )
        if file != "":
            os.chdir(file)
            LoggingSetup.LoggingSetup.Instance().initLogging(location=file)
            logging.info("CWD set to '%s'" % file)

    def openTempDir(self):
        subprocess.Popen('explorer "' + local_folder)

    def aboutMe(self):
        dialog = AboutDialog(
            parent=self,
            app_name="MetExtract",
            version=MetExtractVersion,
            institute_name="iBAM",
        )
        dialog.exec()

    # open local MetExtract documentation
    # taken from http://stackoverflow.com/questions/4216985/call-to-operating-system-to-open-url
    def helpMe(self):
        import sys
        import webbrowser

        url = get_main_dir() + "/documentation/index.html"
        if sys.platform == "darwin":  # in case of OS X
            subprocess.Popen(["open", url])
        else:
            webbrowser.open_new_tab(url)

    def openOboDownloadPage(self):
        import webbrowser

        webbrowser.open_new_tab(OBO_DOWNLOAD_URL)

    # helper method for multiprocessing module of LC-HRMS file processing
    def _send_desktop_notification(self, title, message):
        """Send a non-blocking desktop notification using desktop-notifier."""

        def _runner():
            try:
                from desktop_notifier import DesktopNotifier
            except Exception as ex:
                logging.warning(f"desktop-notifier import failed: {ex}")
                return

            async def _send_async():
                notifier = DesktopNotifier(app_name="MetExtract II")
                await notifier.send(title=title, message=message)

            try:
                asyncio.run(_send_async())
            except RuntimeError:
                # Fallback for environments where an event loop is already running.
                loop = asyncio.new_event_loop()
                try:
                    loop.run_until_complete(_send_async())
                finally:
                    loop.close()
            except Exception as ex:
                logging.warning(f"Failed to send desktop notification: {ex}")

        threading.Thread(target=_runner, daemon=True).start()

    def runProcess(self, dontSave=False, askStarting=True):
        self.terminateJobs = False

        self.closeCurrentOpenResultsFile()
        self.closeLoadedGroupsResultsFile()

        # check certain requirements for LC-HRMS file processing
        if str(self.ui.negativeScanEvent.currentText()) == "None" and str(self.ui.positiveScanEvent.currentText()) == "None":
            QtWidgets.QMessageBox.information(
                self,
                "MetExtract",
                "Please select at least one scan event prior to calculations",
                QtWidgets.QMessageBox.Ok,
            )
            return 1

        if str(self.ui.negativeScanEvent.currentText()) == str(self.ui.positiveScanEvent.currentText()):
            QtWidgets.QMessageBox.information(
                self,
                "MetExtract",
                "It seems two identical scan events are selected for the positive and negative mode. Select only one mode if just a single ionisation mode is present in your files.",
            )
            return 1

        if askStarting:
            if self.ui.isotopePatternCountA.value() == 1 and self.ui.isotopePatternCountB.value() == 1:
                if (
                    QtWidgets.QMessageBox.question(
                        self,
                        "MetExtract",
                        "Are you sure you want to search for metabolites without using isotopic peaks?",
                        QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    )
                    == QtWidgets.QMessageBox.No
                ):
                    return 1

        if self.ui.intensityCutoff.value() > self.ui.intensityThreshold.value() and self.ui.intensityCutoff.value() > 0:
            QtWidgets.QMessageBox.information(
                self,
                "MetExtract",
                "The intensity cutoff needs to be less than the intensity threshold. Adjust the two parameters accordingly",
            )
            return 1

        ma, ea = getIsotopeMass(self.ui.isotopeAText.text())
        mb, eb = getIsotopeMass(self.ui.isotopeBText.text())

        if not (self.labellingExperiment == TRACER) and (len(ea) < 0 or ea != eb) and not (ea == "Hydrogen" and eb == "Deuterium"):
            QtWidgets.QMessageBox.question(
                self,
                "MetExtract",
                "You cannot use two different elements for the labelling process. Please enter isotopes of the same element.\nCalculation aborted",
                QtWidgets.QMessageBox.Ok,
            )
            return 1

        ## TODO copy DB files to working directory..
        ## here66
        if False:
            if not os.path.exists("./DBs"):
                os.makedirs("./DBs")

            for entryInd in range(self.ui.dbList_listView.model().rowCount()):
                dbFile = str(self.ui.dbList_listView.model().item(entryInd, 0).data())
                dbName = dbFile[dbFile.rfind("/") + 1 : dbFile.rfind(".")]

                try:
                    print(dbName, dbFile)
                    if dbFile != "./DBs/%s.tsv" % dbName:
                        z = QtWidgets.QMessageBox.question(
                            self,
                            "MetExtract",
                            "Do you want to copy the database (%s) to the current directory for documentation?" % (dbName),
                            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No | QtWidgets.QMessageBox.Abort,
                        )

                        if z == QtWidgets.QMessageBox.Yes:
                            shutil.copyfile(dbFile, "./DBs/%s.tsv" % dbName)
                            self.ui.dbList_listView.model().item(entryInd, 0).setData("./DBs/%s.tsv")
                        elif z == QtWidgets.QMessageBox.Abort:
                            return 1
                except Exception:
                    pass

        if not dontSave:
            z = QtWidgets.QMessageBox.question(
                self,
                "MetExtract",
                "Do you want to save the loaded files, the group compilation and the settings?",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No | QtWidgets.QMessageBox.Abort,
            )
            if z == QtWidgets.QMessageBox.Yes:
                self.saveGroupsClicked()
            elif z == QtWidgets.QMessageBox.Abort:
                return 1

        if askStarting:
            z = QtWidgets.QMessageBox.question(
                self,
                "MetExtract",
                "Do you want to start the processing?",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            )
            if z == QtWidgets.QMessageBox.No:
                return 1

        definedGroups = self.getAllSampleGroups()
        for group in definedGroups:
            for i in range(len(group.files)):
                group.files[i] = str(group.files[i]).replace("\\", "/")

        # get all individual LC-HRMS files for processing
        files = []

        # --- Unified peak picker and filter config ---
        from .chromPeakPicking.peakpickers import make_peak_picker, make_peak_filter_config

        picker = make_peak_picker(self._peakPickingSettings)
        filter_config = make_peak_filter_config(self._peakPickingSettings)

        indGroups = {}
        for group in definedGroups:
            grName = str(group.name)
            indGroups[grName] = []
            for file in natSort(group.files):
                indGroups[grName].append(str(file))
                if file not in files:
                    files.append(file)

        errorCount = 0

        # Per-step tracking for the final summary dialog
        _ST_SKIPPED_USER = "skipped_user"
        _ST_OK = "ok"
        _ST_ERROR = "error"
        _ST_SKIPPED_PREV = "skipped_prev_error"
        _step_status = {k: _ST_SKIPPED_USER for k in ("individual_files", "bracketing", "reintegration", "convolution", "annotation")}
        _step_elapsed = {k: 0.0 for k in _step_status}
        _step_details = {k: "" for k in _step_status}
        _bracketing_failed = False
        _reintegration_failed = False
        _convolution_failed = False

        cpus = min(cpu_count(), self.ui.cpuCores.value())

        if self.terminateJobs:
            return
        overallStart = time.time()

        writeMZXMLOptions = 0
        if self.ui.saveMZXML.isChecked():
            if self.ui.wm_ia.checkState() == QtCore.Qt.Checked:
                writeMZXMLOptions |= 1
            if self.ui.wm_iap.checkState() == QtCore.Qt.Checked:
                writeMZXMLOptions |= 2
            if self.ui.wm_imb.checkState() == QtCore.Qt.Checked:
                writeMZXMLOptions |= 4
            if self.ui.wm_ib.checkState() == QtCore.Qt.Checked:
                writeMZXMLOptions |= 8

        start = time.time()

        # process individual files
        if self.ui.processIndividualFiles.isChecked():
            _ind_step_start = time.time()
            logging.info("")
            logging.info("Processing %d individual LC-HRMS data files on %d CPU core(s).." % (len(files), min(len(files), cpus)))

            # Initialize multiprocessing pool
            try:
                set_start_method("spawn")
            except RuntimeError:
                pass  # context already set
            p = Pool(processes=min(len(files), cpus), maxtasksperchild=1)
            manager = Manager()
            lock = manager.Lock()
            queue = manager.Queue()

            pIds = {}
            for i in range(len(files)):
                pIds[i + 1] = files[i]

            # start the multiprocessing
            runObjects = [
                FindIsoPairs(
                    files[i],
                    exOperator=str(self.ui.exOperator_LineEdit.text()),
                    exExperimentID=str(self.ui.exExperimentID_LineEdit.text()),
                    exComments=str(self.ui.exComments_TextEdit.toPlainText()),
                    exExperimentName=str(self.ui.exExperimentName_LineEdit.text()),
                    metabolisationExperiment=self.labellingExperiment == TRACER,
                    labellingisotopeA=str(self.ui.isotopeAText.text()),
                    labellingisotopeB=str(self.ui.isotopeBText.text()),
                    xOffset=self.isotopeBmass - self.isotopeAmass,
                    useRatio=self.ui.useRatio.isChecked(),
                    minRatio=self.ui.minRatio.value(),
                    maxRatio=self.ui.maxRatio.value(),
                    useCIsotopePatternValidation=int(self.ui.useCValidation.checkState().value),
                    configuredTracer=self.configuredTracer,
                    intensityThreshold=self.ui.intensityThreshold.value(),
                    intensityCutoff=self.ui.intensityCutoff.value(),
                    startTime=self.ui.scanStartTime.value(),
                    stopTime=self.ui.scanEndTime.value(),
                    maxLoading=self.ui.maxLoading.value(),
                    xCounts=str(self.ui.xCountSearch.text()),
                    ppm=self.ui.ppmRangeIdentification.value(),
                    isotopicPatternCountLeft=self.ui.isotopePatternCountA.value(),
                    isotopicPatternCountRight=self.ui.isotopePatternCountB.value(),
                    lowAbundanceIsotopeCutoff=self.ui.isoAbundance.checkState() == QtCore.Qt.Checked,
                    intensityThresholdIsotopologs=self.ui.intensityThresholdIsotopologs.value(),
                    purityN=self.ui.isotopicAbundanceA.value() / 100.0,
                    purityL=self.ui.isotopicAbundanceB.value() / 100.0,
                    intensityErrorN=self.ui.baseRange.value() / 100.0,
                    intensityErrorL=self.ui.isotopeRange.value() / 100.0,
                    scanIndexOffset=self.ui.scanIndexOffset.value(),
                    minSpectraCount=self.ui.minSpectraCount.value(),
                    clustPPM=self.ui.clustPPM.value(),
                    chromPeakPPM=self.ui.wavelet_EICppm.value(),
                    eicSmoothingWindow=str(self.ui.eicSmoothingWindow.currentText()),
                    eicSmoothingWindowSize=self.ui.eicSmoothingWindowSize.value(),
                    eicSmoothingPolynom=self.ui.smoothingPolynom_spinner.value(),
                    peakCenterError=self.ui.peak_centerError.value(),
                    artificialMPshift_start=self.ui.spinBox_artificialMPshift_start.value(),
                    artificialMPshift_stop=self.ui.spinBox_artificialMPshift_stop.value(),
                    calcIsoRatioNative=self.ui.calcIsoRatioNative_spinBox.value(),
                    calcIsoRatioLabelled=self.ui.calcIsoRatioLabelled_spinBox.value(),
                    calcIsoRatioMoiety=self.ui.calcIsoRatioMoiety_spinBox.value(),
                    minCorrelationConnections=self.ui.minCorrelationConnections.value() / 100.0,
                    positiveScanEvent=str(self.ui.positiveScanEvent.currentText()),
                    negativeScanEvent=str(self.ui.negativeScanEvent.currentText()),
                    correctCCount=self.ui.correctcCount.checkState() == QtCore.Qt.Checked,
                    minCorrelation=self.ui.minCorrelation.value() / 100.0,
                    hAIntensityError=self.ui.hAIntensityError.value() / 100.0,
                    hAMinScans=self.ui.hAMinScans.value(),
                    adducts=self.adducts,
                    elements=self.elementsForNL,
                    heteroAtoms=self.heteroElements,
                    simplifyInSourceFragments=self.ui.checkBox_simplifyInSourceFragments.isChecked(),
                    lock=lock,
                    queue=queue,
                    pID=i + 1,
                    meVersion="MetExtract (%s)" % MetExtractVersion,
                    peak_picker=picker,
                    peak_filter_config=filter_config,
                )
                for i in range(len(files))
            ]
            res = p.imap_unordered(
                findSILFeatures,
                runObjects,
            )

            pw = ProgressWrapper(
                pwCount=min(len(files), cpus) + 1,
                showIndProgress=True,
                indGroups=indGroups,
                parent=self,
                closeCallback=CallBackMethod(_target=interruptIndividualFilesProcessing, selfObj=self, pool=p).getRunMethod(),
            )

            for w in range(1, min(len(files), cpus) + 1):
                pw.setUntils(
                    {
                        0.25: "#E3E9E2",
                        0.35: "#D3D9D2",
                        0.65: "#52626D",
                        0.70: "#6C7B8B",
                        0.75: "#92A1AB",
                        0.8: "#CACDC9",
                        1.0: "darkorange",
                    },
                    w,
                )
                pw.setMax(1.0, w)

            pw.show()
            pwMain = pw.getCallingFunction(0)
            pwMain("max")(len(files))
            pwMain("value")(0)

            failedFiles = []

            # monitor processing of individual LC-HRMS files and report to the user
            loop = True
            freeSlots = [1 + i for i in list(range(min(len(files), cpus)))]
            assignedThreads = {}
            messages_to_print = defaultdict(list)
            while loop and not self.terminateJobs:
                try:
                    completed = res._index
                    if completed == len(files):
                        loop = False

                    pwMain("value")(completed)

                    mess = {}
                    while not (queue.empty()):
                        mes = queue.get(block=False, timeout=1)
                        if mes.pid not in mess:
                            mess[mes.pid] = {}
                        mess[mes.pid][mes.mes] = mes

                    for v in mess.values():
                        if "start" in v.keys():
                            mes = v["start"]
                            if len(freeSlots) > 0:
                                w = freeSlots.pop()
                                assignedThreads[mes.pid] = w

                                pw.getCallingFunction()("statuscolor")(pIds[mes.pid], "orange")
                                pw.getCallingFunction()("statustext")(
                                    pIds[mes.pid],
                                    text="File: %s\nStatus: %s\nProcess ID: %d" % (pIds[mes.pid], "processing", mes.pid),
                                )
                            else:
                                logging.error("Something went wrong..")
                                logging.error('Progress bars do not work correctly, but files will be processed and "finished.." will be printed..')

                    for v in mess.values():
                        for mes in v.values():
                            if mes.mes == "log":
                                messages_to_print[mes.pid].append(mes.val)
                            elif mes.mes in ["text", "max", "value"]:
                                if mes.pid in assignedThreads:
                                    pw.getCallingFunction(assignedThreads[mes.pid])(mes.mes)(mes.val)
                                else:
                                    logging.error("Error in messaging pipeline of subprocess id %d" % mes.pid)
                            elif mes.mes in ["end", "failed", "start"]:
                                pass
                            else:
                                logging.error(f"Received unknown message {mes.mes} with payload {mes.__dict__}")

                    for v in mess.values():
                        if "end" in v.keys() or "failed" in v.keys():
                            mes = None
                            if "end" in v.keys():
                                mes = v["end"]
                            elif "failed" in v.keys():
                                mes = v["failed"]
                                failedFiles.append(pIds[mes.pid])
                            freeS = assignedThreads[mes.pid]
                            pw.getCallingFunction(assignedThreads[mes.pid])("text")("")
                            pw.getCallingFunction(assignedThreads[mes.pid])("value")(0)

                            logging.info("\n" + "##############################################################\n" + "\n".join(messages_to_print[mes.pid]) + "\n" + "##############################################################\n" + "")

                            pw.getCallingFunction()("statuscolor")(
                                pIds[mes.pid],
                                "olivedrab" if mes.mes == "end" else "firebrick",
                            )
                            pw.getCallingFunction()("statustext")(
                                pIds[mes.pid],
                                text="File: %s\nStatus: %s"
                                % (
                                    pIds[mes.pid],
                                    "finished" if mes.mes == "end" else "failed",
                                ),
                            )

                            if freeS == -1:
                                logging.error("Something went wrong..")
                                logging.error('Progress bars do not work correctly, but files will be processed and "finished.." will be printed..')
                            else:
                                assignedThreads[mes.pid] = -1
                                freeSlots.append(freeS)

                        elapsed = (time.time() - start) / 60.0
                        hours = ""
                        if elapsed >= 60.0:
                            hours = "%d hours " % (elapsed // 60)
                        mins = "%.2f mins" % (elapsed % 60.0)
                        pwMain("text")(
                            "<p align='right' >%s%s elapsed</p>\n\n%d / %d files done (%d parallel)"
                            % (
                                hours,
                                mins,
                                completed,
                                len(files),
                                min(cpus, len(files)),
                            )
                        )
                        time.sleep(0.5)
                except Exception:
                    traceback.print_exc()

            # Log time used for processing of individual files
            elapsed = (time.time() - start) / 60.0
            hours = ""
            if elapsed >= 60.0:
                hours = "%d hours " % (elapsed // 60)
            mins = "%.2f mins" % (elapsed % 60.0)

            if not self.terminateJobs:
                logging.info("Individual files processed (%s%s).." % (hours, mins))

            if len(failedFiles) > 0:
                logging.info("Some files failed to process correctly (%s)" % (", ".join(failedFiles)))

            pw.setSkipCallBack(True)
            pw.hide()

            p.close()
            p.terminate()
            p.join()

            _step_elapsed["individual_files"] = (time.time() - _ind_step_start) / 60.0
            if not self.terminateJobs:
                _step_status["individual_files"] = _ST_OK
                _step_details["individual_files"] = f"{len(files)} file(s): {len(files) - len(failedFiles)} finished, {len(failedFiles)} failed"

        if self.terminateJobs:
            return

        resFileFull = str(self.ui.groupsSave.text())
        resFilePath, resFileName = os.path.split(os.path.abspath(resFileFull))
        excel_file = resFileFull.replace(".xlsx", ".tsv").replace(".tsv", ".txt").replace(".txt", "") + ".xlsx"

        # Determine the best available input sheet for annotation when only the annotation step is run
        # (will be overridden later if bracketing/grouping/re-integration runs as part of this execution)
        annotation_input_sheet = "2_StatColumns"
        if os.path.isfile(excel_file):
            try:
                import openpyxl as _opxl_detect

                _wb_detect = _opxl_detect.load_workbook(excel_file, read_only=True)
                for _candidate in ("4_Convoluted", "3_Reintegrated", "2_StatColumns"):
                    if _candidate in _wb_detect.sheetnames:
                        annotation_input_sheet = _candidate
                        break
                _wb_detect.close()
            except Exception:
                pass

        # bracket/group from individual LC-HRMS data / re-integrate missed peaks
        if self.ui.processMultipleFiles.checkState() == QtCore.Qt.Checked:
            pw = ProgressWrapper(1, parent=self)
            pw.show()
            pw.getCallingFunction()("text")("Bracketing results")
            pw.getCallingFunction()("header")("Bracketing..")

            # bracket/group results from individual LC-HRMS data
            if self.ui.groupResults.isChecked():
                logging.info("\n\n##############################################################")
                logging.info("Bracketing of individual LC-HRMS results..")

                _bracketing_step_start = time.time()
                try:
                    if True:
                        # Group results
                        generalProcessingParams = Bunch(
                            exOperator=str(self.ui.exOperator_LineEdit.text()),
                            exExperimentID=str(self.ui.exExperimentID_LineEdit.text()),
                            exComments=str(self.ui.exComments_TextEdit.toPlainText()),
                            exExperimentName=str(self.ui.exExperimentName_LineEdit.text()),
                            metabolisationExperiment=self.labellingExperiment == TRACER,
                            intensityThreshold=self.ui.intensityThreshold.value(),
                            intensityCutoff=self.ui.intensityCutoff.value(),
                            labellingisotopeA=str(self.ui.isotopeAText.text()),
                            labellingisotopeB=str(self.ui.isotopeBText.text()),
                            xOffset=self.isotopeBmass - self.isotopeAmass,
                            useRatio=self.ui.useRatio.isChecked(),
                            minRatio=self.ui.minRatio.value(),
                            maxRatio=self.ui.maxRatio.value(),
                            useCValidation=int(self.ui.useCValidation.checkState().value),
                            configuredTracer=str(self.configuredTracer),
                            startTime=self.ui.scanStartTime.value(),
                            stopTime=self.ui.scanEndTime.value(),
                            maxLoading=self.ui.maxLoading.value(),
                            xCounts=str(self.ui.xCountSearch.text()),
                            ppm=self.ui.ppmRangeIdentification.value(),
                            isotopicPatternCountLeft=self.ui.isotopePatternCountA.value(),
                            isotopicPatternCountRight=self.ui.isotopePatternCountB.value(),
                            lowAbundanceIsotopeCutoff=self.ui.isoAbundance.checkState() == QtCore.Qt.Checked,
                            intensityThresholdIsotopologs=self.ui.intensityThresholdIsotopologs.value(),
                            purityN=self.ui.isotopicAbundanceA.value() / 100.0,
                            purityL=self.ui.isotopicAbundanceB.value() / 100.0,
                            intensityErrorN=self.ui.baseRange.value() / 100.0,
                            intensityErrorL=self.ui.isotopeRange.value() / 100.0,
                            scanIndexOffset=self.ui.scanIndexOffset.value(),
                            minSpectraCount=self.ui.minSpectraCount.value(),
                            clustPPM=self.ui.clustPPM.value(),
                            chromPeakPPM=self.ui.wavelet_EICppm.value(),
                            eicSmoothingWindow=str(self.ui.eicSmoothingWindow.currentText()),
                            eicSmoothingWindowSize=self.ui.eicSmoothingWindowSize.value(),
                            eicSmoothingPolynom=self.ui.smoothingPolynom_spinner.value(),
                            artificialMPshift_start=self.ui.spinBox_artificialMPshift_start.value(),
                            artificialMPshift_stop=self.ui.spinBox_artificialMPshift_stop.value(),
                            scales=[
                                self.ui.wavelet_minScale.value(),
                                self.ui.wavelet_maxScale.value(),
                            ],
                            snrTh=self.ui.wavelet_SNRThreshold.value(),
                            peakCenterError=self.ui.peak_centerError.value(),
                            peakScaleError=self.ui.peak_scaleError.value(),
                            minPeakCorr=self.ui.minPeakCorr.value() / 100.0,
                            checkPeaksRatio=self.ui.checkBox_checkPeakRatio.isChecked(),
                            minPeaksRatio=self.ui.doubleSpinBox_minPeakRatio.value(),
                            maxPeaksRatio=self.ui.doubleSpinBox_maxPeakRatio.value(),
                            checkPeakWidthFilter=self._peakPickingSettings.get("pf_enabled", False),
                            minPeakWidth=self._peakPickingSettings.get("pf_min_peak_width", 0.0),
                            maxPeakWidth=self._peakPickingSettings.get("pf_max_peak_width", 9999.0),
                            minFWHM=self._peakPickingSettings.get("pf_min_fwhm", 0.0),
                            maxFWHM=self._peakPickingSettings.get("pf_max_fwhm", 9999.0),
                            minApexToFlankFactor=self._peakPickingSettings.get("pf_min_apex_to_flank_factor", 0.0),
                            minApexToFlankIncrease=self._peakPickingSettings.get("pf_min_apex_to_flank_increase", 0.0),
                            minSnr=self._peakPickingSettings.get("pf_min_snr", 0.0),
                            calcIsoRatioNative=self.ui.calcIsoRatioNative_spinBox.value(),
                            calcIsoRatioLabelled=self.ui.calcIsoRatioLabelled_spinBox.value(),
                            calcIsoRatioMoiety=self.ui.calcIsoRatioMoiety_spinBox.value(),
                            minCorrelationConnections=self.ui.minCorrelationConnections.value() / 100.0,
                            positiveScanEvent=str(self.ui.positiveScanEvent.currentText()),
                            negativeScanEvent=str(self.ui.negativeScanEvent.currentText()),
                            correctCCount=self.ui.correctcCount.checkState() == QtCore.Qt.Checked,
                            minCorrelation=self.ui.minCorrelation.value() / 100.0,
                            hAIntensityError=self.ui.hAIntensityError.value() / 100.0,
                            hAMinScans=self.ui.hAMinScans.value(),
                            adducts="[%s]" % ",".join([str(a) for a in self.adducts]),
                            elements="[%s]" % ",".join([str(e) for e in self.elementsForNL]),
                            heteroAtoms="[%s]" % ",".join([str(h) for h in self.heteroElements]),
                            simplifyInSourceFragments=self.ui.checkBox_simplifyInSourceFragments.isChecked(),
                            meVersion="MetExtract (%s)" % MetExtractVersion,
                        )
                        procProc = FuncProcess(
                            _target=bracketResults,
                            indGroups=indGroups,
                            xCounts=str(self.ui.xCountSearch.text()),
                            groupSizePPM=self.ui.groupPpm.value(),
                            maxTimeDeviation=self.ui.groupingRT.value() * 60.0,
                            maxLoading=self.ui.maxLoading.value(),
                            positiveScanEvent=str(self.ui.positiveScanEvent.currentText()),
                            negativeScanEvent=str(self.ui.negativeScanEvent.currentText()),
                            file=resFileFull,
                            align=(self.ui.alignChromatograms.checkState() == QtCore.Qt.Checked),
                            nPolynom=self.ui.polynomValue.value(),
                            meVersion="MetExtract (%s)" % MetExtractVersion,
                            generalProcessingParams=generalProcessingParams,
                            start=start,
                        )
                        procProc.addKwd("pwMaxSet", procProc.getQueue())
                        procProc.addKwd("pwValSet", procProc.getQueue())
                        procProc.addKwd("pwTextSet", procProc.getQueue())
                        procProc.start()

                        pw.setCloseCallback(
                            closeCallBack=CallBackMethod(
                                _target=interruptBracketingOfFeaturePairs,
                                selfObj=self,
                                funcProc=procProc,
                            ).getRunMethod()
                        )

                        # check for status updates
                        while procProc.isAlive():
                            QtWidgets.QApplication.processEvents()
                            while not (procProc.getQueue().empty()):
                                mes = procProc.getQueue().get(block=False, timeout=1)

                                # No idea why / where there are sometimes other objects than Bunch(mes, val), but they occur
                                if isinstance(mes, Bunch) and hasattr(mes, "mes") and (hasattr(mes, "val") or hasattr(mes, "text")):
                                    pw.getCallingFunction()(mes.mes)(mes.val)
                                elif mes == (None, None):
                                    ## I have no idea where this object comes from
                                    pass
                                else:
                                    logging.critical("UNKNONW OBJECT IN PROCESSING QUEUE: %s" % str(mes))

                            time.sleep(0.5)

                        # Log time used for bracketing
                        elapsed = (time.time() - start) / 60.0
                        hours = ""
                        if elapsed >= 60.0:
                            hours = "%d hours " % (elapsed // 60)
                        mins = "%.2f mins" % (elapsed % 60.0)

                        if self.terminateJobs:
                            return
                        else:
                            logging.info("Bracketing finished (%s%s).." % (hours, mins))

                        # Arrange grouped results and add statistics columns
                        groups = {}
                        outputOrder = []

                        pw.getCallingFunction()("text")("Adding statistics columns\n")
                        grpStats = []
                        for group in definedGroups:
                            preFix = "_Stat"
                            grpName = group.name + preFix
                            grpAdd(
                                groups,
                                group.name + preFix,
                                group.minFound,
                                [
                                    grp[
                                        (grp.rfind("/") + 1) : max(
                                            grp.lower().rfind(".mzxml"),
                                            grp.lower().rfind(".mzml"),
                                        )
                                    ]
                                    + ("_FID")
                                    for grp in natSort(group.files)
                                ],
                            )
                            outputOrder.append(grpName)
                            grpStats.append(
                                (
                                    str(group.name + "_Stat"),
                                    group.minFound,
                                    group.omitFeatures,
                                    group.removeAsFalsePositive,
                                )
                            )

                        print("Processing Table " + excel_file)
                        addStatsColumnToResults(excel_file, groups, outputOrder, sheet_name="1_Bracketed", new_sheet_name="2_StatColumns")
                        # remove feature pairs not found more than n times (according to user specified omit value)
                        grpOmit(excel_file, grpStats, sheet_name="2_StatColumns", new_sheet_name="2_StatColumns")
                        logging.info("Statistic columns added (and feature pairs omitted)..")

                        _step_elapsed["bracketing"] = (time.time() - _bracketing_step_start) / 60.0
                        _step_status["bracketing"] = _ST_OK
                        try:
                            import openpyxl as _opxl_br

                            _wb_br = _opxl_br.load_workbook(excel_file, read_only=True)
                            _n_feat_br = max(0, _wb_br["2_StatColumns"].max_row - 1) if "2_StatColumns" in _wb_br.sheetnames else 0
                            _wb_br.close()
                            _step_details["bracketing"] = f"{_n_feat_br} feature(s) detected"
                        except Exception:
                            _step_details["bracketing"] = "features detected"

                except Exception as ex:
                    traceback.print_exc()
                    logging.error(str(traceback))

                    _step_elapsed["bracketing"] = (time.time() - _bracketing_step_start) / 60.0
                    _step_status["bracketing"] = _ST_ERROR
                    _step_details["bracketing"] = str(ex)
                    _bracketing_failed = True

                    QtWidgets.QMessageBox.warning(
                        self,
                        "MetExtract",
                        "Error during bracketing of files: '%s'" % str(ex),
                        QtWidgets.QMessageBox.Ok,
                    )
                    errorCount += 1

                logging.info("##############################################################")

            pw.setSkipCallBack(True)
            pw.hide()

            if self.terminateJobs:
                return

            convolution_input_sheet = "2_StatColumns"
            # tracked separately because grouping may run after re-integration and update only annotation input afterwards
            annotation_input_sheet = "2_StatColumns"

            # re-integrate missed peaks (run before grouping, if enabled)
            if self.ui.integratedMissedPeaks.isChecked():
                if _bracketing_failed:
                    _step_status["reintegration"] = _ST_SKIPPED_PREV
                    logging.info("Skipping re-integration: bracketing step failed.")
                else:
                    logging.info("\n\n##############################################################")
                    logging.info("Re-integrating of individual LC-HRMS results..")

                    _reintegration_step_start = time.time()
                    pw = ProgressWrapper(min(len(files), cpus) + 1, showLog=False, parent=self)
                    pw.show()
                    pw.getCallingFunction()("text")("Integrating..")
                    pw.getCallingFunction()("header")("Integrating..")

                    try:
                        # Reintegrate missed peaks in files
                        fDict = {}
                        for group in definedGroups:
                            for grp in natSort(group.files):
                                f = grp
                                f = f.replace("\\", "/")
                                fDict[f] = f[(f.rfind("/") + 1) : max(f.lower().rfind(".mzxml"), f.lower().rfind(".mzml"))]

                        reIntegrateResultsFile(
                            excel_file,
                            "2_StatColumns",
                            "3_Reintegrated",
                            fDict,
                            addPeakArea=bool(self.ui.checkBox_expPeakArea.checkState() == QtCore.Qt.Checked),
                            addPeakAbundance=bool(self.ui.checkBox_expPeakApexIntensity.checkState() == QtCore.Qt.Checked),
                            addPeakSNR=bool(self.ui.checkBox_expPeakSNR.checkState() == QtCore.Qt.Checked),
                            ppm=self.ui.groupPpm.value(),
                            maxRTShift=self.ui.integrationMaxTimeDifference.value(),
                            scales=[
                                self.ui.wavelet_minScale.value(),
                                self.ui.wavelet_maxScale.value(),
                            ],
                            reintegrateIntensityCutoff=self.ui.reintegrateIntensityCutoff.value(),
                            snrTH=self.ui.wavelet_SNRThreshold.value(),
                            smoothingWindow=str(self.ui.eicSmoothingWindow.currentText()),
                            smoothingWindowSize=self.ui.eicSmoothingWindowSize.value(),
                            smoothingWindowPolynom=self.ui.smoothingPolynom_spinner.value(),
                            positiveScanEvent=str(self.ui.positiveScanEvent.currentText()),
                            negativeScanEvent=str(self.ui.negativeScanEvent.currentText()),
                            pw=pw,
                            selfObj=self,
                            cpus=min(len(files), cpus),
                            start=start,
                            peak_filter_config=filter_config,
                            peak_picker=picker,
                        )
                        convolution_input_sheet = "3_Reintegrated"
                        annotation_input_sheet = "3_Reintegrated"

                        _step_elapsed["reintegration"] = (time.time() - _reintegration_step_start) / 60.0
                        _step_status["reintegration"] = _ST_OK
                        try:
                            import openpyxl as _opxl_ri

                            _wb_ri = _opxl_ri.load_workbook(excel_file, read_only=True)
                            _n_feat_ri = max(0, _wb_ri["3_Reintegrated"].max_row - 1) if "3_Reintegrated" in _wb_ri.sheetnames else 0
                            _wb_ri.close()
                            _step_details["reintegration"] = f"{_n_feat_ri} feature(s) re-integrated"
                        except Exception:
                            _step_details["reintegration"] = "re-integration complete"

                        elapsed = (time.time() - start) / 60.0
                        hours = ""
                        if elapsed >= 60.0:
                            hours = "%d hours " % (elapsed // 60)
                        mins = "%.2f mins" % (elapsed % 60.0)
                        logging.info("Re-integrating finished (%s%s).." % (hours, mins))

                    except Exception as e:
                        traceback.print_exc()
                        logging.error(str(traceback))

                        _step_elapsed["reintegration"] = (time.time() - _reintegration_step_start) / 60.0
                        _step_status["reintegration"] = _ST_ERROR
                        _step_details["reintegration"] = str(e)
                        _reintegration_failed = True

                        QtWidgets.QMessageBox.warning(
                            self,
                            "MetExtract",
                            "Error during reintegrating files: '%s'" % str(e),
                            QtWidgets.QMessageBox.Ok,
                        )
                        errorCount += 1
                    finally:
                        pw.setSkipCallBack(True)
                        pw.hide()
                    logging.info("##############################################################")

            if self.terminateJobs:
                pw.hide()
                return

            # Calculate metabolite groups
            if self.ui.convoluteResults.isChecked():
                if _bracketing_failed or _reintegration_failed:
                    _step_status["convolution"] = _ST_SKIPPED_PREV
                    logging.info("Skipping convolution: a preceding step failed.")
                else:
                    logging.info("\n\n##############################################################")
                    _convolution_step_start = time.time()
                    try:
                        pw = ProgressWrapper(1, parent=self)
                        pw.show()
                        pw.getCallingFunction()("text")("Convoluting feature pairs")

                        findIsoPairsInstance = FindIsoPairs(
                            files[0],
                            exOperator=str(self.ui.exOperator_LineEdit.text()),
                            exExperimentID=str(self.ui.exExperimentID_LineEdit.text()),
                            exComments=str(self.ui.exComments_TextEdit.toPlainText()),
                            exExperimentName=str(self.ui.exExperimentName_LineEdit.text()),
                            metabolisationExperiment=self.labellingExperiment == TRACER,
                            labellingisotopeA=str(self.ui.isotopeAText.text()),
                            labellingisotopeB=str(self.ui.isotopeBText.text()),
                            xOffset=self.isotopeBmass - self.isotopeAmass,
                            useRatio=self.ui.useRatio.isChecked(),
                            minRatio=self.ui.minRatio.value(),
                            maxRatio=self.ui.maxRatio.value(),
                            useCIsotopePatternValidation=int(self.ui.useCValidation.checkState().value),
                            configuredTracer=self.configuredTracer,
                            intensityThreshold=self.ui.intensityThreshold.value(),
                            intensityCutoff=self.ui.intensityCutoff.value(),
                            startTime=self.ui.scanStartTime.value(),
                            stopTime=self.ui.scanEndTime.value(),
                            maxLoading=self.ui.maxLoading.value(),
                            xCounts=str(self.ui.xCountSearch.text()),
                            isotopicPatternCountLeft=self.ui.isotopePatternCountA.value(),
                            isotopicPatternCountRight=self.ui.isotopePatternCountB.value(),
                            lowAbundanceIsotopeCutoff=self.ui.isoAbundance.checkState() == QtCore.Qt.Checked,
                            purityN=self.ui.isotopicAbundanceA.value() / 100.0,
                            purityL=self.ui.isotopicAbundanceB.value() / 100.0,
                            intensityErrorN=self.ui.baseRange.value() / 100.0,
                            intensityErrorL=self.ui.isotopeRange.value() / 100.0,
                            scanIndexOffset=self.ui.scanIndexOffset.value(),
                            minSpectraCount=self.ui.minSpectraCount.value(),
                            clustPPM=self.ui.clustPPM.value(),
                            chromPeakPPM=self.ui.wavelet_EICppm.value(),
                            eicSmoothingWindow=str(self.ui.eicSmoothingWindow.currentText()),
                            eicSmoothingWindowSize=self.ui.eicSmoothingWindowSize.value(),
                            eicSmoothingPolynom=self.ui.smoothingPolynom_spinner.value(),
                            artificialMPshift_start=self.ui.spinBox_artificialMPshift_start.value(),
                            artificialMPshift_stop=self.ui.spinBox_artificialMPshift_stop.value(),
                            calcIsoRatioNative=self.ui.calcIsoRatioNative_spinBox.value(),
                            calcIsoRatioLabelled=self.ui.calcIsoRatioLabelled_spinBox.value(),
                            calcIsoRatioMoiety=self.ui.calcIsoRatioMoiety_spinBox.value(),
                            minCorrelationConnections=self.ui.minCorrelationConnections.value() / 100.0,
                            positiveScanEvent=str(self.ui.positiveScanEvent.currentText()),
                            negativeScanEvent=str(self.ui.negativeScanEvent.currentText()),
                            correctCCount=self.ui.correctcCount.checkState() == QtCore.Qt.Checked,
                            minCorrelation=self.ui.minCorrelation.value() / 100.0,
                            hAIntensityError=self.ui.hAIntensityError.value() / 100.0,
                            hAMinScans=self.ui.hAMinScans.value(),
                            adducts=self.adducts,
                            elements=self.elementsForNL,
                            heteroAtoms=self.heteroElements,
                            simplifyInSourceFragments=self.ui.checkBox_simplifyInSourceFragments.isChecked(),
                            lock=None,
                            queue=None,
                            pID=1,
                            meVersion="MetExtract (%s)" % MetExtractVersion,
                            peak_picker=picker,
                            peak_filter_config=filter_config,
                        )

                        procProc = FuncProcess(
                            _target=calculateMetaboliteGroups,
                            file=excel_file,
                            toFile=excel_file,
                            sheet_name=convolution_input_sheet,
                            new_sheet_name="4_Convoluted",
                            groups=definedGroups,
                            eicPPM=self.ui.wavelet_EICppm.value(),
                            maxAnnotationTimeWindow=self.ui.maxAnnotationTimeWindow.value(),
                            minConnectionsInFiles=self.ui.metaboliteClusterMinConnections.value(),
                            minConnectionRate=self.ui.minConnectionRate.value() / 100.0,
                            minPeakCorrelation=self.ui.minCorrelation.value() / 100.0,
                            useAbundanceSimilarity=self.ui.useAbundanceSimilarityForConvolution.checkState() == QtCore.Qt.Checked,
                            abundanceSimilarityThreshold=self.ui.abundanceSimilarityThreshold.value() / 100.0,
                            useRatio=self.ui.useSILRatioForConvolution.checkState() == QtCore.Qt.Checked,
                            cpus=min(len(files), cpus),
                        )

                        # Create a shared Queue and Lock and attach to the FindIsoPairs instance
                        q = procProc.getQueue()
                        lock = Lock()
                        findIsoPairsInstance.queue = q
                        findIsoPairsInstance.lock = lock

                        procProc.addKwd("pwMaxSet", q)
                        procProc.addKwd("pwValSet", q)
                        procProc.addKwd("pwTextSet", q)
                        procProc.addKwd("runIdentificationInstance", findIsoPairsInstance)
                        procProc.start()

                        pw.setCloseCallback(
                            closeCallBack=CallBackMethod(
                                _target=interruptConvolutingOfFeaturePairs,
                                selfObj=self,
                                funcProc=procProc,
                            ).getRunMethod()
                        )

                        # check for status updates
                        while procProc.isAlive():
                            QtWidgets.QApplication.processEvents()
                            while not (procProc.getQueue().empty()):
                                mes = procProc.getQueue().get(block=False, timeout=1)

                                # No idea why / where there are sometimes other objects than Bunch(mes, val), but they occur
                                if isinstance(mes, Bunch) and hasattr(mes, "mes") and (hasattr(mes, "val") or hasattr(mes, "text")):
                                    pw.getCallingFunction()(mes.mes)(mes.val)
                                elif mes == (None, None):
                                    ## I have no idea where this object comes from
                                    pass
                                else:
                                    logging.critical("UNKNONW OBJECT IN PROCESSING QUEUE: %s" % str(mes))

                            time.sleep(0.5)

                        elapsed = (time.time() - start) / 60.0
                        hours = ""
                        if elapsed >= 60.0:
                            hours = "%d hours " % (elapsed // 60)
                        mins = "%.2f mins" % (elapsed % 60.0)

                        if self.terminateJobs:
                            return
                        else:
                            logging.info("Convoluting feature pairs finished (%s%s).." % (hours, mins))
                            annotation_input_sheet = "4_Convoluted"

                        _step_elapsed["convolution"] = (time.time() - _convolution_step_start) / 60.0
                        _step_status["convolution"] = _ST_OK
                        try:
                            import openpyxl as _opxl_cv

                            _wb_cv = _opxl_cv.load_workbook(excel_file, read_only=True)
                            if "4_Convoluted" in _wb_cv.sheetnames:
                                _ws_cv = _wb_cv["4_Convoluted"]
                                _headers_cv = [c.value for c in next(_ws_cv.iter_rows(max_row=1))]
                                if "OGroup" in _headers_cv:
                                    _og_col = _headers_cv.index("OGroup")
                                    _ogroups = {row[_og_col] for row in _ws_cv.iter_rows(min_row=2, values_only=True) if row[_og_col] is not None}
                                    _n_groups = len(_ogroups)
                                    _n_feats_cv = max(0, _ws_cv.max_row - 1)
                                else:
                                    _n_groups = 0
                                    _n_feats_cv = max(0, _ws_cv.max_row - 1)
                            else:
                                _n_groups = 0
                                _n_feats_cv = 0
                            _wb_cv.close()
                            _step_details["convolution"] = f"{_n_feats_cv} feature(s) in {_n_groups} group(s)"
                        except Exception:
                            _step_details["convolution"] = "grouping complete"

                    except Exception as ex:
                        traceback.print_exc()
                        logging.error(str(traceback))

                        _step_elapsed["convolution"] = (time.time() - _convolution_step_start) / 60.0
                        _step_status["convolution"] = _ST_ERROR
                        _step_details["convolution"] = str(ex)
                        _convolution_failed = True

                        QtWidgets.QMessageBox.warning(
                            self,
                            "MetExtract",
                            "Error during convolution of feature pairs: '%s'" % str(ex),
                            QtWidgets.QMessageBox.Ok,
                        )
                        errorCount += 1
                    finally:
                        pw.setSkipCallBack(True)
                    logging.info("##############################################################")

            pw.setSkipCallBack(True)
            pw.hide()

            if self.terminateJobs:
                return

        annotationColumns = []
        ## annotate results
        if self.ui.annotateMetabolites_CheckBox.isChecked():
            if _bracketing_failed or _convolution_failed:
                _step_status["annotation"] = _ST_SKIPPED_PREV
                logging.info("Skipping annotation: a preceding step failed.")
            else:
                logging.info("\n\n##############################################################")
                logging.info("Annotation of detected metabolites..")

                _annotation_step_start = time.time()
                _annotation_error = False

                useAdducts = []
                for adduct in self.adducts:
                    useAdducts.append(
                        [
                            str(adduct.name),
                            adduct.mzoffset,
                            str(adduct.polarity),
                            adduct.charge,
                            adduct.mCount,
                        ]
                    )

                pw = ProgressWrapper(1, parent=self)
                pw.show()
                pw.getCallingFunction()("text")("Annotation of detected metabolites")
                pw.getCallingFunction()("header")("Annotating..")

                if self.ui.searchDB_checkBox.isChecked():
                    pw.getCallingFunction()("text")("Searching hits in databases..")

                    # Collect database files
                    dbFiles = []
                    for entryInd in range(self.ui.dbList_listView.model().rowCount()):
                        dbFile = str(self.ui.dbList_listView.model().item(entryInd, 0).data())
                        dbFiles.append(dbFile)

                    # Prepare parameters
                    useExactXn = str(self.ui.sumFormulasUseExactXn_ComboBox.currentText())
                    if useExactXn.lower() == "plusminus":
                        useExactXn = "PlusMinus_%d" % (self.ui.sumFormulasPlusMinus_spinBox.value())

                    try:
                        db_info_messages = []
                        smiles_mismatches = []
                        addedColumns = annotateResultMatrix.annotateWithDatabases(
                            file=excel_file,
                            sheet_name=annotation_input_sheet,
                            new_sheet_name="5_Annotated",
                            dbFiles=dbFiles,
                            useAdducts=useAdducts,
                            ppm=self.ui.annotationMaxPPM_doubleSpinBox.value(),
                            correctppmPosMode=self.ui.annotation_correctMassByPPMposMode.value(),
                            correctppmNegMode=self.ui.annotation_correctMassByPPMnegMode.value(),
                            rtError=self.ui.maxRTErrorInHits_spinnerBox.value(),
                            useRt=self.ui.checkRTInHits_checkBox.isChecked(),
                            checkXnInHits=useExactXn,
                            processedElement=getElementOfIsotope(str(self.ui.isotopeAText.text())),
                            pwMaxSet=pw.getCallingFunction()("max"),
                            pwValSet=pw.getCallingFunction()("value"),
                            db_info_messages=db_info_messages,
                            smiles_mismatches=smiles_mismatches,
                        )
                        annotationColumns.extend(addedColumns)

                        # Inform the user about SMILES / sum formula mismatches
                        if smiles_mismatches:
                            logging.warning(f"{len(smiles_mismatches)} database entries had a SMILES code that does not match the provided sum formula.\nThese entries were not used for the annotation. See the console / DB_info sheet for details.")

                        if False:
                            logging.info(
                                "## Database search: checkXn %s, ppm: %.5f, correctppm: pos.mode: %.5f / neg.mode: %.5f, Adducts: %s"
                                % (
                                    useExactXn,
                                    self.ui.annotationMaxPPM_doubleSpinBox.value(),
                                    self.ui.annotation_correctMassByPPMposMode.value(),
                                    self.ui.annotation_correctMassByPPMnegMode.value(),
                                    str(useAdducts),
                                )
                            )

                    except Exception as e:
                        traceback.print_exc()
                        logging.error(f"Error during database search: {e}")
                        QtWidgets.QMessageBox.warning(
                            self,
                            "MetExtract",
                            f"Error during database search annotation: {str(e)}",
                            QtWidgets.QMessageBox.Ok,
                        )
                        _annotation_error = True
                        errorCount += 1
                    logging.info("##############################################################\n\n")

                if self.ui.generateSumFormulas_CheckBox.isChecked():
                    pw.getCallingFunction()("text")("Generating sum formulas..")
                    pw.getCallingFunction()("value")(0)

                    try:
                        fT = formulaTools()
                        elemsMin = fT.parseFormula(str(self.ui.sumFormulasMinimumElements_lineEdit.text()))
                        elemsMax = fT.parseFormula(str(self.ui.sumFormulasMaximumElements_lineEdit.text()))

                        useAtoms = []
                        if getElementOfIsotope(str(self.ui.isotopeAText.text())) in elemsMax.keys():
                            useAtoms.append(getElementOfIsotope(str(self.ui.isotopeAText.text())))

                        if "C" in elemsMax.keys() and "C" not in useAtoms:
                            useAtoms.append("C")
                        if "H" in elemsMax.keys() and "H" not in useAtoms:
                            useAtoms.append("H")
                        if "N" in elemsMax.keys() and "N" not in useAtoms:
                            useAtoms.append("N")
                        if "O" in elemsMax.keys() and "O" not in useAtoms:
                            useAtoms.append("O")
                        if "S" in elemsMax.keys() and "S" not in useAtoms:
                            useAtoms.append("S")

                        for atom in elemsMax.keys():
                            if atom not in useAtoms:
                                useAtoms.append(atom)

                        atomsRange = []
                        for atom in useAtoms:
                            minE = 0
                            if atom in elemsMin.keys():
                                minE = elemsMin[atom]
                            maxE = elemsMax[atom]
                            atomsRange.append([minE, maxE])

                        useExactXn = str(self.ui.sumFormulasUseExactXn_ComboBox.currentText())
                        if useExactXn.lower() == "plusminus":
                            useExactXn = "PlusMinus_%d" % (self.ui.sumFormulasPlusMinus_spinBox.value())

                        addedColumns = annotateResultMatrix.annotateWithSumFormulas(
                            file=excel_file,
                            sheet_name="5_Annotated",
                            useAtoms=useAtoms,
                            atomsRange=atomsRange,
                            processedElement=getElementOfIsotope(str(self.ui.isotopeAText.text())),
                            useExactXn=useExactXn,
                            ppm=self.ui.annotationMaxPPM_doubleSpinBox.value(),
                            ppmCorrectionPosMode=self.ui.annotation_correctMassByPPMposMode.value(),
                            ppmCorrectionNegMode=self.ui.annotation_correctMassByPPMnegMode.value(),
                            useAdducts=useAdducts,
                            pwMaxSet=pw.getCallingFunction()("max"),
                            pwValSet=pw.getCallingFunction()("value"),
                            nCores=min(len(files), cpus),
                        )
                        annotationColumns.extend(addedColumns)

                    except Exception as e:
                        traceback.print_exc()
                        logging.error(f"Error during sum formula generation: {e}")
                        QtWidgets.QMessageBox.warning(
                            self,
                            "MetExtract",
                            f"Error during sum formula generation: {str(e)}",
                            QtWidgets.QMessageBox.Ok,
                        )
                        _annotation_error = True
                        errorCount += 1
                    logging.info("\n\n##############################################################")

                print("\n\n")
                pw.hide()

                _step_elapsed["annotation"] = (time.time() - _annotation_step_start) / 60.0
                if _annotation_error:
                    _step_status["annotation"] = _ST_ERROR
                    _step_details["annotation"] = "annotation encountered errors"
                else:
                    _step_status["annotation"] = _ST_OK
                    try:
                        import openpyxl as _opxl_an

                        _wb_an = _opxl_an.load_workbook(excel_file, read_only=True)
                        _n_feat_an = max(0, _wb_an["5_Annotated"].max_row - 1) if "5_Annotated" in _wb_an.sheetnames else 0
                        _wb_an.close()
                        _step_details["annotation"] = f"{_n_feat_an} feature(s) annotated"
                    except Exception:
                        _step_details["annotation"] = "annotation complete"

                logging.info("##############################################################")

        ## Process MSMS info
        if self.ui.generateMSMSInfo_CheckBox.isChecked():
            logging.info("\n\n##############################################################")
            pw = ProgressWrapper(1, parent=self)
            pw.show()
            pw.getCallingFunction()("text")("Preparing MSMS import and list generation")

            inFile = resFileFull
            outFile = resFileFull.replace(".tsv", "_withMSMS.tsv")

            definedGroups = self.getAllSampleGroups()
            samplesForMSMS = []
            for group in definedGroups:
                if group.useAsMSMSTarget:
                    for i in range(len(group.files)):
                        fi = group.files[i]
                        fi = fi.replace("\\", "/")
                        fi = fi[fi.rfind("/") + 1 :]
                        fi = fi.replace(".mzxml", "").replace(".mzXML", "")
                        if str.isdigit(fi[0]):
                            fi = "_" + fi

                        samplesForMSMS.append(fi)

            if len(samplesForMSMS) == 0:
                QtWidgets.QMessageBox.warning(
                    self,
                    "MetExtract",
                    "Error: No sample(s) selected to be used as MSMS targets",
                    QtWidgets.QMessageBox.Ok,
                )
            else:
                opt = optimizeMSMSTargets.OptimizeMSMSTargetList()

                opt.readTargetsFromFile(file=inFile, samplesToUse=samplesForMSMS)

                opt.getMostAbundantFileList(minCounts=self.ui.msms_minCounts.value())

                opt.generateMSMSLists(
                    samplesForMSMS,
                    fileTo=inFile,
                    minCounts=self.ui.msms_minCounts.value(),
                    numberOfFiles=self.ui.msms_numberOfSamples.value(),
                    rtPlusMinus=self.ui.msms_rtWindow.value(),
                    maxParallelTargets=self.ui.msms_maxParallelTargets.value(),
                    noffsprings=self.ui.nOffsprings.value(),
                    ngenerations=self.ui.nGenerations.value(),
                    pwSetText=pw.getCallingFunction()("text"),
                    pwSetMax=pw.getCallingFunction()("max"),
                    pwSetValue=pw.getCallingFunction()("value"),
                )

                opt.writeTargetsToFile(inFile, outFile)

                shutil.copyfile(
                    resFileFull.replace(".tsv", "_withMSMS.tsv"),
                    resFilePath + "/xxx_results__9_withMSMSInfo.tsv",
                )

            pw.hide()
            logging.info("##############################################################")

        self.updateLCMSSampleSettings()

        # Log overall time
        _overall_elapsed = (time.time() - overallStart) / 60.0
        elapsed = _overall_elapsed
        hours = ""
        if elapsed >= 60.0:
            hours = "%d hours " % (elapsed // 60)
        mins = "%.2f mins" % (elapsed % 60.0)

        if errorCount == 0:
            logging.info("Processing successfully finished (%s%s)..\n" % (hours, mins))
        else:
            logging.warning("Processing finished with %d errors (%s%s)..\n" % (errorCount, hours, mins))

        notification_msg = "Processing finished in %s%s" % (hours, mins)
        if errorCount > 0:
            notification_msg = "Processing finished with %d errors in %s%s" % (errorCount, hours, mins)
        self._send_desktop_notification("MetExtract II", notification_msg)

        # Build and show the processing summary dialog
        def _fmt_elapsed(m):
            if m <= 0:
                return "—"
            if m < 1:
                return f"{m * 60:.1f} sec"
            if m < 60:
                return f"{m:.2f} min"
            return f"{int(m // 60)}h {m % 60:.1f} min"

        _step_names = {
            "individual_files": "1. Individual file processing",
            "bracketing": "2. Bracketing",
            "reintegration": "3. Re-integration",
            "convolution": "4. Feature grouping",
            "annotation": "5. Annotation",
        }
        _status_style = {
            _ST_OK: ("olivedrab", "OK"),
            _ST_ERROR: ("firebrick", "Error"),
            _ST_SKIPPED_USER: ("gray", "Skipped (not enabled)"),
            _ST_SKIPPED_PREV: ("darkorange", "Skipped (previous step failed)"),
        }

        rows_html = ""
        for _key in ("individual_files", "bracketing", "reintegration", "convolution", "annotation"):
            _color, _label = _status_style.get(_step_status[_key], ("gray", _step_status[_key]))
            _det = _step_details[_key] or "—"
            _dur = _fmt_elapsed(_step_elapsed[_key])
            rows_html += f"<tr><td style='padding:4px 10px;'>{_step_names[_key]}</td><td style='padding:4px 10px; color:{_color}; font-weight:bold;'>{_label}</td><td style='padding:4px 10px;'>{_det}</td><td style='padding:4px 10px; text-align:right;'>{_dur}</td></tr>"

        _summary_html = f"""<html><body style='font-family:sans-serif; font-size:10pt;'>
<h3 style='color:#333; margin-bottom:6px;'>Processing Summary</h3>
<table border='0' cellspacing='0' cellpadding='0'
       style='border-collapse:collapse; width:100%;'>
  <tr style='background:#ddd;'>
    <th align='left'  style='padding:5px 10px;'>Step</th>
    <th align='left'  style='padding:5px 10px;'>Status</th>
    <th align='left'  style='padding:5px 10px;'>Details</th>
    <th align='right' style='padding:5px 10px;'>Duration</th>
  </tr>
  {rows_html}
</table>
<p style='margin-top:10px;'>
  <b>Total time:</b> {_fmt_elapsed(_overall_elapsed)}
</p>
<p>Results can be viewed:</p>
<ul>
  <li>In this application via the <b>Experimental results</b> tab</li>
  <li>In the Excel file at:<br><tt>{excel_file}</tt></li>
</ul>
</body></html>"""

        _dlg = QtWidgets.QDialog(self)
        _dlg.setWindowTitle("MetExtract II – Processing Summary")
        _dlg.setMinimumSize(800, 420)
        _dlg_layout = QtWidgets.QVBoxLayout(_dlg)
        _browser = QtWidgets.QTextBrowser(_dlg)
        _browser.setHtml(_summary_html)
        _browser.setOpenExternalLinks(False)
        _dlg_layout.addWidget(_browser)
        _btn_box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok, parent=_dlg)
        _btn_box.accepted.connect(_dlg.accept)
        _dlg_layout.addWidget(_btn_box)
        _dlg.exec()

        self.loadGroupsResultsFile(str(self.ui.groupsSave.text()))

    def showResultsSummary(self):
        from .mePyGuis.ResultsSummaryDialog import ResultsSummaryDialog

        definedGroups = self.getAllSampleGroups()
        for group in definedGroups:
            for i in range(len(group.files)):
                group.files[i] = str(group.files[i]).replace("\\", "/")

        # ── Collect per-file rows ──────────────────────────────────────
        file_rows = []
        for group in definedGroups:
            grName = str(group.name)
            grColor = group.color if group.color else ""
            for file in natSort(group.files):
                if not os.path.exists(file + getDBSuffix()):
                    for ionMode in ["+", "-"]:
                        file_rows.append(
                            {
                                "group_name": grName,
                                "group_color": grColor,
                                "file": file,
                                "ion_mode": ionMode,
                                "error": "File not processed",
                            }
                        )
                    continue
                for ionMode in ["+", "-"]:
                    try:
                        file_db_con = PolarsDB(file + getDBSuffix(), format=getDBFormat())

                        nMZs = len(file_db_con.tables["MZs"].filter(pl.col("ionMode") == ionMode))
                        ppm_df = file_db_con.tables["MZs"].filter(pl.col("ionMode") == ionMode).with_columns(((pl.col("lmz") - pl.col("mz") - pl.col("tmz")) * 1_000_000 / pl.col("mz")).alias("ppm_delta"))
                        nMZsPPMDelta = ppm_df["ppm_delta"].mean()
                        nMZsPPMDeltaStd = ppm_df["ppm_delta"].std()

                        nMZBins = file_db_con.tables["MZBins"].filter(pl.col("ionMode") == ionMode).shape[0]
                        nFeatures = file_db_con.tables["chromPeaks"].filter(pl.col("ionMode") == ionMode).shape[0]
                        nMetabolites = file_db_con.tables["featureGroups"].shape[0]

                        ratio_df = file_db_con.tables["MZs"].filter(pl.col("ionMode") == ionMode).select((pl.col("intensity") / pl.col("intensityL")).alias("ratio"))
                        avgRatioSignals = ratio_df["ratio"].mean()
                        avgRatioSignalsStd = ratio_df["ratio"].std()

                        avgRatioFeaturesArea = file_db_con.tables["chromPeaks"].filter(pl.col("ionMode") == ionMode).select((pl.col("NPeakArea") / pl.col("LPeakArea")).alias("area_ratio"))["area_ratio"].mean()

                        avgRatioFeaturesAbundance = file_db_con.tables["chromPeaks"].filter(pl.col("ionMode") == ionMode).select((pl.col("NPeakAbundance") / pl.col("LPeakAbundance")).alias("abundance_ratio"))["abundance_ratio"].mean()

                        enr_df = file_db_con.tables["chromPeaks"].filter((pl.col("peaksRatioMPm1") > 0) & (pl.col("ionMode") == ionMode)).select((pl.col("xcount") / (pl.col("xcount") + pl.col("peaksRatioMPm1"))).alias("enrichment"))
                        avgEnrichmentL = enr_df["enrichment"].mean()
                        avgEnrichmentLStd = enr_df["enrichment"].std()

                        file_rows.append(
                            {
                                "group_name": grName,
                                "group_color": grColor,
                                "file": file,
                                "ion_mode": ionMode,
                                "n_mzs": nMZs if nMZs > 0 else None,
                                "n_mz_bins": nMZBins if nMZBins > 0 else None,
                                "n_features": nFeatures if nFeatures > 0 else None,
                                "n_metabolites": nMetabolites if nMetabolites > 0 else None,
                                "mz_delta_mean": nMZsPPMDelta,
                                "mz_delta_std": nMZsPPMDeltaStd,
                                "avg_ratio_signals": avgRatioSignals,
                                "avg_ratio_signals_std": avgRatioSignalsStd,
                                "avg_ratio_area": avgRatioFeaturesArea,
                                "avg_ratio_abundance": avgRatioFeaturesAbundance,
                                "avg_enrichment": avgEnrichmentL,
                                "avg_enrichment_std": avgEnrichmentLStd,
                                "error": None,
                            }
                        )
                    except Exception as ex:
                        file_rows.append(
                            {
                                "group_name": grName,
                                "group_color": grColor,
                                "file": file,
                                "ion_mode": ionMode,
                                "error": str(ex),
                            }
                        )

        # ── Collect bracketing / convoluted summary ────────────────────
        summary_data = {}
        resFileFull = str(self.ui.groupsSave.text())
        if os.path.exists(resFileFull):
            try:
                res_db = PolarsDB(resFileFull, format="xlsx", load_all_tables=True)

                convoluted_sheet = None
                for candidate in ("3_Reintegrated", "2_StatColumns"):
                    if candidate in res_db.tables:
                        convoluted_sheet = candidate
                        break

                if convoluted_sheet is not None:
                    table_df = res_db.tables[convoluted_sheet]

                    features = set()
                    metabolites = {}
                    metabolitesIonMode = {}

                    for row in table_df.iter_rows(named=True):
                        num = row["Num"]
                        features.add(num)
                        ogroup = row["OGroup"]
                        if ogroup not in metabolites:
                            metabolites[ogroup] = []
                        metabolites[ogroup].append(num)
                        if ogroup not in metabolitesIonMode:
                            metabolitesIonMode[ogroup] = set()
                        metabolitesIonMode[ogroup].add(row["Ionisation_Mode"])

                    summary_data["n_features"] = len(features)
                    summary_data["n_metabolites"] = len(metabolites)
                    summary_data["features_1"] = len([1 for k in metabolites if len(metabolites[k]) == 1])
                    summary_data["features_2"] = len([1 for k in metabolites if len(metabolites[k]) == 2])
                    summary_data["features_3"] = len([1 for k in metabolites if len(metabolites[k]) == 3])
                    summary_data["features_4"] = len([1 for k in metabolites if len(metabolites[k]) == 4])
                    summary_data["features_5"] = len([1 for k in metabolites if len(metabolites[k]) == 5])
                    summary_data["features_5_to_10"] = len([1 for k in metabolites if 5 < len(metabolites[k]) < 11])
                    summary_data["features_10_to_20"] = len([1 for k in metabolites if 10 < len(metabolites[k]) < 21])
                    summary_data["features_gt20"] = len([1 for k in metabolites if 20 < len(metabolites[k])])
                    summary_data["pos_only"] = len([1 for k in metabolitesIonMode if metabolitesIonMode[k] == {"+"}])
                    summary_data["neg_only"] = len([1 for k in metabolitesIonMode if metabolitesIonMode[k] == {"-"}])
                    summary_data["both_modes"] = len([1 for k in metabolitesIonMode if {"+", "-"}.issubset(metabolitesIonMode[k])])

            except Exception as ex:
                import traceback as _tb

                logging.warning(f"Error reading convoluted results for summary: {ex}\n{_tb.format_exc()}")

            try:
                res_db2 = PolarsDB(resFileFull, format="xlsx", load_all_tables=True)
                omitted_sheet = "2_StatColumns_Omitted"
                if omitted_sheet in res_db2.tables:
                    omitted_df = res_db2.tables[omitted_sheet]
                    summary_data["omitted_features"] = len(set(omitted_df["Num"].to_list()))
            except Exception as ex:
                logging.warning(f"Error reading omitted features for summary: {ex}")

        dlg = ResultsSummaryDialog(parent=self, file_rows=file_rows, summary_data=summary_data)
        dlg.exec()

    def groupFilesChanges(self, sta):
        self.ui.label_26.setEnabled(sta)
        self.ui.groupPpm.setEnabled(sta)
        self.ui.alignChromatograms.setEnabled(sta)
        self.ui.label_18.setEnabled(sta)
        self.ui.groupingRT.setEnabled(sta)

    def _configureExperimentalGroupsTableColumns(self):
        header = self.ui.groupsList.horizontalHeader()
        header.setStretchLastSection(False)
        header.setMinimumSectionSize(55)
        header.setSectionsMovable(False)

        # Keep MSMS target column for backward compatibility but hide it from the UI.
        self.ui.groupsList.setColumnHidden(7, True)

        # Keep all columns user-resizable and apply sensible defaults:
        # Name much wider, other metadata columns compact.
        for col in range(self.ui.groupsList.columnCount()):
            header.setSectionResizeMode(col, QtWidgets.QHeaderView.Interactive)

        self.ui.groupsList.setColumnWidth(0, 160)  # Name
        self.ui.groupsList.setColumnWidth(1, 50)  # Files
        self.ui.groupsList.setColumnWidth(2, 50)  # Color
        self.ui.groupsList.setColumnWidth(3, 70)  # Min Found
        self.ui.groupsList.setColumnWidth(4, 60)  # Omit Features
        self.ui.groupsList.setColumnWidth(5, 60)  # Metabolite Grouping
        self.ui.groupsList.setColumnWidth(6, 80)  # False Positive
        self.ui.groupsList.setColumnWidth(7, 90)  # MSMS Target (hidden)

    def saveMZXMLChanged(self, sta):
        self.ui.wm_ia.setEnabled(sta)
        self.ui.wm_iap.setEnabled(sta)
        self.ui.wm_imb.setEnabled(sta)
        self.ui.wm_ib.setEnabled(sta)

    def procIndFilesChanges(self, sta):
        self.ui.frame_procIndFiles.setVisible(sta)

    def processMultipleFilesChanged(self, sta):
        self.ui.frame_bracketResults.setVisible(sta)

    def annotateMetabolitesChanged(self, sta):
        self.ui.frame_annotateMetabolites.setVisible(sta)

    def genrateMSMSListsChanged(self, sta):
        self.ui.frame_generateMSMSTargetLists.setVisible(sta)

    # <editor-fold desc="### visualisation of results of single sample">
    def closeCurrentOpenResultsFile(self):
        if hasattr(self, "currentOpenResultsFile") and self.currentOpenResultsFile is not None:
            self.currentOpenResultsFile.file = None
            self.currentOpenResultsFile = None
        if hasattr(self, "currentOpenRawFile") and self.currentOpenRawFile is not None:
            self.currentOpenRawFile.freeMe()
            self.currentOpenRawFile = None

    def openFileAsCurrentOpenResultsFile(self, file):
        db_path = file + getDBSuffix()
        if os.path.exists(db_path) and os.path.isfile(db_path):
            self.currentOpenResultsFile = Bunch()
            self.currentOpenResultsFile.file = file
            # Load PolarsDB using ZIP archive format
            self.currentOpenResultsFile.db_con = PolarsDB(db_path, format=getDBFormat())
            return True
        else:
            return False

    def selectedResultChanged(self, ind):
        self.ui.res_ExtractedData.clear()
        self.ui.chromPeakName.setText("")

        sortOrder = str(self.ui.sortOrderResults.currentText())

        cInd = self.ui.processedFilesComboBox.currentIndex()
        b = self.ui.processedFilesComboBox.itemData(cInd)

        if not hasattr(b, "file") or b.file is None:
            return -1

        self.closeCurrentOpenResultsFile()
        if self.openFileAsCurrentOpenResultsFile(b.file):
            it = QtWidgets.QTreeWidgetItem(["MZs"])
            self.ui.res_ExtractedData.addTopLevelItem(it)
            it.myType = "MZs"
            count = 0
            children = []
            pw = ProgressWrapper(pwCount=5)

            pw.setText("Reading raw data", i=0)
            pw.setText("", i=1)
            pw.setText("", i=2)
            pw.setTextu("", i=3)
            pw.setTextu("", i=4)
            pw.show()

            try:
                ## Parse raw data
                pw.setMaxu(1, i=0)
                self.currentOpenRawFile = Chromatogram()
                self.currentOpenRawFile.parse_file(b.file, ignoreCharacterData=False)
                pw.setText("Raw data imported (MS1: %d, MS2: %d)" % (len(self.currentOpenRawFile.MS1_list), len(self.currentOpenRawFile.MS2_list)), i=0)
                pw.setValueu(1, i=0)

            except Exception as e:
                logging.warning(f"Could not parse raw file: {e}")
                self.currentOpenRawFile = None
                pw.setText("Raw data import failed", i=0)

            try:
                ## Fetched matched MZ signal pairs
                pw.setText("Fetching mzs", i=1)
                numberOfMZs = 0
                mzs_df = self.currentOpenResultsFile.db_con.tables.get("MZs", pl.DataFrame())
                if len(mzs_df) > 0:
                    numberOfMZs = len(mzs_df)
                pw.setMaxu(numberOfMZs, i=1)

                maxMZsFetch = 50000

                if numberOfMZs < maxMZsFetch:
                    pw.setText("Fetching mzs (%d)" % numberOfMZs, i=1)

                    # Sort by scanid and iterate
                    mzs_sorted = mzs_df.sort("scanid")
                    for row_dict in mzs_sorted.to_dicts():
                        d = QtWidgets.QTreeWidgetItem(
                            it,
                            [
                                str(s)
                                for s in [
                                    row_dict["mz"],
                                    row_dict["xcount"],
                                    row_dict["scanid"],
                                    "%.2f min / %.2f sec" % (row_dict["scantime"] / 60.0, row_dict["scantime"]),
                                    row_dict["loading"],
                                    "%.1f / %.1f / %.3f"
                                    % (
                                        row_dict["intensity"],
                                        row_dict["intensityL"],
                                        row_dict["intensity"] / row_dict["intensityL"],
                                    ),
                                ]
                            ],
                        )
                        d.myType = "mz"
                        d.myData = Bunch(**row_dict)
                        d.myID = int(row_dict["id"])
                        children.append(d)
                        count += 1

                        pw.setValueu(count, i=1)
                    pw.setText("%d MZs fetched" % numberOfMZs, i=1)
                else:
                    pw.setTextu("Mzs not fetched (too many; %d)" % numberOfMZs, i=1)
                it.addChildren(children)
                it.setText(1, "%d" % numberOfMZs)

            except Exception:
                logging.warning("Error loading MZs for sample results", exc_info=True)

            try:
                it = QtWidgets.QTreeWidgetItem(["MZ bins"])
                it.myType = "MZBins"
                self.ui.res_ExtractedData.addTopLevelItem(it)
                mzbins_df = self._get_result_table("MZBins", "mzbins")

                count = 0
                children = []

                ## Load mz bins — build all Qt items first, then add them to the tree in one shot.
                pw.setText("Fetching MZBins (%d)" % len(mzbins_df), i=2)
                pw.setMax(len(mzbins_df), i=2)

                mzbins_kids_df = self._get_result_table("MZBinsKids", "mzbinskids")
                mzbin_id_col = self._resolve_col_name(mzbins_kids_df, "mzbinID", "mzbinid")
                mz_id_col = self._resolve_col_name(mzbins_kids_df, "mzID", "mzid")
                mz_id_join_col = self._resolve_col_name(mzs_df, "id")
                can_load_child_mzs = numberOfMZs < maxMZsFetch

                # Build lookup dictionaries in Python using foreign keys.
                mzbins_rows = mzbins_df.sort("mz").to_dicts() if len(mzbins_df) > 0 else []
                kids_rows = []
                if len(mzbins_kids_df) > 0 and mzbin_id_col is not None:
                    if mz_id_col is not None:
                        kids_rows = mzbins_kids_df.select([mzbin_id_col, mz_id_col]).to_dicts()
                    else:
                        kids_rows = mzbins_kids_df.select([mzbin_id_col]).to_dicts()

                mzs_by_bin = {}
                kids_count_by_bin = {}
                kids_by_bin = {}
                for kid in kids_rows:
                    bid = kid[mzbin_id_col]
                    kids_by_bin.setdefault(bid, []).append(kid)
                for bid, bid_kids in kids_by_bin.items():
                    kids_count_by_bin[bid] = len(bid_kids)

                if can_load_child_mzs and len(mzs_df) > 0 and mz_id_col is not None and mz_id_join_col is not None:
                    mz_rows = mzs_df.select([mz_id_join_col, "mz", "xcount", "scanid", "scantime", "loading", "intensity"]).to_dicts()
                    mz_by_id = {row[mz_id_join_col]: row for row in mz_rows}
                    for bid, bid_kids in kids_by_bin.items():
                        mapped_rows = []
                        for kid in bid_kids:
                            mz_row = mz_by_id.get(kid[mz_id_col])
                            if mz_row is not None:
                                mapped_rows.append(mz_row)
                        if mapped_rows:
                            mapped_rows.sort(key=lambda row: row["scanid"])
                            mzs_by_bin[bid] = mapped_rows

                self.ui.res_ExtractedData.setUpdatesEnabled(False)
                try:
                    for mzbin_row in mzbins_rows:
                        mzbin_id = mzbin_row.get("id") or mzbin_row.get("ID")
                        mzbin_mz = mzbin_row.get("mz") or mzbin_row.get("MZ") or 0.0
                        if mzbin_id is None:
                            continue

                        mz_rows = mzs_by_bin.get(mzbin_id, [])
                        countinner = kids_count_by_bin.get(mzbin_id, len(mz_rows))
                        min_inner = 1000000.0
                        max_inner = 0.0
                        xcount = 0

                        d = QtWidgets.QTreeWidgetItem()
                        d.myType = "mzbin"
                        d.myID = int(mzbin_id)
                        children.append(d)

                        if mz_rows:
                            child_items = []
                            for mz_row in mz_rows:
                                mz_mz = mz_row["mz"]
                                if mz_mz < min_inner:
                                    min_inner = mz_mz
                                if mz_mz > max_inner:
                                    max_inner = mz_mz
                                xcount = mz_row["xcount"]
                                dd = QtWidgets.QTreeWidgetItem(
                                    [
                                        str(mz_mz),
                                        str(xcount),
                                        str(mz_row["scanid"]),
                                        "%.2f min / %.2f sec" % (mz_row["scantime"] / 60.0, mz_row["scantime"]),
                                        str(mz_row["loading"]),
                                        "%.1f" % mz_row["intensity"],
                                    ]
                                )
                                dd.myType = "mz"
                                dd.myData = Bunch(mz=mz_mz, scantime=mz_row["scantime"], id=mz_row[mz_id_join_col])
                                dd.myID = int(mz_row[mz_id_join_col])
                                child_items.append(dd)
                            d.addChildren(child_items)

                        d.setText(0, "%.5f (%d)" % (mzbin_mz, countinner))
                        d.setText(1, "%.4f" % ((max_inner - min_inner) * 1000000.0 / min_inner) if min_inner < 1000000.0 else "0.0")
                        d.setText(2, "%s" % xcount)
                        count += 1
                finally:
                    self.ui.res_ExtractedData.setUpdatesEnabled(True)

                if can_load_child_mzs:
                    pw.setText("%d MZBins fetched" % count, i=2)
                else:
                    pw.setTextu("%d MZBins fetched (child MZ details skipped; too many MZs: %d)" % (count, numberOfMZs), i=2)

                it.addChildren(children)
                it.setText(1, "%d" % len(mzbins_df))

            except Exception:
                logging.warning("Error loading MZBins for sample results", exc_info=True)

            try:
                ## Load feature pairs
                it = QtWidgets.QTreeWidgetItem(["Feature pairs"])
                self.ui.res_ExtractedData.addTopLevelItem(it)
                it.myType = "Features"

                chromPeaks_df = self.currentOpenResultsFile.db_con.tables.get("chromPeaks", pl.DataFrame())
                numberOfFeaturePairs = len(chromPeaks_df)

                pw.setTextu("Fetching feature pairs (%d)" % numberOfFeaturePairs, i=3)
                pw.setMaxu(numberOfFeaturePairs, i=3)

                count = 0
                children = []

                if len(chromPeaks_df) > 0:
                    # LEFT JOIN with tracerConfiguration
                    tracer_df = self.currentOpenResultsFile.db_con.tables.get("tracerConfiguration", pl.DataFrame())
                    if len(tracer_df) > 0:
                        joined_df = chromPeaks_df.join(tracer_df.select(["id", "name"]), left_on="tracer", right_on="id", how="left", suffix="_tracer")
                    else:
                        joined_df = chromPeaks_df.with_columns(pl.lit(None).alias("name"))

                    # Determine sort column based on sortOrder
                    sort_map = {
                        "M/Z": "mz",
                        "RT": "NPeakCenter",
                        "Intensity": "NPeakArea",
                        "Peaks correlation": "peaksCorr",
                    }
                    sort_col = sort_map.get(sortOrder, "mz")

                    # Sort (descending for Intensity and Peaks correlation)
                    if sortOrder in ["Intensity", "Peaks correlation"]:
                        joined_df = joined_df.sort(["tracer", sort_col, "NPeakCenter", "mz", "xcount"], descending=[False, True, False, False, False])
                    else:
                        joined_df = joined_df.sort(["tracer", sort_col, "NPeakCenter", "mz", "xcount"])

                    for row_dict in joined_df.to_dicts():
                        # Convert row_dict to Bunch for backwards compatibility
                        row = Bunch(**row_dict)
                        row.cpID = row_dict["id"]
                        row.tracerName = row_dict.get("name", "")

                        adducts = ""
                        try:
                            lk = loads(base64.b64decode(row.adducts.encode("utf-8")))
                            if len(lk) > 0:
                                adducts = ", ".join(lk)
                        except (ImportError, ModuleNotFoundError, SyntaxError, EOFError) as e:
                            logging.warning(f"Could not load adducts from row data: {e}")
                            adducts = ""

                        fDesc = ""
                        try:
                            lk = loads(base64.b64decode(row.fDesc.encode("utf-8")))
                            if len(lk) > 0:
                                fDesc = ", ".join(lk)
                        except (ImportError, ModuleNotFoundError, SyntaxError, EOFError) as e:
                            logging.warning(f"Could not load feature description from row data: {e}")
                            fDesc = ""

                        heteroAtoms = []
                        try:
                            lk = loads(base64.b64decode(row.heteroAtoms.encode("utf-8")))
                        except (ImportError, ModuleNotFoundError, SyntaxError) as e:
                            logging.warning(f"Could not load hetero atoms from results: {e}")
                            logging.warning("This may be due to cached data from an older version. Using empty dict.")
                            lk = {}
                        for hetAtom in lk:
                            pIso = lk[hetAtom]
                            for hetAtomCount in pIso:
                                heteroAtoms.append("{%s}%d" % (hetAtom, hetAtomCount))
                        heteroAtoms = ", ".join(heteroAtoms)

                        try:
                            assignedMZs = loads(base64.b64decode(row.assignedMZs.encode("utf-8")))
                        except (ImportError, ModuleNotFoundError, SyntaxError) as e:
                            logging.warning(f"Could not load assigned MZs from results: {e}")
                            logging.warning("This may be due to cached data from an older version. Using empty list.")
                            assignedMZs = []

                        try:
                            row.xcount = int(row.xcount)
                        except Exception:
                            traceback.print_exc()  ## double labeling experiment. That's fine

                        xp = ChromPeakPair(
                            NPeakCenter=int(row.NPeakCenter),
                            LPeakScale=float(row.LPeakScale),
                            LPeakCenter=int(row.LPeakCenter),
                            NPeakScale=float(row.NPeakScale),
                            NSNR=0,
                            NPeakArea=-1,
                            mz=float(row.mz),
                            lmz=float(row.lmz),
                            xCount=row.xcount,
                            NBorderLeft=float(row.NBorderLeft),
                            NBorderRight=float(row.NBorderRight),
                            LBorderLeft=float(row.LBorderLeft),
                            LBorderRight=float(row.LBorderRight),
                            NPeakCenterMin=float(row.NPeakCenterMin),
                            LPeakCenterMin=float(row.LPeakCenterMin),
                            eicID=int(row.eicID),
                            massSpectrumID=int(row.massSpectrumID),
                            assignedName=str(row.assignedName),
                            id=int(row.cpID),
                            loading=int(row.Loading),
                            peaksCorr=float(row.peaksCorr),
                            peaksRatio=float(row.peaksRatio),
                            tracer=str(row.tracerName),
                            ionMode=str(row.ionMode),
                            heteroAtoms=heteroAtoms,
                            adducts=adducts,
                            fDesc=fDesc,
                            assignedMZs=assignedMZs,
                            N_startRT=row_dict.get("N_startRT"),
                            N_endRT=row_dict.get("N_endRT"),
                            L_startRT=row_dict.get("L_startRT"),
                            L_endRT=row_dict.get("L_endRT"),
                        )

                        # Check if MSMS spectra exist for this feature
                        try:
                            ppm = float(self.getParametersFromCurrentRes("Mass deviation (+/- ppm)"))
                        except Exception:
                            ppm = 5.0
                        n_start_rt = row_dict.get("N_startRT")
                        n_end_rt = row_dict.get("N_endRT")
                        l_start_rt = row_dict.get("L_startRT")
                        l_end_rt = row_dict.get("L_endRT")
                        if n_start_rt is None:
                            n_start_rt = xp.NPeakCenterMin - xp.NPeakScale
                            n_end_rt = xp.NPeakCenterMin + xp.NPeakScale
                            l_start_rt = xp.LPeakCenterMin - xp.LPeakScale
                            l_end_rt = xp.LPeakCenterMin + xp.LPeakScale
                        rt_min = min(float(n_start_rt), float(l_start_rt))
                        rt_max = max(float(n_end_rt), float(l_end_rt))
                        has_msms = self.hasMSMSSpectra(xp.mz, rt_min, rt_max, ppm) or self.hasMSMSSpectra(xp.lmz, rt_min, rt_max, ppm)
                        msms_marker = "* " if has_msms else ""

                        d = QtWidgets.QTreeWidgetItem(
                            [
                                msms_marker + "%.5f" % xp.mz + " (/" + str(row.ionMode) + str(row.Loading) + ") ",
                                "%.2f / %.2f"
                                % (
                                    float(row.NPeakCenterMin) / 60.0,
                                    float(row.LPeakCenterMin) / 60.0,
                                ),
                                str(row.xcount),
                                "%s / %s / %s " % (adducts, fDesc, heteroAtoms),
                                "%.0f / %.0f" % (float(row.NPeakScale), float(row.LPeakScale)),
                                "%.3f%s"
                                % (
                                    row.peaksCorr,
                                    "" if xp.artificialEICLShift == 0 else " (%d)" % (xp.artificialEICLShift),
                                ),
                                "%.3f / %.3f" % (row.peaksRatio, row.NPeakArea / row.LPeakArea),
                                "%.1f / %.1f" % (row.NPeakArea, row.LPeakArea),
                                "%d" % len(assignedMZs),
                                "%.5f" % (xp.lmz),
                                str(row.tracerName),
                            ]
                        )

                        d.myType = "feature"
                        d.myID = int(row.cpID)
                        d.myData = xp
                        children.append(d)
                        count += 1

                        pw.setValueu(count, i=3)
                pw.setTextu("%d feature pairs fetched" % numberOfFeaturePairs, i=3)

                it.addChildren(children)
                it.setExpanded(False)
                it.setText(1, "%d" % count)

            except Exception:
                traceback.print_exc()

            try:
                it = QtWidgets.QTreeWidgetItem(["Metabolites" if self.labellingExperiment == METABOLOME else "Biotransformation products"])
                self.ui.res_ExtractedData.addTopLevelItem(it)
                it.myType = "Feature Groups"
                it.setExpanded(True)
                count = 0
                fGs = []

                children = []

                # Load Feature Groups with INNER JOIN on tracerConfiguration
                featureGroups_df = self.currentOpenResultsFile.db_con.tables.get("featureGroups", pl.DataFrame())
                if len(featureGroups_df) > 0 and len(tracer_df) > 0:
                    fg_joined = featureGroups_df.join(tracer_df.select(["id", "name"]), left_on="tracer", right_on="id", how="inner", suffix="_tracer").sort(["tracer", "id"])

                    for fg_dict in fg_joined.to_dicts():
                        fg_row = Bunch(fgID=fg_dict["id"], featureName=fg_dict["featureName"], tracerName=fg_dict["name"])
                        fGs.append(fg_row)

                ## Load Feature groups
                pw.setText("Fetching feature groups (%d)" % len(fGs), i=4)
                pw.setMax(len(fGs), i=4)
                for fG in fGs:
                    d = QtWidgets.QTreeWidgetItem(
                        [
                            str(fG.featureName),
                            "",
                            str(fG.fgID),
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            str(fG.tracerName),
                            "",
                        ]
                    )
                    d.myType = "featureGroup"
                    d.myID = fG.fgID
                    d.myData = fG
                    d.setExpanded(True)
                    cpCount = 0
                    sumRt = 0.0
                    has_msms_in_group = False

                    # Get feature group features for this group
                    featureGroupFeatures_df = self.currentOpenResultsFile.db_con.tables.get("featureGroupFeatures", pl.DataFrame())
                    if len(featureGroupFeatures_df) > 0 and len(chromPeaks_df) > 0 and len(tracer_df) > 0:
                        # Filter for this feature group
                        fg_features = featureGroupFeatures_df.filter(pl.col("fGroupID") == fG.fgID)

                        if len(fg_features) > 0:
                            # JOIN chromPeaks with featureGroupFeatures on id=fID
                            cp_fg_join = chromPeaks_df.join(fg_features.select(["fID", "fDesc"]), left_on="id", right_on="fID", how="inner", suffix="_fg")

                            # INNER JOIN with tracerConfiguration
                            cp_fg_tracer = cp_fg_join.join(tracer_df.select(["id", "name"]), left_on="tracer", right_on="id", how="inner", suffix="_tracer")

                            # Always sort feature pairs in feature groups by Area_N abundance.
                            cp_fg_tracer = cp_fg_tracer.sort(["NPeakArea", "mz", "xcount"], descending=[True, False, False])

                            # Pre-compute max NPeakArea so we can set relative bar ratios below
                            group_max_area = cp_fg_tracer["NPeakArea"].max() or 0.0

                            for row_dict in cp_fg_tracer.to_dicts():
                                row = Bunch(**row_dict)
                                row.cpID = row_dict["id"]
                                row.tracerName = row_dict["name"]
                                row.fDesc_fg = row_dict.get("fDesc_fg", row_dict.get("fDesc", ""))
                                adducts = ""
                                try:
                                    lk = loads(base64.b64decode(row.adducts.encode("utf-8")))
                                    if len(lk) > 0:
                                        adducts = ", ".join(lk)
                                except (ImportError, ModuleNotFoundError, SyntaxError) as e:
                                    logging.warning(f"Could not load adducts from row data: {e}")
                                    adducts = ""

                                fDesc = ""
                                try:
                                    lk = loads(base64.b64decode(row.fDesc.encode("utf-8")))
                                    if len(lk) > 0:
                                        fDesc = ", ".join(lk)
                                except (ImportError, ModuleNotFoundError, SyntaxError) as e:
                                    logging.warning(f"Could not load feature description from row data: {e}")
                                    fDesc = ""

                                heteroAtoms = []
                                try:
                                    lk = loads(base64.b64decode(row.heteroAtoms))
                                except (ImportError, ModuleNotFoundError, SyntaxError) as e:
                                    logging.warning(f"Could not load hetero atoms from row data: {e}")
                                    lk = {}
                                for hetAtom in lk:
                                    pIso = lk[hetAtom]
                                    for hetAtomCount in pIso:
                                        heteroAtoms.append("%s%d" % (hetAtom, hetAtomCount))
                                heteroAtoms = ", ".join(heteroAtoms)

                                try:
                                    assignedMZs = loads(base64.b64decode(row.assignedMZs))
                                except (ImportError, ModuleNotFoundError, SyntaxError) as e:
                                    logging.warning(f"Could not load assigned MZs from row data: {e}")
                                    assignedMZs = []

                                xp = ChromPeakPair(
                                    NPeakCenter=int(row.NPeakCenter),
                                    loading=int(row.Loading),
                                    LPeakScale=float(row.LPeakScale),
                                    LPeakCenter=int(row.LPeakCenter),
                                    NPeakScale=float(row.NPeakScale),
                                    NSNR=0,
                                    NPeakArea=-1,
                                    mz=float(row.mz),
                                    lmz=float(row.lmz),
                                    xCount=row.xcount,
                                    NPeakCenterMin=float(row.NPeakCenterMin),
                                    NBorderLeft=float(row.NBorderLeft),
                                    NBorderRight=float(row.NBorderRight),
                                    LBorderLeft=float(row.LBorderLeft),
                                    LBorderRight=float(row.LBorderRight),
                                    LPeakCenterMin=float(row.LPeakCenterMin),
                                    eicID=int(row.eicID),
                                    massSpectrumID=int(row.massSpectrumID),
                                    assignedName=str(row.assignedName),
                                    id=int(row.cpID),
                                    tracer=str(row.tracerName),
                                    ionMode=str(row.ionMode),
                                    adducts=adducts,
                                    heteroAtoms=heteroAtoms,
                                    assignedMZs=assignedMZs,
                                    artificialEICLShift=int(row.artificialEICLShift),
                                )
                                xp.fDesc = str(row.fDesc)
                                xp.peaksCorr = float(row.peaksCorr)
                                xp.peaksRatio = float(row.peaksRatio)

                                # Check if MSMS spectra exist for this feature
                                try:
                                    ppm = float(self.getParametersFromCurrentRes("Mass deviation (+/- ppm)"))
                                except Exception:
                                    ppm = 5.0
                                rt_min = xp.NPeakCenterMin - xp.NPeakScale
                                rt_max = xp.NPeakCenterMin + xp.NPeakScale
                                rt_min = min(rt_min, xp.LPeakCenterMin - xp.LPeakScale)
                                rt_max = max(rt_max, xp.LPeakCenterMin + xp.LPeakScale)
                                has_msms = self.hasMSMSSpectra(xp.mz, rt_min, rt_max, ppm) or self.hasMSMSSpectra(xp.lmz, rt_min, rt_max, ppm)
                                msms_marker = "* " if has_msms else ""

                                # Track if any feature in this group has MSMS
                                if has_msms:
                                    has_msms_in_group = True

                                g = QtWidgets.QTreeWidgetItem(
                                    [
                                        msms_marker + "%.5f" % xp.mz + " (/" + str(row.ionMode) + str(row.Loading) + ") ",
                                        "%.2f / %.2f"
                                        % (
                                            float(row.NPeakCenterMin) / 60.0,
                                            float(row.LPeakCenterMin) / 60.0,
                                        ),
                                        str(row.xcount),
                                        "%s / %s / %s " % (adducts, fDesc, heteroAtoms),
                                        "%.0f / %.0f" % (float(row.NPeakScale), float(row.LPeakScale)),
                                        "%.3f%s"
                                        % (
                                            row.peaksCorr,
                                            "" if xp.artificialEICLShift == 0 else " (%d)" % (xp.artificialEICLShift),
                                        ),
                                        "%.3f / %.3f" % (row.peaksRatio, row.NPeakArea / row.LPeakArea),
                                        "%.1f / %.1f" % (row.NPeakArea, row.LPeakArea),
                                        "%d" % len(assignedMZs),
                                        "%.5f" % (xp.lmz),
                                        str(row.tracerName),
                                    ]
                                )

                                g.myType = "feature"
                                g.myID = int(row.cpID)
                                g.myData = xp
                                if group_max_area > 0.0:
                                    g.setData(0, _RELATIVE_BAR_ROLE, float(row.NPeakArea) / group_max_area)
                                d.addChild(g)

                                sumRt = sumRt + xp.NPeakCenterMin
                                cpCount += 1

                    d.setText(2, str(cpCount))
                    if cpCount > 0:
                        d.setText(1, "%.2f" % (sumRt / cpCount / 60.0))
                    else:
                        d.setText(1, "")

                    # Add asterisk to feature group name if any feature has MSMS
                    if has_msms_in_group:
                        d.setText(0, "* " + str(fG.featureName))

                    children.append(d)
                    count += 1
                    pw.setValueu(count, i=4)
                it.addChildren(children)
                it.setText(1, "%d" % count)

                pw.hide()

            except Exception:
                pass

            try:
                ## Load parameters
                it = QtWidgets.QTreeWidgetItem(["Parameters"])
                it.myType = "parameter"
                self.ui.res_ExtractedData.addTopLevelItem(it)
                config = {}
                config_df = self.currentOpenResultsFile.db_con.tables["config"]
                for row_dict in config_df.to_dicts():
                    config[str(row_dict["key"])] = str(row_dict["value"])

                if config["metabolisationExperiment"] == "True":
                    itl = QtWidgets.QTreeWidgetItem(["Tracers"])
                    it.addChild(itl)
                    itl.myType = "parameter"
                    itl.myType = "parameter"
                    for tracer in loads(base64.b64decode(config["configuredTracers"])):
                        itle = QtWidgets.QTreeWidgetItem([tracer.name])
                        itl.addChild(itle)
                        itle.myType = "parameter"
                        itlee = QtWidgets.QTreeWidgetItem(["Elements", "%d" % tracer.elementCount])
                        itle.addChild(itlee)
                        itlee.myType = "parameter"
                        itlee = QtWidgets.QTreeWidgetItem(["Labelling", "%s/%s" % (tracer.isotopeA, tracer.isotopeB)])
                        itle.addChild(itlee)
                        itlee.myType = "parameter"
                        itlee = QtWidgets.QTreeWidgetItem(
                            [
                                "Delta mz",
                                "%.5f" % (getIsotopeMass(tracer.isotopeB)[0] - getIsotopeMass(tracer.isotopeA)[0]),
                            ]
                        )
                        itle.addChild(itlee)
                        itlee.myType = "parameter"
                        itlee = QtWidgets.QTreeWidgetItem(
                            [
                                "%s purity" % tracer.isotopeA,
                                "%.2f%%" % (float(tracer.enrichmentA) * 100.0),
                            ]
                        )
                        itle.addChild(itlee)
                        itlee.myType = "parameter"
                        itlee = QtWidgets.QTreeWidgetItem(
                            [
                                "%s purity" % tracer.isotopeB,
                                "%.2f%%" % (float(tracer.enrichmentB) * 100.0),
                            ]
                        )
                        itle.addChild(itlee)
                        itlee.myType = "parameter"
                        itlee = QtWidgets.QTreeWidgetItem(["%s amount" % tracer.isotopeA, "%.2f" % tracer.amountA])
                        itle.addChild(itlee)
                        itlee.myType = "parameter"
                        itlee = QtWidgets.QTreeWidgetItem(["%s amount" % tracer.isotopeB, "%.2f" % tracer.amountB])
                        itle.addChild(itlee)
                        itlee.myType = "parameter"
                        itlee = QtWidgets.QTreeWidgetItem(
                            [
                                "%s/%s ratio" % (tracer.isotopeA, tracer.isotopeB),
                                "%.4f" % tracer.monoisotopicRatio,
                            ]
                        )
                        itle.addChild(itlee)
                        itlee.myType = "parameter"
                        itlee = QtWidgets.QTreeWidgetItem(
                            [
                                "Upper error",
                                "%.2f%%" % (float(tracer.maxRelNegBias) * 100.0),
                            ]
                        )
                        itle.addChild(itlee)
                        itlee.myType = "parameter"
                        itlee = QtWidgets.QTreeWidgetItem(
                            [
                                "Lower error",
                                "%.2f%%" % (float(tracer.maxRelPosBias) * 100.0),
                            ]
                        )
                        itle.addChild(itlee)
                        itlee.myType = "parameter"
                else:
                    itl = QtWidgets.QTreeWidgetItem(["Full metabolome labelling experiment"])
                    it.addChild(itl)
                    itl.myType = "parameter"
                    itle = QtWidgets.QTreeWidgetItem(
                        [
                            "%d%s/%d%s"
                            % (
                                float(config["isotopeA"]),
                                config["labellingElement"],
                                float(config["isotopeB"]),
                                config["labellingElement"],
                            )
                        ]
                    )
                    itl.addChild(itle)
                    itle.myType = "parameter"
                    itle = QtWidgets.QTreeWidgetItem(
                        [
                            "%d%s purity" % (float(config["isotopeA"]), config["labellingElement"]),
                            "%.2f%%" % (float(config["purityN"]) * 100.0),
                        ]
                    )
                    itl.addChild(itle)
                    itle.myType = "parameter"
                    itle = QtWidgets.QTreeWidgetItem(
                        [
                            "%d%s purity" % (float(config["isotopeB"]), config["labellingElement"]),
                            "%.2f%%" % (float(config["purityL"]) * 100.0),
                        ]
                    )
                    itl.addChild(itle)
                    itle.myType = "parameter"

                    if "useRatio" in config.keys() and config["useRatio"]:
                        if "minRatio" in config.keys():
                            itle = QtWidgets.QTreeWidgetItem(["Min. ratio", config["minRatio"]])
                            itl.addChild(itle)
                            itle.myType = "parameter"
                        if "maxRatio" in config.keys():
                            itle = QtWidgets.QTreeWidgetItem(["Max. ratio", config["maxRatio"]])
                            itl.addChild(itle)
                            itle.myType = "parameter"

                itl = QtWidgets.QTreeWidgetItem(["MZ picking"])
                it.addChild(itl)
                itl.myType = "parameter"
                if "positiveScanEvent" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(["ScanEvent (positive)", config["positiveScanEvent"]])
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "negativeScanEvent" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(["ScanEvent (negative)", config["negativeScanEvent"]])
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "xMin" in config.keys() and "xMax" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(
                        [
                            "Atoms range",
                            "%d - %d" % (int(config["xMin"]), int(config["xMax"])),
                        ]
                    )
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "startTime" in config.keys() and "stopTime" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(
                        [
                            "Scan range",
                            "%.2f - %.2f min" % (float(config["startTime"]), float(config["stopTime"])),
                        ]
                    )
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "intensityThreshold" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(
                        [
                            "Intensity threshold",
                            "%.0f" % float(config["intensityThreshold"]),
                        ]
                    )
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "intensityCutoff" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(["Intensity cutoff", "%.0f" % float(config["intensityCutoff"])])
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "maxLoading" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(["Max. charge", "%d" % int(config["maxLoading"])])
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "ppm" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(["Mass deviation (+/- ppm)", "%.1f" % (float(config["ppm"]))])
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "isotopicPatternCountLeft" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(
                        [
                            "Isotopic pattern count (A)",
                            config["isotopicPatternCountLeft"],
                        ]
                    )
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "isotopicPatternCountRight" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(
                        [
                            "Isotopic pattern count (B)",
                            config["isotopicPatternCountRight"],
                        ]
                    )
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "lowAbundanceIsotopeCutoff" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(
                        [
                            "Check M±1 / M'±1 isotopolog abundance",
                            config["lowAbundanceIsotopeCutoff"],
                        ]
                    )
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "Intensity abundance error" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(["Intensity abundance error"])
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "intensityErrorN" in config.keys():
                    itlee = QtWidgets.QTreeWidgetItem(["Non-labelled ion", config["intensityErrorN"]])
                    itle.addChild(itlee)
                    itlee.myType = "parameter"
                if "intensityErrorL" in config.keys():
                    itlee = QtWidgets.QTreeWidgetItem(["Labelled ion", config["intensityErrorL"]])
                    itle.addChild(itlee)
                    itlee.myType = "parameter"

                itl = QtWidgets.QTreeWidgetItem(["MZ clustering"])
                it.addChild(itl)
                itl.myType = "parameter"
                if "clustPPM" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(["Clustering ppm", config["clustPPM"]])
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "minSpectraCount" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(["Min. spectra", config["minSpectraCount"]])
                    itl.addChild(itle)
                    itle.myType = "parameter"

                itl = QtWidgets.QTreeWidgetItem(["Chromatographic peak picking"])
                it.addChild(itl)
                itl.myType = "parameter"
                if "chromPeakPPM" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(["EIC ppm", config["chromPeakPPM"]])
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "eicSmoothing" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(["EIC smoothing", config["eicSmoothing"]])
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "eicSmoothing" in config.keys() and bool(config["eicSmoothing"]) and "eicSmoothingWindowSize" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(["Smoothing window size", config["eicSmoothingWindowSize"]])
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "scales" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(
                        [
                            "Min. scale",
                            "%.0f" % (loads(base64.b64decode(config["scales"]))[0]),
                        ]
                    )
                    itl.addChild(itle)
                    itle.myType = "parameter"
                    itle = QtWidgets.QTreeWidgetItem(
                        [
                            "Max. scale",
                            "%.0f" % (loads(base64.b64decode(config["scales"]))[1]),
                        ]
                    )
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "snrTh" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(["SNR threshold", config["snrTh"]])
                    itl.addChild(itle)
                    itle.myType = "parameter"

                itl = QtWidgets.QTreeWidgetItem(["Peak matching"])
                it.addChild(itl)
                itl.myType = "parameter"
                if "peakCenterError" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(["Center error", config["peakCenterError"]])
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "minPeakCorr" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(["Min. corr", config["minPeakCorr"]])
                    itl.addChild(itle)
                    itle.myType = "parameter"

                itl = QtWidgets.QTreeWidgetItem(["Group features"])
                it.addChild(itl)
                itl.myType = "parameter"
                if "minCorrelation" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(["Min. corr", config["minCorrelation"]])
                    itl.addChild(itle)
                    itle.myType = "parameter"
                if "minCorrelationConnections" in config.keys():
                    itle = QtWidgets.QTreeWidgetItem(
                        [
                            "Min. correlation connections",
                            config["minCorrelationConnections"],
                        ]
                    )
                    itl.addChild(itle)
                    itle.myType = "parameter"
                itle = QtWidgets.QTreeWidgetItem(["Adducts"])
                itl.addChild(itle)
                itle.myType = "parameter"
                if "adducts" in config.keys():
                    for adduct in loads(base64.b64decode(config["adducts"])):
                        itlee = QtWidgets.QTreeWidgetItem(
                            [
                                str(adduct.name),
                                str(adduct.mzoffset),
                                str(adduct.polarity),
                            ]
                        )
                        itle.addChild(itlee)
                        itlee.myType = "parameter"
                itle = QtWidgets.QTreeWidgetItem(["Neutral loss (elements)"])
                itl.addChild(itle)
                itle.myType = "parameter"
                if "elements" in config.keys():
                    for elem, elemDetails in loads(base64.b64decode(config["elements"])).items():
                        itlee = QtWidgets.QTreeWidgetItem(
                            [
                                str(elem),
                                "%.4f" % elemDetails.weight,
                                str(elemDetails.numberValenzElectrons),
                                "%d-%d" % (elemDetails.minCount, elemDetails.maxCount),
                            ]
                        )
                        itle.addChild(itlee)
                        itlee.myType = "parameter"

            except Exception:
                pass

            self.sortTreeChildren(self.ui.res_ExtractedData.topLevelItem(3), 1)

            self.filterEdited(str(self.ui.dataFilter.text()))

            sectionSizes = [235, 65, 50, 105, 110, 70, 85, 70, 50, 60, 75]
            for i in range(11):
                self.ui.res_ExtractedData.header().resizeSection(i, sectionSizes[i])

            ## Setup diagnostics
            it = QtWidgets.QTreeWidgetItem(["Diagnostics"])
            it.myType = "diagnostic"
            self.ui.res_ExtractedData.addTopLevelItem(it)

            itl = QtWidgets.QTreeWidgetItem(["Observed intensities"])
            it.addChild(itl)
            itl.myType = "diagnostic - observed intensities"

            itl = QtWidgets.QTreeWidgetItem(["Native vs. labeled signal intensity"])
            it.addChild(itl)
            itl.myType = "diagnostic - observed intensities comparison"

            itl = QtWidgets.QTreeWidgetItem(["Relative mz error"])
            it.addChild(itl)
            itl.myType = "diagnostic - relative mz error"

            itl = QtWidgets.QTreeWidgetItem(["Relative mz error vs. signal intensity"])
            it.addChild(itl)
            itl.myType = "diagnostic - relative mz error vs intensity"

            itl = QtWidgets.QTreeWidgetItem(["Relative mzbin deviation"])
            it.addChild(itl)
            itl.myType = "diagnostic - relative mzbin deviation"

            itl = QtWidgets.QTreeWidgetItem(["Feature pair correlations"])
            it.addChild(itl)
            itl.myType = "diagnostic - feature pair correlations"

            itl = QtWidgets.QTreeWidgetItem(["Feature pair M+1/M ratio absolute"])
            it.addChild(itl)
            itl.myType = "diagnostic - feature pair mp1 to m ratio abs"

            itl = QtWidgets.QTreeWidgetItem(["Feature pair M+1/M RIA"])
            it.addChild(itl)
            itl.myType = "diagnostic - feature pair mp1 to m ratio rel"

            itl = QtWidgets.QTreeWidgetItem(["Feature pair M+1/M RIA vs intensity"])
            it.addChild(itl)
            itl.myType = "diagnostic - feature pair mp1 to m IRA vs intensity"

            itl = QtWidgets.QTreeWidgetItem(["Feature pair M'-1/M' ratio absolute"])
            it.addChild(itl)
            itl.myType = "diagnostic - feature pair mPp1 to mP ratio abs"

            itl = QtWidgets.QTreeWidgetItem(["Feature pair M'-1/M' RIA"])
            it.addChild(itl)
            itl.myType = "diagnostic - feature pair mPp1 to mP ratio rel"

            itl = QtWidgets.QTreeWidgetItem(["Feature pair M'-1/M' RIA vs intensity"])
            it.addChild(itl)
            itl.myType = "diagnostic - feature pair mPp1 to mP IRA vs intensity"

            itl = QtWidgets.QTreeWidgetItem(["Feature pair assigned MZs"])
            it.addChild(itl)
            itl.myType = "diagnostic - feature pair assigned mzs"

            itl = QtWidgets.QTreeWidgetItem(["Mean feature pair MZ deviation"])
            it.addChild(itl)
            itl.myType = "diagnostic - feature pair mz deviation mean"

            itl = QtWidgets.QTreeWidgetItem(["Feature pair MZ deviation"])
            it.addChild(itl)
            itl.myType = "diagnostic - feature pair mz deviation"

            itl = QtWidgets.QTreeWidgetItem(["Peak MZ deviation"])
            it.addChild(itl)
            itl.myType = "diagnostic - EIC mz deviation"

            itl = QtWidgets.QTreeWidgetItem(["Peak FWHM deviation"])
            it.addChild(itl)
            itl.myType = "diagnostic - peak fwhm"

        else:
            QtWidgets.QMessageBox.warning(
                self,
                "MetExtract",
                "No results are available for file %s",
                QtWidgets.QMessageBox.Ok,
            )

    def deColorQTreeWidgetItem(self, item):
        item.setBackground(0, QColor("white"))
        item.setBackground(1, QColor("white"))
        item.setBackground(2, QColor("white"))
        item.setBackground(3, QColor("white"))
        item.setBackground(4, QColor("white"))
        item.setBackground(5, QColor("white"))
        item.setBackground(6, QColor("white"))
        item.setBackground(7, QColor("white"))
        item.setBackground(8, QColor("white"))
        item.setBackground(9, QColor("white"))

        for i in range(item.childCount()):
            self.deColorQTreeWidgetItem(item.child(i))

    def getParametersFromCurrentRes(self, parameter):
        p = self.ui.res_ExtractedData.topLevelItem(4)
        childs = [p]
        while len(childs) > 0:
            child = childs.pop(0)
            for i in range(child.childCount()):
                childs.append(child.child(i))
            if child.text(0) == parameter:
                return child.text(1)

    def isTracerMetabolisationExperiment(self):
        config_df = self.currentOpenResultsFile.db_con.tables["config"].filter(pl.col("key") == "metabolisationExperiment")
        rows = config_df.to_dicts()
        assert len(rows) == 1
        return str(rows[0]["value"]).lower() == "true"

    def getAllowedIsotopeRatioErrorsForResult(self):
        config_df = self.currentOpenResultsFile.db_con.tables["config"]

        rows_n = config_df.filter(pl.col("key") == "intensityErrorN").to_dicts()
        assert len(rows_n) == 1
        intErrN = float(rows_n[0]["value"])

        rows_l = config_df.filter(pl.col("key") == "intensityErrorL").to_dicts()
        assert len(rows_l) == 1
        intErrL = float(rows_l[0]["value"])

        return intErrN, intErrL

    def getTICsForResult(self):
        if hasattr(self.currentOpenResultsFile, "tics"):
            return self.currentOpenResultsFile.tics

        tics = {}
        tics_df = self.currentOpenResultsFile.db_con.tables["tics"]
        for row_dict in tics_df.to_dicts():
            id = int(row_dict["id"])
            loading = str(row_dict["loading"])
            scanevent = str(row_dict["scanevent"])
            times = [float(t) / 60.0 for t in row_dict["times"].split(";")]
            intensities = [float(t) for t in row_dict["intensities"].split(";")]

            tics[loading] = Bunch(
                id=id,
                loading=loading,
                scanEvent=scanevent,
                times=times,
                intensities=intensities,
            )

        setattr(self.currentOpenResultsFile, "tics", tics)

        return tics

    def getLabellingParametersForResult(self, featureID):
        if not (self.isTracerMetabolisationExperiment()):
            config_df = self.currentOpenResultsFile.db_con.tables["config"]

            rows_a = config_df.filter(pl.col("key") == "isotopeA").to_dicts()
            assert len(rows_a) == 1
            isoA = float(rows_a[0]["value"])

            rows_b = config_df.filter(pl.col("key") == "isotopeB").to_dicts()
            assert len(rows_b) == 1
            isoB = float(rows_b[0]["value"])

            rows_purn = config_df.filter(pl.col("key") == "purityN").to_dicts()
            assert len(rows_purn) == 1
            purN = float(rows_purn[0]["value"])

            rows_purl = config_df.filter(pl.col("key") == "purityL").to_dicts()
            assert len(rows_purl) == 1
            purL = float(rows_purl[0]["value"])

            return isoB - isoA, purN, purL
        else:
            chrompeaks_df = self.currentOpenResultsFile.db_con.tables["chromPeaks"].filter(pl.col("id") == featureID)
            rows_cp = chrompeaks_df.to_dicts()
            assert len(rows_cp) == 1
            trcid = int(rows_cp[0]["tracer"])

            tracer_config_df = self.currentOpenResultsFile.db_con.tables["tracerConfiguration"].filter(pl.col("id") == trcid)
            rows_tc = tracer_config_df.to_dicts()
            assert len(rows_tc) == 1
            dmz = float(rows_tc[0]["deltaMZ"])
            purN = float(rows_tc[0]["purityN"])
            purL = float(rows_tc[0]["purityL"])

            return dmz, purN, purL

    def _get_result_table(self, *table_names):
        """Return the first matching results table, case-insensitive, or an empty DataFrame."""
        tables = self.currentOpenResultsFile.db_con.tables
        for table_name in table_names:
            if table_name in tables:
                return tables[table_name]

        lowered = {str(name).lower(): name for name in tables.keys()}
        for table_name in table_names:
            match = lowered.get(str(table_name).lower())
            if match is not None:
                return tables[match]

        return pl.DataFrame()

    def _resolve_col_name(self, df, *candidate_names):
        """Resolve a column name in a DataFrame, case-insensitive."""
        if len(df) == 0:
            return None

        col_map = {str(col).lower(): col for col in df.columns}
        for name in candidate_names:
            if name in df.columns:
                return name
            match = col_map.get(str(name).lower())
            if match is not None:
                return match

        return None

    def selectedResChanged(self):
        annotationPPM = self.ui.doubleSpinBox_isotopologAnnotationPPM.value()

        # Always restore "Result name" row; individual branches may hide it again
        self.ui.label_22.setVisible(True)
        self.ui.chromPeakName.setVisible(True)
        self.ui.setChromPeakName.setVisible(True)

        for i in range(self.ui.res_ExtractedData.topLevelItemCount()):
            self.deColorQTreeWidgetItem(self.ui.res_ExtractedData.topLevelItem(i))

        selectedItems = self.ui.res_ExtractedData.selectedItems()

        changePlots = False
        for item in selectedItems:
            if hasattr(item, "myType"):
                if item.myType == "parameter":
                    continue
                else:
                    changePlots = True
        if not changePlots:
            return

        self.clearPlot(self.ui.pl1)
        self.clearPlot(self.ui.pl2A)
        self.clearPlot(self.ui.pl2B)
        self.clearPlot(self.ui.pl3)

        useColi = 0
        maxIntX = 0
        maxIntY = 0
        minIntY = 0
        minMZ = 1000000
        maxMZ = 0
        minMZH = 0
        maxMZH = 0
        minTime = 1000
        maxTime = 1
        x_vals = []
        y_vals = []
        mzs = []
        peaks = []
        plotTypes = set()
        selFeatureGroups = []

        featuresPosSelected = False
        for item in selectedItems:
            if not (hasattr(item, "myType")):
                continue

            if item.myType == "MZs" or item.myType == "mz":
                plotTypes.add("MZs")
            if item.myType == "MZBins" or item.myType == "mzbin":
                plotTypes.add("MZBins")
            if item.myType == "Features" or item.myType == "feature":
                plotTypes.add("Features")
            if item.myType == "Feature Groups" or item.myType == "feature group":
                plotTypes.add("Feature Groups")
            if item.myType.lower().startswith("diagnostic"):
                plotTypes.add("diagnostic")

            if item.myType == "Features" or item.myType == "feature":
                if hasattr(item, "myData"):
                    cp = item.myData
                    if cp.ionMode == "-":
                        pass
                    else:
                        featuresPosSelected = True

        if len(plotTypes) > 1:
            QtWidgets.QMessageBox.warning(
                self,
                "MetExtract",
                "Selecting different result types in not supported. Please select only one or multipel MZs, MZBins, FeaturePairs or FeatureGroups at a time",
                QtWidgets.QMessageBox.Ok,
            )
            self.clearPlot(self.ui.pl1)
            return

        for selIndex, item in enumerate(selectedItems):
            if not (hasattr(item, "myType")):
                continue

            # <editor-fold desc="#mz result">
            if item.myType == "MZs" or item.myType == "mz":
                plotTypes.add("MZs")
                self.ui.res_ExtractedData.setHeaderLabels(
                    [
                        "MZ",
                        "Xn",
                        "Scan id",
                        "Rt",
                        "Charge",
                        "Intensity",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ]
                )

                t = item
                if len(x_vals) == 0:
                    x = []
                    y = []
                    if t.myType == "mz":
                        t = t.parent()
                    for j in range(t.childCount()):
                        child = t.child(j)
                        assert child.myType == "mz"
                        x.append(child.myData.scantime / 60.0)
                        y.append(child.myData.mz)

                    maxIntY = max(y)
                    maxIntX = max(x)

                    x_vals = [x]
                    y_vals = [y]

                if item.myType == "mz":
                    if len(x_vals) == 1:
                        x_vals.append([])
                        y_vals.append([])
                    x_vals[1].append(item.myData.scantime / 60.0)
                    y_vals[1].append(item.myData.mz)
            # </editor-fold>

            # <editor-fold desc="#mzbin results">
            elif item.myType == "MZBins" or item.myType == "mzbin":
                self.ui.res_ExtractedData.setHeaderLabels(["MZ", "Delta ppm", "Xn", "", "", "", "", "", "", "", ""])
                if item.myType == "MZBins":
                    plotTypes.add("MZBins")
                    for i in range(item.childCount()):
                        kid = item.child(i)
                        assert kid.myType == "mzbin"
                        x = []
                        y = []
                        for j in range(kid.childCount()):
                            child = kid.child(j)
                            assert child.myType == "mz"
                            x.append(child.myData.scantime / 60.0)
                            y.append(child.myData.mz)
                        x_vals.append(x)
                        y_vals.append(y)
                    if len(x_vals) > 0:
                        maxIntX = max(max(x_vals), maxIntX)
                        maxIntY = max(max(y_vals), maxIntY)
                    else:
                        maxIntX = 1
                        maxIntY = 1

                elif item.myType == "mzbin":
                    plotTypes.add("mzbin")
                    x = []
                    y = []
                    if item.myType == "mzbin" and len(x_vals) == 0:
                        t = item.parent()
                        for o in range(t.childCount()):
                            childo = t.child(o)
                            for j in range(childo.childCount()):
                                childj = childo.child(j)

                                assert childj.myType == "mz"
                                x.append(childj.myData.scantime / 60.0)
                                y.append(childj.myData.mz)
                        x_vals.append(x)
                        y_vals.append(y)

                    x = []
                    y = []
                    for j in range(item.childCount()):
                        child = item.child(j)
                        assert child.myType == "mz"
                        x.append(child.myData.scantime / 60.0)
                        y.append(child.myData.mz)
                    x_vals.append(x)
                    y_vals.append(y)
                if len(x_vals) > 0:
                    maxIntX = max(max(v) for v in x_vals if v)
                    maxIntY = max(max(v) for v in y_vals if v)
                else:
                    maxIntX = 1
                    maxIntY = 1
            # </editor-fold>

            # <editor-fold desc="#feature results">
            elif item.myType == "Features" or item.myType == "feature":
                self.ui.chromPeakName.setText("")
                if item.myType == "Features":
                    self.ui.label_22.setVisible(False)
                    self.ui.chromPeakName.setVisible(False)
                    self.ui.setChromPeakName.setVisible(False)

                self.ui.res_ExtractedData.setHeaderLabels(
                    [
                        "MZ (/Ionmode Z)",
                        "Rt min",
                        "Xn",
                        "Adducts / in-source fragments / hetero atoms",
                        "Scale M / M'",
                        "Peak cor",
                        "M:M' peaks ratio / area ratio",
                        "Area M / M'",
                        "Scans",
                        "LMZ",
                        "Tracer",
                    ]
                )

                if item.myType == "Features":
                    _fm_rts = []
                    _fm_mzs = []
                    _fm_areas = []
                    _fm_point_data = []
                    plotTypes.add("Features")
                    plotTypes.add("FeatureMap")
                    for _fmi in range(item.childCount()):
                        _fmc = item.child(_fmi)
                        assert _fmc.myType == "feature"
                        _rt = _fmc.myData.NPeakCenterMin / 60.0
                        _mz = _fmc.myData.mz
                        try:
                            _area = max(1.0, float(_fmc.text(7).split(" / ")[0]))
                        except Exception:
                            _area = 1.0
                        _fm_rts.append(_rt)
                        _fm_mzs.append(_mz)
                        _fm_areas.append(_area)
                        _fm_point_data.append(
                            {
                                "rt": _rt,
                                "mz": _mz,
                                "native_area": _area,
                                "id": _fmc.myData.id,
                                "ogroup": "N/A",
                                "polarity": _fmc.myData.ionMode,
                                "charge": _fmc.myData.loading,
                                "xcount": _fmc.myData.xCount,
                                "tree_item": _fmc,
                            }
                        )
                    if _fm_areas:
                        import math as _math

                        _log_areas = [_math.log10(a) for a in _fm_areas]
                        _min_log = min(_log_areas)
                        _max_log = max(_log_areas)
                        _range_log = max(_max_log - _min_log, 1.0)
                        _fm_sizes = [20 + 200 * (la - _min_log) / _range_log for la in _log_areas]
                    else:
                        _fm_sizes = []
                    _ax_fm = self.ui.pl1.twinxs[0]
                    _ax_fm._fm_point_data = _fm_point_data
                    _ax_fm.scatter(
                        _fm_rts,
                        _fm_mzs,
                        s=_fm_sizes,
                        c=[predefinedColors[0]] * len(_fm_rts),
                        alpha=0.6,
                    )
                    _ax_fm.set_xlabel("Retention time (min)")
                    _ax_fm.set_ylabel("m/z")
                    _ax_fm.set_title(f"Feature map ({len(_fm_rts)} feature pairs)")
                    mzs = _fm_mzs

                if item.myType == "feature":
                    cp = item.myData
                    plotTypes.add("feature")
                    mzs.append(cp.mz)
                    peaks.append(cp.NPeakCenterMin / 60.0)
                    xic = []
                    xicL = []
                    times = []
                    maxE = 1

                    invert = 1
                    if self.ui.negEIC.isChecked():
                        invert = -1

                    xics_df = self.currentOpenResultsFile.db_con.tables["XICs"].filter(pl.col("id") == cp.eicID)
                    for row_dict in xics_df.to_dicts():
                        xic = [float(t) for t in row_dict["xic"].split(";")]
                        xicL = [float(t) for t in row_dict["xicL"].split(";")]
                        xicfirstiso = [float(t) for t in row_dict["xicfirstiso"].split(";")]
                        xicLfirstiso = [float(t) for t in row_dict["xicLfirstiso"].split(";")]
                        xicLfirstisoconjugate = [float(t) for t in row_dict["xicLfirstisoconjugate"].split(";")]
                        xic_smoothed = [float(t) for t in row_dict["xic_smoothed"].split(";")]
                        xicL_smoothed = [float(t) for t in row_dict["xicL_smoothed"].split(";")]
                        offset = cp.NPeakCenterMin / 60.0 if self.ui.setPeakCentersToZero.isChecked() else 0
                        times = [float(t) / 60.0 - offset for t in row_dict["times"].split(";")]
                        try:
                            allPeaks = loads(base64.b64decode(str(row_dict["allPeaks"])))
                        except (ImportError, ModuleNotFoundError, SyntaxError) as e:
                            logging.warning(f"Could not load peak data from results: {e}")
                            logging.warning("This may be due to cached data from an older version. Using empty list.")
                            allPeaks = []
                        xic_baseline = [float(t) for t in row_dict["xic_baseline"].split(";")]
                        xicL_baseline = [float(t) for t in row_dict["xicL_baseline"].split(";")]

                    if self.ui.scaleFeatures.isChecked():
                        s = int(cp.NPeakCenter - cp.NBorderLeft * 1)
                        e = int(cp.NPeakCenter + cp.NBorderRight * 1)
                        maxP = s + max(range(e - s), key=lambda x: xic[s + x])
                        maxE = mean(xic[(maxP - 3) : (maxP + 3)])
                        maxEL = mean(xicL[(maxP - 3) : (maxP + 3)])
                        if maxE == 0:
                            maxE = 1
                        if maxEL == 0:
                            maxEL = 1
                        if not self.ui.scaleLabelledFeatures.isChecked():
                            maxEL = maxE

                        xic = [u / maxE for u in xic]
                        xicfirstiso = [u / maxE for u in xicfirstiso]
                        xicL = [u / maxEL for u in xicL]
                        xicLfirstiso = [u / maxEL for u in xicLfirstiso]
                        xicLfirstisoconjugate = [u / maxEL for u in xicLfirstisoconjugate]
                        xic_smoothed = [u / maxE for u in xic_smoothed]
                        xicL_smoothed = [u / maxEL for u in xicL_smoothed]
                        xic_baseline = [u / maxE for u in xic_baseline]
                        xicL_baseline = [u / maxEL for u in xicL_baseline]

                    xicL = [invert * u for u in xicL]
                    xicLfirstiso = [invert * u for u in xicLfirstiso]
                    xicLfirstisoconjugate = [invert * u for u in xicLfirstisoconjugate]
                    xicL_smoothed = [invert * u for u in xicL_smoothed]
                    xicL_baseline = [invert * u for u in xicL_baseline]

                    if self.ui.flattenXIC.isChecked():
                        ps = min(
                            int(cp.NPeakCenter - cp.NPeakScale * 2),
                            int(cp.LPeakCenter - cp.LPeakScale * 2),
                        )
                        pe = max(
                            int(cp.NPeakCenter + cp.NPeakScale * 2),
                            int(cp.LPeakCenter + cp.LPeakScale * 2),
                        )
                    else:
                        ps = 0
                        pe = len(times)
                    try:
                        minTime = min(
                            minTime,
                            min(times[int(cp.NPeakCenter - cp.NPeakScale * 1) : int(cp.NPeakCenter + cp.NPeakScale * 1)]),
                        )
                        maxTime = max(
                            maxTime,
                            max(times[int(cp.NPeakCenter - cp.NPeakScale * 1) : int(cp.NPeakCenter + cp.NPeakScale * 1)]),
                        )
                        maxIntY = max(
                            maxIntY,
                            max(xic[int(cp.NPeakCenter - cp.NPeakScale * 1) : int(cp.NPeakCenter + cp.NPeakScale * 1)]),
                        )
                        minIntY = min(
                            minIntY,
                            min(xicL[int(cp.LPeakCenter - cp.LPeakScale * 1) : int(cp.LPeakCenter + cp.LPeakScale * 1)]),
                        )
                    except Exception:
                        pass

                    item.setBackground(0, QColor(predefinedColors[useColi % len(predefinedColors)]))
                    item.setBackground(1, QColor(predefinedColors[useColi % len(predefinedColors)]))
                    item.setBackground(2, QColor(predefinedColors[useColi % len(predefinedColors)]))
                    item.setBackground(3, QColor(predefinedColors[useColi % len(predefinedColors)]))
                    item.setBackground(4, QColor(predefinedColors[useColi % len(predefinedColors)]))
                    item.setBackground(5, QColor(predefinedColors[useColi % len(predefinedColors)]))
                    item.setBackground(6, QColor(predefinedColors[useColi % len(predefinedColors)]))
                    item.setBackground(7, QColor(predefinedColors[useColi % len(predefinedColors)]))

                    self.drawPlot(
                        self.ui.pl1,
                        plotIndex=0,
                        x=times[ps:pe],
                        y=xic[ps:pe],
                        fill=[
                            int(cp.NPeakCenter - cp.NBorderLeft),
                            int(cp.NPeakCenter + cp.NBorderRight),
                        ],
                        rearrange=len(selectedItems) == 1,
                        label="M: %.4f (%s, %s)" % (cp.mz, cp.xCount, cp.ionMode),
                        useCol=useColi,
                    )

                    if self.ui.checkBox_showBaseline.isChecked():
                        self.drawPlot(
                            self.ui.pl1,
                            plotIndex=0,
                            x=times[ps:pe],
                            y=xic_baseline[ps:pe],
                            fill=[
                                int(cp.LPeakCenter - cp.LBorderLeft),
                                int(cp.LPeakCenter + cp.LBorderRight),
                            ],
                            rearrange=len(selectedItems) == 1,
                            label=None,
                            useCol=useColi,
                            linestyle="--",
                        )

                    if self.ui.showSmoothedEIC_checkBox.isChecked():
                        self.drawPlot(
                            self.ui.pl1,
                            plotIndex=0,
                            x=times[ps:pe],
                            y=xic_smoothed[ps:pe],
                            fill=[
                                int(cp.NPeakCenter - cp.NBorderLeft),
                                int(cp.NPeakCenter + cp.NBorderRight),
                            ],
                            rearrange=len(selectedItems) == 1,
                            label=None,
                            useCol=useColi,
                            linestyle="--",
                        )

                    self.drawPoints(
                        self.ui.pl1,
                        x=[times[a.scanIndex] for a in cp.assignedMZs],
                        y=[xic[a.scanIndex] for a in cp.assignedMZs],
                    )

                    if self.ui.showIsotopologues.isChecked():
                        self.drawPlot(
                            self.ui.pl1,
                            plotIndex=0,
                            x=times[ps:pe],
                            y=xicfirstiso[ps:pe],
                            fill=[
                                int(cp.NPeakCenter - cp.NBorderLeft),
                                int(cp.NPeakCenter + cp.NBorderRight),
                            ],
                            rearrange=len(selectedItems) == 1,
                            label=None,
                            useCol=useColi,
                        )

                    if self.ui.showArtificialShoft_checkBox.isChecked():
                        pst = ps
                        pet = pe
                        ps = ps + cp.artificialEICLShift
                        pe = pe + cp.artificialEICLShift

                        if pe > len(times):
                            pe = pet
                            pet = pet + cp.artificialEICLShift
                        if ps < 0:
                            ps = pst
                            pst = pst - cp.artificialEICLShift
                    else:
                        pst = ps
                        pet = pe

                    self.drawPlot(
                        self.ui.pl1,
                        plotIndex=0,
                        x=times[pst:pet],
                        y=xicL[ps:pe],
                        fill=[
                            int(cp.LPeakCenter - cp.LBorderLeft),
                            int(cp.LPeakCenter + cp.LBorderRight),
                        ],
                        rearrange=len(selectedItems) == 1,
                        label="M': %.4f (%s, %s)" % (cp.mz, cp.xCount, cp.ionMode),
                        useCol=useColi,
                    )

                    if self.ui.checkBox_showBaseline.isChecked():
                        self.drawPlot(
                            self.ui.pl1,
                            plotIndex=0,
                            x=times[pst:pet],
                            y=xicL_baseline[ps:pe],
                            fill=[
                                int(cp.LPeakCenter - cp.LBorderLeft),
                                int(cp.LPeakCenter + cp.LBorderRight),
                            ],
                            rearrange=len(selectedItems) == 1,
                            label=None,
                            useCol=useColi,
                            linestyle="--",
                        )

                    if self.ui.showSmoothedEIC_checkBox.isChecked():
                        self.drawPlot(
                            self.ui.pl1,
                            plotIndex=0,
                            x=times[pst:pet],
                            y=xicL_smoothed[ps:pe],
                            fill=[
                                int(cp.LPeakCenter - cp.LBorderLeft),
                                int(cp.LPeakCenter + cp.LBorderRight),
                            ],
                            rearrange=len(selectedItems) == 1,
                            label=None,
                            useCol=useColi,
                            linestyle="--",
                        )

                    if self.ui.showIsotopologues.isChecked():
                        self.drawPlot(
                            self.ui.pl1,
                            plotIndex=0,
                            x=times[pst:pet],
                            y=xicLfirstiso[ps:pe],
                            fill=[
                                int(cp.LPeakCenter - cp.LBorderLeft),
                                int(cp.LPeakCenter + cp.LBorderRight),
                            ],
                            rearrange=len(selectedItems) == 1,
                            label=None,
                            useCol=useColi,
                        )

                    maxindex, maxvalue = max(
                        enumerate(
                            xic[int(cp.NPeakCenter - 1) : int(cp.NPeakCenter + 1)],
                            start=int(cp.NPeakCenter - 1),
                        ),
                        key=itemgetter(1),
                    )
                    if self.ui.plotAddLabels.checkState() == QtCore.Qt.Checked:
                        self.addAnnotation(
                            self.ui.pl1,
                            "%.5f\n%.2f min" % (cp.mz, cp.NPeakCenterMin / 60.0),
                            (times[maxindex], xic[maxindex]),
                            (times[maxindex], xic[maxindex]),
                            0,
                            fcColor=predefinedColors[useColi % len(predefinedColors)],
                            ecColor=predefinedColors[useColi % len(predefinedColors)],
                            arrowColor=predefinedColors[useColi % len(predefinedColors)],
                            alpha=0.25,
                        )

                    if self.ui.showDiagnostics.isChecked():
                        for pa in allPeaks["peaksN"]:
                            if pa.apex_index != cp.NPeakCenter:
                                self.addAnnotation(
                                    self.ui.pl1,
                                    "",
                                    (times[pa.apex_index], xic[pa.apex_index]),
                                    (times[pa.apex_index], xic[pa.apex_index]),
                                    0,
                                    fcColor="slategrey",
                                    ecColor="slategrey",
                                    arrowColor="slategrey",
                                    alpha=0.15,
                                    add=30,
                                )
                        for pa in allPeaks["peaksL"]:
                            self.addAnnotation(
                                self.ui.pl1,
                                "",
                                (times[pa.apex_index], xicL[pa.apex_index]),
                                (times[pa.apex_index], xicL[pa.apex_index]),
                                0,
                                fcColor="slategrey",
                                ecColor="slategrey",
                                arrowColor="slategrey",
                                alpha=0.15,
                                up=False,
                                add=30,
                            )

                        cp_loading = int(cp.loading)
                        cp_xcount = int(cp.xCount)

                        for row_dict in (
                            self.currentOpenResultsFile.db_con.tables["chromPeaks"]
                            .join(self.currentOpenResultsFile.db_con.tables["featureGroupFeatures"], left_on="id", right_on="fID", how="left")
                            .join(self.currentOpenResultsFile.db_con.tables["featureGroups"], left_on="fGroupID", right_on="id", how="left")
                            .filter((pl.col("mz") >= cp.mz * (1 - 15 / 1000000.0)) & (pl.col("mz") <= cp.mz * (1 + 15 / 1000000.0)) & (pl.col("Loading").cast(pl.Int64, strict=False) == cp_loading) & (pl.col("xcount").cast(pl.Int64, strict=False) == cp_xcount))
                            .select([pl.col("id").alias("c_id"), pl.col("eicID"), pl.col("NPeakCenterMin"), pl.col("NPeakCenter"), pl.col("mz"), pl.col("xcount"), pl.col("Loading").alias("loading"), pl.col("featureName").alias("FGroupID")])
                            .to_dicts()
                        ):
                            row = (row_dict["c_id"], row_dict["eicID"], row_dict["NPeakCenterMin"], row_dict["NPeakCenter"], row_dict["mz"], row_dict["xcount"], row_dict["loading"], row_dict["FGroupID"])
                            if cp.NPeakCenter != row[3]:
                                self.addAnnotation(
                                    self.ui.pl1,
                                    "%s\n%.5f\n%.2f min" % (str(row[7]), row[4], row[2] / 60.0),
                                    (times[row[3]], xic[row[3]]),
                                    (times[row[3]], xic[row[3]]),
                                    0,
                                    fcColor="slategrey",
                                    ecColor="slategrey",
                                    arrowColor="slategrey",
                                    alpha=0.15,
                                )

                    self.isTracerMetabolisationExperiment()
                    mzD, purN, purL = self.getLabellingParametersForResult(cp.id)

                    toDrawMzs = []
                    toDrawInts = []
                    ms_df = self.currentOpenResultsFile.db_con.tables["massspectrum"].filter(pl.col("mID") == cp.massSpectrumID)
                    for row_dict in ms_df.to_dicts():
                        mzs = [float(t) for t in row_dict["mzs"].split(";")]
                        intensities = [float(t) for t in row_dict["intensities"].split(";")]

                        for mz in mzs:
                            toDrawMzs.append(mz)
                            toDrawMzs.append(mz)
                            toDrawMzs.append(mz)

                        for intensity in intensities:
                            toDrawInts.append(0)
                            if cp.ionMode == "-" and featuresPosSelected:
                                toDrawInts.append(-intensity)
                            else:
                                toDrawInts.append(intensity)
                            toDrawInts.append(0)

                    self.drawPlot(
                        self.ui.pl3,
                        plotIndex=0,
                        x=toDrawMzs,
                        y=toDrawInts,
                        useCol="lightgrey",
                        multipleLocator=None,
                        alpha=0.1,
                        title="",
                        xlab="MZ",
                    )

                useColi += 1
            # </editor-fold>

            # <editor-fold desc="#featureGroup results">
            elif item.myType == "featureGroup" or item.myType == "Feature Groups":
                if item.myType == "Feature Groups":
                    self.ui.label_22.setVisible(False)
                    self.ui.chromPeakName.setVisible(False)
                    self.ui.setChromPeakName.setVisible(False)
                self.ui.res_ExtractedData.setHeaderLabels(
                    [
                        "Feature group / MZ (/Ionmode Z)",
                        "Rt min",
                        "Xn",
                        "Adducts / in-source fragments / hetero atoms",
                        "Scale M / M'",
                        "Peak cor",
                        "M:M' peaks ratio / area ratio",
                        "Area M / M'",
                        "Scans",
                        "Tracer",
                    ]
                )

                item.setBackground(0, QColor(predefinedColors[(useColi) % len(predefinedColors)]))
                item.setBackground(1, QColor(predefinedColors[(useColi) % len(predefinedColors)]))
                item.setBackground(2, QColor(predefinedColors[(useColi) % len(predefinedColors)]))
                item.setBackground(3, QColor(predefinedColors[(useColi) % len(predefinedColors)]))
                item.setBackground(4, QColor(predefinedColors[(useColi) % len(predefinedColors)]))
                item.setBackground(5, QColor(predefinedColors[(useColi) % len(predefinedColors)]))
                item.setBackground(6, QColor(predefinedColors[(useColi) % len(predefinedColors)]))
                item.setBackground(7, QColor(predefinedColors[(useColi) % len(predefinedColors)]))

                if item.myType == "Feature Groups":
                    import math as _math

                    plotTypes.add("Feature Groups")
                    plotTypes.add("FeatureMap")
                    _all_fm_rts = []
                    _all_fm_mzs = []
                    _all_fm_sizes = []
                    _all_fm_colors = []
                    _all_fm_point_data = []
                    _fm_group_lines = []
                    for _gi in range(item.childCount()):
                        _gc = item.child(_gi)
                        _g_rts = []
                        _g_mzs = []
                        _g_areas = []
                        _g_group_name = _gc.myData.featureName if hasattr(_gc, "myData") and hasattr(_gc.myData, "featureName") else str(_gc.text(0))
                        for _gfi in range(_gc.childCount()):
                            _gff = _gc.child(_gfi)
                            assert _gff.myType == "feature"
                            _rt = _gff.myData.NPeakCenterMin / 60.0
                            _mz = _gff.myData.mz
                            try:
                                _area = max(1.0, float(_gff.text(7).split(" / ")[0]))
                            except Exception:
                                _area = 1.0
                            _g_rts.append(_rt)
                            _g_mzs.append(_mz)
                            _g_areas.append(_area)
                            _all_fm_point_data.append(
                                {
                                    "rt": _rt,
                                    "mz": _mz,
                                    "native_area": _area,
                                    "id": _gff.myData.id,
                                    "ogroup": _g_group_name,
                                    "polarity": _gff.myData.ionMode,
                                    "charge": _gff.myData.loading,
                                    "xcount": _gff.myData.xCount,
                                    "tree_item": _gff,
                                }
                            )
                        if _g_areas:
                            _g_log = [_math.log10(a) for a in _g_areas]
                            _g_min = min(_g_log)
                            _g_max = max(_g_log)
                            _g_range = max(_g_max - _g_min, 1.0)
                            _g_sizes = [20 + 200 * (la - _g_min) / _g_range for la in _g_log]
                        else:
                            _g_sizes = []
                        _col = predefinedColors[_gi % len(predefinedColors)]
                        _all_fm_rts.extend(_g_rts)
                        _all_fm_mzs.extend(_g_mzs)
                        _all_fm_sizes.extend(_g_sizes)
                        _all_fm_colors.extend([_col] * len(_g_rts))
                        _fm_group_lines.append((_g_rts, _g_mzs, _col))
                    _ax_fg = self.ui.pl1.twinxs[0]
                    _ax_fg._fm_point_data = _all_fm_point_data
                    _ax_fg.scatter(
                        _all_fm_rts,
                        _all_fm_mzs,
                        s=_all_fm_sizes,
                        c=_all_fm_colors,
                        alpha=0.6,
                    )
                    for _line_rts, _line_mzs, _line_col in _fm_group_lines:
                        if len(_line_rts) >= 2:
                            _sorted_pts = sorted(zip(_line_mzs, _line_rts))
                            _sx = [_rt for _, _rt in _sorted_pts]
                            _sy = [_mz for _mz, _ in _sorted_pts]
                            _ax_fg.plot(_sx, _sy, color=_line_col, linewidth=0.8, alpha=0.5, zorder=1)
                    _ax_fg.set_xlabel("Retention time (min)")
                    _ax_fg.set_ylabel("m/z")
                    _n_groups = item.childCount()
                    _n_feat = len(_all_fm_rts)
                    _ax_fg.set_title(f"Feature map ({_n_feat} feature pairs in {_n_groups} metabolite groups)")

                if item.myType == "featureGroup":
                    selFeatureGroups.append(item)
                    self.ui.chromPeakName.setText(item.myData.featureName)
                    plotTypes.add("feature")
                    meanRT = 0
                    countK = 0

                    self.clearPlot(self.ui.pl3)

                    toDrawIntsPos = []
                    toDrawIntsNeg = []
                    toDrawMZsPos = []
                    toDrawMZsNeg = []
                    hasPos = False
                    hasNeg = False
                    massSpectraAvailable = False

                    msi = 0
                    mzs = {}
                    intensities = {}
                    mstime = {}

                    try:
                        # Get mass spectrum data from massspectrum table
                        ms_df = self.currentOpenResultsFile.db_con.tables["massspectrum"].filter(pl.col("fgID") == item.myID)
                        for row_dict in ms_df.to_dicts():
                            msi += 1
                            ionMode = str(row_dict["ionMode"])
                            mzs[ionMode] = [float(u) for u in str(row_dict["mzs"]).split(";")]
                            intensities[ionMode] = [float(u) for u in str(row_dict["intensities"]).split(";")]
                            mstime[ionMode] = float(row_dict["time"])
                    except Exception:
                        pass

                    hasPos = "+" in mzs.keys()
                    hasNeg = "-" in mzs.keys()

                    massSpectraAvailable = hasPos or hasNeg

                    if hasPos:
                        for mz in mzs["+"]:
                            toDrawMZsPos.append(mz)
                            toDrawMZsPos.append(mz)
                            toDrawMZsPos.append(mz)
                        for intensity in intensities["+"]:
                            toDrawIntsPos.append(0)
                            toDrawIntsPos.append(intensity)
                            toDrawIntsPos.append(0)

                        self.drawPlot(
                            self.ui.pl3,
                            plotIndex=0,
                            x=toDrawMZsPos,
                            y=toDrawIntsPos,
                            useCol="lightgrey",
                            multipleLocator=None,
                            alpha=0.1,
                            title="",
                            xlab="MZ",
                        )

                    if hasNeg:
                        negInt = 1.0

                        if "+" in mzs.keys():
                            negInt = -1.0

                        for mz in mzs["-"]:
                            toDrawMZsNeg.append(mz)
                            toDrawMZsNeg.append(mz)
                            toDrawMZsNeg.append(mz)
                        for intensity in intensities["-"]:
                            toDrawIntsNeg.append(0)
                            toDrawIntsNeg.append(intensity * negInt)
                            toDrawIntsNeg.append(0)

                        self.drawPlot(
                            self.ui.pl3,
                            plotIndex=0,
                            x=toDrawMZsNeg,
                            y=toDrawIntsNeg,
                            useCol="lightgrey",
                            multipleLocator=None,
                            alpha=0.1,
                            title="",
                            xlab="MZ",
                        )
                    mzs = []
                    childIDs = []
                    maxInt = 0

                    for childi in range(item.childCount()):
                        if not (item.child(childi).isHidden()):
                            cp = item.child(childi).myData

                            if cp.ionMode == "-":
                                max(toDrawIntsNeg)
                                toDrawInts = toDrawIntsNeg
                                toDrawMzs = toDrawMZsNeg
                            else:
                                max(toDrawIntsPos)
                                toDrawInts = toDrawIntsPos
                                toDrawMzs = toDrawMZsPos

                            if massSpectraAvailable:
                                self.isTracerMetabolisationExperiment()
                                mzD, purN, purL = self.getLabellingParametersForResult(cp.id)

                                bm = (
                                    min(
                                        range(len(toDrawMzs)),
                                        key=lambda i: abs(toDrawMzs[i] - cp.mz),
                                    )
                                    + 1
                                )
                                bml = (
                                    min(
                                        range(len(toDrawMzs)),
                                        key=lambda i: abs(toDrawMzs[i] - (cp.mz + mzD * cp.xCount / cp.loading)),
                                    )
                                    + 1
                                )

                                intLeft = toDrawInts[bm]
                                intRight = toDrawInts[bml]

                                h = 0
                                if cp.ionMode == "-" and hasPos:
                                    h = min(intLeft, intRight)
                                else:
                                    h = max(intLeft, intRight)

                                if self.ui.MSLabels.checkState() == QtCore.Qt.Checked:
                                    self.addAnnotation(
                                        self.ui.pl3,
                                        "mz: %.5f\nl-mz: %.5f\nd-mz: %.5f\nXn: %d Z: %s%d"
                                        % (
                                            cp.mz,
                                            cp.lmz,
                                            mzD * cp.xCount,
                                            cp.xCount,
                                            cp.ionMode,
                                            cp.loading,
                                        ),
                                        (
                                            cp.mz + mzD * cp.xCount / cp.loading / 2.0,
                                            h * 1.1,
                                        ),
                                        (10, 120),
                                        rotation=0,
                                        up=not (cp.ionMode == "-" and hasPos),
                                    )

                                self.addArrow(
                                    self.ui.pl3,
                                    (cp.mz, toDrawInts[bm]),
                                    (cp.mz, h * 1.1),
                                    drawArrowHead=True,
                                )
                                self.addArrow(
                                    self.ui.pl3,
                                    (cp.mz, h * 1.1),
                                    (cp.mz + mzD * cp.xCount / cp.loading, h * 1.1),
                                    ecColor="slategrey",
                                )
                                self.addArrow(
                                    self.ui.pl3,
                                    (
                                        cp.mz + mzD * cp.xCount / cp.loading,
                                        toDrawInts[bml],
                                    ),
                                    (cp.mz + mzD * cp.xCount / cp.loading, h * 1.1),
                                    drawArrowHead=True,
                                )

                                if self.ui.MSIsos.checkState() == QtCore.Qt.Checked:
                                    bml = (
                                        min(
                                            range(len(toDrawMzs)),
                                            key=lambda w: abs(toDrawMzs[w] - (cp.mz + 1.00335 / cp.loading)),
                                        )
                                        + 1
                                    )

                                    if cp.ionMode == "-" and hasPos:
                                        h = min(toDrawInts[bm], toDrawInts[bml])
                                    else:
                                        h = max(toDrawInts[bm], toDrawInts[bml])

                                    bm = (
                                        min(
                                            range(len(toDrawMzs)),
                                            key=lambda w: abs(toDrawMzs[w] - (cp.mz + 1.00335 * (cp.xCount - 1) / cp.loading)),
                                        )
                                        + 1
                                    )
                                    bml = (
                                        min(
                                            range(len(toDrawMzs)),
                                            key=lambda w: abs(toDrawMzs[w] - (cp.mz + 1.00335 * cp.xCount / cp.loading)),
                                        )
                                        + 1
                                    )

                                    if cp.ionMode == "-" and hasPos:
                                        h = min(toDrawInts[bm], toDrawInts[bml])
                                    else:
                                        h = max(toDrawInts[bm], toDrawInts[bml])

                                    intErrN, intErrL = self.getAllowedIsotopeRatioErrorsForResult()
                                    self.ui.pl3.twinxs[0].add_patch(
                                        patches.Rectangle(
                                            (
                                                cp.mz * (1.0 - annotationPPM / 1000000.0),
                                                intLeft * 0,
                                            ),
                                            cp.mz * (2 * annotationPPM / 1000000.0),
                                            0.01 * intLeft,
                                            edgecolor="none",
                                            facecolor="purple",
                                            alpha=0.2,
                                        )
                                    )
                                    self.ui.pl3.twinxs[0].add_patch(
                                        patches.Rectangle(
                                            (
                                                (cp.mz + (1.00335 * cp.xCount) / cp.loading) * (1.0 - annotationPPM / 1000000.0),
                                                intLeft * 0,
                                            ),
                                            (cp.mz + (1.00335 * cp.xCount) / cp.loading) * (2 * annotationPPM / 1000000.0),
                                            0.01 * intLeft,
                                            edgecolor="none",
                                            facecolor="purple",
                                            alpha=0.2,
                                        )
                                    )

                                    for iso in [1, 2, 3]:
                                        self.ui.pl3.twinxs[0].add_patch(
                                            patches.Rectangle(
                                                (
                                                    (cp.mz + (1.00335 * iso) / cp.loading) * (1.0 - annotationPPM / 1000000.0),
                                                    intLeft * 0,
                                                ),
                                                (cp.mz + (1.00335 * iso) / cp.loading) * (2 * annotationPPM / 1000000.0),
                                                0.01 * intLeft,
                                                edgecolor="none",
                                                facecolor="purple",
                                                alpha=0.2,
                                            )
                                        )
                                        self.ui.pl3.twinxs[0].add_patch(
                                            patches.Rectangle(
                                                (
                                                    (cp.mz + (1.00335 * (cp.xCount - iso)) / cp.loading) * (1.0 - annotationPPM / 1000000.0),
                                                    intLeft * 0,
                                                ),
                                                (cp.mz + (1.00335 * (cp.xCount - iso)) / cp.loading) * (2 * annotationPPM / 1000000.0),
                                                0.01 * intLeft,
                                                edgecolor="none",
                                                facecolor="purple",
                                                alpha=0.2,
                                            )
                                        )
                                        self.ui.pl3.twinxs[0].add_patch(
                                            patches.Rectangle(
                                                (
                                                    (cp.mz + (1.00335 * (cp.xCount + iso)) / cp.loading) * (1.0 - annotationPPM / 1000000.0),
                                                    intLeft * 0,
                                                ),
                                                (cp.mz + (1.00335 * (cp.xCount + iso)) / cp.loading) * (2 * annotationPPM / 1000000.0),
                                                0.01 * intLeft,
                                                edgecolor="none",
                                                facecolor="purple",
                                                alpha=0.2,
                                            )
                                        )

                                        ratioN = getNormRatio(purN, cp.xCount, iso)
                                        ratioL = getNormRatio(purL, cp.xCount, iso)
                                        self.addArrow(
                                            self.ui.pl3,
                                            (
                                                cp.mz + (1.00335 * iso) / cp.loading,
                                                intLeft * max(0, (ratioN - intErrN) - 0.005),
                                            ),
                                            (
                                                cp.mz + (1.00335 * iso) / cp.loading,
                                                intLeft * max(0, (ratioN - intErrN) + 0.005),
                                            ),
                                            linewidth=5,
                                            alpha=2,
                                            ecColor="DarkSeaGreen",
                                        )
                                        self.addArrow(
                                            self.ui.pl3,
                                            (
                                                cp.mz + (1.00335 * iso) / cp.loading,
                                                intLeft * max(0, (ratioN + intErrN) - 0.005),
                                            ),
                                            (
                                                cp.mz + (1.00335 * iso) / cp.loading,
                                                intLeft * (ratioN + intErrN) + 0.005,
                                            ),
                                            linewidth=5,
                                            alpha=2,
                                            ecColor="DarkSeaGreen",
                                        )
                                        self.addArrow(
                                            self.ui.pl3,
                                            (
                                                cp.mz + (1.00335 * iso) / cp.loading,
                                                intLeft * max(0, (ratioN - 0.005)),
                                            ),
                                            (
                                                cp.mz + (1.00335 * iso) / cp.loading,
                                                intLeft * max(0, (ratioN + 0.005)),
                                            ),
                                            linewidth=5,
                                            alpha=0.02,
                                            ecColor="Orange",
                                        )

                                        self.addArrow(
                                            self.ui.pl3,
                                            (
                                                cp.mz + 1.00335 * (cp.xCount - iso) / cp.loading,
                                                intRight
                                                * max(
                                                    0,
                                                    max(0, (ratioL - intErrL) - 0.005),
                                                ),
                                            ),
                                            (
                                                cp.mz + 1.00335 * (cp.xCount - iso) / cp.loading,
                                                intRight * max(0, (ratioL - intErrL) + 0.005),
                                            ),
                                            linewidth=5,
                                            alpha=0.02,
                                            ecColor="DarkSeaGreen",
                                        )
                                        self.addArrow(
                                            self.ui.pl3,
                                            (
                                                cp.mz + 1.00335 * (cp.xCount - iso) / cp.loading,
                                                intRight * max(0, (ratioL + intErrL) - 0.005),
                                            ),
                                            (
                                                cp.mz + 1.00335 * (cp.xCount - iso) / cp.loading,
                                                intRight * (ratioL + intErrL) + 0.005,
                                            ),
                                            linewidth=5,
                                            alpha=0.02,
                                            ecColor="DarkSeaGreen",
                                        )
                                        self.addArrow(
                                            self.ui.pl3,
                                            (
                                                cp.mz + 1.00335 * (cp.xCount - iso) / cp.loading,
                                                intRight * max(0, (ratioL - 0.005)),
                                            ),
                                            (
                                                cp.mz + 1.00335 * (cp.xCount - iso) / cp.loading,
                                                intRight * max(0, (ratioL + 0.005)),
                                            ),
                                            linewidth=5,
                                            alpha=0.1,
                                            ecColor="Orange",
                                        )

                                    if self.ui.drawFPIsotopologues.checkState() == QtCore.Qt.Checked:
                                        for iso in [1, 2, 3]:
                                            self.addArrow(
                                                self.ui.pl3,
                                                (
                                                    cp.mz - (1.00335 * iso) / cp.loading,
                                                    0,
                                                ),
                                                (
                                                    cp.mz - (1.00335 * iso) / cp.loading,
                                                    intLeft * 0.1,
                                                ),
                                                linewidth=5,
                                                ecColor="yellow",
                                            )
                                            self.addArrow(
                                                self.ui.pl3,
                                                (
                                                    cp.mz + 1.00335 * (cp.xCount + iso) / cp.loading,
                                                    0,
                                                ),
                                                (
                                                    cp.mz + 1.00335 * (cp.xCount + iso) / cp.loading,
                                                    intRight * 0.1,
                                                ),
                                                linewidth=5,
                                                ecColor="yellow",
                                            )

                                        self.addArrow(
                                            self.ui.pl3,
                                            (cp.mz - 1.00335 / (cp.loading * 2), 0),
                                            (
                                                cp.mz - 1.00335 / (cp.loading * 2),
                                                intLeft * 0.05,
                                            ),
                                            linewidth=5,
                                            ecColor="yellow",
                                        )
                                        self.addArrow(
                                            self.ui.pl3,
                                            (cp.mz + 1.00335 / (cp.loading * 2), 0),
                                            (
                                                cp.mz + 1.00335 / (cp.loading * 2),
                                                intLeft * 0.05,
                                            ),
                                            linewidth=5,
                                            ecColor="yellow",
                                        )
                                        self.addArrow(
                                            self.ui.pl3,
                                            (
                                                cp.mz + 1.00335 * cp.xCount / cp.loading + 1.00335 / (cp.loading * 2),
                                                0,
                                            ),
                                            (
                                                cp.mz + 1.00335 * cp.xCount / cp.loading + 1.00335 / (cp.loading * 2),
                                                intRight * 0.05,
                                            ),
                                            linewidth=5,
                                            ecColor="yellow",
                                        )
                                        self.addArrow(
                                            self.ui.pl3,
                                            (
                                                cp.mz + 1.00335 * cp.xCount / cp.loading - 1.00335 / (cp.loading * 2),
                                                0,
                                            ),
                                            (
                                                cp.mz + 1.00335 * cp.xCount / cp.loading - 1.00335 / (cp.loading * 2),
                                                intRight * 0.05,
                                            ),
                                            linewidth=5,
                                            ecColor="yellow",
                                        )

                                for i, mz in enumerate(toDrawMzs):
                                    for inc in range(0, cp.xCount + 1):
                                        if (
                                            abs(mz - (cp.mz + mzD * inc / cp.loading)) * 1000000.0 / (cp.mz + mzD * inc / cp.loading) <= annotationPPM
                                            or abs(mz - (cp.mz + mzD * (cp.xCount - inc) / cp.loading)) * 1000000.0 / (cp.mz + mzD * (cp.xCount - inc) / cp.loading) <= annotationPPM
                                            or abs(mz - (cp.mz + mzD * (cp.xCount + inc) / cp.loading)) * 1000000.0 / (cp.mz + mzD * (cp.xCount + inc) / cp.loading) <= annotationPPM
                                        ):
                                            ttoDrawMzs = []
                                            ttoDrawInts = []

                                            ttoDrawMzs.append(toDrawMzs[i])
                                            ttoDrawMzs.append(toDrawMzs[i])
                                            ttoDrawMzs.append(toDrawMzs[i])

                                            ttoDrawInts.append(0)
                                            ttoDrawInts.append(toDrawInts[i])
                                            ttoDrawInts.append(0)

                                            minMZ = min(minMZ, toDrawMzs[i])
                                            maxMZ = max(maxMZ, toDrawMzs[i])
                                            minMZH = min(minMZH, toDrawInts[i])
                                            maxMZH = max(maxMZH, toDrawInts[i])

                                            self.drawPlot(
                                                self.ui.pl3,
                                                plotIndex=0,
                                                x=ttoDrawMzs,
                                                y=ttoDrawInts,
                                                useCol="black",
                                                multipleLocator=None,
                                                alpha=0.1,
                                                title="",
                                                xlab="MZ",
                                            )

                    for childi in range(item.childCount()):
                        if not (item.child(childi).isHidden()):
                            child = item.child(childi).myData

                            childIDs.append(child.id)

                            mzs.append(child.mz)
                            peaks.append(child.NPeakCenterMin / 60.0)

                            xic = []
                            xicL = []
                            times = []

                            invert = 1
                            if self.ui.negEIC.isChecked():
                                invert = -1

                            xics_df = self.currentOpenResultsFile.db_con.tables["XICs"].filter(pl.col("id") == child.eicID)
                            for row_dict in xics_df.to_dicts():
                                xic = [float(t) for t in row_dict["xic"].split(";")]
                                xicL = [float(t) for t in row_dict["xicL"].split(";")]
                                xicfirstiso = [float(t) for t in row_dict["xicfirstiso"].split(";")]
                                xicLfirstiso = [float(t) for t in row_dict["xicLfirstiso"].split(";")]
                                xicLfirstisoconjugate = [float(t) for t in row_dict["xicLfirstisoconjugate"].split(";")]
                                xic_smoothed = [float(t) for t in row_dict["xic_smoothed"].split(";")]
                                xicL_smoothed = [float(t) for t in row_dict["xicL_smoothed"].split(";")]
                                times = [float(t) / 60.0 for t in row_dict["times"].split(";")]
                                xic_baseline = [float(t) for t in row_dict["xic_baseline"].split(";")]
                                xicL_baseline = [float(t) for t in row_dict["xicL_baseline"].split(";")]

                            minTime = min(
                                minTime,
                                min(times[int(child.NPeakCenter - child.NPeakScale * 1) : int(child.NPeakCenter + child.NPeakScale * 1)]),
                            )
                            maxTime = max(
                                maxTime,
                                max(times[int(child.NPeakCenter - child.NPeakScale * 1) : int(child.NPeakCenter + child.NPeakScale * 1)]),
                            )

                            if self.ui.scaleFeatures.isChecked():
                                s = int(child.NPeakCenter - child.NBorderLeft * 1)
                                e = int(child.NPeakCenter + child.NBorderRight * 1)
                                maxP = s + max(range(e - s), key=lambda x: xic[s + x])
                                maxE = mean(xic[(maxP - 3) : (maxP + 3)])
                                maxEL = mean(xicL[(maxP - 3) : (maxP + 3)])

                                if maxE == 0:
                                    maxE = 1
                                if maxEL == 0:
                                    maxEL = 1
                                if not self.ui.scaleLabelledFeatures.isChecked():
                                    maxEL = maxE

                                xic = [u / maxE for u in xic]
                                xic_smoothed = [u / maxE for u in xic_smoothed]
                                xicfirstiso = [u / maxE for u in xicfirstiso]
                                xic_baseline = [u / maxE for u in xic_baseline]

                                xicL = [u / maxEL for u in xicL]
                                xicL_smoothed = [u / maxEL for u in xicL_smoothed]
                                xicLfirstiso = [u / maxEL for u in xicLfirstiso]
                                xicLfirstisoconjugate = [u / maxEL for u in xicLfirstisoconjugate]
                                xicL_baseline = [u / maxEL for u in xicL_baseline]

                            xicL = [invert * u for u in xicL]
                            xicL_smoothed = [invert * u for u in xicL_smoothed]
                            xicLfirstiso = [invert * u for u in xicLfirstiso]
                            xicLfirstisoconjugate = [invert * u for u in xicLfirstisoconjugate]

                            if self.ui.flattenXIC.isChecked():
                                ps = min(
                                    int(child.NPeakCenter - child.NPeakScale * 2),
                                    int(child.LPeakCenter - child.LPeakScale * 2),
                                )
                                pe = max(
                                    int(child.NPeakCenter + child.NPeakScale * 2),
                                    int(child.LPeakCenter + child.LPeakScale * 2),
                                )
                            else:
                                ps = 0
                                pe = len(times)
                                # times=times[ps:pe]
                                # xic=xic[ps:pe]
                                # xic_smoothed=xic_smoothed[ps:pe]
                                # xicfirstiso=xicfirstiso[ps:pe]
                                # xicL=xicL[ps:pe]
                                # xicL_smoothed=xicL_smoothed[ps:pe]
                                # xicLfirstiso=xicLfirstiso[ps:pe]
                                # xicLfirstisoconjugate=xicLfirstisoconjugate[ps:pe]

                            maxInt = max(
                                maxInt,
                                max(xic[int(child.NPeakCenter - child.NPeakScale * 1) : int(child.NPeakCenter + child.NPeakScale * 1)]),
                            )

                            maxIntY = max(
                                maxIntY,
                                max(xic[int(child.NPeakCenter - child.NPeakScale * 1) : int(child.NPeakCenter + child.NPeakScale * 1)]),
                            )
                            minIntY = min(
                                minIntY,
                                min(xicL[int(child.LPeakCenter - child.LPeakScale * 1) : int(child.LPeakCenter + child.LPeakScale * 1)]),
                            )

                            self.drawPlot(
                                self.ui.pl1,
                                plotIndex=0,
                                x=times[ps:pe],
                                y=xic[ps:pe],
                                fill=[
                                    max(
                                        int(child.LPeakCenter - child.LBorderLeft),
                                        int(child.NPeakCenter - child.NBorderLeft),
                                    ),
                                    min(
                                        int(child.NPeakCenter + child.NBorderRight),
                                        int(child.LPeakCenter + child.LBorderRight),
                                    ),
                                ],
                                rearrange=len(selectedItems) == 1,
                                label="M: %.4f (%d, %s)" % (child.mz, child.xCount, child.ionMode),
                                useCol=useColi,
                            )  # useCol=selIndex*2)

                            if self.ui.checkBox_showBaseline.isChecked():
                                self.drawPlot(
                                    self.ui.pl1,
                                    plotIndex=0,
                                    x=times[ps:pe],
                                    y=xic_baseline[ps:pe],
                                    fill=[
                                        max(
                                            int(child.LPeakCenter - child.LBorderLeft),
                                            int(child.NPeakCenter - child.NBorderLeft),
                                        ),
                                        min(
                                            int(child.NPeakCenter + child.NBorderRight),
                                            int(child.LPeakCenter + child.LBorderRight),
                                        ),
                                    ],
                                    rearrange=len(selectedItems) == 1,
                                    label="",
                                    useCol=useColi,
                                    linestyle="--",
                                )
                                self.drawPlot(
                                    self.ui.pl1,
                                    plotIndex=0,
                                    x=times[ps:pe],
                                    y=xicL_baseline[ps:pe],
                                    fill=[
                                        max(
                                            int(child.LPeakCenter - child.LBorderLeft),
                                            int(child.NPeakCenter - child.NBorderLeft),
                                        ),
                                        min(
                                            int(child.NPeakCenter + child.NBorderRight),
                                            int(child.LPeakCenter + child.LBorderRight),
                                        ),
                                    ],
                                    rearrange=len(selectedItems) == 1,
                                    label="",
                                    useCol=useColi,
                                    linestyle="--",
                                )

                            if self.ui.showSmoothedEIC_checkBox.isChecked():
                                self.drawPlot(
                                    self.ui.pl1,
                                    plotIndex=0,
                                    x=times[ps:pe],
                                    y=xic_smoothed[ps:pe],
                                    fill=[
                                        max(
                                            int(child.LPeakCenter - child.LBorderLeft),
                                            int(child.NPeakCenter - child.NBorderLeft),
                                        ),
                                        min(
                                            int(child.NPeakCenter + child.NBorderRight),
                                            int(child.LPeakCenter + child.LBorderRight),
                                        ),
                                    ],
                                    rearrange=len(selectedItems) == 1,
                                    label="",
                                    useCol=useColi,
                                    linestyle="--",
                                )

                            if self.ui.showIsotopologues.isChecked():
                                self.drawPlot(
                                    self.ui.pl1,
                                    plotIndex=0,
                                    x=times[ps:pe],
                                    y=xicfirstiso[ps:pe],
                                    fill=[
                                        max(
                                            int(child.LPeakCenter - child.LBorderLeft),
                                            int(child.NPeakCenter - child.NBorderLeft),
                                        ),
                                        min(
                                            int(child.NPeakCenter + child.NBorderRight),
                                            int(child.LPeakCenter + child.LBorderRight),
                                        ),
                                    ],
                                    rearrange=len(selectedItems) == 1,
                                    label=None,
                                    useCol=useColi,
                                )

                            self.drawPlot(
                                self.ui.pl1,
                                plotIndex=0,
                                x=times[ps:pe],
                                y=xicL[ps:pe],
                                fill=[
                                    max(
                                        int(child.LPeakCenter - child.LBorderLeft),
                                        int(child.NPeakCenter - child.NBorderLeft),
                                    ),
                                    min(
                                        int(child.NPeakCenter + child.NBorderRight),
                                        int(child.LPeakCenter + child.LBorderRight),
                                    ),
                                ],
                                rearrange=len(selectedItems) == 1,
                                label="M': %.4f (%d, %s)" % (child.mz, child.xCount, child.ionMode),
                                useCol=useColi,
                            )

                            if self.ui.showSmoothedEIC_checkBox.isChecked():
                                self.drawPlot(
                                    self.ui.pl1,
                                    plotIndex=0,
                                    x=times[ps:pe],
                                    y=xicL_smoothed[ps:pe],
                                    fill=[
                                        max(
                                            int(child.LPeakCenter - child.LBorderLeft),
                                            int(child.NPeakCenter - child.NBorderLeft),
                                        ),
                                        min(
                                            int(child.NPeakCenter + child.NBorderRight),
                                            int(child.LPeakCenter + child.LBorderRight),
                                        ),
                                    ],
                                    rearrange=len(selectedItems) == 1,
                                    label=None,
                                    useCol=useColi,
                                    linestyle="--",
                                )

                            if self.ui.showIsotopologues.isChecked():
                                self.drawPlot(
                                    self.ui.pl1,
                                    plotIndex=0,
                                    x=times[ps:pe],
                                    y=xicLfirstiso[ps:pe],
                                    fill=[
                                        max(
                                            int(child.LPeakCenter - child.LBorderLeft),
                                            int(child.NPeakCenter - child.NBorderLeft),
                                        ),
                                        min(
                                            int(child.NPeakCenter + child.NBorderRight),
                                            int(child.LPeakCenter + child.LBorderRight),
                                        ),
                                    ],
                                    rearrange=len(selectedItems) == 1,
                                    label=None,
                                    useCol=useColi,
                                )

                            maxindex, maxvalue = max(
                                enumerate(
                                    xic[int(child.NPeakCenter - 1) : int(child.NPeakCenter + 1)],
                                    start=int(child.NPeakCenter - 1),
                                ),
                                key=itemgetter(1),
                            )
                            meanRT += child.NPeakCenterMin / 60.0
                            countK += 1

                    meanRT /= countK

                    if self.ui.plotAddLabels.checkState() == QtCore.Qt.Checked:
                        self.addAnnotation(
                            self.ui.pl1,
                            "%s\n@ %.2f" % (item.myData.featureName, meanRT),
                            (meanRT, maxInt),
                            (times[maxindex], xic[maxindex]),
                            0,
                        )
                useColi += 1

            if len(selFeatureGroups) >= 1:
                childIDs = []
                for selFeatureGroup in selFeatureGroups:
                    for childi in range(selFeatureGroup.childCount()):
                        child = selFeatureGroup.child(childi).myData
                        childIDs.append(child.id)
                if len(childIDs) >= 1:
                    try:
                        self.clearPlot(self.ui.pl2A)
                        self.clearPlot(self.ui.pl2B)
                        plt.cla()
                        self.ui.pl2A.fig.subplots_adjust(left=0.15, bottom=0.05, right=0.99, top=0.85)
                        self.ui.pl2B.fig.subplots_adjust(left=0.15, bottom=0.05, right=0.99, top=0.85)

                        fRows = 0
                        minCorr = 1
                        config_df = self.currentOpenResultsFile.db_con.tables["config"].filter(pl.col("key") == "minCorrelation")
                        for row_dict in config_df.to_dicts():
                            fRows += 1
                            minCorr = float(row_dict["value"])
                        assert 0 < fRows <= 1, "Min Correlation not found or found multiple times in settings"

                        dataCorr = []
                        dataSILRatios = []
                        featureIDToColsNum = {}
                        texts = []
                        for i in range(len(childIDs)):
                            arow = []
                            brow = []
                            for j in range(len(childIDs)):
                                arow.append(0)
                                brow.append(0)
                            dataCorr.append(arow)
                            dataSILRatios.append(brow)
                            texts.append("")
                            featureIDToColsNum[childIDs[i]] = i

                        ",".join(["%d" % f for f in childIDs])

                        chrompeaks_df = self.currentOpenResultsFile.db_con.tables["chromPeaks"].filter(pl.col("id").is_in(childIDs))
                        for row_dict in chrompeaks_df.to_dicts():
                            id = row_dict["id"]
                            mz = row_dict["mz"]
                            xcount = row_dict["xcount"]
                            ionMode = row_dict["ionMode"]

                            fI1 = featureIDToColsNum[id]
                            texts[fI1] = "%s%.4f/%d" % (ionMode, mz, xcount)

                        minCorrCurrent = 1
                        maxSilRatioCurrent = 0
                        ff_df = self.currentOpenResultsFile.db_con.tables["featurefeatures"].filter(pl.col("fID1").is_in(childIDs) & pl.col("fID2").is_in(childIDs))
                        for row_dict in ff_df.to_dicts():
                            fID1 = row_dict["fID1"]
                            fID2 = row_dict["fID2"]
                            correlation = row_dict["corr"]
                            silRatioValue = row_dict["silRatioValue"]

                            minCorrCurrent = min(minCorrCurrent, correlation)
                            maxSilRatioCurrent = max(maxSilRatioCurrent, silRatioValue - 1)

                            fI1 = featureIDToColsNum[fID1]
                            fI2 = featureIDToColsNum[fID2]

                            if fI1 > fI2:
                                a = fI2
                                fI2 = fI1
                                fI1 = a

                            dataCorr[fI1][fI2] = correlation
                            dataCorr[fI2][fI1] = correlation

                            dataSILRatios[fI1][fI2] = silRatioValue - 1
                            dataSILRatios[fI2][fI1] = silRatioValue - 1

                        dv = []
                        for i in range(len(dataCorr)):
                            dataCorr[i][i] = 1
                            dv.extend(dataCorr[i])

                        hc = HCA_general.HCA_generic()
                        tree = hc.generateTree(dataCorr)

                        datOrd = range(len(dataCorr))

                        datOrd = hc.getObjsOrderInTree(tree)
                        dnew = []
                        dnew2 = []
                        for i in datOrd:
                            dnew.append([dataCorr[i][j] for j in datOrd])
                            dnew2.append([dataSILRatios[i][j] for j in datOrd])
                        dataCorr = dnew
                        dataSILRatios = dnew2

                        dataCorr[len(dataCorr) - 1][len(dataCorr) - 1] = 1

                        colorDict = {
                            "red": (
                                (0.0, 0, convertXaToX(47 / 255.0, 0.2)),
                                (
                                    0.5,
                                    convertXaToX(47 / 255.0, 0.5),
                                    convertXaToX(178 / 255.0, 0.225),
                                ),
                                (
                                    minCorr + (1 - minCorr) / 2.0,
                                    convertXaToX(178 / 255.0, 0.525),
                                    convertXaToX(154 / 255.0, 0.5),
                                ),
                                (1.0, convertXaToX(154 / 255.0, 0.9), 0),
                            ),
                            "green": (
                                (0.0, 0, convertXaToX(79 / 255.0, 0.2)),
                                (
                                    0.5,
                                    convertXaToX(79 / 255.0, 0.5),
                                    convertXaToX(34 / 255.0, 0.225),
                                ),
                                (
                                    minCorr + (1 - minCorr) / 2.0,
                                    convertXaToX(34 / 255.0, 0.525),
                                    convertXaToX(205 / 255.0, 0.5),
                                ),
                                (1.0, convertXaToX(205 / 255.0, 0.9), 0),
                            ),
                            "blue": (
                                (0.0, 0, convertXaToX(79 / 255.0, 0.2)),
                                (
                                    0.5,
                                    convertXaToX(79 / 255.0, 0.5),
                                    convertXaToX(34 / 255.0, 0.225),
                                ),
                                (
                                    minCorr + (1 - minCorr) / 2.0,
                                    convertXaToX(34 / 255.0, 0.525),
                                    convertXaToX(50 / 255.0, 0.5),
                                ),
                                (1.0, convertXaToX(50 / 255.0, 0.9), 0.0),
                            ),
                        }

                        from matplotlib.colors import LinearSegmentedColormap

                        custom_cmap = LinearSegmentedColormap("custom_colormap", colorDict)
                        cax = self.ui.pl2A.twinxs[0].matshow(dataCorr, cmap=custom_cmap)

                        self.ui.pl2A.axes.set_xticks([i for i in range(len(dataCorr))])
                        self.ui.pl2A.axes.set_xticklabels([texts[i] for i in datOrd], rotation=90)
                        self.ui.pl2A.axes.set_yticks([i for i in range(len(dataCorr))])
                        self.ui.pl2A.axes.set_yticklabels([texts[i] for i in datOrd])

                        self.ui.pl2A.axes.set_yticks([i - 0.5 for i in range(len(dataCorr) + 1)], minor=True)
                        self.ui.pl2A.axes.set_xticks([i - 0.5 for i in range(len(dataCorr) + 1)], minor=True)

                        self.ui.pl2A.axes.grid(ls="solid", which="minor", color="white", linewidth=2)

                        cax.set_clim(-1, 1)
                        cb = self.ui.pl2A.fig.colorbar(cax, ticks=[-1, 0, minCorr, minCorrCurrent])
                        cb.outline.set_linewidth(0)

                        colorDict = {
                            "red": (
                                (0.0, 0, convertXaToX(47 / 255.0, 0.2)),
                                (
                                    0.5,
                                    convertXaToX(47 / 255.0, 0.5),
                                    convertXaToX(154 / 255.0, 0.9),
                                ),
                                (
                                    0.6,
                                    convertXaToX(154 / 255.0, 0.5),
                                    convertXaToX(178 / 255.0, 0.225),
                                ),
                                (1.0, convertXaToX(178 / 255.0, 0.525), 0),
                            ),
                            "green": (
                                (0.0, 0, convertXaToX(79 / 255.0, 0.2)),
                                (
                                    0.5,
                                    convertXaToX(79 / 255.0, 0.5),
                                    convertXaToX(34 / 255.0, 0.225),
                                ),
                                (
                                    0.6,
                                    convertXaToX(205 / 255.0, 0.5),
                                    convertXaToX(205 / 255.0, 0.9),
                                ),
                                (1.0, convertXaToX(34 / 255.0, 0.525), 0),
                            ),
                            "blue": (
                                (0.0, 0, convertXaToX(79 / 255.0, 0.2)),
                                (
                                    0.5,
                                    convertXaToX(79 / 255.0, 0.5),
                                    convertXaToX(50 / 255.0, 0.9),
                                ),
                                (
                                    0.6,
                                    convertXaToX(50 / 255.0, 0.5),
                                    convertXaToX(34 / 255.0, 0.225),
                                ),
                                (1.0, convertXaToX(34 / 255.0, 0.525), 0.0),
                            ),
                        }

                        custom_cmap2 = LinearSegmentedColormap("custom_colormap2", colorDict)
                        cbx = self.ui.pl2B.twinxs[0].matshow(dataSILRatios, cmap=custom_cmap2)
                        self.ui.pl2B.axes.set_xticks([i for i in range(len(dataSILRatios))])
                        self.ui.pl2B.axes.set_xticklabels([texts[i] for i in datOrd], rotation=90)
                        self.ui.pl2B.axes.set_yticks([i for i in range(len(dataSILRatios))])
                        self.ui.pl2B.axes.set_yticklabels([texts[i] for i in datOrd])

                        self.ui.pl2B.axes.set_yticks([i - 0.5 for i in range(len(dataSILRatios) + 1)], minor=True)
                        self.ui.pl2B.axes.set_xticks([i - 0.5 for i in range(len(dataSILRatios) + 1)], minor=True)

                        self.ui.pl2B.axes.grid(ls="solid", which="minor", color="white", linewidth=2)

                        cbx.set_clim(-1, 1)
                        cb = self.ui.pl2B.fig.colorbar(cbx, ticks=[0, 0.1, 0.2, maxSilRatioCurrent])
                        cb.outline.set_linewidth(0)

                        self.ui.pl2A.fig.canvas.draw()
                        self.ui.pl2B.fig.canvas.draw()

                    except Exception:
                        traceback.print_exc()
                        logging.error(str(traceback))

            # </editor-fold>

        useColi = 0
        for selIndex, item in enumerate(selectedItems):
            if not (hasattr(item, "myType")):
                continue

            # <editor-fold desc="#feature results">
            if item.myType == "Features" or item.myType == "feature":
                self.ui.chromPeakName.setText("")

                self.ui.res_ExtractedData.setHeaderLabels(
                    [
                        "MZ (/Ionmode Z)",
                        "Rt min",
                        "Xn",
                        "Adducts / in-source fragments / hetero atoms",
                        "Scale M / M'",
                        "Peak cor",
                        "M:M' peaks ratio / area ratio",
                        "Area M / M'",
                        "Scans",
                        "LMZ",
                        "Tracer",
                    ]
                )

                if item.myType == "feature":
                    cp = item.myData
                    plotTypes.add("feature")
                    mzs.append(cp.mz)
                    peaks.append(cp.NPeakCenterMin / 60.0)
                    xic = []
                    xicL = []
                    times = []
                    maxE = 1

                    invert = 1
                    if self.ui.negEIC.isChecked():
                        invert = -1

                    self.isTracerMetabolisationExperiment()
                    mzD, purN, purL = self.getLabellingParametersForResult(cp.id)

                    toDrawMzs = []
                    toDrawInts = []
                    ms_df = self.currentOpenResultsFile.db_con.tables["massspectrum"].filter(pl.col("mID") == cp.massSpectrumID)
                    for row_dict in ms_df.to_dicts():
                        mzs = [float(t) for t in row_dict["mzs"].split(";")]
                        intensities = [float(t) for t in row_dict["intensities"].split(";")]

                        for mz in mzs:
                            toDrawMzs.append(mz)
                            toDrawMzs.append(mz)
                            toDrawMzs.append(mz)

                        for intensity in intensities:
                            toDrawInts.append(0)
                            if cp.ionMode == "-" and featuresPosSelected:
                                toDrawInts.append(-intensity)
                            else:
                                toDrawInts.append(intensity)
                            toDrawInts.append(0)

                    # self.drawPlot(self.ui.pl3, plotIndex=0, x=toDrawMzs, y=toDrawInts, useCol="lightgrey", multipleLocator=None,  alpha=0.1, title="", xlab="MZ")

                    if not toDrawMzs:
                        continue

                    bm = (
                        min(
                            range(len(toDrawMzs)),
                            key=lambda i: abs(toDrawMzs[i] - cp.mz),
                        )
                        + 1
                    )
                    bml = (
                        min(
                            range(len(toDrawMzs)),
                            key=lambda i: abs(toDrawMzs[i] - (cp.lmz)),
                        )
                        + 1
                    )

                    intLeft = toDrawInts[bm]
                    intRight = toDrawInts[bml]

                    h = 0
                    if cp.ionMode == "-" and featuresPosSelected:
                        h = min(intLeft, intRight)
                    else:
                        h = max(intLeft, intRight)

                    if self.ui.MSLabels.checkState() == QtCore.Qt.Checked:
                        self.addAnnotation(
                            self.ui.pl3,
                            "mz: %.5f\nl-mz: %.5f\nd-mz: %.5f\nXn: %s Z: %s%d" % (cp.mz, cp.lmz, mzD, cp.xCount, cp.ionMode, cp.loading),
                            (cp.mz + (cp.lmz - cp.mz) / 2.0, h * 1.1),
                            (10, 120),
                            rotation=0,
                            up=not (cp.ionMode == "-" and featuresPosSelected),
                        )

                    self.addArrow(
                        self.ui.pl3,
                        (cp.mz, toDrawInts[bm]),
                        (cp.mz, h * 1.1),
                        drawArrowHead=True,
                    )
                    self.addArrow(
                        self.ui.pl3,
                        (cp.mz, h * 1.1),
                        (cp.lmz, h * 1.1),
                        ecColor="slategrey",
                    )
                    self.addArrow(
                        self.ui.pl3,
                        (cp.lmz, toDrawInts[bml]),
                        (cp.lmz, h * 1.1),
                        drawArrowHead=True,
                    )

                    if self.ui.MSIsos.checkState() == QtCore.Qt.Checked:
                        bml = (
                            min(
                                range(len(toDrawMzs)),
                                key=lambda w: abs(toDrawMzs[w] - (cp.mz + 1.00335 / cp.loading)),
                            )
                            + 1
                        )

                        if cp.ionMode == "-" and featuresPosSelected:
                            h = min(toDrawInts[bm], toDrawInts[bml])
                        else:
                            h = max(toDrawInts[bm], toDrawInts[bml])

                        bm = (
                            min(
                                range(len(toDrawMzs)),
                                key=lambda w: abs(toDrawMzs[w] - (cp.lmz - 1.00335 / cp.loading)),
                            )
                            + 1
                        )
                        bml = (
                            min(
                                range(len(toDrawMzs)),
                                key=lambda w: abs(toDrawMzs[w] - (cp.lmz + 1.00335 / cp.loading)),
                            )
                            + 1
                        )

                        if cp.ionMode == "-" and featuresPosSelected:
                            h = min(toDrawInts[bm], toDrawInts[bml])
                        else:
                            h = max(toDrawInts[bm], toDrawInts[bml])

                        intErrN, intErrL = self.getAllowedIsotopeRatioErrorsForResult()

                        self.ui.pl3.twinxs[0].add_patch(
                            patches.Rectangle(
                                (
                                    cp.mz * (1.0 - annotationPPM / 1000000.0),
                                    intLeft * 0,
                                ),
                                cp.mz * (2 * annotationPPM / 1000000.0),
                                0.01 * intLeft,
                                edgecolor="none",
                                facecolor="purple",
                                alpha=0.2,
                            )
                        )
                        self.ui.pl3.twinxs[0].add_patch(
                            patches.Rectangle(
                                (
                                    (cp.lmz + 1.00335 / cp.loading) * (1.0 - annotationPPM / 1000000.0),
                                    intLeft * 0,
                                ),
                                (cp.lmz + 1.00335 / cp.loading) * (2 * annotationPPM / 1000000.0),
                                0.01 * intLeft,
                                edgecolor="none",
                                facecolor="purple",
                                alpha=0.2,
                            )
                        )

                        for iso in [1, 2, 3]:
                            self.ui.pl3.twinxs[0].add_patch(
                                patches.Rectangle(
                                    (
                                        (cp.mz + (1.00335 * iso) / cp.loading) * (1.0 - annotationPPM / 1000000.0),
                                        intLeft * 0,
                                    ),
                                    (cp.mz + (1.00335 * iso) / cp.loading) * (2 * annotationPPM / 1000000.0),
                                    0.01 * intLeft,
                                    edgecolor="none",
                                    facecolor="purple",
                                    alpha=0.2,
                                )
                            )
                            self.ui.pl3.twinxs[0].add_patch(
                                patches.Rectangle(
                                    (
                                        (cp.lmz - (1.00335 * iso) / cp.loading) * (1.0 - annotationPPM / 1000000.0),
                                        intLeft * 0,
                                    ),
                                    (cp.lmz - (1.00335 * iso) / cp.loading) * (2 * annotationPPM / 1000000.0),
                                    0.01 * intLeft,
                                    edgecolor="none",
                                    facecolor="purple",
                                    alpha=0.2,
                                )
                            )
                            self.ui.pl3.twinxs[0].add_patch(
                                patches.Rectangle(
                                    (
                                        (cp.lmz + (1.00335 * iso) / cp.loading) * (1.0 - annotationPPM / 1000000.0),
                                        intLeft * 0,
                                    ),
                                    (cp.lmz + (1.00335 * iso) / cp.loading) * (2 * annotationPPM / 1000000.0),
                                    0.01 * intLeft,
                                    edgecolor="none",
                                    facecolor="purple",
                                    alpha=0.2,
                                )
                            )

                            if not (isinstance(cp.xCount, str)):
                                ratioN = getNormRatio(purN, cp.xCount, iso)
                                ratioL = getNormRatio(purL, cp.xCount, iso)
                                self.addArrow(
                                    self.ui.pl3,
                                    (
                                        cp.mz + (1.00335 * iso) / cp.loading,
                                        intLeft * max(0, (ratioN - intErrN) - 0.005),
                                    ),
                                    (
                                        cp.mz + (1.00335 * iso) / cp.loading,
                                        intLeft * max(0, (ratioN - intErrN) + 0.005),
                                    ),
                                    linewidth=5,
                                    alpha=2,
                                    ecColor="DarkSeaGreen",
                                )

                                self.addArrow(
                                    self.ui.pl3,
                                    (
                                        cp.mz + (1.00335 * iso) / cp.loading,
                                        intLeft * max(0, (ratioN + intErrN) - 0.005),
                                    ),
                                    (
                                        cp.mz + (1.00335 * iso) / cp.loading,
                                        intLeft * (ratioN + intErrN) + 0.005,
                                    ),
                                    linewidth=5,
                                    alpha=2,
                                    ecColor="DarkSeaGreen",
                                )
                                self.addArrow(
                                    self.ui.pl3,
                                    (
                                        cp.mz + (1.00335 * iso) / cp.loading,
                                        intLeft * max(0, (ratioN - 0.005)),
                                    ),
                                    (
                                        cp.mz + (1.00335 * iso) / cp.loading,
                                        intLeft * max(0, (ratioN + 0.005)),
                                    ),
                                    linewidth=5,
                                    alpha=0.02,
                                    ecColor="Orange",
                                )

                                self.addArrow(
                                    self.ui.pl3,
                                    (
                                        cp.lmz - 1.00335 * iso / cp.loading,
                                        intRight * max(0, max(0, (ratioL - intErrL) - 0.005)),
                                    ),
                                    (
                                        cp.lmz - 1.00335 * iso / cp.loading,
                                        intRight * max(0, (ratioL - intErrL) + 0.005),
                                    ),
                                    linewidth=5,
                                    alpha=0.02,
                                    ecColor="DarkSeaGreen",
                                )
                                self.addArrow(
                                    self.ui.pl3,
                                    (
                                        cp.lmz - 1.00335 * iso / cp.loading,
                                        intRight * max(0, (ratioL + intErrL) - 0.005),
                                    ),
                                    (
                                        cp.lmz - 1.00335 * iso / cp.loading,
                                        intRight * (ratioL + intErrL) + 0.005,
                                    ),
                                    linewidth=5,
                                    alpha=0.02,
                                    ecColor="DarkSeaGreen",
                                )
                                self.addArrow(
                                    self.ui.pl3,
                                    (
                                        cp.mz + 1.00335 * (cp.xCount - iso) / cp.loading,
                                        intRight * max(0, (ratioL - 0.005)),
                                    ),
                                    (
                                        cp.mz + 1.00335 * (cp.xCount - iso) / cp.loading,
                                        intRight * max(0, (ratioL + 0.005)),
                                    ),
                                    linewidth=5,
                                    alpha=0.1,
                                    ecColor="Orange",
                                )

                        if self.ui.drawFPIsotopologues.checkState() == QtCore.Qt.Checked:
                            for iso in [1, 2, 3]:
                                self.addArrow(
                                    self.ui.pl3,
                                    (cp.mz - (1.00335 * iso) / cp.loading, 0),
                                    (
                                        cp.mz - (1.00335 * iso) / cp.loading,
                                        intLeft * 0.1,
                                    ),
                                    linewidth=5,
                                    ecColor="yellow",
                                )
                                self.addArrow(
                                    self.ui.pl3,
                                    (cp.lmz + 1.00335 * iso / cp.loading, 0),
                                    (
                                        cp.lmz + 1.00335 * iso / cp.loading,
                                        intRight * 0.1,
                                    ),
                                    linewidth=5,
                                    ecColor="yellow",
                                )

                            self.addArrow(
                                self.ui.pl3,
                                (cp.mz - 1.00335 / (cp.loading * 2), 0),
                                (cp.mz - 1.00335 / (cp.loading * 2), intLeft * 0.05),
                                linewidth=5,
                                ecColor="yellow",
                            )
                            self.addArrow(
                                self.ui.pl3,
                                (cp.mz + 1.00335 / (cp.loading * 2), 0),
                                (cp.mz + 1.00335 / (cp.loading * 2), intLeft * 0.05),
                                linewidth=5,
                                ecColor="yellow",
                            )
                            self.addArrow(
                                self.ui.pl3,
                                (cp.lmz + 1.00335 / (cp.loading * 2), 0),
                                (cp.lmz + 1.00335 / (cp.loading * 2), intRight * 0.05),
                                linewidth=5,
                                ecColor="yellow",
                            )
                            self.addArrow(
                                self.ui.pl3,
                                (cp.lmz - 1.00335 / (cp.loading * 2), 0),
                                (cp.lmz - 1.00335 / (cp.loading * 2), intRight * 0.05),
                                linewidth=5,
                                ecColor="yellow",
                            )

                    ms_df = self.currentOpenResultsFile.db_con.tables["massspectrum"].filter(pl.col("mID") == cp.massSpectrumID)
                    for row_dict in ms_df.to_dicts():
                        mzs = [float(t) for t in row_dict["mzs"].split(";")]
                        intensities = [float(t) for t in row_dict["intensities"].split(";")]

                        for i, mz in enumerate(mzs):
                            for inc in range(0, int((cp.lmz - cp.mz) // 1.00335484) * cp.loading):
                                if (
                                    abs(mz - (cp.mz + 1.00335484 * inc / cp.loading)) * 1000000.0 / (cp.mz + 1.00335484 * inc / cp.loading) <= annotationPPM
                                    or abs(mz - (cp.lmz + 1.00335484 * inc / cp.loading)) * 1000000.0 / (cp.lmz + 1.00335484 * inc / cp.loading) <= annotationPPM
                                    or abs(mz - (cp.lmz - 1.00335484 * inc / cp.loading)) * 1000000.0 / (cp.lmz - 1.00335484 * inc / cp.loading) <= annotationPPM
                                ):
                                    toDrawMzs = []
                                    toDrawInts = []

                                    toDrawMzs.append(mzs[i])
                                    toDrawMzs.append(mzs[i])
                                    toDrawMzs.append(mzs[i])

                                    toDrawInts.append(0)
                                    if cp.ionMode == "-" and featuresPosSelected:
                                        toDrawInts.append(-intensities[i])
                                        minMZH = min(minMZH, -intensities[i])
                                    else:
                                        toDrawInts.append(intensities[i])
                                        maxMZH = max(maxMZH, intensities[i])
                                    toDrawInts.append(0)

                                    minMZ = min(minMZ, mzs[i])
                                    maxMZ = max(maxMZ, mzs[i])

                                    self.drawPlot(
                                        self.ui.pl3,
                                        plotIndex=0,
                                        x=toDrawMzs,
                                        y=toDrawInts,
                                        useCol=predefinedColors[useColi % len(predefinedColors)],
                                        multipleLocator=None,
                                        alpha=0.1,
                                        title="",
                                        xlab="MZ",
                                    )
                                    # self.drawPlot(self.ui.pl3, plotIndex=0, x=toDrawMzs, y=toDrawInts, useCol="black", multipleLocator=None,  alpha=0.1, title="", xlab="MZ")
                useColi += 1
            # </editor-fold>

        # <editor-fold desc="#featureGroup results">

        if len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - observed intensities":
            intensities = []
            mzs_df = self.currentOpenResultsFile.db_con.tables["mzs"]
            for row_dict in mzs_df.to_dicts():
                intensities.append(log10(row_dict["intensity"]))
            self.ui.pl1.twinxs[0].hist(intensities, 50, facecolor="green", alpha=0.5)
            self.ui.pl1.axes.set_title("Histogram of matched signal pairs - intensity of native signals")
            self.ui.pl1.axes.set_xlabel("Log10(Signal intensity)")
            self.ui.pl1.axes.set_ylabel("Frequency")
            self.drawCanvas(self.ui.pl1)

        elif len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - observed intensities comparison":
            intensities = []
            intensitiesL = []
            text, ok = QtWidgets.QInputDialog.getText(
                self.parentWidget(),
                "Observed intensities comparions",
                "Please enter the expected ratio of native to labeled signal pair abundances (e.g. 1)",
            )
            expRatio = 1
            if ok:
                expRatio = float(text)
            mzs_df = self.currentOpenResultsFile.db_con.tables["mzs"]
            for row_dict in mzs_df.to_dicts():
                intensities.append(log10(row_dict["intensity"]))
                intensitiesL.append(log10(row_dict["intensityL"]))
            minInt = min(min(intensities), min(intensitiesL))
            maxInt = max(max(intensities), max(intensitiesL))

            self.ui.pl1.twinxs[0].plot(
                [log10(10**minInt), log10((10**maxInt) * 2)],
                [
                    log10((10**minInt) * (1 / expRatio)),
                    log10((10**maxInt) * 2 * (1 / expRatio)),
                ],
            )
            self.ui.pl1.twinxs[0].plot(intensities, intensitiesL, "ro")
            self.ui.pl1.axes.set_title("Histogram of matched signal pairs - intensity of native and labeled signals")
            self.ui.pl1.axes.set_xlabel("Log10(Native signal intensity)")
            self.ui.pl1.axes.set_ylabel("Log10(Labeled signal intensity)")
            self.drawCanvas(self.ui.pl1)

        elif len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - relative mz error":
            mzdifferrors = []
            mzs_df = self.currentOpenResultsFile.db_con.tables["mzs"]
            for row_dict in mzs_df.to_dicts():
                mz = row_dict["mz"]
                lmz = row_dict["lmz"]
                tmz = row_dict["tmz"]
                mzdifferrors.append((lmz - mz - tmz) * 1000000 / mz)
            self.ui.pl1.twinxs[0].hist(mzdifferrors, 50, facecolor="green", alpha=0.5)
            self.ui.pl1.axes.set_title("Histogram of matched signal pairs - relative m/z error (ppm)")
            self.ui.pl1.axes.set_xlabel("Relative m/z error (ppm)")
            self.ui.pl1.axes.set_ylabel("Frequency")
            self.drawCanvas(self.ui.pl1)

        elif len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - relative mz error vs intensity":
            mzdifferrors = []
            intensities = []
            mzs_df = self.currentOpenResultsFile.db_con.tables["mzs"]
            for row_dict in mzs_df.to_dicts():
                mz = row_dict["mz"]
                lmz = row_dict["lmz"]
                tmz = row_dict["tmz"]
                intensity = row_dict["intensity"]
                mzdifferrors.append((lmz - mz - tmz) * 1000000 / mz)
                intensities.append(log10(intensity))
            self.ui.pl1.twinxs[0].plot(mzdifferrors, intensities, "ro")
            self.ui.pl1.axes.set_title("Matched signal pairs - relative m/z error (ppm) vs. signal intensity")
            self.ui.pl1.axes.set_xlabel("Relative m/z error (ppm)")
            self.ui.pl1.axes.set_ylabel("Log10(Signal intensity)")
            self.drawCanvas(self.ui.pl1)

        elif len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - relative mzbin deviation":
            mzdeviations = []
            mzbinskids_df = self.currentOpenResultsFile.db_con.tables["mzbinskids"]
            mzs_df = self.currentOpenResultsFile.db_con.tables["mzs"]
            joined = mzbinskids_df.join(mzs_df, left_on="mzid", right_on="id", how="inner")
            grouped = joined.group_by("mzbinid").agg([pl.col("mz").min().alias("mzmin"), pl.col("mz").max().alias("mzmax"), pl.col("mz").count().alias("n")])
            for row_dict in grouped.to_dicts():
                row_dict["mzbinid"]
                mzmin = row_dict["mzmin"]
                mzmax = row_dict["mzmax"]
                n = row_dict["n"]
                if n > 1:
                    mzdeviations.append((mzmax - mzmin) * 1000000 / mzmin)
            self.ui.pl1.twinxs[0].hist(mzdeviations, 50, facecolor="green", alpha=0.5)
            self.ui.pl1.axes.set_title("Histogram of binned signal pairs - relative m/z variance (ppm)")
            self.ui.pl1.axes.set_xlabel("Relative m/z deviation (ppm)")
            self.ui.pl1.axes.set_ylabel("Frequency")
            self.drawCanvas(self.ui.pl1)

        elif len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - feature pair correlations":
            peaksCorr = []
            chrompeaks_df = self.currentOpenResultsFile.db_con.tables["chromPeaks"]
            for row_dict in chrompeaks_df.to_dicts():
                peaksCorr.append(row_dict["peaksCorr"])
            self.ui.pl1.twinxs[0].hist(
                peaksCorr,
                [i / 100.0 for i in range(-100, 100, 1)],
                facecolor="green",
                alpha=0.5,
            )
            self.ui.pl1.axes.set_title("Histogram of matched feature pairs - peak correlations")
            self.ui.pl1.axes.set_xlabel("Peak correlation")
            self.ui.pl1.axes.set_ylabel("Frequency")
            self.drawCanvas(self.ui.pl1)

        elif len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - feature pair mp1 to m ratio abs":
            peaksRatio = []
            text, ok = QtWidgets.QInputDialog.getText(
                self.parentWidget(),
                "Isotopic enrichment",
                "Please enter the isotopic enrichment to be used for the diagnostics plot (e.g. 0.9893)",
                text="0.9893",
            )
            isoEnr = 0.9893
            if ok:
                text = text.replace(",", ".")
                isoEnr = float(text)
            chrompeaks_df = self.currentOpenResultsFile.db_con.tables["chromPeaks"]
            for row_dict in chrompeaks_df.to_dicts():
                peakRatio = row_dict["peaksRatioMp1"]
                xcount = row_dict["xcount"]
                if peakRatio == -1:
                    peaksRatio.append(-100.0)
                else:
                    peaksRatio.append(100.0 * max(-100, min(100, peakRatio - getNormRatio(isoEnr, xcount, 1))))
            self.ui.pl1.twinxs[0].hist(
                peaksRatio,
                [i for i in [i / 10.0 for i in range(-1000, 1000, 25)]],
                facecolor="green",
                alpha=0.5,
            )
            self.ui.pl1.axes.set_title("Histogram of isotopolog ratio error - observed minus theoretical ratio for M+1 to M")
            self.ui.pl1.axes.set_xlabel("Ratio error (%)")
            self.ui.pl1.axes.set_ylabel("Frequency")
            self.drawCanvas(self.ui.pl1)

        elif len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - feature pair mp1 to m ratio rel":
            peaksRatio = []
            text, ok = QtWidgets.QInputDialog.getText(
                self.parentWidget(),
                "Isotopic enrichment",
                "Please enter the isotopic enrichment to be used for the diagnostics plot (e.g. 0.9893)",
                text="0.9893",
            )
            isoEnr = 0.9893
            if ok:
                text = text.replace(",", ".")
                isoEnr = float(text)
            chrompeaks_df = self.currentOpenResultsFile.db_con.tables["chromPeaks"]
            for row_dict in chrompeaks_df.to_dicts():
                peakRatio = row_dict["peaksRatioMp1"]
                xcount = row_dict["xcount"]
                if peakRatio == -1:
                    peaksRatio.append(-100.0)
                else:
                    peaksRatio.append(
                        max(
                            -100,
                            min(
                                100,
                                100 * (peakRatio / getNormRatio(isoEnr, xcount, 1) - 1),
                            ),
                        )
                    )
            self.ui.pl1.twinxs[0].hist(
                peaksRatio,
                [i for i in [i / 10.0 for i in range(-1000, 1000, 25)]],
                facecolor="green",
                alpha=0.5,
            )
            self.ui.pl1.axes.set_title("Histogram of RIA  for M+1 to M")
            self.ui.pl1.axes.set_xlabel("RIA (%)")
            self.ui.pl1.axes.set_ylabel("Frequency")
            self.drawCanvas(self.ui.pl1)

        elif len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - feature pair mp1 to m IRA vs intensity":
            peaksRatio = []
            areas = []
            text, ok = QtWidgets.QInputDialog.getText(
                self.parentWidget(),
                "Isotopic enrichment",
                "Please enter the isotopic enrichment to be used for the diagnostics plot (e.g. 0.9893)",
                text="0.9893",
            )
            isoEnr = 0.9893
            if ok:
                text = text.replace(",", ".")
                isoEnr = float(text)
            chrompeaks_df = self.currentOpenResultsFile.db_con.tables["chromPeaks"]
            for row_dict in chrompeaks_df.to_dicts():
                peakRatio = row_dict["peaksRatioMp1"]
                xcount = row_dict["xcount"]
                area = row_dict["LPeakArea"]
                area = log10(area)
                if peakRatio == -1:
                    peaksRatio.append(-100.0)
                    areas.append(area)
                else:
                    peaksRatio.append(
                        max(
                            -100,
                            min(
                                100,
                                100 * (peakRatio / getNormRatio(isoEnr, xcount, 1) - 1),
                            ),
                        )
                    )
                    areas.append(area)
            self.ui.pl1.twinxs[0].plot(peaksRatio, areas, "ro")
            self.ui.pl1.axes.set_title("RIA for M+1 to M vs peak intensity")
            self.ui.pl1.axes.set_xlabel("RIA (%)")
            self.ui.pl1.axes.set_ylabel("Peak Intensity")
            self.drawCanvas(self.ui.pl1)

        elif len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - feature pair mPp1 to mP ratio abs":
            peaksRatio = []
            text, ok = QtWidgets.QInputDialog.getText(
                self.parentWidget(),
                "Isotopic enrichment",
                "Please enter the isotopic enrichment to be used for the diagnostics plot (e.g. 0.986)",
                text="0.986",
            )
            isoEnr = 0.986
            if ok:
                text = text.replace(",", ".")
                isoEnr = float(text)
            chrompeaks_df = self.currentOpenResultsFile.db_con.tables["chromPeaks"]
            for row_dict in chrompeaks_df.to_dicts():
                peakRatio = row_dict["peaksRatioMPm1"]
                xcount = row_dict["xcount"]
                if peakRatio == -1:
                    peaksRatio.append(-100.0)
                else:
                    peaksRatio.append(100.0 * max(-100, min(100, peakRatio - getNormRatio(isoEnr, xcount, 1))))
            self.ui.pl1.twinxs[0].hist(
                peaksRatio,
                [i for i in [i / 10.0 for i in range(-1000, 1000, 25)]],
                facecolor="green",
                alpha=0.5,
            )
            self.ui.pl1.axes.set_title("Histogram of isotopolog ratio error - observed minus theoretical ratio for M'-1 to M'\nAssumend enrichment: %.2f%%" % (isoEnr * 100))
            self.ui.pl1.axes.set_xlabel("Ratio error (%)")
            self.ui.pl1.axes.set_ylabel("Frequency")
            self.drawCanvas(self.ui.pl1)

        elif len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - feature pair mPp1 to mP ratio rel":
            peaksRatio = []
            text, ok = QtWidgets.QInputDialog.getText(
                self.parentWidget(),
                "Isotopic enrichment",
                "Please enter the isotopic enrichment to be used for the diagnostics plot (e.g. 0.986)",
                text="0.986",
            )
            isoEnr = 0.986
            if ok:
                text = text.replace(",", ".")
                isoEnr = float(text)
            chrompeaks_df = self.currentOpenResultsFile.db_con.tables["chromPeaks"]
            for row_dict in chrompeaks_df.to_dicts():
                peakRatio = row_dict["peaksRatioMPm1"]
                xcount = row_dict["xcount"]
                if peakRatio == -1:
                    peaksRatio.append(-100.0)
                else:
                    peaksRatio.append(
                        max(
                            -100,
                            min(
                                100,
                                100 * (peakRatio / getNormRatio(isoEnr, xcount, 1) - 1),
                            ),
                        )
                    )
            self.ui.pl1.twinxs[0].hist(
                peaksRatio,
                [i for i in [i / 10.0 for i in range(-1000, 1000, 25)]],
                facecolor="green",
                alpha=0.5,
            )
            self.ui.pl1.axes.set_title("Histogram of RIA for M'-1 to M'\nAssumend enrichment: %.2f%%" % (isoEnr * 100))
            self.ui.pl1.axes.set_xlabel("RIA (%)")
            self.ui.pl1.axes.set_ylabel("Frequency")
            self.drawCanvas(self.ui.pl1)

        elif len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - feature pair mPp1 to mP IRA vs intensity":
            peaksRatio = []
            areas = []
            text, ok = QtWidgets.QInputDialog.getText(
                self.parentWidget(),
                "Isotopic enrichment",
                "Please enter the isotopic enrichment to be used for the diagnostics plot (e.g. 0.986)",
                text="0.986",
            )
            isoEnr = 0.986
            if ok:
                text = text.replace(",", ".")
                isoEnr = float(text)
            chrompeaks_df = self.currentOpenResultsFile.db_con.tables["chromPeaks"]
            for row_dict in chrompeaks_df.to_dicts():
                peakRatio = row_dict["peaksRatioMPm1"]
                xcount = row_dict["xcount"]
                area = row_dict["LPeakArea"]
                area = log10(area)
                if peakRatio == -1:
                    peaksRatio.append(-100.0)
                    areas.append(area)
                else:
                    peaksRatio.append(
                        max(
                            -100,
                            min(
                                100,
                                100 * (peakRatio / getNormRatio(isoEnr, xcount, 1) - 1),
                            ),
                        )
                    )
                    areas.append(area)
            self.ui.pl1.twinxs[0].plot(peaksRatio, areas, "ro")
            self.ui.pl1.axes.set_title("RIA for M'-1 to M' vs peak intensity\nAssumend enrichment: %.2f%%" % (isoEnr * 100))
            self.ui.pl1.axes.set_xlabel("RIA (%)")
            self.ui.pl1.axes.set_ylabel("Peak Intensity")
            self.drawCanvas(self.ui.pl1)

        elif len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - feature pair assigned mzs":
            assignedmzs = []
            chrompeaks_df = self.currentOpenResultsFile.db_con.tables["chromPeaks"]
            for row_dict in chrompeaks_df.to_dicts():
                n = row_dict["assignedMZs"]
                assignedmzs.append(len(loads(base64.b64decode(n))))
            self.ui.pl1.twinxs[0].hist(assignedmzs, 30, facecolor="green", alpha=0.5)
            self.ui.pl1.axes.set_title("Histogram of signal pairs assigned to feature pairs")
            self.ui.pl1.axes.set_xlabel("Number of signal pairs for a feature pair")
            self.ui.pl1.axes.set_ylabel("Frequency")
            self.drawCanvas(self.ui.pl1)

        elif len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - feature pair mz deviation mean":
            devMeans = []
            devVals = []
            chrompeaks_df = self.currentOpenResultsFile.db_con.tables["chromPeaks"]
            for row_dict in chrompeaks_df.to_dicts():
                mzdifferrors = loads(base64.b64decode(row_dict["mzdifferrors"]))
                devMeans.append(mzdifferrors.mean if mzdifferrors.mean is not None else -1)
            self.ui.pl1.twinxs[0].hist(devMeans, 30, facecolor="green", alpha=0.5, label="Means")
            self.ui.pl1.twinxs[0].legend(loc="upper right")
            self.ui.pl1.axes.set_title("Histogram of feature pairs (mean m/z) - mean relative m/z error (ppm)")
            self.ui.pl1.axes.set_xlabel("Mean, relative m/z error (ppm)")
            self.ui.pl1.axes.set_ylabel("Frequency")
            self.drawCanvas(self.ui.pl1)

            removeids = []
            chrompeaks_sorted = self.currentOpenResultsFile.db_con.tables["chromPeaks"].sort("nPeakCenterMin")
            for row_dict in chrompeaks_sorted.to_dicts():
                id = row_dict["id"]
                mzdifferrors = loads(base64.b64decode(row_dict["mzdifferrors"]))
                mz = row_dict["mz"]
                row_dict["nPeakCenterMin"] / 60.0
                if mzdifferrors.mean is None or mzdifferrors.mean < 1:
                    removeids.append(id)

            if False:
                if (
                    QtWidgets.QMessageBox.question(
                        self,
                        "MetExtract",
                        "Are you sure you want to delete the some results?\nThis action cannot be undone",
                        QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    )
                    == QtWidgets.QMessageBox.Yes
                ):
                    print("ids to remove", removeids)
                    self.currentOpenResultsFile.curs.execute("DELETE FROM chromPeaks WHERE id in (%s)" % (",".join(str(s) for s in removeids)))
                    self.currentOpenResultsFile.curs.execute("DELETE FROM featureGroupFeatures WHERE fID in (%s)" % (",".join(str(s) for s in removeids)))
                    self.currentOpenResultsFile.conn.commit()
                    print("ids successfully removed")

        elif len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - feature pair mz deviation":
            devMeans = []
            devVals = []
            chrompeaks_df = self.currentOpenResultsFile.db_con.tables["chromPeaks"]
            for row_dict in chrompeaks_df.to_dicts():
                mzdifferrors = loads(base64.b64decode(row_dict["mzdifferrors"]))
                devVals.extend([v for v in mzdifferrors.vals if v is not None])
            self.ui.pl1.twinxs[0].hist(
                devVals,
                30,
                facecolor="blue",
                alpha=0.5,
                label="Scan-level",
            )
            self.ui.pl1.twinxs[0].legend(loc="upper right")
            self.ui.pl1.axes.set_title("Histogram of feature pairs (all scans) - mean relative m/z error (ppm)")
            self.ui.pl1.axes.set_xlabel("Mean, relative m/z error (ppm)")
            self.ui.pl1.axes.set_ylabel("Frequency")
            self.drawCanvas(self.ui.pl1)

        elif len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - EIC mz deviation":
            devVals = []
            chrompeaks_df = self.currentOpenResultsFile.db_con.tables["chromPeaks"]
            xics_df = self.currentOpenResultsFile.db_con.tables["XICs"]
            joined = chrompeaks_df.join(xics_df, left_on="eicID", right_on="id", how="inner")
            for row_dict in joined.to_dicts():
                id = row_dict["id"]
                mz = row_dict["mz"]
                center = row_dict["NPeakCenter"]
                scale = row_dict["NPeakScale"]
                row_dict["NPeakCenterMin"] / 60.0
                mzs = row_dict["mzs"]
                row_dict["mzsL"]
                mzs = [float(t) for t in mzs.split(";")]
                # mzsL = [float(t) for t in mzsL.split(";")]
                mzsPeak = mzs[max(0, center - int(1.5 * scale)) : min(len(mzs) - 1, center + int(1.5 * scale))]
                # mzsLPeak = mzsL[max(0, center - int(1.5 * scale)):min(len(mzs) - 1, center + int(1.5 * scale))]

                mzsPeak = [t for t in mzsPeak if t >= 0]
                m = mean(mzsPeak)
                mzsPeak = [t - m for t in mzsPeak]
                devVals.append(sd(mzsPeak) * 1000000 / m)

            self.ui.pl1.twinxs[0].hist(
                devVals,
                30,
                facecolor="blue",
                alpha=0.5,
                label="Peak-level",
            )
            self.ui.pl1.twinxs[0].legend(loc="upper right")
            self.ui.pl1.axes.set_title("Histogram of standard deviations of peak mz deviations")
            self.ui.pl1.axes.set_xlabel("SD(mz deviation of peak) (ppm)")
            self.ui.pl1.axes.set_ylabel("Frequency")
            self.drawCanvas(self.ui.pl1)

        elif len(selectedItems) == 1 and selectedItems[0].myType == "diagnostic - peak fwhm":
            fwhmVals = []
            chrompeaks_df = self.currentOpenResultsFile.db_con.tables["chromPeaks"]
            for row_dict in chrompeaks_df.to_dicts():
                new_FWHM_M = row_dict["new_FWHM_M"]
                new_FWHM_Mp = row_dict["new_FWHM_Mp"]
                fwhmVals.append(new_FWHM_M)
                fwhmVals.append(new_FWHM_Mp)
            print(fwhmVals)
            self.ui.pl1.twinxs[0].hist(fwhmVals, 30, facecolor="blue", alpha=0.5, label="FWHM")
            self.ui.pl1.twinxs[0].legend(loc="upper right")
            self.ui.pl1.axes.set_title("Histogram of FWHM of M and Mp peaks")
            self.ui.pl1.axes.set_xlabel("FWHM [sec]")
            self.ui.pl1.axes.set_ylabel("Frequency")
            self.drawCanvas(self.ui.pl1)

        # </editor-fold>

        # <editor-fold desc="#feature pair plotting">
        elif "Features" in plotTypes or "Feature Groups" in plotTypes or "feature" in plotTypes:
            if "FeatureMap" in plotTypes and "feature" not in plotTypes:
                # Top-level feature map: axes already labelled; just draw canvas
                self.drawCanvas(self.ui.pl1)
                # Connect hover/click handlers so the user can inspect scatter points
                _fm_ax = self.ui.pl1.twinxs[0]
                self._setup_feature_map_hover(self.ui.pl1, _fm_ax)
            else:
                if self.ui.scaleFeatures.checkState() == QtCore.Qt.Checked:
                    self.ui.pl1.axes.set_ylabel("Intensity [counts; normalised]")
                else:
                    self.ui.pl1.axes.set_ylabel("Intensity [counts]")
                if self.ui.autoZoomPlot.checkState() == QtCore.Qt.Checked:
                    self.drawCanvas(
                        self.ui.pl1,
                        ylim=(minIntY * 1.6, maxIntY * 1.6),
                        xlim=(minTime * 0.85, maxTime * 1.15),
                    )
                else:
                    self.drawCanvas(self.ui.pl1)
        # </editor-fold>

        # <editor-fold desc="#mz and mzbin plotting">
        elif "MZs" in plotTypes or "mz" in plotTypes or "mzbin" in plotTypes or "MZBins" in plotTypes:
            colour_LUT = [
                "#0000FF",
                "#00FF00",
                "#FF00FF",
                "#00FFFF",
                "#FFFF00",
                "#FFFFFF",
                "#F0F0F0",
                "#0F0F0F",
            ]

            # the scatter plot:
            maxIntX = max([max(x) for x in x_vals])
            maxIntY = max([max(y) for y in y_vals])
            colors = []
            axScatter = self.ui.pl1.axes

            for i in range(self.ui.res_ExtractedData.topLevelItem(2).childCount()):
                child = self.ui.res_ExtractedData.topLevelItem(2).child(i)
                assert child.myType == "feature"

                tppm = float(self.getParametersFromCurrentRes("Mass deviation (+/- ppm)"))
                ppmupper = 1 + tppm / 1000000.0
                ppmlower = 1 - tppm / 1000000.0

                axScatter.plot(
                    [
                        child.myData.NPeakCenterMin / 60.0 - child.myData.NPeakScale / 60.0,
                        child.myData.NPeakCenterMin / 60.0 + child.myData.NPeakScale / 60.0,
                        child.myData.NPeakCenterMin / 60.0 + child.myData.NPeakScale / 60.0,
                        child.myData.NPeakCenterMin / 60.0 - child.myData.NPeakScale / 60.0,
                        child.myData.NPeakCenterMin / 60.0 - child.myData.NPeakScale / 60.0,
                    ],
                    [
                        child.myData.mz * ppmlower,
                        child.myData.mz * ppmlower,
                        child.myData.mz * ppmupper,
                        child.myData.mz * ppmupper,
                        child.myData.mz * ppmlower,
                    ],
                    color="red",
                )

                axScatter.plot(
                    [
                        child.myData.NPeakCenterMin / 60.0 - child.myData.NPeakScale / 60.0,
                        child.myData.NPeakCenterMin / 60.0 + child.myData.NPeakScale / 60.0,
                    ],
                    [child.myData.mz, child.myData.mz],
                    color="red",
                )

            for i in range(len(x_vals)):
                colour = colour_LUT[i % len(colour_LUT)]
                axScatter.scatter(x_vals[i], y_vals[i], color=colour, linewidth=0)
                colors.append(colour)

            self.ui.pl1.twinxs[0].set_ylabel("m/z")
            self.ui.pl1.twinxs[0].set_xlabel("Retention time [minutes]")
            self.drawCanvas(self.ui.pl1, ylim=[0, maxIntY], xlim=[0, maxIntX])
        # </editor-fold>

        else:
            self.drawCanvas(self.ui.pl1)

        self.drawCanvas(
            self.ui.pl3,
            xlim=(
                minMZ - max((maxMZ - minMZ) * 0.1, 4),
                maxMZ + max((maxMZ - minMZ) * 0.1, 4),
            ),
            ylim=(minMZH * 1.35, maxMZH * 1.35),
        )

        # Filter and display MSMS spectra for selected features
        self.updateMSMSList(selectedItems)

    # </editor-fold>

    # <editor-fold desc="### MSMS spectrum functions">
    def hasMSMSSpectra(self, mz, rt_min, rt_max, ppm=5.0):
        """Check if MSMS spectra exist for a given m/z and RT range"""
        if not hasattr(self, "currentOpenRawFile") or self.currentOpenRawFile is None:
            return False

        if not hasattr(self.currentOpenRawFile, "MS2_list") or len(self.currentOpenRawFile.MS2_list) == 0:
            return False

        # Calculate m/z range with tolerance
        mz_min = mz * (1 - ppm / 1000000.0)
        mz_max = mz * (1 + ppm / 1000000.0)

        # Check if any MS2 scans match the criteria
        for ms2_scan in self.currentOpenRawFile.MS2_list:
            if rt_min <= ms2_scan.retention_time <= rt_max:
                if mz_min <= ms2_scan.precursor_mz <= mz_max:
                    return True

        return False

    def updateMSMSList(self, selectedItems):
        """Filter and populate MSMS spectra table based on selected features"""

        class _NSItem(QTableWidgetItem):
            def __lt__(self, other):
                try:
                    return float(self.text()) < float(other.text())
                except (ValueError, TypeError):
                    return self.text() < other.text()

        tbl = self.ui.msms_SpectraList
        tbl.setSortingEnabled(False)
        tbl.setRowCount(0)

        if not hasattr(self, "currentOpenRawFile") or self.currentOpenRawFile is None:
            return

        if not hasattr(self.currentOpenRawFile, "MS2_list") or len(self.currentOpenRawFile.MS2_list) == 0:
            return

        # Collect RT and m/z ranges from selected features
        feature_ranges = []
        for item in selectedItems:
            if hasattr(item, "myType") and (item.myType == "feature" or item.myType == "Features"):
                if item.myType == "feature":
                    cp = item.myData
                    # Prefer stored peak-boundary RTs (in seconds); fall back to XIC times lookup
                    n_start_rt = getattr(cp, "N_startRT", None)
                    n_end_rt = getattr(cp, "N_endRT", None)
                    l_start_rt = getattr(cp, "L_startRT", None)
                    l_end_rt = getattr(cp, "L_endRT", None)

                    if n_start_rt is None or n_end_rt is None or l_start_rt is None or l_end_rt is None:
                        try:
                            xics_rows = self.currentOpenResultsFile.db_con.tables["XICs"].filter(pl.col("id") == cp.eicID).to_dicts()
                            if xics_rows:
                                times_raw = [float(t) for t in xics_rows[0]["times"].split(";")]
                                n_s = max(0, int(cp.NPeakCenter) - int(cp.NBorderLeft))
                                n_e = min(len(times_raw) - 1, int(cp.NPeakCenter) + int(cp.NBorderRight))
                                l_s = max(0, int(cp.LPeakCenter) - int(cp.LBorderLeft))
                                l_e = min(len(times_raw) - 1, int(cp.LPeakCenter) + int(cp.LBorderRight))
                                n_start_rt = times_raw[n_s]
                                n_end_rt = times_raw[n_e]
                                l_start_rt = times_raw[l_s]
                                l_end_rt = times_raw[l_e]
                            else:
                                n_start_rt = cp.NPeakCenterMin - cp.NPeakScale
                                n_end_rt = cp.NPeakCenterMin + cp.NPeakScale
                                l_start_rt = cp.LPeakCenterMin - cp.LPeakScale
                                l_end_rt = cp.LPeakCenterMin + cp.LPeakScale
                        except Exception:
                            n_start_rt = cp.NPeakCenterMin - cp.NPeakScale
                            n_end_rt = cp.NPeakCenterMin + cp.NPeakScale
                            l_start_rt = cp.LPeakCenterMin - cp.LPeakScale
                            l_end_rt = cp.LPeakCenterMin + cp.LPeakScale

                    rt_min = min(float(n_start_rt), float(l_start_rt))
                    rt_max = max(float(n_end_rt), float(l_end_rt))

                    try:
                        ppm = float(self.getParametersFromCurrentRes("Mass deviation (+/- ppm)"))
                    except Exception:
                        ppm = 5.0

                    native_mz_min = cp.mz * (1 - ppm / 1000000.0)
                    native_mz_max = cp.mz * (1 + ppm / 1000000.0)
                    feature_ranges.append({"rt_min": rt_min, "rt_max": rt_max, "mz_min": native_mz_min, "mz_max": native_mz_max, "ionMode": cp.ionMode, "feature_name": "%.4f @ %.2f min" % (cp.mz, cp.NPeakCenterMin / 60.0), "form": "native"})

                    if hasattr(cp, "lmz"):
                        labeled_mz_min = cp.lmz * (1 - ppm / 1000000.0)
                        labeled_mz_max = cp.lmz * (1 + ppm / 1000000.0)
                        feature_ranges.append({"rt_min": rt_min, "rt_max": rt_max, "mz_min": labeled_mz_min, "mz_max": labeled_mz_max, "ionMode": cp.ionMode, "feature_name": "%.4f @ %.2f min" % (cp.lmz, cp.LPeakCenterMin / 60.0), "form": "labeled"})

        if not feature_ranges:
            return

        matching_ms2 = []
        for ms2_scan in self.currentOpenRawFile.MS2_list:
            for fr in feature_ranges:
                if fr["rt_min"] <= ms2_scan.retention_time <= fr["rt_max"]:
                    if ms2_scan.polarity == fr["ionMode"]:
                        if fr["mz_min"] <= ms2_scan.precursor_mz <= fr["mz_max"]:
                            matching_ms2.append({"scan": ms2_scan, "feature_name": fr["feature_name"], "form": fr["form"]})
                            break

        _native_color = QtGui.QColor(30, 144, 255, 60)  # dodgerblue
        _labeled_color = QtGui.QColor(178, 34, 34, 60)  # firebrick

        for ms2_info in sorted(matching_ms2, key=lambda x: x["scan"].precursor_intensity):
            scan = ms2_info["scan"]
            form_label = "M" if ms2_info["form"] == "native" else "M\u2032"
            row_idx = tbl.rowCount()
            tbl.insertRow(row_idx)

            nl_item = _NSItem(form_label)
            nl_item.setData(QtCore.Qt.UserRole, scan)
            nl_item.setData(QtCore.Qt.UserRole + 1, ms2_info["form"])
            tbl.setItem(row_idx, 0, nl_item)
            tbl.setItem(row_idx, 1, _NSItem(f"{scan.precursor_intensity:.4g}"))
            tbl.setItem(row_idx, 2, _NSItem(f"{scan.precursor_mz:.4f}"))
            tbl.setItem(row_idx, 3, _NSItem(f"{scan.retention_time / 60.0:.2f}"))

            row_color = _labeled_color if ms2_info["form"] == "labeled" else _native_color
            for col in range(4):
                tbl.item(row_idx, col).setBackground(row_color)

        tbl.setSortingEnabled(True)
        tbl.resizeColumnsToContents()

        total_rows = tbl.rowCount()
        if total_rows > 0:
            tbl.selectRow(total_rows - 1)

    def _setup_msms_hover(self, plot_obj):
        """Connect hover and click handlers to an MSMS canvas.

        Hover: when the cursor is close to a fragment peak (using normalised 2-D
        distance in m/z and intensity), the vline is redrawn 3× thicker and a
        popup annotation shows its m/z value.

        Click: same proximity test; if close enough the annotation is *pinned*
        and will survive zoom/pan.  Pinned annotations are cleared the next time
        this method is called (i.e. when new spectra are selected).

        State on plot_obj:
            _hover_cid      – mpl connection id for motion_notify_event
            _click_cid      – mpl connection id for button_press_event
            _hover_artists  – temporary artists removed on next move event
            _pinned_artists – persistent artists cleared on next setup call
        """
        canvas = plot_obj.canvas

        # Disconnect any previous handlers
        for attr in ("_hover_cid", "_click_cid"):
            cid = getattr(plot_obj, attr, None)
            if cid is not None:
                try:
                    canvas.mpl_disconnect(cid)
                except Exception:
                    pass
        plot_obj._hover_cid = None
        plot_obj._click_cid = None
        plot_obj._hover_artists = []

        # Clear pinned annotations from previous spectra
        for artist in list(getattr(plot_obj, "_pinned_artists", [])):
            try:
                artist.remove()
            except Exception:
                pass
        plot_obj._pinned_artists = []

        if not plot_obj.axes:
            return

        def _find_closest(ax, mx, my):
            """Return (cidx, norm_dist) for the nearest peak using 2-D normalised distance."""
            peaks = getattr(ax, "_msms_peaks", None)
            if peaks is None or len(peaks[0]) == 0:
                return None, None
            mz_arr, int_arr, _ = peaks
            xlim = ax.get_xlim()
            ylim = ax.get_ylim()
            x_range = xlim[1] - xlim[0] or 1.0
            y_range = ylim[1] - ylim[0] or 1.0
            best_idx, best_dist = 0, float("inf")
            for i, (mz, iv) in enumerate(zip(mz_arr, int_arr)):
                dx = abs(float(mz) - mx) / x_range
                dy = abs(float(iv) - my) / y_range
                d = (dx**2 + dy**2) ** 0.5
                if d < best_dist:
                    best_dist = d
                    best_idx = i
            return best_idx, best_dist

        def _on_hover(event):
            for artist in list(plot_obj._hover_artists):
                try:
                    artist.remove()
                except Exception:
                    pass
            plot_obj._hover_artists.clear()

            if event.inaxes is None or event.xdata is None or event.ydata is None:
                canvas.draw_idle()
                return

            ax = event.inaxes
            peaks = getattr(ax, "_msms_peaks", None)
            if peaks is None or len(peaks[0]) == 0:
                canvas.draw_idle()
                return

            mz_arr, int_arr, spec_color = peaks
            cidx, nd = _find_closest(ax, event.xdata, event.ydata)
            if nd is None or nd > 0.03:
                canvas.draw_idle()
                return

            mz_val = float(mz_arr[cidx])
            int_val = float(int_arr[cidx])

            vl = ax.vlines(mz_val, 0, int_val, colors=spec_color, linewidth=4.5, zorder=5)
            plot_obj._hover_artists.append(vl)

            ann = ax.annotate(
                "m/z  %.4f" % mz_val,
                xy=(mz_val, int_val),
                xytext=(8, 8),
                textcoords="offset points",
                fontsize=11,
                color="#202124",
                bbox=dict(boxstyle="round,pad=0.4", facecolor="lightyellow", edgecolor="#aaaaaa", alpha=0.95),
                zorder=10,
            )
            plot_obj._hover_artists.append(ann)
            canvas.draw_idle()

        def _on_click(event):
            if event.inaxes is None or event.xdata is None or event.ydata is None:
                return

            ax = event.inaxes
            peaks = getattr(ax, "_msms_peaks", None)
            if peaks is None or len(peaks[0]) == 0:
                return

            mz_arr, int_arr, spec_color = peaks
            cidx, nd = _find_closest(ax, event.xdata, event.ydata)
            if nd is None or nd > 0.03:
                return

            mz_val = float(mz_arr[cidx])
            int_val = float(int_arr[cidx])

            # Pin a thick vline and labelled annotation that survive zoom/pan
            vl = ax.vlines(mz_val, 0, int_val, colors=spec_color, linewidth=4.5, zorder=5)
            plot_obj._pinned_artists.append(vl)

            ann = ax.annotate(
                "m/z  %.4f" % mz_val,
                xy=(mz_val, int_val),
                xytext=(8, 8),
                textcoords="offset points",
                fontsize=11,
                color="#202124",
                bbox=dict(boxstyle="round,pad=0.4", facecolor="lightyellow", edgecolor="#888888", alpha=0.98),
                zorder=11,
            )
            plot_obj._pinned_artists.append(ann)
            canvas.draw_idle()

        plot_obj._hover_cid = canvas.mpl_connect("motion_notify_event", _on_hover)
        plot_obj._click_cid = canvas.mpl_connect("button_press_event", _on_click)

    def _setup_feature_map_hover(self, plot_obj, ax):
        """Connect hover and click handlers to a feature-map scatter canvas.

        On hover: find the closest scatter point (normalised 2-D RT/m/z distance)
        and display a popup showing id, metabolite-group, polarity, charge, m/z,
        retention time and native abundance.

        On click: same proximity test; if close enough, navigate the sample-results
        tree to the corresponding tree item.

        Feature data is read from ``ax._fm_point_data``, a list of dicts with keys:
            rt, mz, native_area, id, ogroup, polarity, charge, tree_item
        """
        canvas = plot_obj.canvas

        # Disconnect any previous feature-map handlers
        for attr in ("_fm_hover_cid", "_fm_click_cid"):
            cid = getattr(plot_obj, attr, None)
            if cid is not None:
                try:
                    canvas.mpl_disconnect(cid)
                except Exception:
                    pass
        plot_obj._fm_hover_cid = None
        plot_obj._fm_click_cid = None
        plot_obj._fm_hover_artists = []

        point_data = getattr(ax, "_fm_point_data", [])
        if not point_data:
            return

        def _find_closest_fm(event_ax, ex, ey):
            xlim = event_ax.get_xlim()
            ylim = event_ax.get_ylim()
            x_range = (xlim[1] - xlim[0]) or 1.0
            y_range = (ylim[1] - ylim[0]) or 1.0
            best_idx, best_dist = 0, float("inf")
            for i, pt in enumerate(point_data):
                dx = abs(pt["rt"] - ex) / x_range
                dy = abs(pt["mz"] - ey) / y_range
                d = (dx**2 + dy**2) ** 0.5
                if d < best_dist:
                    best_dist = d
                    best_idx = i
            return best_idx, best_dist

        def _on_fm_hover(event):
            for artist in list(plot_obj._fm_hover_artists):
                try:
                    artist.remove()
                except Exception:
                    pass
            plot_obj._fm_hover_artists.clear()

            if event.inaxes is not ax or event.xdata is None or event.ydata is None:
                canvas.draw_idle()
                return

            cidx, nd = _find_closest_fm(ax, event.xdata, event.ydata)
            if nd > 0.04:
                canvas.draw_idle()
                return

            pt = point_data[cidx]
            label = f"ID: {pt['id']}\nGroup: {pt['ogroup']}\nPolarity: {pt['polarity']}\nCharge: {pt['charge']}\nXn: {pt.get('xcount', 'N/A')}\nm/z: {pt['mz']:.5f}\nRT: {pt['rt']:.2f} min\nNative abundance: {pt['native_area']:.1f}"
            ann = ax.annotate(
                label,
                xy=(pt["rt"], pt["mz"]),
                xytext=(12, 8),
                textcoords="offset points",
                fontsize=18,
                color="#202124",
                bbox=dict(boxstyle="round,pad=0.7", facecolor="lightyellow", edgecolor="#aaaaaa", alpha=0.95),
                zorder=10,
            )
            plot_obj._fm_hover_artists.append(ann)
            canvas.draw_idle()

        def _on_fm_dblclick(event):
            if event.dblclick is False or event.inaxes is not ax or event.xdata is None or event.ydata is None:
                return

            cidx, nd = _find_closest_fm(ax, event.xdata, event.ydata)
            if nd > 0.04:
                return

            tree_item = point_data[cidx].get("tree_item")
            if tree_item is None:
                return
            tree = self.ui.res_ExtractedData
            parent = tree_item.parent()
            if parent is not None:
                tree.expandItem(parent)
            tree.setCurrentItem(tree_item)
            tree.scrollToItem(tree_item)

        plot_obj._fm_hover_cid = canvas.mpl_connect("motion_notify_event", _on_fm_hover)
        plot_obj._fm_click_cid = canvas.mpl_connect("button_press_event", _on_fm_dblclick)

    def plotSelectedMSMSSpectra(self):
        """Plot selected MSMS spectra as subplots with shared x-axis"""
        selected_rows = sorted(set(item.row() for item in self.ui.msms_SpectraList.selectedItems()))

        if not selected_rows:
            # Clear the plot
            self.ui.plMSMS.fig.clear()
            self.ui.plMSMS.axes = []
            self.ui.plMSMS.canvas.draw()
            return

        # Clear previous plots
        self.ui.plMSMS.fig.clear()
        self.ui.plMSMS.axes = []

        # Calculate subplot grid
        n_spectra = len(selected_rows)
        n_cols = 1 if n_spectra == 2 else min(2, n_spectra)
        n_rows = (n_spectra + n_cols - 1) // n_cols

        # Create subplots with shared x-axis
        first_ax = None
        for idx, row_idx in enumerate(selected_rows):
            col0 = self.ui.msms_SpectraList.item(row_idx, 0)
            if col0 is None:
                continue
            scan = col0.data(QtCore.Qt.UserRole)
            form_type = col0.data(QtCore.Qt.UserRole + 1)

            # Determine color based on form type
            if form_type == "labeled":
                spectrum_color = "firebrick"
                label_color = "darkred"
            else:  # native
                spectrum_color = "dodgerblue"
                label_color = "darkblue"

            # Create subplot with shared x-axis
            if idx == 0:
                ax = self.ui.plMSMS.fig.add_subplot(n_rows, n_cols, idx + 1)
                first_ax = ax
            else:
                ax = self.ui.plMSMS.fig.add_subplot(n_rows, n_cols, idx + 1, sharex=first_ax)

            self.ui.plMSMS.axes.append(ax)

            # Plot MS/MS spectrum as stem plot with form-specific color
            if len(scan.mz_list) > 0:
                ax.vlines(scan.mz_list, 0, scan.intensity_list, colors=spectrum_color, linewidth=1.5)
                ax.plot(scan.mz_list, scan.intensity_list, "o", markersize=3, color=spectrum_color)
                # Store per-axis peak data for hover interactivity
                ax._msms_peaks = (scan.mz_list.copy(), scan.intensity_list.copy(), spectrum_color)

                if len(scan.intensity_list) > 0:
                    intensity_with_idx = [(intensity, i) for i, intensity in enumerate(scan.intensity_list)]
                    intensity_with_idx.sort(reverse=True)
                    top_indices = [i for _, i in intensity_with_idx[:10]]
                    for peak_idx in top_indices:
                        mz_val = scan.mz_list[peak_idx]
                        intensity_val = scan.intensity_list[peak_idx]
                        ax.text(mz_val, intensity_val * 1.01, "%.4f" % mz_val, fontsize=9, ha="center", va="bottom", rotation=0, color=label_color, alpha=0.6)

            ax.set_xlabel("m/z", fontsize=12)
            ax.set_ylabel("Intensity", fontsize=12)
            ax.set_title("Scan %d: %.4f m/z | RT %.2f min | I %.3g" % (scan.id, scan.precursor_mz, scan.retention_time / 60.0, scan.precursor_intensity), fontsize=11)
            ax.tick_params(labelsize=12)
            ax.grid(True, alpha=0.3)

        try:
            self.ui.plMSMS.fig.tight_layout()
        except Exception:
            self.ui.plMSMS.fig.subplots_adjust(left=0.08, bottom=0.08, right=0.98, top=0.95, hspace=0.4, wspace=0.3)

        self._setup_msms_hover(self.ui.plMSMS)
        self.ui.plMSMS.canvas.draw()

    def _get_msms_filter_string(self, scan):
        """Return the filter string for an MS/MS scan.

        Prioritizes the original mzML "filter string" (cvParam MS:1000512),
        preserved on the scan as ``filter_string`` during parsing, and falls
        back to the cvParams list or the scan's filter_line attribute."""
        if scan is None:
            return ""
        fs = getattr(scan, "filter_string", None)
        if fs and ("N/A" not in fs and "Unknown" not in fs):
            return fs
        if hasattr(scan, "cvParams") and scan.cvParams:
            for cv in scan.cvParams:
                if cv.get("accession") == "MS:1000512":
                    return cv.get("value", "") or ""
        fl = getattr(scan, "filter_line", None)
        if fl and ("N/A" in fl or "Unknown" in fl or fl.startswith("NA //") or fl.startswith("MSn ")):
            return ""
        return fl or ""

    def _compile_msms_filter_regex(self, pattern):
        """Compile the MS/MS filter-string regex. Returns None for empty/invalid patterns."""
        if not pattern:
            return None
        try:
            return re.compile(pattern)
        except re.error:
            return None

    def _msms_filter_match(self, filter_string, compiled_regex):
        """Test a filter string against a compiled regex.

        Returns (matches, replacement) where replacement is the first capturing
        group (or None if no groups) when the regex matches, else (False, None)."""
        if compiled_regex is None:
            return True, None
        m = compiled_regex.search(filter_string or "")
        if not m:
            return False, None
        replacement = None
        if m.groups():
            replacement = next((g for g in m.groups() if g is not None), None)
        return True, replacement

    def updateMSMSList_exp(self, selectedItems):
        """Filter and populate MSMS spectra table for experimental results panel"""

        class _NSItem(QTableWidgetItem):
            def __lt__(self, other):
                try:
                    return float(self.text()) < float(other.text())
                except (ValueError, TypeError):
                    return self.text() < other.text()

        tbl = self.ui.msms_SpectraList_exp
        tbl.setSortingEnabled(False)
        tbl.setRowCount(0)

        if not hasattr(self, "loadedMZXMLs") or self.loadedMZXMLs is None:
            return

        try:
            msms_rt_window = self.ui.doubleSpinBox_resultsExperiment_MSMSRTWindow.value()
        except Exception:
            msms_rt_window = 0.5

        try:
            prec_intens_percent = self.ui.doubleSpinBox_resultsExperiment_MSMSPrecIntensPercent.value() / 100.0
        except Exception:
            prec_intens_percent = 0.5

        try:
            abs_intens_threshold = self.ui.doubleSpinBox_resultsExperiment_MSMSAbsIntensThreshold.value()
        except Exception:
            abs_intens_threshold = 0.0

        try:
            ppm = self.ui.doubleSpinBox_resultsExperiment_EICppm.value()
        except Exception:
            ppm = 5.0

        try:
            filter_regex_pattern = str(self.ui.lineEdit_msms_filter_regex.text()).strip()
        except Exception:
            filter_regex_pattern = ""
        filter_regex = self._compile_msms_filter_regex(filter_regex_pattern)

        # Build file-path -> group-name mapping (loadedMZXMLs is keyed by both path and group)
        file_to_group = {}
        group_color_map = {}
        for grp in self.getAllSampleGroups():
            for fpath in grp.files:
                file_to_group[fpath] = grp.name
            qc = QtGui.QColor(str(grp.color))
            if qc.isValid():
                qc.setAlpha(90)
                group_color_map[str(grp.name)] = qc

        # Only consider keys that are actual file paths (not group-name aliases)
        file_keys = [k for k in self.loadedMZXMLs if k.lower().endswith(".mzxml") or k.lower().endswith(".mzml")]

        feature_ranges = []
        # Build per-feature, per-sample RT bounds from the results DB (used for precursor intensity check)
        rows_by_num = {}
        if hasattr(self, "experimentResults") and self.experimentResults is not None and self.experimentResults.db_con is not None:
            for sheet in ["4_Convoluted", "3_Reintegrated", "1_Bracketed"]:
                try:
                    tdf = self.experimentResults.db_con.get_table(sheet)
                    if tdf is not None and not tdf.is_empty():
                        rows_by_num = {r["Num"]: r for r in tdf.to_dicts()}
                        break
                except Exception:
                    pass

        for item in selectedItems:
            if hasattr(item, "bunchData"):
                bd = item.bunchData
                if bd.type == "featurePair":
                    native_mz_min = bd.mz * (1 - ppm / 1000000.0)
                    native_mz_max = bd.mz * (1 + ppm / 1000000.0)
                    labeled_mz_min = bd.lmz * (1 - ppm / 1000000.0) if bd.lmz is not None else None
                    labeled_mz_max = bd.lmz * (1 + ppm / 1000000.0) if bd.lmz is not None else None

                    # RT window around the feature's average RT (in seconds)
                    rt_min_s = bd.rt - msms_rt_window * 60.0
                    rt_max_s = bd.rt + msms_rt_window * 60.0

                    # Per-file peak RT bounds for precursor intensity check
                    per_file_peak_rt = {}
                    row_data = rows_by_num.get(getattr(bd, "id", None))
                    if row_data is not None:
                        for file_key in file_keys:
                            fname = os.path.basename(file_key)
                            for ext in [".mzxml", ".mzml"]:
                                if fname.lower().endswith(ext):
                                    fname = fname[: -len(ext)]
                                    break
                            sv = row_data.get(f"{fname}_N_startRT")
                            ev = row_data.get(f"{fname}_N_endRT")
                            lsv = row_data.get(f"{fname}_L_startRT")
                            lev = row_data.get(f"{fname}_L_endRT")
                            try:
                                peak_rt_min = min(
                                    float(str(sv).split(";")[0]) * 60.0 if sv is not None else float("inf"),
                                    float(str(lsv).split(";")[0]) * 60.0 if lsv is not None else float("inf"),
                                )
                                peak_rt_max = max(
                                    float(str(ev).split(";")[0]) * 60.0 if ev is not None else float("-inf"),
                                    float(str(lev).split(";")[0]) * 60.0 if lev is not None else float("-inf"),
                                )
                                if peak_rt_min != float("inf") and peak_rt_max != float("-inf"):
                                    per_file_peak_rt[file_key] = (peak_rt_min, peak_rt_max)
                            except (TypeError, ValueError):
                                pass

                    # Determine scan event for EIC extraction
                    scan_event = getattr(bd, "scanEvent", None)

                    feature_ranges.append(
                        {
                            "native_mz": bd.mz,
                            "native_mz_min": native_mz_min,
                            "native_mz_max": native_mz_max,
                            "labeled_mz_min": labeled_mz_min,
                            "labeled_mz_max": labeled_mz_max,
                            "rt_min_s": rt_min_s,
                            "rt_max_s": rt_max_s,
                            "per_file_peak_rt": per_file_peak_rt,
                            "global_rt": bd.rt,  # seconds
                            "feature_num": getattr(bd, "id", None),
                            "metaboliteGroupID": getattr(bd, "metaboliteGroupID", None),
                            "xn": getattr(bd, "xn", None),
                            "scan_event": scan_event,
                            "ionisation_mode": getattr(bd, "ionisationMode", None),
                        }
                    )

        if not feature_ranges:
            return

        # Cache for EIC max intensities: (file_key, feature_num) -> max_intensity_in_peak
        _eic_max_cache = {}

        def _get_eic_peak_max(file_key, fr):
            """Return max EIC intensity in the peak RT region for a given file/feature, or None if unavailable."""
            cache_key = (file_key, fr.get("feature_num"))
            if cache_key in _eic_max_cache:
                return _eic_max_cache[cache_key]

            peak_rt = fr["per_file_peak_rt"].get(file_key)
            if peak_rt is None:
                # Use global RT +/- MSMS window as fallback bounds
                peak_rt = (fr["global_rt"] - msms_rt_window * 60.0, fr["global_rt"] + msms_rt_window * 60.0)

            mzxml_file = self.loadedMZXMLs.get(file_key)
            if mzxml_file is None:
                _eic_max_cache[cache_key] = None
                return None

            try:
                scan_event = fr.get("scan_event")
                if scan_event is None:
                    ion_mode = fr.get("ionisation_mode")
                    filter_lines = mzxml_file.getFilterLines(includeMS1=True, includeMS2=False, includePosPolarity=True, includeNegPolarity=True)
                    if filter_lines:
                        if ion_mode and "+" in str(ion_mode):
                            scan_event = next((fl for fl in filter_lines if "+" in fl), filter_lines[0])
                        elif ion_mode and "-" in str(ion_mode):
                            scan_event = next((fl for fl in filter_lines if "-" in fl), filter_lines[0])
                        else:
                            scan_event = filter_lines[0]

                if scan_event is None:
                    _eic_max_cache[cache_key] = None
                    return None

                eic, times, _, _ = mzxml_file.getEIC(fr["native_mz"], ppm=ppm, filterLine=scan_event)
                peak_max = 0.0
                for intensity, t in zip(eic, times):
                    if peak_rt[0] <= t <= peak_rt[1]:
                        if intensity > peak_max:
                            peak_max = intensity
                result = peak_max if peak_max > 0.0 else None
            except Exception:
                result = None

            _eic_max_cache[cache_key] = result
            return result

        all_ms2_scans = []
        for file_key in file_keys:
            mzxml_file = self.loadedMZXMLs[file_key]
            if not (hasattr(mzxml_file, "MS2_list") and len(mzxml_file.MS2_list) > 0):
                continue
            for ms2_scan in mzxml_file.MS2_list:
                for fr in feature_ranges:
                    # RT window filter: window around average feature RT
                    if not (fr["rt_min_s"] <= ms2_scan.retention_time <= fr["rt_max_s"]):
                        continue

                    # Check polarity matches the feature's ionisation mode
                    ion_mode = fr.get("ionisation_mode")
                    if ion_mode and ms2_scan.polarity and ms2_scan.polarity != ion_mode:
                        continue

                    form = None
                    # Check precursor m/z against native form
                    if fr["native_mz_min"] <= ms2_scan.precursor_mz <= fr["native_mz_max"]:
                        form = "native"
                    # Check against labeled form
                    elif fr["labeled_mz_min"] is not None and fr["labeled_mz_min"] <= ms2_scan.precursor_mz <= fr["labeled_mz_max"]:
                        form = "labeled"

                    if form is None:
                        continue

                    # Filter-string regex filter (cvParam MS:1000512 / filter_line)
                    filter_string = self._get_msms_filter_string(ms2_scan)
                    fs_match, fs_replacement = self._msms_filter_match(filter_string, filter_regex)
                    if not fs_match:
                        continue

                    # Precursor intensity percent check (applied per file)
                    if prec_intens_percent > 0.0:
                        peak_max = _get_eic_peak_max(file_key, fr)
                        if peak_max is not None and ms2_scan.precursor_intensity < prec_intens_percent * peak_max:
                            continue

                    # Absolute intensity threshold check
                    if abs_intens_threshold > 0.0 and ms2_scan.precursor_intensity < abs_intens_threshold:
                        continue

                    all_ms2_scans.append(
                        {
                            "scan": ms2_scan,
                            "form": form,
                            "file": file_key,
                            "feature_num": fr.get("feature_num"),
                            "o_group": fr.get("metaboliteGroupID"),
                            "xn": fr.get("xn"),
                            "filter_string": filter_string,
                            "filter_replacement": fs_replacement,
                        }
                    )
                    break

        sorted_scans = [(s["scan"], s["form"], s["file"], s.get("feature_num"), s.get("o_group"), s.get("xn"), s.get("filter_string"), s.get("filter_replacement")) for s in all_ms2_scans]
        sorted_scans = natSort(sorted_scans, key=lambda x: x[0].precursor_intensity)

        # Build feature_num -> set of forms for the MS2 filter
        _msms_feature_forms = {}
        for s in all_ms2_scans:
            fn = s.get("feature_num")
            if fn is not None:
                if fn not in _msms_feature_forms:
                    _msms_feature_forms[fn] = set()
                _msms_feature_forms[fn].add(s["form"])
        self._exp_msms_feature_forms = _msms_feature_forms

        _native_color = QtGui.QColor(30, 144, 255, 60)
        _labeled_color = QtGui.QColor(178, 34, 34, 60)

        for scan, form, file_key, feature_num, o_group, xn, filter_string, filter_replacement in sorted_scans:
            form_label = "M\u2032" if form == "labeled" else "M"
            if filter_replacement:
                form_label = f"{form_label} [{filter_replacement}]"
            row_idx = tbl.rowCount()
            tbl.insertRow(row_idx)

            group_name = file_to_group.get(file_key, "")
            filename = os.path.basename(file_key)

            nl_item = _NSItem(form_label)
            nl_item.setData(QtCore.Qt.UserRole, scan)
            nl_item.setData(QtCore.Qt.UserRole + 1, form)
            nl_item.setData(QtCore.Qt.UserRole + 2, feature_num)
            nl_item.setData(QtCore.Qt.UserRole + 3, o_group)
            nl_item.setData(QtCore.Qt.UserRole + 4, file_key)
            nl_item.setData(QtCore.Qt.UserRole + 5, xn)
            tbl.setItem(row_idx, 0, nl_item)
            tbl.setItem(row_idx, 1, _NSItem(f"{scan.precursor_intensity:.4g}"))
            tbl.setItem(row_idx, 2, _NSItem(f"{scan.precursor_mz:.4f}"))
            tbl.setItem(row_idx, 3, _NSItem(f"{scan.retention_time / 60.0:.2f}"))
            tbl.setItem(row_idx, 4, _NSItem(group_name))
            tbl.setItem(row_idx, 5, _NSItem(filename))
            tbl.setItem(row_idx, 6, _NSItem(filter_string or ""))

            row_color = _labeled_color if form == "labeled" else _native_color
            for col in range(7):
                tbl.item(row_idx, col).setBackground(row_color)

            # Color the Group/File columns with the experimental group's color, if known
            grp_color = group_color_map.get(group_name)
            if grp_color is not None:
                tbl.item(row_idx, 4).setBackground(grp_color)
                tbl.item(row_idx, 5).setBackground(grp_color)

        tbl.setSortingEnabled(True)
        tbl.resizeColumnsToContents()

        total_rows = tbl.rowCount()
        if total_rows > 0:
            tbl.selectRow(total_rows - 1)

        self._update_msms_similarity_stats_exp()

    def _update_msms_similarity_stats_exp(self):
        """Compute and display, above the MSMS list, the pairwise similarity percentiles
        (min/P10/P50/P90/max) for spectra of the same "type" (native/labeled x filter-string
        regex group 1) currently listed for the selected feature(s)."""
        label = getattr(self.ui, "label_msms_exp_typestats", None)
        if label is None:
            return
        if not MATCHMS_AVAILABLE:
            label.setText("")
            return

        algorithm_name, rel_intensity_pct, mz_tolerance = self._get_msms_similarity_settings()
        algorithm = self._get_msms_similarity_algorithm(algorithm_name, mz_tolerance)
        try:
            filter_regex_pattern = str(self.ui.lineEdit_msms_filter_regex.text()).strip()
        except Exception:
            filter_regex_pattern = ""
        compiled_regex = self._compile_msms_filter_regex(filter_regex_pattern)

        by_type = defaultdict(list)
        for row in self._iter_exp_msms_rows():
            type_key = self._msms_type_key(row["form"], row["filter_string"], compiled_regex)
            by_type[type_key].append(row["scan"])

        lines = []
        for type_key in sorted(by_type.keys(), key=lambda k: (k[0] or "", k[1] or "")):
            scans = by_type[type_key]
            if len(scans) < 2:
                continue
            scores = self._compute_pairwise_similarity_scores(scans, algorithm, rel_intensity_pct)
            stats = self._format_percentile_stats(scores)
            if stats:
                lines.append(f"<b>{self._format_msms_type_label(type_key)}</b>: {stats}")

        if lines:
            label.setText(f"<b>Similarity ({algorithm_name}) per type:</b><br>" + "<br>".join(lines))
        else:
            label.setText("")

    def _update_msms_selected_similarity_exp(self, selected_rows):
        """Compute and display, above the MSMS list, the similarity of the currently
        user-selected spectra in msms_SpectraList_exp."""
        label = getattr(self.ui, "label_msms_exp_selected_similarity", None)
        if label is None:
            return
        if not MATCHMS_AVAILABLE or len(selected_rows) < 2:
            label.setText("")
            return

        algorithm_name, rel_intensity_pct, mz_tolerance = self._get_msms_similarity_settings()
        algorithm = self._get_msms_similarity_algorithm(algorithm_name, mz_tolerance)

        scans = []
        for row_idx in selected_rows:
            col0 = self.ui.msms_SpectraList_exp.item(row_idx, 0)
            if col0 is None:
                continue
            scan = col0.data(QtCore.Qt.UserRole)
            if scan is not None:
                scans.append(scan)

        if len(scans) < 2:
            label.setText("")
            return

        scores = self._compute_pairwise_similarity_scores(scans, algorithm, rel_intensity_pct)
        if not scores:
            label.setText("")
            return

        if len(scans) == 2:
            label.setText(f"<b>Selected spectra similarity ({algorithm_name}):</b> {scores[0]:.3f}")
        else:
            arr = np.asarray(scores, dtype=float)
            label.setText(f"<b>Selected spectra similarity ({algorithm_name}), n={len(scans)}:</b> min={arr.min():.3f}, mean={arr.mean():.3f}, max={arr.max():.3f} (pairs={len(scores)})")

    def updatePeakDetailsTab(self, plotItems):
        """Populate the peak details tab tables for the selected features."""

        class _SortableItem(QtWidgets.QTableWidgetItem):
            """QTableWidgetItem that sorts numerically when the cell contains a number.
            For scientific-notation and semicolon-separated cells the first numeric
            value is used as the sort key."""

            def __lt__(self, other):
                def _key(text):
                    # Take the first semicolon-separated token
                    token = str(text).split(";")[0].strip()
                    try:
                        return (0, float(token))
                    except (ValueError, TypeError):
                        return (1, str(text).lower())

                return _key(self.text()) < _key(other.text())

        def _make_item(text):
            return _SortableItem(text)

        per_sample_table = self.ui.tableWidget_peakDetails_perSample
        per_group_table = self.ui.tableWidget_peakDetails_perGroup

        per_sample_table.clear()
        per_group_table.clear()

        if not plotItems or not hasattr(self, "experimentResults") or self.experimentResults is None:
            per_sample_table.setRowCount(0)
            per_group_table.setRowCount(0)
            return

        # Get sample file names from defined groups
        definedGroups = self.getAllSampleGroups()

        # Collect file names in order
        file_entries = []  # list of (group_name, file_path, file_name)
        for group in definedGroups:
            for fi in group.files:
                fi = str(fi)
                fname = fi[max(fi.rfind("/") + 1, fi.rfind("\\") + 1) :]
                for ext in [".mzxml", ".mzml", ".mzXML", ".mzML"]:
                    if fname.lower().endswith(ext.lower()):
                        fname = fname[: -len(ext)]
                        break
                file_entries.append((str(group.name), fi, fname))

        # Try to get results dataframe
        # Use cached DataFrame to avoid re-reading the Excel file on every feature selection change
        results_df = getattr(self.experimentResults, "_peak_details_df", None)
        if results_df is None:
            sheet_candidates = ["4_Convoluted", "3_Reintegrated", "1_Bracketed"]
            for sheet_name in sheet_candidates:
                try:
                    tbl = self.experimentResults.db_con.get_table(sheet_name)
                    if tbl is not None and not tbl.is_empty():
                        results_df = tbl
                        self.experimentResults._peak_details_df = results_df
                        break
                except Exception:
                    continue

        if results_df is None:
            per_sample_table.setRowCount(0)
            per_group_table.setRowCount(0)
            return

        # Per-sample columns to show
        per_sample_col_suffixes = [
            "_Found",
            "_Area_N",
            "_Area_L",
            "_SNR_N",
            "_SNR_L",
            "_N_startRT",
            "_N_apexRT",
            "_N_endRT",
            "_L_startRT",
            "_L_apexRT",
            "_L_endRT",
            "_N_FWHM",
            "_L_FWHM",
            "_N_PeakWidth",
            "_L_PeakWidth",
            "_N_ApexToFlankFactor",
            "_L_ApexToFlankFactor",
            "_N_ApexToFlankIncrease",
            "_L_ApexToFlankIncrease",
        ]
        per_sample_col_labels = [
            "Found",
            "Area (N)",
            "Area (L)",
            "SNR (N)",
            "SNR (L)",
            "Start RT (N) [min]",
            "Apex RT (N) [min]",
            "End RT (N) [min]",
            "Start RT (L) [min]",
            "Apex RT (L) [min]",
            "End RT (L) [min]",
            "FWHM (N) [min]",
            "FWHM (L) [min]",
            "PeakWidth (N) [min]",
            "PeakWidth (L) [min]",
            "ApexFlankFactor (N)",
            "ApexFlankFactor (L)",
            "ApexFlankIncrease (N)",
            "ApexFlankIncrease (L)",
        ]

        # Build group-name → semi-transparent QColor map (30% opacity = alpha 77)
        group_color_map = {}
        for grp in definedGroups:
            qc = QtGui.QColor(str(grp.color))
            if qc.isValid():
                qc.setAlpha(77)
                group_color_map[str(grp.name)] = qc

        # Filter the DataFrame ONCE for all selected feature IDs so we never call
        # results_df.filter() inside the per-file / per-column loops.
        selected_ids = [getattr(pi, "id", None) for pi in plotItems]
        selected_ids_known = [i for i in selected_ids if i is not None]
        if selected_ids_known:
            selected_df = results_df.filter(pl.col("Num").is_in(selected_ids_known))
        else:
            # Fall back to MZ-based filter for items without an id
            mz_vals = [pi.mz for pi in plotItems if getattr(pi, "id", None) is None]
            if mz_vals:
                expr = pl.lit(False)
                for mz in mz_vals:
                    expr = expr | ((pl.col("MZ").cast(pl.Float64) - mz).abs() < 0.001)
                selected_df = results_df.filter(expr)
            else:
                selected_df = results_df.clear()
        # Build a dict keyed by Num → row dict for O(1) access later
        selected_rows_by_num = {row["Num"]: row for row in selected_df.to_dicts()}

        # Build per-sample table
        all_sample_headers = ["Sample", "Group"] + per_sample_col_labels
        per_sample_table.setColumnCount(len(all_sample_headers))
        per_sample_table.setHorizontalHeaderLabels(all_sample_headers)

        sample_rows = []
        for group_name, fi, fname in file_entries:
            row_data = {"sample": fname, "group": group_name}
            for suffix, label in zip(per_sample_col_suffixes, per_sample_col_labels):
                col_name = fname + suffix
                if col_name in results_df.columns:
                    values = []
                    for pi in plotItems:
                        num = getattr(pi, "id", None)
                        row = selected_rows_by_num.get(num) if num is not None else None
                        if row is not None:
                            val = row.get(col_name)
                            if val is not None:
                                values.append(str(val))
                    row_data[label] = "; ".join(values) if values else None
                else:
                    row_data[label] = None
            sample_rows.append(row_data)

        # Labels that should be formatted in scientific notation
        sci_notation_labels = {
            "Area (N)",
            "Area (L)",
            "SNR (N)",
            "SNR (L)",
            "ApexFlankFactor (N)",
            "ApexFlankFactor (L)",
            "ApexFlankIncrease (N)",
            "ApexFlankIncrease (L)",
        }

        per_sample_table.setSortingEnabled(False)  # disable while populating
        per_sample_table.setRowCount(len(sample_rows))
        for row_idx, row_data in enumerate(sample_rows):
            row_color = group_color_map.get(row_data["group"])

            def _cell(text, color=row_color):
                item = _make_item(text)
                if color is not None:
                    item.setBackground(color)
                return item

            per_sample_table.setItem(row_idx, 0, _cell(str(row_data["sample"])))
            per_sample_table.setItem(row_idx, 1, _cell(str(row_data["group"])))
            for col_idx, label in enumerate(per_sample_col_labels):
                val = row_data.get(label)
                if val is not None and label in sci_notation_labels:
                    # Format each semicolon-separated value in scientific notation
                    parts = []
                    for part in str(val).split("; "):
                        try:
                            parts.append(f"{float(part):.3e}")
                        except ValueError:
                            parts.append(part)
                    display = "; ".join(parts)
                else:
                    display = str(val) if val is not None else ""
                per_sample_table.setItem(row_idx, col_idx + 2, _cell(display))
        per_sample_table.setSortingEnabled(True)
        per_sample_table.horizontalHeader().setSectionsMovable(True)
        per_sample_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        per_sample_table.resizeColumnsToContents()

        # Build per-group aggregation table
        import statistics as stats_mod

        numeric_suffixes = [
            "_Area_N",
            "_Area_L",
            "_SNR_N",
            "_SNR_L",
            "_N_startRT",
            "_N_apexRT",
            "_N_endRT",
            "_L_startRT",
            "_L_apexRT",
            "_L_endRT",
            "_N_FWHM",
            "_L_FWHM",
            "_N_PeakWidth",
            "_L_PeakWidth",
            "_N_ApexToFlankFactor",
            "_L_ApexToFlankFactor",
            "_N_ApexToFlankIncrease",
            "_L_ApexToFlankIncrease",
        ]
        numeric_labels = [
            "Area (N)",
            "Area (L)",
            "SNR (N)",
            "SNR (L)",
            "Start RT (N) [min]",
            "Apex RT (N) [min]",
            "End RT (N) [min]",
            "Start RT (L) [min]",
            "Apex RT (L) [min]",
            "End RT (L) [min]",
            "FWHM (N) [min]",
            "FWHM (L) [min]",
            "PeakWidth (N) [min]",
            "PeakWidth (L) [min]",
            "ApexFlankFactor (N)",
            "ApexFlankFactor (L)",
            "ApexFlankIncrease (N)",
            "ApexFlankIncrease (L)",
        ]

        # Collect values per group per metric
        group_metric_values = {}
        group_detection_counts = {}
        for group_name, fi, fname in file_entries:
            if group_name not in group_metric_values:
                group_metric_values[group_name] = {lbl: [] for lbl in numeric_labels}
                group_detection_counts[group_name] = 0
            # Count detections
            found_col = fname + "_Found"
            if found_col in results_df.columns:
                for pi in plotItems:
                    num = getattr(pi, "id", None)
                    row = selected_rows_by_num.get(num) if num is not None else None
                    if row is not None:
                        val = row.get(found_col)
                        if val is not None:
                            group_detection_counts[group_name] += 1
            for suffix, label in zip(numeric_suffixes, numeric_labels):
                col_name = fname + suffix
                if col_name in results_df.columns:
                    for pi in plotItems:
                        num = getattr(pi, "id", None)
                        row = selected_rows_by_num.get(num) if num is not None else None
                        if row is not None:
                            val = row.get(col_name)
                            if val is not None:
                                try:
                                    for v in str(val).split(";"):
                                        group_metric_values[group_name][label].append(float(v.strip()))
                                except Exception:
                                    pass

        group_names = list(group_metric_values.keys())

        def _rsd(vals):
            if len(vals) < 2:
                return None
            m = stats_mod.mean(vals)
            if m == 0:
                return None
            return stats_mod.stdev(vals) / abs(m) * 100.0

        # Labels shown in scientific notation in the aggregated table
        sci_agg_labels = {
            "Area (N)",
            "Area (L)",
            "SNR (N)",
            "SNR (L)",
            "ApexFlankFactor (N)",
            "ApexFlankFactor (L)",
            "ApexFlankIncrease (N)",
            "ApexFlankIncrease (L)",
        }

        def _fmt(label, val):
            if val is None:
                return ""
            if label in sci_agg_labels:
                return f"{val:.3e}"
            return f"{val:.4g}"

        # Build ordered column layout:
        # Group | N detections | Mean Area N | RSD Area N | Mean Area L | RSD Area L | then mean/min/max for rest
        remaining_labels = [lbl for lbl in numeric_labels if lbl not in ("Area (N)", "Area (L)")]
        agg_headers = ["Group", "N detections", "Area (N) mean", "Area (N) RSD%", "Area (L) mean", "Area (L) RSD%"] + [f"{lbl} mean" for lbl in remaining_labels] + [f"{lbl} min" for lbl in remaining_labels] + [f"{lbl} max" for lbl in remaining_labels]
        per_group_table.setColumnCount(len(agg_headers))
        per_group_table.setHorizontalHeaderLabels(agg_headers)
        per_group_table.setSortingEnabled(False)  # disable while populating
        per_group_table.setRowCount(len(group_names))

        for row_idx, group_name in enumerate(group_names):
            grp_color = group_color_map.get(group_name)

            def _grp_cell(text, color=grp_color):
                item = _make_item(text)
                if color is not None:
                    item.setBackground(color)
                return item

            per_group_table.setItem(row_idx, 0, _grp_cell(group_name))
            per_group_table.setItem(row_idx, 1, _grp_cell(str(group_detection_counts[group_name])))
            col_offset = 2
            # Area N mean + RSD
            area_n_vals = group_metric_values[group_name]["Area (N)"]
            area_n_mean = stats_mod.mean(area_n_vals) if area_n_vals else None
            area_n_rsd = _rsd(area_n_vals)
            per_group_table.setItem(row_idx, col_offset, _grp_cell(_fmt("Area (N)", area_n_mean)))
            per_group_table.setItem(row_idx, col_offset + 1, _grp_cell(f"{area_n_rsd:.2f}" if area_n_rsd is not None else ""))
            col_offset += 2
            # Area L mean + RSD
            area_l_vals = group_metric_values[group_name]["Area (L)"]
            area_l_mean = stats_mod.mean(area_l_vals) if area_l_vals else None
            area_l_rsd = _rsd(area_l_vals)
            per_group_table.setItem(row_idx, col_offset, _grp_cell(_fmt("Area (L)", area_l_mean)))
            per_group_table.setItem(row_idx, col_offset + 1, _grp_cell(f"{area_l_rsd:.2f}" if area_l_rsd is not None else ""))
            col_offset += 2
            # Mean for remaining labels
            for label in remaining_labels:
                vals = group_metric_values[group_name][label]
                mean_val = stats_mod.mean(vals) if vals else None
                per_group_table.setItem(row_idx, col_offset, _grp_cell(_fmt(label, mean_val)))
                col_offset += 1
            # Min for remaining labels
            for label in remaining_labels:
                vals = group_metric_values[group_name][label]
                min_val = min(vals) if vals else None
                per_group_table.setItem(row_idx, col_offset, _grp_cell(_fmt(label, min_val)))
                col_offset += 1
            # Max for remaining labels
            for label in remaining_labels:
                vals = group_metric_values[group_name][label]
                max_val = max(vals) if vals else None
                per_group_table.setItem(row_idx, col_offset, _grp_cell(_fmt(label, max_val)))
                col_offset += 1
        per_group_table.setSortingEnabled(True)
        per_group_table.horizontalHeader().setSectionsMovable(True)
        per_group_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        per_group_table.resizeColumnsToContents()

    def plotSelectedMSMSSpectra_exp(self):
        """Plot selected MSMS spectra from experimental results panel"""
        selected_rows = sorted(set(item.row() for item in self.ui.msms_SpectraList_exp.selectedItems()))
        self._update_msms_selected_similarity_exp(selected_rows)

        if not selected_rows:
            self.ui.plMSMS_exp.fig.clear()
            self.ui.plMSMS_exp.axes = []
            self.ui.plMSMS_exp.canvas.draw()
            return

        self.ui.plMSMS_exp.fig.clear()
        self.ui.plMSMS_exp.axes = []

        n_spectra = len(selected_rows)
        n_cols = 1 if n_spectra == 2 else min(2, n_spectra)
        n_rows = (n_spectra + n_cols - 1) // n_cols

        first_ax = None
        for idx, row_idx in enumerate(selected_rows):
            col0 = self.ui.msms_SpectraList_exp.item(row_idx, 0)
            if col0 is None:
                continue
            scan = col0.data(QtCore.Qt.UserRole)
            form_type = col0.data(QtCore.Qt.UserRole + 1)

            if form_type == "labeled":
                spectrum_color = "firebrick"
                label_color = "darkred"
            else:
                spectrum_color = "dodgerblue"
                label_color = "darkblue"

            if idx == 0:
                ax = self.ui.plMSMS_exp.fig.add_subplot(n_rows, n_cols, idx + 1)
                first_ax = ax
            else:
                ax = self.ui.plMSMS_exp.fig.add_subplot(n_rows, n_cols, idx + 1, sharex=first_ax)

            self.ui.plMSMS_exp.axes.append(ax)

            if len(scan.mz_list) > 0:
                ax.vlines(scan.mz_list, 0, scan.intensity_list, colors=spectrum_color, linewidth=1.5)
                ax.plot(scan.mz_list, scan.intensity_list, "o", markersize=3, color=spectrum_color)
                # Store per-axis peak data for hover interactivity
                ax._msms_peaks = (scan.mz_list.copy(), scan.intensity_list.copy(), spectrum_color)

                if len(scan.intensity_list) > 0:
                    intensity_with_idx = [(intensity, i) for i, intensity in enumerate(scan.intensity_list)]
                    intensity_with_idx.sort(reverse=True)
                    top_indices = [i for _, i in intensity_with_idx[:10]]
                    for peak_idx in top_indices:
                        mz_val = scan.mz_list[peak_idx]
                        intensity_val = scan.intensity_list[peak_idx]
                        ax.text(mz_val, intensity_val * 1.01, "%.4f" % mz_val, fontsize=9, ha="center", va="bottom", rotation=0, color=label_color, alpha=0.6)

            ax.set_xlabel("m/z", fontsize=12)
            ax.set_ylabel("Intensity", fontsize=12)
            ax.set_title("Scan %d: %.4f m/z | RT %.2f min | I %.3g" % (scan.id, scan.precursor_mz, scan.retention_time / 60.0, scan.precursor_intensity), fontsize=11)
            ax.tick_params(labelsize=12)
            ax.grid(True, alpha=0.3)

        try:
            self.ui.plMSMS_exp.fig.tight_layout()
        except Exception:
            self.ui.plMSMS_exp.fig.subplots_adjust(left=0.08, bottom=0.08, right=0.98, top=0.95, hspace=0.4, wspace=0.3)

        self._setup_msms_hover(self.ui.plMSMS_exp)
        self.ui.plMSMS_exp.canvas.draw()

    def _iter_exp_msms_rows(self):
        tbl = self.ui.msms_SpectraList_exp
        for row_idx in range(tbl.rowCount()):
            col0 = tbl.item(row_idx, 0)
            if col0 is None:
                continue
            scan = col0.data(QtCore.Qt.UserRole)
            form = col0.data(QtCore.Qt.UserRole + 1)
            feature_num = col0.data(QtCore.Qt.UserRole + 2)
            o_group = col0.data(QtCore.Qt.UserRole + 3)
            file_key = col0.data(QtCore.Qt.UserRole + 4)
            xn = col0.data(QtCore.Qt.UserRole + 5)
            if scan is None:
                continue
            yield {
                "row": row_idx,
                "scan": scan,
                "form": form,
                "feature_num": feature_num,
                "o_group": o_group,
                "file_key": file_key,
                "xn": xn,
                "filter_string": self._get_msms_filter_string(scan),
            }

    def _to_matchms_spectrum(self, scan, rel_intensity_pct=None):
        """Convert an MSScan/MS2Scan to a normalized matchms Spectrum.

        rel_intensity_pct, if given (0-100), removes fragments whose intensity
        is below that percentage of the spectrum's most abundant fragment."""
        if not MATCHMS_AVAILABLE:
            return None
        if scan is None or len(scan.mz_list) == 0:
            return None
        # matchms requires mz values to be sorted in ascending order
        mz = np.asarray(scan.mz_list, dtype=float)
        intens = np.asarray(scan.intensity_list, dtype=float)
        order = np.argsort(mz)
        spec = MatchmsSpectrum(
            mz=mz[order],
            intensities=intens[order],
            metadata={"precursor_mz": float(scan.precursor_mz), "retention_time": float(scan.retention_time)},
        )
        spec = matchms_normalize_intensities(spec)
        if rel_intensity_pct is not None and rel_intensity_pct > 0.0 and spec is not None:
            spec = matchms_select_by_relative_intensity(spec, intensity_from=float(rel_intensity_pct) / 100.0, intensity_to=1.0)
        return spec

    def _get_msms_similarity_settings(self):
        """Read the similarity-scoring options (algorithm, relative intensity threshold
        in %, m/z tolerance in Da) from the "Show options" panel. Falls back to
        defaults (ModifiedCosineHungarian, 1%, 0.01 Da) if the widgets are unavailable."""
        try:
            algorithm_name = str(self.ui.comboBox_msms_similarity_algorithm.currentText())
        except Exception:
            algorithm_name = "ModifiedCosineHungarian"
        try:
            rel_intensity_pct = float(self.ui.doubleSpinBox_msms_similarity_relIntensity.value())
        except Exception:
            rel_intensity_pct = 1.0
        try:
            mz_tolerance = float(self.ui.doubleSpinBox_msms_similarity_mzTolerance.value())
        except Exception:
            mz_tolerance = 0.01
        return algorithm_name, rel_intensity_pct, mz_tolerance

    def _get_msms_similarity_algorithm(self, algorithm_name, mz_tolerance):
        """Instantiate a matchms similarity algorithm by name with the given m/z tolerance."""
        if not MATCHMS_AVAILABLE:
            return None
        cls = MSMS_SIMILARITY_ALGORITHMS.get(algorithm_name, MSMS_SIMILARITY_ALGORITHMS.get("ModifiedCosineHungarian"))
        if cls is None:
            return None
        try:
            return cls(tolerance=mz_tolerance)
        except TypeError:
            return cls()

    def _msms_pair_score(self, algorithm, spec_a, spec_b):
        """Return a similarity score (float) between two matchms spectra, or None on failure."""
        if algorithm is None or spec_a is None or spec_b is None:
            return None
        try:
            result = algorithm.pair(spec_a, spec_b)
        except Exception:
            return None
        try:
            return float(result["score"])
        except (TypeError, IndexError, ValueError, KeyError):
            try:
                return float(result)
            except Exception:
                return None

    @staticmethod
    def _restrict_to_common_peaks(spec_a, spec_b, mz_tolerance):
        """Return copies of spec_a/spec_b containing only the fragments that have a
        matching counterpart (within mz_tolerance) in the other spectrum. Used to score
        spectra while ignoring any fragments that are missing in the other spectrum."""
        mz_a = spec_a.peaks.mz
        int_a = spec_a.peaks.intensities
        mz_b = spec_b.peaks.mz
        int_b = spec_b.peaks.intensities
        keep_a = np.zeros(len(mz_a), dtype=bool)
        keep_b = np.zeros(len(mz_b), dtype=bool)
        for i, mz in enumerate(mz_a):
            diffs = np.abs(mz_b - mz)
            matches = np.where(diffs <= mz_tolerance)[0]
            if matches.size > 0:
                keep_a[i] = True
                keep_b[matches] = True
        if not keep_a.any() or not keep_b.any():
            return None, None
        restricted_a = MatchmsSpectrum(mz=mz_a[keep_a], intensities=int_a[keep_a], metadata=spec_a.metadata)
        restricted_b = MatchmsSpectrum(mz=mz_b[keep_b], intensities=int_b[keep_b], metadata=spec_b.metadata)
        return restricted_a, restricted_b

    def _msms_pair_score_and_matches(self, algorithm, spec_a, spec_b, ignore_unmatched=False, mz_tolerance=0.01):
        """Return (score, n_matches, n_fragments_a, n_fragments_b) for a pair of matchms
        spectra, or (None, None, n_fragments_a, n_fragments_b) on failure. If
        ignore_unmatched is True, the score is computed only from fragments that have a
        matching counterpart in the other spectrum (i.e. missing fragments are ignored
        instead of penalizing the score)."""
        n_fragments_a = len(spec_a.peaks.mz) if spec_a is not None else 0
        n_fragments_b = len(spec_b.peaks.mz) if spec_b is not None else 0
        if algorithm is None or spec_a is None or spec_b is None:
            return None, None, n_fragments_a, n_fragments_b

        score_spec_a, score_spec_b = spec_a, spec_b
        if ignore_unmatched:
            score_spec_a, score_spec_b = self._restrict_to_common_peaks(spec_a, spec_b, mz_tolerance)
            if score_spec_a is None or score_spec_b is None:
                return 0.0, 0, n_fragments_a, n_fragments_b

        try:
            result = algorithm.pair(score_spec_a, score_spec_b)
        except Exception:
            return None, None, n_fragments_a, n_fragments_b

        try:
            score = float(result["score"])
        except (TypeError, IndexError, ValueError, KeyError):
            try:
                score = float(result)
            except Exception:
                return None, None, n_fragments_a, n_fragments_b

        n_matches = None
        try:
            n_matches = int(result["matches"])
        except (TypeError, IndexError, ValueError, KeyError):
            n_matches = None

        return score, n_matches, n_fragments_a, n_fragments_b

    def _msms_type_key(self, form, filter_string, compiled_regex):
        """Return the (form, regex-group-1) 'type' key used to group spectra for
        similarity statistics, e.g. ("native", "hcd25.0") or ("labeled", None)."""
        _, replacement = self._msms_filter_match(filter_string, compiled_regex)
        return (form, replacement)

    def _format_msms_type_label(self, type_key):
        form, replacement = type_key
        label = "M\u2032" if form == "labeled" else "M"
        if replacement:
            label = f"{label} [{replacement}]"
        return label

    def _compute_pairwise_similarity_scores(self, scans, algorithm, rel_intensity_pct):
        """Compute all pairwise similarity scores among a list of MSScan/MS2Scan objects.

        Returns a list of floats (one per unique pair, i<j). Scans that cannot be
        converted to a valid matchms spectrum are skipped."""
        if algorithm is None or len(scans) < 2:
            return []
        spectra = [self._to_matchms_spectrum(s, rel_intensity_pct=rel_intensity_pct) for s in scans]
        scores = []
        for i in range(len(spectra)):
            if spectra[i] is None:
                continue
            for j in range(i + 1, len(spectra)):
                if spectra[j] is None:
                    continue
                score = self._msms_pair_score(algorithm, spectra[i], spectra[j])
                if score is not None:
                    scores.append(score)
        return scores

    @staticmethod
    def _format_percentile_stats(scores):
        """Return an HTML-ish summary string with min/p10/p50/p90/max for a list of scores."""
        if not scores:
            return None
        arr = np.asarray(scores, dtype=float)
        p10, p50, p90 = np.percentile(arr, [10, 50, 90])
        return f"min={arr.min():.3f}, P10={p10:.3f}, P50={p50:.3f}, P90={p90:.3f}, max={arr.max():.3f} (n={len(arr)})"

    def _show_msms_similarity_dialog(self, form_filter):
        if not MATCHMS_AVAILABLE:
            QtWidgets.QMessageBox.warning(self, "MS/MS similarity", "matchms is not available in this environment.")
            return

        rows = [r for r in self._iter_exp_msms_rows() if r.get("form") == form_filter and r.get("feature_num") is not None]
        by_feature = defaultdict(list)
        for row in rows:
            by_feature[row["feature_num"]].append(row)
        feature_ids = sorted(by_feature.keys())
        if len(feature_ids) < 2:
            QtWidgets.QMessageBox.information(self, "MS/MS similarity", "At least two features with MS/MS spectra are required.")
            return

        repr_scans = {}
        for fid in feature_ids:
            repr_scans[fid] = max(by_feature[fid], key=lambda x: float(getattr(x["scan"], "precursor_intensity", 0.0)))["scan"]

        algorithm_name, rel_intensity_pct, mz_tolerance = self._get_msms_similarity_settings()

        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle(f"MS/MS similarity ({'native' if form_filter == 'native' else 'labeled'}) - {algorithm_name}")
        dlg.resize(980, 760)
        layout = QtWidgets.QVBoxLayout(dlg)

        ctrl = QtWidgets.QHBoxLayout()
        ctrl.addWidget(QtWidgets.QLabel("Similarity threshold:"))
        threshold_spin = QtWidgets.QDoubleSpinBox()
        threshold_spin.setRange(0.0, 1.0)
        threshold_spin.setDecimals(3)
        threshold_spin.setSingleStep(0.05)
        threshold_spin.setValue(0.7)
        ctrl.addWidget(threshold_spin)
        ctrl.addStretch(1)
        layout.addLayout(ctrl)

        matrix = QtWidgets.QTableWidget()
        matrix.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        matrix.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectItems)
        matrix.setRowCount(len(feature_ids))
        matrix.setColumnCount(len(feature_ids))
        labels = [f"Num {fid}" for fid in feature_ids]
        matrix.setVerticalHeaderLabels(labels)
        matrix.setHorizontalHeaderLabels(labels)
        layout.addWidget(matrix, 1)

        mirror = QtCore.QObject()
        mirror.fig = Figure((5.0, 3.0), dpi=80, facecolor="white")
        mirror.canvas = FigureCanvas(mirror.fig)
        mirror.axes = mirror.fig.add_subplot(111)
        layout.addWidget(mirror.canvas, 1)

        cosine = self._get_msms_similarity_algorithm(algorithm_name, mz_tolerance)
        pair_cache = {}

        def _paint():
            threshold = threshold_spin.value()
            for i, fid_a in enumerate(feature_ids):
                for j, fid_b in enumerate(feature_ids):
                    if j < i:
                        continue
                    if fid_a == fid_b:
                        score = 1.0
                        pair_cache[(i, j)] = (repr_scans[fid_a], repr_scans[fid_b], score)
                    else:
                        sp_a = self._to_matchms_spectrum(repr_scans[fid_a], rel_intensity_pct=rel_intensity_pct)
                        sp_b = self._to_matchms_spectrum(repr_scans[fid_b], rel_intensity_pct=rel_intensity_pct)
                        if sp_a is None or sp_b is None:
                            score = 0.0
                        else:
                            try:
                                score = float(cosine.pair(sp_a, sp_b).get("score", 0.0))
                            except Exception:
                                score = 0.0
                        pair_cache[(i, j)] = (repr_scans[fid_a], repr_scans[fid_b], score)
                    pair_cache[(j, i)] = pair_cache[(i, j)]

                    item = QtWidgets.QTableWidgetItem(f"{score:.3f}")
                    item.setTextAlignment(QtCore.Qt.AlignCenter)
                    red = int(255 * (1.0 - score))
                    green = int(255 * score)
                    item.setBackground(QtGui.QColor(red, green, 80))
                    if score >= threshold:
                        item.setForeground(QtGui.QBrush(QtGui.QColor("black")))
                    else:
                        item.setForeground(QtGui.QBrush(QtGui.QColor("white")))
                    matrix.setItem(i, j, item)
                    if i != j:
                        sym_item = QtWidgets.QTableWidgetItem(f"{score:.3f}")
                        sym_item.setTextAlignment(QtCore.Qt.AlignCenter)
                        sym_item.setBackground(QtGui.QColor(red, green, 80))
                        if score >= threshold:
                            sym_item.setForeground(QtGui.QBrush(QtGui.QColor("black")))
                        else:
                            sym_item.setForeground(QtGui.QBrush(QtGui.QColor("white")))
                        matrix.setItem(j, i, sym_item)
            matrix.resizeColumnsToContents()

        def _show_selected_pair():
            idxs = matrix.selectedIndexes()
            if len(idxs) == 0:
                return
            i, j = idxs[0].row(), idxs[0].column()
            if (i, j) not in pair_cache:
                return
            scan_a, scan_b, score = pair_cache[(i, j)]
            mirror.fig.clear()
            ax = mirror.fig.add_subplot(111)
            ax.vlines(scan_a.mz_list, 0, scan_a.intensity_list, colors="dodgerblue", linewidth=1.2)
            ax.vlines(scan_b.mz_list, 0, -np.asarray(scan_b.intensity_list), colors="firebrick", linewidth=1.2)
            ax.axhline(0, color="black", linewidth=0.8)
            ax.set_xlabel("m/z")
            ax.set_ylabel("Intensity (mirror)")
            ax.set_title(f"Num {feature_ids[i]} vs Num {feature_ids[j]} | similarity={score:.3f}")
            ax.grid(True, alpha=0.2)
            mirror.fig.tight_layout()
            mirror.canvas.draw()

        threshold_spin.valueChanged.connect(_paint)
        matrix.itemSelectionChanged.connect(_show_selected_pair)
        _paint()
        if matrix.rowCount() > 0 and matrix.columnCount() > 0:
            matrix.setCurrentCell(0, 0)
            _show_selected_pair()

        dlg.exec()

    def _select_msms_fragments(self, scan, selection):
        """Return (mz_list, intensity_list) filtered according to a fragment selection.

        selection is one of:
        - ("all", None)
        - ("top", n): keep the n most intense fragments
        - ("percent", p): keep fragments with intensity >= p% of the base peak
        """
        mzs = list(scan.mz_list)
        its = list(scan.intensity_list)
        if not mzs:
            return [], []
        mode, value = selection
        if mode == "top" and value is not None:
            order = sorted(range(len(its)), key=lambda i: its[i], reverse=True)[: int(value)]
            keep = sorted(order)
            return [mzs[i] for i in keep], [its[i] for i in keep]
        if mode == "percent" and value is not None:
            base = max(its) if its else 0.0
            if base <= 0:
                return mzs, its
            threshold = base * float(value) / 100.0
            keep = [i for i in range(len(its)) if its[i] >= threshold]
            return [mzs[i] for i in keep], [its[i] for i in keep]
        return mzs, its

    def _copy_msms_spectrum(self, scan, fmt, selection=("all", None)):
        if scan is None:
            return
        mzs, its = self._select_msms_fragments(scan, selection)
        if fmt == "tsv":
            lines = ["mz\tintensity"] + [f"{float(mz):.6f}\t{float(it):.6f}" for mz, it in zip(mzs, its)]
        else:  # list / massbank-like (use space delimiter)
            lines = [f"{float(mz):.6f} {float(it):.6f}" for mz, it in zip(mzs, its)]
        QtWidgets.QApplication.clipboard().setText("\n".join(lines))

    def _ask_msms_fragment_selection(self, mode):
        """Prompt for the number of top fragments or the percentage. Returns a selection tuple or None."""
        if mode == "top":
            n, ok = QtWidgets.QInputDialog.getInt(self, "Top-x fragments", "Number of most intense fragments to copy:", 10, 1, 100000, 1)
            if not ok:
                return None
            return ("top", n)
        if mode == "percent":
            p, ok = QtWidgets.QInputDialog.getDouble(self, "Fragments >= x%", "Minimum intensity (% of base peak):", 5.0, 0.0, 100.0, 2)
            if not ok:
                return None
            return ("percent", p)
        return ("all", None)

    def _show_msms_context_menu(self, table_widget, pos):
        item = table_widget.itemAt(pos)
        if item is None:
            return
        row = item.row()
        col0 = table_widget.item(row, 0)
        if col0 is None:
            return
        scan = col0.data(QtCore.Qt.UserRole)
        menu = QtWidgets.QMenu(table_widget)

        fmt_menus = [
            ("Copy spectrum (m/z intensity list)", "list"),
            ("Copy spectrum (TSV)", "tsv"),
            ("Copy spectrum (MassBank style)", "massbank"),
        ]
        action_map = {}
        for label, fmt in fmt_menus:
            submenu = menu.addMenu(label)
            a_all = submenu.addAction("All fragments")
            a_top = submenu.addAction("Top-x fragments")
            a_pct = submenu.addAction("Fragments >=x%")
            action_map[a_all] = (fmt, "all")
            action_map[a_top] = (fmt, "top")
            action_map[a_pct] = (fmt, "percent")

        show_similar_action = None
        if table_widget is getattr(self.ui, "msms_SpectraList_exp", None):
            menu.addSeparator()
            show_similar_action = menu.addAction("Show similar spectra...")

        sel = menu.exec(table_widget.mapToGlobal(pos))
        if sel is None:
            return
        if show_similar_action is not None and sel is show_similar_action:
            self._show_similar_spectra_for_scan(scan)
            return
        if sel not in action_map:
            return
        fmt, mode = action_map[sel]
        selection = self._ask_msms_fragment_selection(mode)
        if selection is None:
            return
        self._copy_msms_spectrum(scan, fmt, selection)

    def _build_feature_match_ranges(self):
        """Build a lightweight list of per-feature RT/mz windows from the FULL experiment
        results table (independent of which features are currently visible/selected in the
        tree). Used to associate an arbitrary MS/MS scan with a detected feature pair."""
        ranges = []
        if not hasattr(self, "experimentResults") or self.experimentResults is None or self.experimentResults.db_con is None:
            return ranges
        try:
            df = self.experimentResults.db_con.tables[self.experimentResults.selected_table]
        except Exception:
            return ranges
        try:
            msms_rt_window = self.ui.doubleSpinBox_resultsExperiment_MSMSRTWindow.value()
        except Exception:
            msms_rt_window = 0.5
        try:
            ppm = self.ui.doubleSpinBox_resultsExperiment_EICppm.value()
        except Exception:
            ppm = 5.0

        def _first_float(val):
            if val is None:
                return None
            try:
                return float(str(val).split(";")[0])
            except (TypeError, ValueError):
                return None

        for r in df.to_dicts():
            num = r.get("Num")
            if num is None:
                continue
            rt_min = _first_float(r.get("RT"))
            if rt_min is None:
                continue
            mz = _first_float(r.get("MZ"))
            lmz = _first_float(r.get("L_MZ"))
            mode = str(r.get("Ionisation_Mode", "+") or "+")
            ranges.append(
                {
                    "num": num,
                    "ogroup": r.get("OGroup"),
                    "rt_min_s": rt_min * 60.0 - msms_rt_window * 60.0,
                    "rt_max_s": rt_min * 60.0 + msms_rt_window * 60.0,
                    "native_mz_min": mz * (1 - ppm / 1e6) if mz is not None else None,
                    "native_mz_max": mz * (1 + ppm / 1e6) if mz is not None else None,
                    "labeled_mz_min": lmz * (1 - ppm / 1e6) if lmz is not None else None,
                    "labeled_mz_max": lmz * (1 + ppm / 1e6) if lmz is not None else None,
                    "mode": mode,
                }
            )
        return ranges

    @staticmethod
    def _match_scan_to_feature(scan, feature_ranges):
        """Return (feature_num, form, ogroup) for the first feature range whose RT window
        and native/labeled m/z window contains the given scan, or (None, None, None) if
        none match."""
        for fr in feature_ranges:
            if not (fr["rt_min_s"] <= scan.retention_time <= fr["rt_max_s"]):
                continue
            mode = fr.get("mode")
            if mode and getattr(scan, "polarity", None) and scan.polarity != mode:
                continue
            if fr["native_mz_min"] is not None and fr["native_mz_min"] <= scan.precursor_mz <= fr["native_mz_max"]:
                return fr["num"], "native", fr.get("ogroup")
            if fr["labeled_mz_min"] is not None and fr["labeled_mz_min"] <= scan.precursor_mz <= fr["labeled_mz_max"]:
                return fr["num"], "labeled", fr.get("ogroup")
        return None, None, None

    def _ask_show_similar_spectra_params(self):
        """Ask the user for the parameters used by the cross-feature 'Show similar
        spectra' search. Returns a dict, or None if the user cancelled."""
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Show similar spectra")
        form = QtWidgets.QFormLayout(dlg)

        info = QtWidgets.QLabel("Searches loaded MS/MS spectra (even from other features) for spectra similar to the selected one. Precursor m/z does not need to match.")
        info.setWordWrap(True)
        form.addRow(info)

        default_algorithm, default_rel_intens, default_mz_tol = self._get_msms_similarity_settings()

        algorithm_combo = QtWidgets.QComboBox()
        algorithm_combo.addItems(list(MSMS_SIMILARITY_ALGORITHMS.keys()))
        idx = algorithm_combo.findText(default_algorithm)
        algorithm_combo.setCurrentIndex(idx if idx >= 0 else 0)
        form.addRow("Similarity algorithm:", algorithm_combo)

        sim_spin = QtWidgets.QDoubleSpinBox()
        sim_spin.setRange(0.0, 1.0)
        sim_spin.setDecimals(3)
        sim_spin.setSingleStep(0.05)
        sim_spin.setValue(0.8)
        form.addRow("Similarity score threshold:", sim_spin)

        min_matches_spin = QtWidgets.QSpinBox()
        min_matches_spin.setRange(0, 1000)
        min_matches_spin.setValue(5)
        form.addRow("Min. number of matched fragments:", min_matches_spin)

        missing_frag_combo = QtWidgets.QComboBox()
        missing_frag_combo.addItems(["Penalize unmatched fragments (standard scoring)", "Ignore unmatched fragments (score matched fragments only)"])
        form.addRow("Unmatched fragments:", missing_frag_combo)

        intens_spin = QtWidgets.QDoubleSpinBox()
        intens_spin.setRange(0.0, 100.0)
        intens_spin.setDecimals(1)
        intens_spin.setValue(default_rel_intens)
        intens_spin.setSuffix("%")
        form.addRow("Min. rel. fragment intensity:", intens_spin)

        mz_spin = QtWidgets.QDoubleSpinBox()
        mz_spin.setRange(0.0001, 1.0)
        mz_spin.setDecimals(4)
        mz_spin.setSingleStep(0.001)
        mz_spin.setValue(default_mz_tol)
        form.addRow("m/z error:", mz_spin)

        scope_all = QtWidgets.QRadioButton("All loaded MS/MS spectra (every loaded file)")
        scope_filtered = QtWidgets.QRadioButton("Only spectra matching the current Show-options filters")
        scope_all.setChecked(True)
        scope_group = QtWidgets.QButtonGroup(dlg)
        scope_group.addButton(scope_all)
        scope_group.addButton(scope_filtered)
        form.addRow(scope_all)
        form.addRow(scope_filtered)

        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        form.addRow(btns)

        if dlg.exec() != QtWidgets.QDialog.Accepted:
            return None

        return {
            "algorithm": algorithm_combo.currentText(),
            "similarity_threshold": sim_spin.value(),
            "min_matches": min_matches_spin.value(),
            "ignore_unmatched_fragments": missing_frag_combo.currentIndex() == 1,
            "rel_intensity_pct": intens_spin.value(),
            "mz_tolerance": mz_spin.value(),
            "scope": "filtered" if scope_filtered.isChecked() else "all",
        }

    def _show_similar_spectra_for_scan(self, reference_scan):
        """Search all loaded MS/MS spectra for spectra similar to ``reference_scan``
        (using ModifiedCosineHungarian, ignoring precursor m/z) and show the matches
        in a non-blocking popup window."""
        if not MATCHMS_AVAILABLE:
            QtWidgets.QMessageBox.warning(self, "Show similar spectra", "matchms is not available in this environment.")
            return
        if reference_scan is None:
            return

        params = self._ask_show_similar_spectra_params()
        if params is None:
            return

        if not hasattr(self, "loadedMZXMLs") or self.loadedMZXMLs is None:
            QtWidgets.QMessageBox.information(self, "Show similar spectra", "No raw MS/MS data loaded.")
            return

        algorithm = self._get_msms_similarity_algorithm(params["algorithm"], params["mz_tolerance"])
        ref_spec = self._to_matchms_spectrum(reference_scan, rel_intensity_pct=params["rel_intensity_pct"])
        if ref_spec is None or len(ref_spec.peaks) == 0:
            QtWidgets.QMessageBox.information(self, "Show similar spectra", "The selected spectrum has no usable fragments.")
            return

        # Gather candidate (scan, file_key) pairs depending on the chosen search scope
        candidates = []
        if params["scope"] == "filtered":
            # Consider spectra of ANY feature pair (not just the currently selected
            # one) that would be visible with the current "Show options" filters
            # (RT window, precursor m/z tolerance, precursor intensity thresholds
            # and the filter-string regex).
            try:
                df = self.experimentResults.db_con.tables[self.experimentResults.selected_table]
                all_rows = df.to_dicts()
            except Exception:
                all_rows = []
            match_result = self._compute_msms_filtered_matches(all_rows) if all_rows else None
            if match_result is not None:
                for m in match_result["matched"]:
                    candidates.append((m["scan"], m["file_key"]))
        else:
            file_keys = [k for k in self.loadedMZXMLs if k.lower().endswith(".mzxml") or k.lower().endswith(".mzml")]
            for file_key in file_keys:
                mzxml_file = self.loadedMZXMLs[file_key]
                ms2_list = getattr(mzxml_file, "MS2_list", None)
                if not ms2_list:
                    continue
                for ms2_scan in ms2_list:
                    candidates.append((ms2_scan, file_key))

        feature_ranges = self._build_feature_match_ranges()

        progress = QtWidgets.QProgressDialog("Searching for similar spectra...", "Cancel", 0, max(len(candidates), 1), self)
        progress.setWindowTitle("Show similar spectra")
        progress.setWindowModality(QtCore.Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)

        matches = []
        seen_scan_ids = set()
        try:
            for idx, (cand_scan, file_key) in enumerate(candidates):
                if idx % 20 == 0 or idx == len(candidates) - 1:
                    progress.setValue(idx)
                    QtWidgets.QApplication.processEvents()
                    if progress.wasCanceled():
                        break
                if cand_scan is reference_scan or id(cand_scan) in seen_scan_ids:
                    continue
                seen_scan_ids.add(id(cand_scan))
                cand_spec = self._to_matchms_spectrum(cand_scan, rel_intensity_pct=params["rel_intensity_pct"])
                if cand_spec is None or len(cand_spec.peaks) == 0:
                    continue
                score, n_matches, n_fragments_ref, n_fragments_query = self._msms_pair_score_and_matches(
                    algorithm,
                    ref_spec,
                    cand_spec,
                    ignore_unmatched=params["ignore_unmatched_fragments"],
                    mz_tolerance=params["mz_tolerance"],
                )
                if score is None or score < params["similarity_threshold"]:
                    continue
                if n_matches is not None and n_matches < params["min_matches"]:
                    continue
                num, form, ogroup = self._match_scan_to_feature(cand_scan, feature_ranges)
                matches.append(
                    {
                        "scan": cand_scan,
                        "file_key": file_key,
                        "score": score,
                        "num": num,
                        "form": form,
                        "ogroup": ogroup,
                        "n_matches": n_matches,
                        "n_fragments_ref": n_fragments_ref,
                        "n_fragments_query": n_fragments_query,
                    }
                )
        finally:
            progress.setValue(max(len(candidates), 1))
            progress.close()

        matches.sort(key=lambda m: m["score"], reverse=True)

        if not matches:
            QtWidgets.QMessageBox.information(self, "Show similar spectra", "No spectra above the similarity threshold were found.")
            return

        self._show_similar_spectra_results_popup(reference_scan, matches, params)

    def _show_similar_spectra_results_popup(self, reference_scan, matches, params):
        """Show a non-blocking (modeless) popup with the matched MS/MS spectra, each
        annotated with its detected feature pair (if any). Double-clicking a match
        navigates to and selects its feature pair in the main experiment-results tree."""
        dlg = QtWidgets.QDialog(self)
        dlg.setAttribute(QtCore.Qt.WA_DeleteOnClose)
        dlg.setWindowModality(QtCore.Qt.NonModal)
        dlg.setWindowTitle(f"Similar spectra ({params['algorithm']} \u2265 {params['similarity_threshold']:.3f}) - {len(matches)} found")
        dlg.resize(1050, 700)
        layout = QtWidgets.QVBoxLayout(dlg)

        info = QtWidgets.QLabel(f"Reference: precursor m/z {reference_scan.precursor_mz:.4f}, polarity {getattr(reference_scan, 'polarity', None) or 'n/a'}, RT {reference_scan.retention_time / 60.0:.2f} min. Double-click a match to select its feature pair in the main tree.")
        info.setWordWrap(True)
        layout.addWidget(info)

        splitter = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        layout.addWidget(splitter, 1)

        list_widget = QtWidgets.QTableWidget()
        list_widget.setColumnCount(11)
        list_widget.setHorizontalHeaderLabels(["Score", "Matched frags.", "Ref. frags.", "Query frags.", "Feature pair", "OGroup", "Polarity", "Prec. m/z", "\u0394 Prec. m/z", "RT (min)", "File"])
        list_widget.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        list_widget.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        list_widget.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        list_widget.setSortingEnabled(True)
        list_widget.setRowCount(len(matches))
        for i, m in enumerate(matches):
            score_item = QTableWidgetItem(f"{m['score']:.3f}")
            score_item.setData(QtCore.Qt.UserRole, m)
            list_widget.setItem(i, 0, score_item)
            n_matches_text = str(m["n_matches"]) if m.get("n_matches") is not None else "n/a"
            list_widget.setItem(i, 1, QTableWidgetItem(n_matches_text))
            list_widget.setItem(i, 2, QTableWidgetItem(str(m.get("n_fragments_ref", ""))))
            list_widget.setItem(i, 3, QTableWidgetItem(str(m.get("n_fragments_query", ""))))
            if m["num"] is not None:
                feat_text = f"Num {m['num']} ({'M\u2032' if m['form'] == 'labeled' else 'M'})"
            else:
                feat_text = "(no detected feature pair)"
            list_widget.setItem(i, 4, QTableWidgetItem(feat_text))
            list_widget.setItem(i, 5, QTableWidgetItem(str(m["ogroup"]) if m.get("ogroup") is not None else ""))
            list_widget.setItem(i, 6, QTableWidgetItem(str(getattr(m["scan"], "polarity", None) or "n/a")))
            list_widget.setItem(i, 7, QTableWidgetItem(f"{m['scan'].precursor_mz:.4f}"))
            list_widget.setItem(i, 8, QTableWidgetItem(f"{m['scan'].precursor_mz - reference_scan.precursor_mz:+.4f}"))
            list_widget.setItem(i, 9, QTableWidgetItem(f"{m['scan'].retention_time / 60.0:.2f}"))
            list_widget.setItem(i, 10, QTableWidgetItem(os.path.basename(m["file_key"])))
        list_widget.resizeColumnsToContents()
        splitter.addWidget(list_widget)

        plot_obj = QtCore.QObject()
        plot_obj.fig = Figure((6.0, 4.5), dpi=80, facecolor="white")
        plot_obj.canvas = FigureCanvas(plot_obj.fig)
        plot_obj.mpl_toolbar = NavigationToolbar(plot_obj.canvas, dlg)
        plot_container = QtWidgets.QWidget()
        plot_container_layout = QtWidgets.QVBoxLayout(plot_container)
        plot_container_layout.setContentsMargins(0, 0, 0, 0)
        plot_container_layout.addWidget(plot_obj.mpl_toolbar)
        plot_container_layout.addWidget(plot_obj.canvas, 1)
        splitter.addWidget(plot_container)
        splitter.setSizes([420, 630])

        def _plot_pair(m):
            plot_obj.fig.clear()
            ax = plot_obj.fig.add_subplot(111)
            ref_intens = np.asarray(reference_scan.intensity_list, dtype=float)
            match_intens = np.asarray(m["scan"].intensity_list, dtype=float)
            ref_max = ref_intens.max() if ref_intens.size else 0.0
            match_max = match_intens.max() if match_intens.size else 0.0
            ref_norm = ref_intens / ref_max * 100.0 if ref_max > 0.0 else ref_intens
            match_norm = match_intens / match_max * 100.0 if match_max > 0.0 else match_intens
            ax.vlines(reference_scan.mz_list, 0, ref_norm, colors="dodgerblue", linewidth=1.2, label="Reference")
            ax.vlines(m["scan"].mz_list, 0, -match_norm, colors="firebrick", linewidth=1.2, label="Match")
            ax.axhline(0, color="black", linewidth=0.8)
            feat_txt = f"Num {m['num']} ({'M\u2032' if m['form'] == 'labeled' else 'M'})" if m["num"] is not None else "no detected feature pair"
            ax.set_title(f"Score={m['score']:.3f} | {feat_txt}", fontsize=10)
            ax.set_xlabel("m/z")
            ax.set_ylabel("Relative intensity (%, each spectrum normalized to its own base peak)")
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.2)
            plot_obj.fig.tight_layout()
            plot_obj.canvas.draw()

        def _on_selection_changed():
            rows = sorted(set(i.row() for i in list_widget.selectedItems()))
            if not rows:
                return
            item = list_widget.item(rows[0], 0)
            if item is None:
                return
            _plot_pair(item.data(QtCore.Qt.UserRole))

        def _on_double_click(item):
            m = list_widget.item(item.row(), 0).data(QtCore.Qt.UserRole)
            if m.get("num") is not None:
                self._showFeatureInExperimentResults(m["num"])

        list_widget.itemSelectionChanged.connect(_on_selection_changed)
        list_widget.itemDoubleClicked.connect(_on_double_click)
        list_widget.selectRow(0)
        _plot_pair(matches[0])

        btn_row = QtWidgets.QHBoxLayout()
        btn_row.addStretch(1)
        close_btn = QtWidgets.QPushButton("Close")
        close_btn.clicked.connect(dlg.close)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

        # Keep a reference alive so the non-modal dialog isn't garbage-collected while open
        if not hasattr(self, "_similar_spectra_dialogs"):
            self._similar_spectra_dialogs = []
        self._similar_spectra_dialogs.append(dlg)

        def _on_destroyed(_obj=None, _dlg=dlg):
            if _dlg in self._similar_spectra_dialogs:
                self._similar_spectra_dialogs.remove(_dlg)

        dlg.destroyed.connect(_on_destroyed)

        dlg.show()

    def _countMSMSPerFeatureForRows(self, rows, min_precursor_intensity):
        """Count filtered MSMS spectra per feature (native and labeled separately).

        Only spectra that would be shown in the experiment MSMS spectra list are
        considered: i.e. those within the configured RT window (deviation from the
        feature apex) whose precursor m/z matches the native/labeled m/z (within the
        EIC ppm tolerance) and whose filter string passes the configured filter-line
        regex. Among those, only spectra with a precursor intensity >=
        ``min_precursor_intensity`` are counted.

        Returns a dict mapping feature Num -> {"native": int, "labeled": int}.
        """
        counts = {}
        if not hasattr(self, "loadedMZXMLs") or self.loadedMZXMLs is None:
            return counts

        try:
            msms_rt_window = self.ui.doubleSpinBox_resultsExperiment_MSMSRTWindow.value()
        except Exception:
            msms_rt_window = 0.5
        try:
            ppm = self.ui.doubleSpinBox_resultsExperiment_EICppm.value()
        except Exception:
            ppm = 5.0
        try:
            filter_regex_pattern = str(self.ui.lineEdit_msms_filter_regex.text()).strip()
        except Exception:
            filter_regex_pattern = ""
        filter_regex = self._compile_msms_filter_regex(filter_regex_pattern)

        file_keys = [k for k in self.loadedMZXMLs if k.lower().endswith(".mzxml") or k.lower().endswith(".mzml")]

        feature_ranges = []
        for r in rows:
            num = r.get("Num")
            if num is None:
                continue
            try:
                rt_s = float(r.get("RT", 0.0)) * 60.0
            except (TypeError, ValueError):
                continue
            try:
                mz = float(r.get("MZ")) if r.get("MZ") is not None else None
            except (TypeError, ValueError):
                mz = None
            try:
                lmz = float(r.get("L_MZ")) if r.get("L_MZ") is not None else None
            except (TypeError, ValueError):
                lmz = None

            counts[num] = {"native": 0, "labeled": 0}
            feature_ranges.append(
                {
                    "num": num,
                    "native_mz_min": mz * (1 - ppm / 1000000.0) if mz is not None else None,
                    "native_mz_max": mz * (1 + ppm / 1000000.0) if mz is not None else None,
                    "labeled_mz_min": lmz * (1 - ppm / 1000000.0) if lmz is not None else None,
                    "labeled_mz_max": lmz * (1 + ppm / 1000000.0) if lmz is not None else None,
                    "rt_min_s": rt_s - msms_rt_window * 60.0,
                    "rt_max_s": rt_s + msms_rt_window * 60.0,
                }
            )

        if not feature_ranges:
            return counts

        for file_key in file_keys:
            mzxml_file = self.loadedMZXMLs[file_key]
            if not (hasattr(mzxml_file, "MS2_list") and len(mzxml_file.MS2_list) > 0):
                continue
            for ms2_scan in mzxml_file.MS2_list:
                # Apply the additional minimum precursor intensity threshold first
                if min_precursor_intensity > 0.0 and ms2_scan.precursor_intensity < min_precursor_intensity:
                    continue

                filter_string = None
                fs_checked = False
                fs_match = True
                for fr in feature_ranges:
                    if not (fr["rt_min_s"] <= ms2_scan.retention_time <= fr["rt_max_s"]):
                        continue

                    form = None
                    if fr["native_mz_min"] is not None and fr["native_mz_min"] <= ms2_scan.precursor_mz <= fr["native_mz_max"]:
                        form = "native"
                    elif fr["labeled_mz_min"] is not None and fr["labeled_mz_min"] <= ms2_scan.precursor_mz <= fr["labeled_mz_max"]:
                        form = "labeled"

                    if form is None:
                        continue

                    # Filter-line regex check (scan-dependent only, evaluated once per scan)
                    if not fs_checked:
                        filter_string = self._get_msms_filter_string(ms2_scan)
                        fs_match, _ = self._msms_filter_match(filter_string, filter_regex)
                        fs_checked = True
                    if not fs_match:
                        break

                    counts[fr["num"]][form] += 1

        return counts

    def _generateInclusionLists(self):
        """Show a dialog to configure and export targeted inclusion lists from the experiment results."""
        if not hasattr(self, "experimentResults") or self.experimentResults is None or self.experimentResults.db_con is None:
            QtWidgets.QMessageBox.information(self, "Generate Inclusion Lists", "No experiment results loaded.")
            return

        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Generate Inclusion Lists")
        layout = QtWidgets.QFormLayout(dlg)

        styleCombo = QtWidgets.QComboBox(dlg)
        styleCombo.addItems(["IQ-X", "QExactive"])
        layout.addRow("Style", styleCombo)

        chkPos = QtWidgets.QCheckBox("Include positive mode")
        chkPos.setChecked(True)
        layout.addRow(chkPos)
        chkNeg = QtWidgets.QCheckBox("Include negative mode")
        chkNeg.setChecked(True)
        layout.addRow(chkNeg)
        chkSepPol = QtWidgets.QCheckBox("Separate polarities")
        layout.addRow(chkSepPol)

        chkNat = QtWidgets.QCheckBox("Include native")
        chkNat.setChecked(True)
        layout.addRow(chkNat)
        chkLab = QtWidgets.QCheckBox("Include labeled")
        chkLab.setChecked(True)
        layout.addRow(chkLab)
        chkSepIso = QtWidgets.QCheckBox("Separate isotopolog form lists")
        layout.addRow(chkSepIso)

        rtOffsetSpin = QtWidgets.QDoubleSpinBox(dlg)
        rtOffsetSpin.setRange(0.0, 1000.0)
        rtOffsetSpin.setDecimals(3)
        rtOffsetSpin.setSingleStep(0.1)
        rtOffsetSpin.setValue(0.5)
        layout.addRow("RT offset (min)", rtOffsetSpin)

        intensSpin = QtWidgets.QDoubleSpinBox(dlg)
        intensSpin.setRange(0.0, 1e15)
        intensSpin.setDecimals(2)
        intensSpin.setValue(0.0)
        layout.addRow("Intensity threshold", intensSpin)

        chkMSMS = QtWidgets.QCheckBox("Only features without MSMS spectra")
        chkMSMS.setChecked(False)
        layout.addRow(chkMSMS)

        msmsMinPrecIntensSpin = QtWidgets.QDoubleSpinBox(dlg)
        msmsMinPrecIntensSpin.setRange(0.0, 1e15)
        msmsMinPrecIntensSpin.setDecimals(2)
        msmsMinPrecIntensSpin.setValue(1e6)
        layout.addRow("MSMS min. precursor intensity", msmsMinPrecIntensSpin)

        msmsMinSpectraSpin = QtWidgets.QSpinBox(dlg)
        msmsMinSpectraSpin.setRange(1, 1000000)
        msmsMinSpectraSpin.setValue(1)
        layout.addRow("MSMS min. number of spectra", msmsMinSpectraSpin)

        # Enable the MSMS thresholds only when the MSMS filter is active
        def _toggleMSMSControls(checked):
            msmsMinPrecIntensSpin.setEnabled(checked)
            msmsMinSpectraSpin.setEnabled(checked)

        chkMSMS.toggled.connect(_toggleMSMSControls)
        _toggleMSMSControls(chkMSMS.isChecked())

        fileRow = QtWidgets.QHBoxLayout()
        fileEdit = QtWidgets.QLineEdit(dlg)
        fileEdit.setReadOnly(True)
        fileBtn = QtWidgets.QPushButton("Choose file...", dlg)
        fileRow.addWidget(fileEdit)
        fileRow.addWidget(fileBtn)
        layout.addRow("Export to", fileRow)

        def _chooseFile():
            fname, _ = QtWidgets.QFileDialog.getSaveFileName(dlg, "Select inclusion list file", "", "CSV files (*.csv)")
            if fname:
                fileEdit.setText(fname)

        fileBtn.clicked.connect(_chooseFile)

        buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.StandardButton.Ok | QtWidgets.QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        layout.addRow(buttons)

        if dlg.exec() != QtWidgets.QDialog.DialogCode.Accepted:
            return

        include_pos = chkPos.isChecked()
        include_neg = chkNeg.isChecked()
        include_nat = chkNat.isChecked()
        include_lab = chkLab.isChecked()
        if not (include_pos or include_neg):
            QtWidgets.QMessageBox.warning(self, "Generate Inclusion Lists", "Select at least one polarity to include.")
            return
        if not (include_nat or include_lab):
            QtWidgets.QMessageBox.warning(self, "Generate Inclusion Lists", "Select at least native or labeled to include.")
            return

        out_file = str(fileEdit.text()).strip()
        if not out_file:
            QtWidgets.QMessageBox.warning(self, "Generate Inclusion Lists", "Please choose an output file.")
            return

        style = styleCombo.currentText()
        rt_offset = rtOffsetSpin.value()
        intens_threshold = intensSpin.value()
        sep_pol = chkSepPol.isChecked()
        sep_iso = chkSepIso.isChecked()
        msms_only = chkMSMS.isChecked()
        msms_min_prec_intens = msmsMinPrecIntensSpin.value()
        msms_min_spectra = msmsMinSpectraSpin.value()

        # Collect feature rows from the loaded experiment results, respecting active tree filters
        try:
            df = self.experimentResults.db_con.tables[self.experimentResults.selected_table]
        except Exception:
            QtWidgets.QMessageBox.warning(self, "Generate Inclusion Lists", "Could not access the loaded results table.")
            return

        rows = df.to_dicts()
        visible_nums = self._getVisibleFeatureNums()
        if visible_nums is not None:
            rows = [r for r in rows if r.get("Num") in visible_nums]

        # When restricting to features without MSMS spectra, count filtered MSMS
        # spectra (native and labeled separately) per feature, using the same RT-window
        # (apex deviation) and filter-line parameters as the experiment MSMS spectra
        # list. Features that reach the minimum number of spectra are excluded.
        msms_counts = {}
        if msms_only:
            msms_counts = self._countMSMSPerFeatureForRows(rows, msms_min_prec_intens)

        # Build polarity groups: (suffix, set_of_modes)
        pol_groups = []
        if sep_pol:
            if include_pos:
                pol_groups.append(("pos", {"+"}))
            if include_neg:
                pol_groups.append(("neg", {"-"}))
        else:
            modes = set()
            if include_pos:
                modes.add("+")
            if include_neg:
                modes.add("-")
            if modes == {"+", "-"}:
                suffix = "posneg"
            elif modes == {"+"}:
                suffix = "pos"
            else:
                suffix = "neg"
            pol_groups.append((suffix, modes))

        # Build isotopolog groups: (suffix, set_of_forms)
        iso_groups = []
        if sep_iso:
            if include_nat:
                iso_groups.append(("nat", {"native"}))
            if include_lab:
                iso_groups.append(("lab", {"labeled"}))
        else:
            forms = set()
            if include_nat:
                forms.add("native")
            if include_lab:
                forms.add("labeled")
            if forms == {"native", "labeled"}:
                suffix = "natlab"
            elif forms == {"native"}:
                suffix = "nat"
            else:
                suffix = "lab"
            iso_groups.append((suffix, forms))

        base, ext = os.path.splitext(out_file)
        ext = ".csv"

        written_files = []
        total_entries = 0
        for pol_suffix, modes in pol_groups:
            for iso_suffix, forms in iso_groups:
                entries = []
                for r in rows:
                    mode = str(r.get("Ionisation_Mode", "+") or "+")
                    if mode not in modes:
                        continue
                    avg = r.get("Average_peakarea")
                    if intens_threshold > 0.0 and avg is not None and float(avg) < intens_threshold:
                        continue
                    rt = float(r.get("RT", 0.0))
                    polarity = "positive" if "+" in mode else "negative"
                    ogroup = r.get("OGroup")
                    num = r.get("Num")
                    feat_msms = msms_counts.get(num, {}) if msms_only else None
                    if "native" in forms and r.get("MZ") is not None:
                        if not msms_only or feat_msms.get("native", 0) < msms_min_spectra:
                            entries.append(
                                {
                                    "compound": "OGroup%s_Num%s_native_%s" % (ogroup, num, polarity),
                                    "mz": float(r["MZ"]),
                                    "polarity": polarity,
                                    "tstart": rt - rt_offset,
                                    "tstop": rt + rt_offset,
                                    "intensityThreshold": intens_threshold,
                                }
                            )
                    if "labeled" in forms and r.get("L_MZ") is not None:
                        if not msms_only or feat_msms.get("labeled", 0) < msms_min_spectra:
                            entries.append(
                                {
                                    "compound": "OGroup%s_Num%s_labeled_%s" % (ogroup, num, polarity),
                                    "mz": float(r["L_MZ"]),
                                    "polarity": polarity,
                                    "tstart": rt - rt_offset,
                                    "tstop": rt + rt_offset,
                                    "intensityThreshold": intens_threshold,
                                }
                            )

                if not entries:
                    continue

                target = "%s_%s_%s%s" % (base, pol_suffix, iso_suffix, ext)
                if style == "IQ-X":
                    writeIQXInclusionList(entries, target)
                else:
                    writeQExactiveInclusionList(entries, target)
                written_files.append(os.path.basename(target))
                total_entries += len(entries)

        if written_files:
            QtWidgets.QMessageBox.information(
                self,
                "Generate Inclusion Lists",
                "Wrote %d entries to %d file(s):\n%s" % (total_entries, len(written_files), "\n".join(written_files)),
            )
        else:
            QtWidgets.QMessageBox.information(self, "Generate Inclusion Lists", "No features matched the selected criteria.")

    def _export_exp_msms_mgf(self):
        """Export filtered MSMS spectra from all features to MGF file."""
        if not hasattr(self, "loadedMZXMLs") or self.loadedMZXMLs is None:
            QtWidgets.QMessageBox.information(self, "Export MS/MS", "No MS/MS data available.")
            return

        # Show export options dialog
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Export MS/MS Spectra")
        layout = QtWidgets.QVBoxLayout(dlg)

        label = QtWidgets.QLabel("Select export mode:")
        layout.addWidget(label)

        mode_group = QtWidgets.QButtonGroup(dlg)
        all_spectra_radio = QtWidgets.QRadioButton("Export all spectra (current filtering applied)")
        all_spectra_radio.setChecked(True)
        abundant_radio = QtWidgets.QRadioButton("Export most abundant spectrum per feature\n(one native, one labeled - highest precursor intensity)")

        mode_group.addButton(all_spectra_radio, 0)
        mode_group.addButton(abundant_radio, 1)

        layout.addWidget(all_spectra_radio)
        layout.addWidget(abundant_radio)
        layout.addSpacing(10)

        separate_nl_checkbox = QtWidgets.QCheckBox("Separate for N/L type")
        separate_nl_checkbox.setToolTip("Write separate MGF files for native and labeled spectra, and for each distinct match of the first capturing group of the filter string regex. The separating factors are appended as underscore-separated suffixes to the output filename.")
        layout.addWidget(separate_nl_checkbox)
        layout.addSpacing(10)

        regex_row = QtWidgets.QHBoxLayout()
        regex_row.addWidget(QtWidgets.QLabel("Filter string regex:"))
        filter_regex_edit = QtWidgets.QLineEdit("(FTMS).*")
        try:
            filter_regex_edit.setText(str(self.ui.lineEdit_msms_filter_regex.text()))
        except Exception:
            pass
        filter_regex_edit.setToolTip("Regular expression applied to each spectrum's filter string. Empty exports all spectra.")
        regex_row.addWidget(filter_regex_edit)
        layout.addLayout(regex_row)

        buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.StandardButton.Ok | QtWidgets.QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        layout.addWidget(buttons)

        if dlg.exec() != QtWidgets.QDialog.DialogCode.Accepted:
            return

        export_mode = "most_abundant" if mode_group.checkedId() == 1 else "all"
        export_filter_regex = self._compile_msms_filter_regex(str(filter_regex_edit.text()).strip())
        separate_nl = separate_nl_checkbox.isChecked()

        # Collect all features from the tree
        all_features = {}
        tree = self.ui.resultsExperiment_TreeWidget
        for i in range(tree.topLevelItemCount()):
            top = tree.topLevelItem(i)
            for c in range(top.childCount()):
                child = top.child(c)
                bd = getattr(child, "bunchData", None)
                if bd and hasattr(bd, "id"):
                    all_features[bd.id] = bd

        if not all_features:
            QtWidgets.QMessageBox.information(self, "Export MS/MS", "No features available.")
            return

        # Get current filter settings
        try:
            msms_rt_window = self.ui.doubleSpinBox_resultsExperiment_MSMSRTWindow.value()
        except Exception:
            msms_rt_window = 0.5

        try:
            prec_intens_percent = self.ui.doubleSpinBox_resultsExperiment_MSMSPrecIntensPercent.value() / 100.0
        except Exception:
            prec_intens_percent = 0.5

        try:
            abs_intens_threshold = self.ui.doubleSpinBox_resultsExperiment_MSMSAbsIntensThreshold.value()
        except Exception:
            abs_intens_threshold = 0.0

        try:
            ppm = self.ui.doubleSpinBox_resultsExperiment_EICppm.value()
        except Exception:
            ppm = 5.0

        # Build file-path -> group-name mapping
        file_to_group = {}
        for grp in self.getAllSampleGroups():
            for fpath in grp.files:
                file_to_group[fpath] = grp.name

        file_keys = [k for k in self.loadedMZXMLs if k.lower().endswith(".mzxml") or k.lower().endswith(".mzml")]

        # Collect feature data for per-file RT bounds
        rows_by_num = {}
        if hasattr(self, "experimentResults") and self.experimentResults is not None and self.experimentResults.db_con is not None:
            for sheet in ["4_Convoluted", "3_Reintegrated", "1_Bracketed"]:
                try:
                    tdf = self.experimentResults.db_con.get_table(sheet)
                    if tdf is not None and not tdf.is_empty():
                        rows_by_num = {r["Num"]: r for r in tdf.to_dicts()}
                        break
                except Exception:
                    pass

        # Cache for EIC max intensities
        _eic_max_cache = {}

        def _get_eic_peak_max(file_key, native_mz, rt_center, rt_window):
            cache_key = (file_key, native_mz)
            if cache_key in _eic_max_cache:
                return _eic_max_cache[cache_key]

            mzxml_file = self.loadedMZXMLs.get(file_key)
            if mzxml_file is None:
                return None

            peak_rt = (rt_center - rt_window * 60.0, rt_center + rt_window * 60.0)
            try:
                eic, times, _, _ = mzxml_file.getEIC(native_mz, ppm=ppm, filterLine=None)
                peak_max = 0.0
                for intensity, t in zip(eic, times):
                    if peak_rt[0] <= t <= peak_rt[1]:
                        if intensity > peak_max:
                            peak_max = intensity
                result = peak_max if peak_max > 0.0 else None
            except Exception:
                result = None

            _eic_max_cache[cache_key] = result
            return result

        # Collect all matching spectra
        all_ms2_scans = []
        for file_key in file_keys:
            mzxml_file = self.loadedMZXMLs[file_key]
            if not (hasattr(mzxml_file, "MS2_list") and len(mzxml_file.MS2_list) > 0):
                continue

            for ms2_scan in mzxml_file.MS2_list:
                for feature_num, bd in all_features.items():
                    # Calculate m/z windows
                    native_mz_min = bd.mz * (1 - ppm / 1000000.0)
                    native_mz_max = bd.mz * (1 + ppm / 1000000.0)
                    labeled_mz_min = bd.lmz * (1 - ppm / 1000000.0) if bd.lmz is not None else None
                    labeled_mz_max = bd.lmz * (1 + ppm / 1000000.0) if bd.lmz is not None else None

                    # RT window around feature
                    rt_min_s = bd.rt - msms_rt_window * 60.0
                    rt_max_s = bd.rt + msms_rt_window * 60.0

                    # Check RT window
                    if not (rt_min_s <= ms2_scan.retention_time <= rt_max_s):
                        continue

                    # Check polarity matches the feature's ionisation mode
                    ion_mode = getattr(bd, "ionisationMode", None)
                    if ion_mode and ms2_scan.polarity and ms2_scan.polarity != ion_mode:
                        continue

                    # Check precursor m/z
                    form = None
                    if native_mz_min <= ms2_scan.precursor_mz <= native_mz_max:
                        form = "native"
                    elif labeled_mz_min is not None and labeled_mz_min <= ms2_scan.precursor_mz <= labeled_mz_max:
                        form = "labeled"

                    if form is None:
                        continue

                    # Filter-string regex filter (cvParam MS:1000512 / filter_line)
                    fs_match, fs_replacement = self._msms_filter_match(self._get_msms_filter_string(ms2_scan), export_filter_regex)
                    if not fs_match:
                        continue

                    # Check percent threshold
                    if prec_intens_percent > 0.0:
                        peak_max = _get_eic_peak_max(file_key, bd.mz, bd.rt, msms_rt_window)
                        if peak_max is not None and ms2_scan.precursor_intensity < prec_intens_percent * peak_max:
                            continue

                    # Check absolute threshold
                    if abs_intens_threshold > 0.0 and ms2_scan.precursor_intensity < abs_intens_threshold:
                        continue

                    all_ms2_scans.append(
                        {
                            "scan": ms2_scan,
                            "form": form,
                            "file": file_key,
                            "feature_num": feature_num,
                            "o_group": getattr(bd, "metaboliteGroupID", None),
                            "xn": getattr(bd, "xn", None),
                            "filter_replacement": fs_replacement,
                        }
                    )
                    break

        # Filter to most abundant if requested
        if export_mode == "most_abundant":
            abundant_scans = {}
            for row in all_ms2_scans:
                feature_num = row["feature_num"]
                form = row["form"]
                key = (feature_num, form)

                if key not in abundant_scans:
                    abundant_scans[key] = row
                else:
                    # Keep the one with higher precursor intensity
                    if row["scan"].precursor_intensity > abundant_scans[key]["scan"].precursor_intensity:
                        abundant_scans[key] = row

            all_ms2_scans = list(abundant_scans.values())

        if len(all_ms2_scans) == 0:
            QtWidgets.QMessageBox.information(self, "Export MS/MS", "No MS/MS spectra match the current filters.")
            return

        # Ask user for file location
        save_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Export MS/MS spectra to MGF", "msms_export.mgf", "MGF files (*.mgf)")
        if not save_path:
            return

        # Group rows for separate output files if requested (native/labeled and, if present,
        # the first capturing group of the filter string regex); otherwise a single group.
        if separate_nl:
            groups = defaultdict(list)
            for row in all_ms2_scans:
                suffix_parts = [row["form"]]
                if row.get("filter_replacement"):
                    suffix_parts.append(re.sub(r"[^A-Za-z0-9.\-]+", "-", str(row["filter_replacement"])))
                groups[tuple(suffix_parts)].append(row)
        else:
            groups = {(): all_ms2_scans}

        base, ext = os.path.splitext(save_path)
        if not ext:
            ext = ".mgf"

        def _write_mgf_rows(rows, target_path):
            with open(target_path, "w", encoding="utf-8") as f:
                for row in rows:
                    scan = row["scan"]
                    form = row["form"]
                    feature_num = row["feature_num"]
                    o_group = row["o_group"]
                    file_key = row["file"]
                    xn = row["xn"]

                    # Get file name
                    file_name = os.path.basename(file_key) if file_key else "unknown"

                    # Get polarity from the scan or other metadata
                    polarity = getattr(scan, "polarity", None) or ""

                    # Extract filter string: prioritize MS:1000512 from cvParams, fall back to filter_line
                    filter_string = None
                    if hasattr(scan, "cvParams") and scan.cvParams:
                        for cv in scan.cvParams:
                            if cv.get("accession") == "MS:1000512":
                                filter_string = cv.get("value", "")
                                break

                    # Fall back to filter_line attribute if no cvParam found
                    if not filter_string:
                        filter_string = getattr(scan, "filter_line", None)
                        # Don't use default "N/A" values
                        if filter_string and ("N/A" in filter_string or "Unknown" in filter_string):
                            filter_string = None

                    # Normalize spectrum (most abundant = 100.0)
                    processing = []
                    if len(scan.intensity_list) > 0:
                        max_intensity = max(scan.intensity_list)
                        if max_intensity > 0:
                            normalized_intensities = [float(i) / max_intensity * 100.0 for i in scan.intensity_list]
                            processing.append("Normalization - relative intensity to most intense peak")
                        else:
                            normalized_intensities = list(scan.intensity_list)
                    else:
                        normalized_intensities = []

                    # Write MGF block
                    f.write("BEGIN IONS\n")
                    f.write(f"TITLE=Feature_{feature_num}_{form}\n")
                    f.write(f"PEPMASS={float(getattr(scan, 'precursor_mz', 0.0)):.6f}\n")
                    f.write(f"RTINSECONDS={float(getattr(scan, 'retention_time', 0.0)):.3f}\n")
                    f.write(f"PRECURSOR_INTENSITY={float(getattr(scan, 'precursor_intensity', 0.0)):.4g}\n")
                    f.write(f"FEATURE_ID={feature_num}\n")
                    f.write(f"OGROUP={o_group}\n")
                    f.write(f"FORM={form}\n")
                    f.write(f"POLARITY={polarity}\n")
                    f.write(f"FILE={file_name}\n")
                    f.write(f"XN={xn}\n")
                    if filter_string:
                        f.write(f"FILTER_LINE={filter_string}\n")
                    if hasattr(scan, "collisionEnergy") and scan.collisionEnergy:
                        f.write(f"COLLISION_ENERGY={scan.collisionEnergy}\n")
                    if hasattr(scan, "setup") and scan.setup:
                        f.write(f"SETUP={scan.setup}\n")
                    if processing:
                        f.write(f"PROCESSING={';'.join(processing)}\n")

                    # Write normalized peaks
                    for mz, intensity in zip(scan.mz_list, normalized_intensities):
                        f.write(f"{float(mz):.6f} {float(intensity):.2f}\n")

                    f.write("END IONS\n\n")

        try:
            written_files = []
            for suffix_parts, rows in groups.items():
                target_path = f"{base}_{'_'.join(suffix_parts)}{ext}" if suffix_parts else save_path
                _write_mgf_rows(rows, target_path)
                written_files.append(target_path)

            QtWidgets.QMessageBox.information(
                self,
                "Export MS/MS",
                "Exported %d spectra to %d file(s):\n%s" % (len(all_ms2_scans), len(written_files), "\n".join(os.path.basename(p) for p in written_files)),
            )

            # Show FragExtract command dialog
            dlg = QtWidgets.QDialog(self)
            dlg.setWindowTitle("MS/MS Export Complete")
            dlg.setMinimumWidth(600)
            layout = QtWidgets.QVBoxLayout(dlg)

            info_label = QtWidgets.QLabel(f"Successfully exported {len(all_ms2_scans)} spectra.")
            layout.addWidget(info_label)

            layout.addWidget(QtWidgets.QLabel("To clean the spectra, run these command(s):"))

            # Build FragExtract command(s), one per written file
            frag_commands = []
            for mgf_path in written_files:
                output_folder = os.path.dirname(mgf_path)
                frag_commands.append(f'.\\FragExtract.exe --isotope-mz-diff 1.00335484 --max-rt-diff 0.1 --grouping-fields FEATURE_ID --input-mgf "{mgf_path}" --output-folder "{output_folder}/plots"')
            frag_command = "\n".join(frag_commands)

            cmd_text = QtWidgets.QTextEdit()
            cmd_text.setPlainText(frag_command)
            cmd_text.setReadOnly(True)
            cmd_text.setMaximumHeight(120)
            layout.addWidget(cmd_text)

            copy_btn = QtWidgets.QPushButton("Copy Command")
            copy_btn.clicked.connect(lambda: QtWidgets.QApplication.clipboard().setText(frag_command))
            layout.addWidget(copy_btn)

            close_btn = QtWidgets.QPushButton("Close")
            close_btn.clicked.connect(dlg.accept)
            layout.addWidget(close_btn)

            dlg.exec()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Export MS/MS error", f"Failed to export: {str(e)}")

    def _export_msms_mgf(self):
        rows = list(self._iter_exp_msms_rows())
        if len(rows) == 0:
            QtWidgets.QMessageBox.information(self, "MS/MS export", "No MS/MS spectra available for export.")
            return

        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Export MS/MS to MGF")
        form = QtWidgets.QVBoxLayout(dlg)
        grid = QtWidgets.QFormLayout()
        mode = QtWidgets.QComboBox()
        mode.addItems(["Raw spectra", "Average spectrum per feature", "Most abundant spectrum per feature", "Cleaned spectrum per feature"])
        grid.addRow("Export mode:", mode)
        include_native = QtWidgets.QCheckBox("Export native spectra")
        include_native.setChecked(True)
        include_labeled = QtWidgets.QCheckBox("Export labeled spectra")
        include_labeled.setChecked(True)
        grid.addRow(include_native)
        grid.addRow(include_labeled)
        allow_zero_label = QtWidgets.QCheckBox("Cleaning: allow 0 labeling atoms")
        allow_zero_label.setChecked(True)
        grid.addRow(allow_zero_label)
        separate_collision = QtWidgets.QCheckBox("Create separate files for each collision setup")
        grid.addRow(separate_collision)
        collision_list = QtWidgets.QListWidget()
        collision_list.setSelectionMode(QtWidgets.QAbstractItemView.MultiSelection)
        collision_keys = sorted({f"{getattr(r['scan'], 'filter_line', '')}|CE={getattr(r['scan'], 'collisionEnergy', '')}" for r in rows})
        for ck in collision_keys:
            it = QtWidgets.QListWidgetItem(ck)
            it.setFlags(it.flags() | QtCore.Qt.ItemIsUserCheckable)
            it.setCheckState(QtCore.Qt.Checked)
            collision_list.addItem(it)
        grid.addRow("Collision setups:", collision_list)
        filter_regex_edit = QtWidgets.QLineEdit()
        try:
            filter_regex_edit.setText(str(self.ui.lineEdit_msms_filter_regex.text()))
        except Exception:
            pass
        filter_regex_edit.setToolTip("Regular expression applied to each spectrum's filter string. Empty exports all displayed spectra.")
        grid.addRow("Filter string regex:", filter_regex_edit)
        form.addLayout(grid)
        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        form.addWidget(btns)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        if dlg.exec() != QtWidgets.QDialog.Accepted:
            return

        forms = set()
        if include_native.isChecked():
            forms.add("native")
        if include_labeled.isChecked():
            forms.add("labeled")
        if len(forms) == 0:
            QtWidgets.QMessageBox.warning(self, "MS/MS export", "Select at least one isotopolog form.")
            return

        save_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save MS/MS MGF", "msms_export.mgf", "MGF file (*.mgf)")
        if save_path == "":
            return

        selected_collision = {collision_list.item(i).text() for i in range(collision_list.count()) if collision_list.item(i).checkState() == QtCore.Qt.Checked}
        export_regex = self._compile_msms_filter_regex(str(filter_regex_edit.text()).strip())
        selected = [r for r in rows if r["form"] in forms and f"{getattr(r['scan'], 'filter_line', '')}|CE={getattr(r['scan'], 'collisionEnergy', '')}" in selected_collision and self._msms_filter_match(r.get("filter_string", ""), export_regex)[0]]
        if len(selected) == 0:
            QtWidgets.QMessageBox.information(self, "MS/MS export", "No spectra match the current export selection.")
            return

        by_key = defaultdict(list)
        for r in selected:
            collision = f"{getattr(r['scan'], 'filter_line', '')}|CE={getattr(r['scan'], 'collisionEnergy', '')}"
            if separate_collision.isChecked():
                by_key[(r["form"], collision)].append(r)
            else:
                by_key[(r["form"], "all")].append(r)

        isotope_mass_offset = getattr(self, "_cached_isotope_mass_offset", None)
        if isotope_mass_offset is None:
            isotope_mass_offset = abs(getIsotopeMass(str(self.ui.isotopeBText.text())) - getIsotopeMass(str(self.ui.isotopeAText.text())))
            self._cached_isotope_mass_offset = isotope_mass_offset

        def _clean_fragextract_like(native_scan, labeled_scan, xn):
            if native_scan is None or labeled_scan is None:
                return native_scan, labeled_scan
            min_atoms = 0 if allow_zero_label.isChecked() else 1
            max_atoms = int(xn) if xn is not None else 12
            max_atoms = max(min_atoms, max_atoms)
            best_n = min_atoms
            best_score = -1
            charge = max(1, int(getattr(native_scan, "precursorCharge", 1) or 1))
            n_mz = np.asarray(native_scan.mz_list, dtype=float)
            l_mz = np.asarray(labeled_scan.mz_list, dtype=float)
            n_it = np.asarray(native_scan.intensity_list, dtype=float)
            l_it = np.asarray(labeled_scan.intensity_list, dtype=float)
            for n in range(min_atoms, max_atoms + 1):
                shift = n * isotope_mass_offset / charge
                score = 0.0
                for mz, inten in zip(n_mz, n_it):
                    if np.any(np.abs((l_mz - mz) - shift) <= 0.01):
                        score += float(inten)
                if score > best_score:
                    best_score = score
                    best_n = n
            shift = best_n * isotope_mass_offset / charge
            keep_n = []
            keep_l = []
            for i, mz in enumerate(n_mz):
                if np.any(np.abs((l_mz - mz) - shift) <= 0.01):
                    keep_n.append(i)
            for i, mz in enumerate(l_mz):
                if np.any(np.abs((mz - n_mz) - shift) <= 0.01):
                    keep_l.append(i)
            cn = deepcopy(native_scan)
            cl = deepcopy(labeled_scan)
            if keep_n:
                cn.mz_list = n_mz[keep_n]
                cn.intensity_list = n_it[keep_n]
            if keep_l:
                cl.mz_list = l_mz[keep_l]
                cl.intensity_list = l_it[keep_l]
            return cn, cl

        def _representative_spectrum(entries):
            if mode.currentText() == "Most abundant spectrum per feature":
                return max(entries, key=lambda x: float(getattr(x["scan"], "precursor_intensity", 0.0)))["scan"]
            if mode.currentText() == "Raw spectra":
                return None
            bins = defaultdict(float)
            for e in entries:
                scan = e["scan"]
                for mz, inten in zip(scan.mz_list, scan.intensity_list):
                    mz_bin = round(float(mz), 3)
                    bins[mz_bin] += float(inten)
            if len(bins) == 0:
                return entries[0]["scan"]
            mzs = sorted(bins.keys())
            ints = [bins[mz] / max(1, len(entries)) for mz in mzs]
            rep = deepcopy(entries[0]["scan"])
            rep.mz_list = np.asarray(mzs, dtype=float)
            rep.intensity_list = np.asarray(ints, dtype=float)
            return rep

        written_files = []
        for (form_key, collision_key), vals in by_key.items():
            out_path = save_path
            if len(by_key) > 1:
                suffix = f"_{form_key}_{sanitize_filename(collision_key)}"
                out_path = save_path.replace(".mgf", f"{suffix}.mgf")
            with open(out_path, "w", encoding="utf-8") as out:
                if mode.currentText() == "Raw spectra":
                    export_entries = [[v] for v in vals]
                else:
                    by_feature = defaultdict(list)
                    for v in vals:
                        by_feature[v["feature_num"]].append(v)
                    export_entries = [fe for fe in by_feature.values()]
                for entries in export_entries:
                    if len(entries) == 0:
                        continue
                    feature_num = entries[0]["feature_num"]
                    o_group = entries[0]["o_group"]
                    scan = entries[0]["scan"] if mode.currentText() == "Raw spectra" else _representative_spectrum(entries)
                    if mode.currentText() == "Cleaned spectrum per feature":
                        native_entries = [x for x in selected if x["feature_num"] == feature_num and x["form"] == "native"]
                        labeled_entries = [x for x in selected if x["feature_num"] == feature_num and x["form"] == "labeled"]
                        n_scan = _representative_spectrum(native_entries) if native_entries else None
                        l_scan = _representative_spectrum(labeled_entries) if labeled_entries else None
                        n_scan, l_scan = _clean_fragextract_like(n_scan, l_scan, entries[0].get("xn"))
                        if form_key == "native" and n_scan is not None:
                            scan = n_scan
                        elif form_key == "labeled" and l_scan is not None:
                            scan = l_scan
                    out.write("BEGIN IONS\n")
                    out.write(f"TITLE=Num_{feature_num}_OGROUP_{o_group}_{form_key}\n")
                    out.write(f"PEPMASS={float(getattr(scan, 'precursor_mz', 0.0)):.6f}\n")
                    out.write(f"RTINSECONDS={float(getattr(scan, 'retention_time', 0.0)):.3f}\n")
                    out.write(f"FEATURE_NUM={feature_num}\n")
                    out.write(f"OGROUP={o_group}\n")
                    out.write(f"FORM={form_key}\n")
                    out.write(f"FILTER_LINE={getattr(scan, 'filter_line', '')}\n")
                    out.write(f"COLLISION_ENERGY={getattr(scan, 'collisionEnergy', '')}\n")
                    for mz, inten in zip(scan.mz_list, scan.intensity_list):
                        out.write(f"{float(mz):.6f} {float(inten):.6f}\n")
                    out.write("END IONS\n\n")
            written_files.append(out_path)

        QtWidgets.QMessageBox.information(self, "MS/MS export", "Exported:\n" + "\n".join(written_files))

    def _show_msms_overview(self):
        rows = list(self._iter_exp_msms_rows())
        by_feature = defaultdict(lambda: {"native": 0, "labeled": 0})
        for r in rows:
            if r["feature_num"] is None:
                continue
            by_feature[r["feature_num"]][r["form"]] += 1
        if len(by_feature) == 0:
            QtWidgets.QMessageBox.information(self, "MS/MS overview", "No feature-linked MS/MS spectra available.")
            return

        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("MS/MS overview")
        dlg.resize(640, 420)
        layout = QtWidgets.QVBoxLayout(dlg)
        table = QtWidgets.QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Feature Num", "Native scans", "Labeled scans", "Total"])
        table.setRowCount(len(by_feature))
        for i, fid in enumerate(sorted(by_feature.keys())):
            n = by_feature[fid]["native"]
            l = by_feature[fid]["labeled"]
            table.setItem(i, 0, QtWidgets.QTableWidgetItem(str(fid)))
            table.setItem(i, 1, QtWidgets.QTableWidgetItem(str(n)))
            table.setItem(i, 2, QtWidgets.QTableWidgetItem(str(l)))
            table.setItem(i, 3, QtWidgets.QTableWidgetItem(str(n + l)))
        table.resizeColumnsToContents()
        layout.addWidget(table)
        fig = Figure((5.0, 2.5), dpi=80, facecolor="white")
        canvas = FigureCanvas(fig)
        ax = fig.add_subplot(111)
        fids = sorted(by_feature.keys())
        native_counts = [by_feature[f]["native"] for f in fids]
        labeled_counts = [by_feature[f]["labeled"] for f in fids]
        x = np.arange(len(fids))
        ax.bar(x - 0.2, native_counts, width=0.4, color="dodgerblue", label="Native")
        ax.bar(x + 0.2, labeled_counts, width=0.4, color="firebrick", label="Labeled")
        ax.set_xticks(x)
        ax.set_xticklabels([str(f) for f in fids], rotation=90)
        ax.set_xlabel("Feature Num")
        ax.set_ylabel("MS/MS scan count")
        ax.legend(loc="upper right")
        fig.tight_layout()
        layout.addWidget(canvas)
        dlg.exec()

    def _compute_msms_filtered_matches(self, rows):
        """Match all loaded MS/MS scans against the given feature rows using the same
        filter parameters as the experiment "Show options" (RT window around the
        feature apex, precursor m/z tolerance, precursor intensity percent / absolute
        thresholds and the filter-string regex).

        Returns a dict with keys "matched", "feature_ranges", "file_keys",
        "file_to_group", "per_feature_form_count", "per_sample_stats" or ``None`` if
        raw MS/MS data is not available.
        """
        if not hasattr(self, "loadedMZXMLs") or self.loadedMZXMLs is None:
            return None

        # Read the "Show options" MSMS filter parameters
        try:
            msms_rt_window = self.ui.doubleSpinBox_resultsExperiment_MSMSRTWindow.value()
        except Exception:
            msms_rt_window = 0.5
        try:
            prec_intens_percent = self.ui.doubleSpinBox_resultsExperiment_MSMSPrecIntensPercent.value() / 100.0
        except Exception:
            prec_intens_percent = 0.0
        try:
            abs_intens_threshold = self.ui.doubleSpinBox_resultsExperiment_MSMSAbsIntensThreshold.value()
        except Exception:
            abs_intens_threshold = 0.0
        try:
            ppm = self.ui.doubleSpinBox_resultsExperiment_EICppm.value()
        except Exception:
            ppm = 5.0
        try:
            filter_regex_pattern = str(self.ui.lineEdit_msms_filter_regex.text()).strip()
        except Exception:
            filter_regex_pattern = ""
        filter_regex = self._compile_msms_filter_regex(filter_regex_pattern)

        file_keys = [k for k in self.loadedMZXMLs if k.lower().endswith(".mzxml") or k.lower().endswith(".mzml")]

        # File-path -> group-name mapping and display sample names
        file_to_group = {}
        for grp in self.getAllSampleGroups():
            for fpath in grp.files:
                file_to_group[fpath] = grp.name

        def _sample_name(file_key):
            name = os.path.basename(file_key)
            return re.sub(r"\.(mzxml|mzml)$", "", name, flags=re.IGNORECASE)

        def _first_float(val):
            if val is None:
                return None
            try:
                return float(str(val).split(";")[0])
            except (TypeError, ValueError):
                return None

        # Build per-feature ranges with apex RT and per-sample apex intensities
        feature_ranges = []
        for r in rows:
            num = r.get("Num")
            if num is None:
                continue
            rt_min = _first_float(r.get("RT"))
            if rt_min is None:
                continue
            mz = _first_float(r.get("MZ"))
            lmz = _first_float(r.get("L_MZ"))
            mode = str(r.get("Ionisation_Mode", "+") or "+")

            per_file_apex_rt = {"native": {}, "labeled": {}}
            per_file_abund = {"native": {}, "labeled": {}}
            per_file_peak_rt = {"native": {}, "labeled": {}}
            for file_key in file_keys:
                sname = _sample_name(file_key)
                n_apex = _first_float(r.get(f"{sname}_N_apexRT"))
                l_apex = _first_float(r.get(f"{sname}_L_apexRT"))
                if n_apex is not None:
                    per_file_apex_rt["native"][file_key] = n_apex
                if l_apex is not None:
                    per_file_apex_rt["labeled"][file_key] = l_apex
                n_ab = _first_float(r.get(f"{sname}_Abundance_N"))
                l_ab = _first_float(r.get(f"{sname}_Abundance_L"))
                if n_ab is not None:
                    per_file_abund["native"][file_key] = n_ab
                if l_ab is not None:
                    per_file_abund["labeled"][file_key] = l_ab
                n_start = _first_float(r.get(f"{sname}_N_startRT"))
                n_end = _first_float(r.get(f"{sname}_N_endRT"))
                if n_start is not None and n_end is not None:
                    per_file_peak_rt["native"][file_key] = (n_start * 60.0, n_end * 60.0)
                l_start = _first_float(r.get(f"{sname}_L_startRT"))
                l_end = _first_float(r.get(f"{sname}_L_endRT"))
                if l_start is not None and l_end is not None:
                    per_file_peak_rt["labeled"][file_key] = (l_start * 60.0, l_end * 60.0)

            max_abund = {
                "native": max(per_file_abund["native"].values()) if per_file_abund["native"] else 0.0,
                "labeled": max(per_file_abund["labeled"].values()) if per_file_abund["labeled"] else 0.0,
            }

            feature_ranges.append(
                {
                    "num": num,
                    "rt_min": rt_min,
                    "rt_min_s": rt_min * 60.0 - msms_rt_window * 60.0,
                    "rt_max_s": rt_min * 60.0 + msms_rt_window * 60.0,
                    "mz": mz,
                    "lmz": lmz,
                    "native_mz_min": mz * (1 - ppm / 1e6) if mz is not None else None,
                    "native_mz_max": mz * (1 + ppm / 1e6) if mz is not None else None,
                    "labeled_mz_min": lmz * (1 - ppm / 1e6) if lmz is not None else None,
                    "labeled_mz_max": lmz * (1 + ppm / 1e6) if lmz is not None else None,
                    "mode": mode,
                    "per_file_apex_rt": per_file_apex_rt,
                    "per_file_abund": per_file_abund,
                    "per_file_peak_rt": per_file_peak_rt,
                    "max_abund": max_abund,
                }
            )

        # EIC peak-max cache: (file_key, num, form) -> intensity (peak apex height in that file)
        _eic_max_cache = {}

        def _eic_peak_max(file_key, fr, form):
            cache_key = (file_key, fr["num"], form)
            if cache_key in _eic_max_cache:
                return _eic_max_cache[cache_key]
            mz_val = fr["mz"] if form == "native" else fr["lmz"]
            if mz_val is None:
                _eic_max_cache[cache_key] = None
                return None
            peak_rt = fr["per_file_peak_rt"][form].get(file_key)
            if peak_rt is None:
                peak_rt = (fr["rt_min"] * 60.0 - msms_rt_window * 60.0, fr["rt_min"] * 60.0 + msms_rt_window * 60.0)
            mzxml_file = self.loadedMZXMLs.get(file_key)
            if mzxml_file is None:
                _eic_max_cache[cache_key] = None
                return None
            try:
                filter_lines = mzxml_file.getFilterLines(includeMS1=True, includeMS2=False, includePosPolarity=True, includeNegPolarity=True)
                scan_event = None
                if filter_lines:
                    mode = fr.get("mode")
                    if mode and "+" in str(mode):
                        scan_event = next((fl for fl in filter_lines if "+" in fl), filter_lines[0])
                    elif mode and "-" in str(mode):
                        scan_event = next((fl for fl in filter_lines if "-" in fl), filter_lines[0])
                    else:
                        scan_event = filter_lines[0]
                if scan_event is None:
                    _eic_max_cache[cache_key] = None
                    return None
                eic, times, _, _ = mzxml_file.getEIC(mz_val, ppm=ppm, filterLine=scan_event)
                peak_max = 0.0
                for intensity, t in zip(eic, times):
                    if peak_rt[0] <= t <= peak_rt[1] and intensity > peak_max:
                        peak_max = intensity
                result = peak_max if peak_max > 0.0 else None
            except Exception:
                result = None
            _eic_max_cache[cache_key] = result
            return result

        # Match MS/MS scans against the feature ranges using the Show-options filters
        matched = []  # one record per matched scan
        per_feature_form_count = defaultdict(int)  # (num, form) -> count
        per_sample_stats = {}  # file_key -> {"obtained": int, "native": int, "labeled": int, "native_features": set, "labeled_features": set}
        for file_key in file_keys:
            mzxml_file = self.loadedMZXMLs[file_key]
            ms2_list = getattr(mzxml_file, "MS2_list", None)
            obtained = len(ms2_list) if ms2_list else 0
            per_sample_stats[file_key] = {"obtained": obtained, "native": 0, "labeled": 0, "native_features": set(), "labeled_features": set()}
            if not ms2_list:
                continue
            for ms2_scan in ms2_list:
                if abs_intens_threshold > 0.0 and ms2_scan.precursor_intensity < abs_intens_threshold:
                    continue
                fs_string = None
                fs_checked = False
                fs_match = True
                fs_replacement = None
                for fr in feature_ranges:
                    if not (fr["rt_min_s"] <= ms2_scan.retention_time <= fr["rt_max_s"]):
                        continue
                    mode = fr.get("mode")
                    if mode and ms2_scan.polarity and ms2_scan.polarity != mode:
                        continue
                    form = None
                    if fr["native_mz_min"] is not None and fr["native_mz_min"] <= ms2_scan.precursor_mz <= fr["native_mz_max"]:
                        form = "native"
                    elif fr["labeled_mz_min"] is not None and fr["labeled_mz_min"] <= ms2_scan.precursor_mz <= fr["labeled_mz_max"]:
                        form = "labeled"
                    if form is None:
                        continue

                    if not fs_checked:
                        fs_string = self._get_msms_filter_string(ms2_scan)
                        fs_match, fs_replacement = self._msms_filter_match(fs_string, filter_regex)
                        fs_checked = True
                    if not fs_match:
                        break

                    apex_intensity = fr["per_file_abund"][form].get(file_key)
                    if apex_intensity is None or apex_intensity <= 0.0:
                        apex_intensity = _eic_peak_max(file_key, fr, form)
                    if prec_intens_percent > 0.0 and apex_intensity is not None and apex_intensity > 0.0:
                        if ms2_scan.precursor_intensity < prec_intens_percent * apex_intensity:
                            continue

                    apex_rt = fr["per_file_apex_rt"][form].get(file_key, fr["rt_min"])
                    matched.append(
                        {
                            "num": fr["num"],
                            "form": form,
                            "file_key": file_key,
                            "prec_intensity": ms2_scan.precursor_intensity,
                            "rt_offset_min": ms2_scan.retention_time / 60.0 - apex_rt,
                            "apex_intensity": apex_intensity,
                            "scan": ms2_scan,
                            "filter_replacement": fs_replacement,
                        }
                    )
                    per_feature_form_count[(fr["num"], form)] += 1
                    per_sample_stats[file_key][form] += 1
                    per_sample_stats[file_key][f"{form}_features"].add(fr["num"])
                    break

        return {
            "matched": matched,
            "feature_ranges": feature_ranges,
            "file_keys": file_keys,
            "file_to_group": file_to_group,
            "per_feature_form_count": per_feature_form_count,
            "per_sample_stats": per_sample_stats,
        }

    def _showMSMSPrecursorOverview(self):
        """Show an overview of the currently filtered MSMS precursor spectra.

        Uses the same MSMS filter parameters as the experiment "Show options"
        (RT window around the feature apex, precursor m/z tolerance, precursor
        intensity percent / absolute thresholds and the filter-string regex) and
        evaluates them against all features currently visible in the results tree.
        """
        if not hasattr(self, "experimentResults") or self.experimentResults is None or self.experimentResults.db_con is None:
            QtWidgets.QMessageBox.information(self, "MSMS precursor overview", "No experiment results loaded.")
            return
        if not hasattr(self, "loadedMZXMLs") or self.loadedMZXMLs is None:
            QtWidgets.QMessageBox.information(self, "MSMS precursor overview", "No raw MS/MS data loaded.")
            return

        try:
            df = self.experimentResults.db_con.tables[self.experimentResults.selected_table]
        except Exception:
            QtWidgets.QMessageBox.warning(self, "MSMS precursor overview", "Could not access the loaded results table.")
            return

        rows = df.to_dicts()
        visible_nums = self._getVisibleFeatureNums()
        if visible_nums is not None:
            rows = [r for r in rows if r.get("Num") in visible_nums]
        if not rows:
            QtWidgets.QMessageBox.information(self, "MSMS precursor overview", "No features available.")
            return

        match_result = self._compute_msms_filtered_matches(rows)
        if match_result is None:
            QtWidgets.QMessageBox.information(self, "MSMS precursor overview", "No raw MS/MS data loaded.")
            return
        matched = match_result["matched"]
        feature_ranges = match_result["feature_ranges"]
        file_to_group = match_result["file_to_group"]
        per_feature_form_count = match_result["per_feature_form_count"]
        per_sample_stats = match_result["per_sample_stats"]

        try:
            msms_rt_window = self.ui.doubleSpinBox_resultsExperiment_MSMSRTWindow.value()
        except Exception:
            msms_rt_window = 0.5

        def _sample_name(file_key):
            name = os.path.basename(file_key)
            return re.sub(r"\.(mzxml|mzml)$", "", name, flags=re.IGNORECASE)

        n_native = sum(1 for m in matched if m["form"] == "native")
        n_labeled = sum(1 for m in matched if m["form"] == "labeled")

        # Determine fragmented vs not-fragmented ions (one ion per feature/form)
        fragmented = {"native": [], "labeled": []}  # dicts with rt, mz, count
        not_fragmented = {"native": [], "labeled": []}  # dicts with rt, mz, max_abund
        features_with_msms = set()
        for fr in feature_ranges:
            for form, mz_key in (("native", "mz"), ("labeled", "lmz")):
                mz_val = fr[mz_key]
                if mz_val is None:
                    continue
                count = per_feature_form_count.get((fr["num"], form), 0)
                if count > 0:
                    features_with_msms.add(fr["num"])
                    fragmented[form].append({"rt": fr["rt_min"], "mz": mz_val, "count": count})
                else:
                    not_fragmented[form].append({"rt": fr["rt_min"], "mz": mz_val, "max_abund": fr["max_abund"][form]})

        n_features = len({fr["num"] for fr in feature_ranges})
        n_features_no_msms = n_features - len(features_with_msms)

        n_features_native = len(fragmented["native"]) + len(not_fragmented["native"])
        n_features_labeled = len(fragmented["labeled"]) + len(not_fragmented["labeled"])
        n_with_msms_native = len(fragmented["native"])
        n_with_msms_labeled = len(fragmented["labeled"])
        pct_native = (n_with_msms_native / n_features_native * 100.0) if n_features_native else 0.0
        pct_labeled = (n_with_msms_labeled / n_features_labeled * 100.0) if n_features_labeled else 0.0

        _native_color = "#1E90FF"
        _labeled_color = "#B22222"

        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("MSMS precursor overview")
        dlg.resize(1100, 860)
        outer_layout = QtWidgets.QVBoxLayout(dlg)
        scroll = QtWidgets.QScrollArea(dlg)
        scroll.setWidgetResizable(True)
        scroll_content = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(scroll_content)
        scroll.setWidget(scroll_content)
        outer_layout.addWidget(scroll)

        summary = QtWidgets.QLabel(
            f"<b>Features with MSMS spectra:</b> {len(features_with_msms)} / {n_features} &nbsp;·&nbsp; "
            f"native (M): {n_with_msms_native} / {n_features_native} ({pct_native:.1f}%) &nbsp;·&nbsp; "
            f"labeled (M\u2032): {n_with_msms_labeled} / {n_features_labeled} ({pct_labeled:.1f}%) &nbsp;·&nbsp; "
            f"without MSMS: {n_features_no_msms}<br>"
            f"<b>Filtered MSMS spectra:</b> {len(matched)} total &nbsp;·&nbsp; "
            f"native (M): {n_native} &nbsp;·&nbsp; labeled (M\u2032): {n_labeled}<br>"
        )
        summary.setTextFormat(QtCore.Qt.RichText)
        layout.addWidget(summary)

        plot = QtCore.QObject()
        plot.fig = Figure((10.0, 6.5), dpi=80, facecolor="white")
        plot.canvas = FigureCanvas(plot.fig)
        plot.mpl_toolbar = NavigationToolbar(plot.canvas, dlg)
        layout.addWidget(plot.mpl_toolbar)
        layout.addWidget(plot.canvas, 1)

        axFrag = plot.fig.add_subplot(2, 2, 1)
        axNotFrag = plot.fig.add_subplot(2, 2, 2)
        axOffset = plot.fig.add_subplot(2, 2, 3)
        axApex = plot.fig.add_subplot(2, 2, 4)

        # Feature map: fragmented ions (dot size ~ number of MSMS spectra)
        for form, color in (("native", _native_color), ("labeled", _labeled_color)):
            pts = fragmented[form]
            if pts:
                counts = [p["count"] for p in pts]
                max_c = max(counts)
                sizes = [20.0 + (c / max_c) * 230.0 for c in counts]
                axFrag.scatter(
                    [p["rt"] for p in pts],
                    [p["mz"] for p in pts],
                    s=sizes,
                    c=color,
                    alpha=0.7,
                    linewidths=0,
                    label=f"{'M' if form == 'native' else 'M' + chr(0x2032)} ({len(pts)})",
                )
        axFrag.set_xlabel("RT (min)")
        axFrag.set_ylabel("m/z")
        axFrag.set_title("Fragmented ions (size = # MSMS)")
        axFrag.grid(True, alpha=0.2)
        if fragmented["native"] or fragmented["labeled"]:
            axFrag.legend(fontsize=8, framealpha=0.7)

        # Feature map: not-fragmented ions (dot size ~ log10 max peak intensity)
        for form, color in (("native", _native_color), ("labeled", _labeled_color)):
            pts = not_fragmented[form]
            if pts:
                logvals = [log10(p["max_abund"]) if p["max_abund"] and p["max_abund"] > 1.0 else 0.0 for p in pts]
                max_l = max(logvals) if logvals else 0.0
                if max_l <= 0.0:
                    sizes = [25.0 for _ in pts]
                else:
                    sizes = [20.0 + (v / max_l) * 230.0 for v in logvals]
                axNotFrag.scatter(
                    [p["rt"] for p in pts],
                    [p["mz"] for p in pts],
                    s=sizes,
                    c=color,
                    alpha=0.7,
                    linewidths=0,
                    label=f"{'M' if form == 'native' else 'M' + chr(0x2032)} ({len(pts)})",
                )
        axNotFrag.set_xlabel("RT (min)")
        axNotFrag.set_ylabel("m/z")
        axNotFrag.set_title("Not fragmented ions (size = log10 max intensity)")
        axNotFrag.grid(True, alpha=0.2)
        if not_fragmented["native"] or not_fragmented["labeled"]:
            axNotFrag.legend(fontsize=8, framealpha=0.7)

        # Scatter: precursor intensity vs RT offset relative to apex
        for form, color in (("native", _native_color), ("labeled", _labeled_color)):
            xs = [m["rt_offset_min"] for m in matched if m["form"] == form]
            ys = [m["prec_intensity"] for m in matched if m["form"] == form]
            if xs:
                axOffset.scatter(xs, ys, c=color, alpha=0.55, s=20, linewidths=0, label="M" if form == "native" else "M\u2032", zorder=3)
        axOffset.axvline(0.0, color="gray", linewidth=0.8, linestyle="--")
        axOffset.set_xlabel("RT offset from peak apex (min)")
        axOffset.set_ylabel("Precursor intensity (log10)")
        axOffset.set_title("Precursor intensity vs. RT offset")
        axOffset.grid(True, alpha=0.2)
        if any(m["prec_intensity"] > 0.0 for m in matched):
            axOffset.set_yscale("log")

        # Mirrored (negative) histogram of the peaks' start/end retention times relative to
        # their apex, so the acquired MSMS spectra (positive side) can be compared against
        # where the chromatographic peaks actually started/ended.
        start_offsets = []
        end_offsets = []
        for fr in feature_ranges:
            for form in ("native", "labeled"):
                for file_key, apex_rt in fr["per_file_apex_rt"][form].items():
                    peak_rt = fr["per_file_peak_rt"][form].get(file_key)
                    if peak_rt is None:
                        continue
                    start_offsets.append(peak_rt[0] / 60.0 - apex_rt)
                    end_offsets.append(peak_rt[1] / 60.0 - apex_rt)

        hist_handles, hist_labels = [], []
        if start_offsets or end_offsets:
            axHist = axOffset.twinx()
            all_offsets = start_offsets + end_offsets
            bin_width = 0.02
            lim = max(max(abs(v) for v in all_offsets), msms_rt_window, bin_width)
            n_bins = int(np.ceil(lim / bin_width))
            bins = np.arange(-n_bins, n_bins + 1) * bin_width
            width = bin_width * 0.9
            centers = (bins[:-1] + bins[1:]) / 2.0
            max_count = 1
            if start_offsets:
                counts_s, _ = np.histogram(start_offsets, bins=bins)
                max_count = max(max_count, int(counts_s.max()))
                axHist.bar(centers, -counts_s, width=width, color="seagreen", alpha=0.4, label="Peak start", zorder=1)
            if end_offsets:
                counts_e, _ = np.histogram(end_offsets, bins=bins)
                max_count = max(max_count, int(counts_e.max()))
                axHist.bar(centers, -counts_e, width=width, color="darkorange", alpha=0.4, label="Peak end", zorder=1)
            axHist.set_ylim(-max_count * 4.0, max_count)
            axHist.set_yticks([0, -max_count])
            axHist.set_yticklabels(["0", str(max_count)])
            axHist.set_ylabel("Peak start/end count")
            axOffset.set_zorder(axHist.get_zorder() + 1)
            axOffset.patch.set_visible(False)
            hist_handles, hist_labels = axHist.get_legend_handles_labels()

        scatter_handles, scatter_labels = axOffset.get_legend_handles_labels()
        if scatter_handles or hist_handles:
            axOffset.legend(scatter_handles + hist_handles, scatter_labels + hist_labels, fontsize=7, framealpha=0.7)

        # Scatter: precursor intensity vs peak apex intensity
        for form, color in (("native", _native_color), ("labeled", _labeled_color)):
            pts = [(m["apex_intensity"], m["prec_intensity"]) for m in matched if m["form"] == form and m["apex_intensity"] is not None and m["apex_intensity"] > 0.0]
            if pts:
                axApex.scatter([p[0] for p in pts], [p[1] for p in pts], c=color, alpha=0.55, s=20, linewidths=0, label="M" if form == "native" else "M\u2032")
        axApex.set_xlabel("Peak apex intensity (log10)")
        axApex.set_ylabel("Precursor intensity (log10)")
        axApex.set_title("Precursor intensity vs. peak apex intensity")
        axApex.grid(True, alpha=0.2)
        apex_pts = [m for m in matched if m["apex_intensity"] is not None and m["apex_intensity"] > 0.0]
        if apex_pts:
            axApex.set_xscale("log")
            axApex.set_yscale("log")
            axApex.legend(fontsize=8, framealpha=0.7)
        else:
            axApex.text(0.5, 0.5, "No peak apex intensities available", ha="center", va="center", transform=axApex.transAxes, fontsize=9, color="gray")

        plot.fig.tight_layout()
        plot.canvas.draw()

        # MS/MS similarity histograms: one subplot per (native/labeled x filter-string
        # regex group 1) "type", pooling pairwise similarity scores computed only between
        # spectra belonging to the same feature (never across different features).
        if MATCHMS_AVAILABLE:
            algorithm_name, rel_intensity_pct, mz_tolerance = self._get_msms_similarity_settings()
            algorithm = self._get_msms_similarity_algorithm(algorithm_name, mz_tolerance)

            by_feature_type = defaultdict(list)  # (num, type_key) -> [scan, ...]
            for m in matched:
                type_key = (m["form"], m.get("filter_replacement"))
                by_feature_type[(m["num"], type_key)].append(m["scan"])

            pooled_scores_by_type = defaultdict(list)
            for (_num, type_key), scans in by_feature_type.items():
                if len(scans) < 2:
                    continue
                pooled_scores_by_type[type_key].extend(self._compute_pairwise_similarity_scores(scans, algorithm, rel_intensity_pct))

            type_keys = sorted(k for k, v in pooled_scores_by_type.items() if v)
            if type_keys:
                layout.addWidget(QtWidgets.QLabel(f"<b>MS/MS similarity ({algorithm_name}) per type (within-feature pairs, pooled)</b>"))
                n_types = len(type_keys)
                hist_cols = min(3, n_types)
                hist_rows = (n_types + hist_cols - 1) // hist_cols
                hist_plot = QtCore.QObject()
                hist_plot.fig = Figure((10.0, 3.2 * hist_rows), dpi=80, facecolor="white")
                hist_plot.canvas = FigureCanvas(hist_plot.fig)
                layout.addWidget(hist_plot.canvas, 1)
                for idx, type_key in enumerate(type_keys):
                    scores = pooled_scores_by_type[type_key]
                    ax = hist_plot.fig.add_subplot(hist_rows, hist_cols, idx + 1)
                    color = _labeled_color if type_key[0] == "labeled" else _native_color
                    ax.hist(scores, bins=np.linspace(0.0, 1.0, 21), color=color, alpha=0.75, edgecolor="black", linewidth=0.5)
                    stats = self._format_percentile_stats(scores)
                    ax.set_title(f"{self._format_msms_type_label(type_key)}\n{stats}", fontsize=8)
                    ax.set_xlabel("Similarity score")
                    ax.set_ylabel("Count")
                    ax.set_xlim(0.0, 1.0)
                    ax.grid(True, alpha=0.2)
                hist_plot.fig.tight_layout()
                hist_plot.canvas.draw()
                self._msms_overview_hist_plot = hist_plot  # keep a reference alive

        # Per-sample table
        layout.addWidget(QtWidgets.QLabel("<b>Per-sample MSMS spectra</b>"))
        table = QtWidgets.QTableWidget()
        table.setColumnCount(7)
        table.setHorizontalHeaderLabels(
            [
                "Sample",
                "Group",
                "Spectra obtained",
                "Assigned native (M)",
                "Assigned labeled (M\u2032)",
                "Features w/ MSMS native (M)",
                "Features w/ MSMS labeled (M\u2032)",
            ]
        )
        sample_keys = sorted(
            (k for k, st in per_sample_stats.items() if st["obtained"] > 0),
            key=lambda k: _sample_name(k).lower(),
        )
        table.setRowCount(len(sample_keys))
        for i, file_key in enumerate(sample_keys):
            st = per_sample_stats[file_key]
            table.setItem(i, 0, QtWidgets.QTableWidgetItem(_sample_name(file_key)))
            table.setItem(i, 1, QtWidgets.QTableWidgetItem(str(file_to_group.get(file_key, ""))))
            table.setItem(i, 2, QtWidgets.QTableWidgetItem(str(st["obtained"])))
            table.setItem(i, 3, QtWidgets.QTableWidgetItem(str(st["native"])))
            table.setItem(i, 4, QtWidgets.QTableWidgetItem(str(st["labeled"])))
            table.setItem(i, 5, QtWidgets.QTableWidgetItem(str(len(st["native_features"]))))
            table.setItem(i, 6, QtWidgets.QTableWidgetItem(str(len(st["labeled_features"]))))
        table.resizeColumnsToContents()
        table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        table.setMaximumHeight(200)
        layout.addWidget(table)

        btn_row = QtWidgets.QHBoxLayout()
        btn_row.addStretch(1)
        close_btn = QtWidgets.QPushButton("Close")
        close_btn.clicked.connect(dlg.accept)
        btn_row.addWidget(close_btn)
        outer_layout.addLayout(btn_row)

        dlg.exec()

    def _show_msms_filter_strings_list(self):
        """Show a scrollable list of all unique filter strings of the loaded MS/MS spectra."""
        filter_strings = {}
        if hasattr(self, "loadedMZXMLs") and self.loadedMZXMLs is not None:
            seen_ids = set()
            for key, mzxml_file in self.loadedMZXMLs.items():
                if id(mzxml_file) in seen_ids:
                    continue
                seen_ids.add(id(mzxml_file))
                if not (hasattr(mzxml_file, "MS2_list") and len(mzxml_file.MS2_list) > 0):
                    continue
                for ms2_scan in mzxml_file.MS2_list:
                    fs = self._get_msms_filter_string(ms2_scan)
                    if fs:
                        filter_strings[fs] = filter_strings.get(fs, 0) + 1

        if len(filter_strings) == 0:
            QtWidgets.QMessageBox.information(self, "Filter strings", "No MS/MS filter strings available. Load raw data first.")
            return

        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Loaded MS/MS filter strings")
        dlg.resize(720, 480)
        layout = QtWidgets.QVBoxLayout(dlg)
        layout.addWidget(QtWidgets.QLabel(f"{len(filter_strings)} unique filter string(s) found across the loaded MS/MS spectra:"))
        listw = QtWidgets.QListWidget()
        for fs in sorted(filter_strings.keys()):
            listw.addItem(f"[{filter_strings[fs]}x]  {fs}")
        layout.addWidget(listw, 1)
        btn_row = QtWidgets.QHBoxLayout()
        copy_btn = QtWidgets.QPushButton("Copy all")
        copy_btn.clicked.connect(lambda: QtWidgets.QApplication.clipboard().setText("\n".join(sorted(filter_strings.keys()))))
        btn_row.addWidget(copy_btn)
        btn_row.addStretch(1)
        close_btn = QtWidgets.QPushButton("Close")
        close_btn.clicked.connect(dlg.accept)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)
        dlg.exec()

    # </editor-fold>

    # <editor-fold desc="### general plotting functions">
    def addArrow(
        self,
        plt,
        point,
        at,
        plotIndex=0,
        fcColor="white",
        ecColor="white",
        arrowColor="slategrey",
        alpha=1,
        arrowAlpha=1,
        drawArrowHead=False,
        linewidth=1,
    ):
        if point != at:
            if not drawArrowHead:
                plt.twinxs[plotIndex].arrow(
                    at[0],
                    at[1],
                    point[0] - at[0],
                    point[1] - at[1],
                    color=ecColor,
                    shape="right",
                    head_width=(point[0] - at[0]) * 0.3,
                    head_length=(point[1] - at[1]) * 0.0,
                    linewidth=linewidth,
                )
            else:
                plt.twinxs[plotIndex].annotate(
                    "",
                    xy=point,
                    xytext=at,
                    xycoords="data",
                    textcoords="data",
                    va="center",
                    ha="center",
                    bbox=dict(boxstyle="round", fc=fcColor, ec=ecColor, alpha=alpha),
                    rotation=0,
                    arrowprops=dict(
                        arrowstyle="->",
                        connectionstyle="arc3",
                        color=arrowColor,
                        alpha=arrowAlpha,
                    ),
                )

    def addCircle(self, plt, at, plotIndex):
        pass

    def addAnnotation(
        self,
        plt,
        text,
        point,
        at,
        plotIndex=0,
        rotation=0,
        arrowAlpha=0.5,
        offset=(-10, 80),
        fcColor="white",
        ecColor="white",
        arrowColor="firebrick",
        alpha=0.8,
        up=True,
        add=80,
    ):
        if not up:
            add = -add
        plt.twinxs[plotIndex].annotate(
            text,
            xy=point,
            xytext=(-10, add),
            xycoords="data",
            textcoords="offset points",
            va="center",
            ha="center",
            bbox=dict(boxstyle="round", fc=fcColor, ec=ecColor, alpha=alpha),
            rotation=rotation,
            arrowprops=dict(arrowstyle="wedge", color=arrowColor, alpha=arrowAlpha),
        )

    def clearPlot(self, plt, setXtoZero=False):
        for ax in plt.twinxs:
            plt.fig.delaxes(ax)
        plt.axes = plt.fig.add_subplot(111)

        simpleaxis(plt.axes)
        if setXtoZero:
            plt.axes.spines["bottom"].set_position("zero")

        plt.axes.ticklabel_format(style="sci", axis="y", scilimits=(0, 0))
        plt.twinxs = [plt.axes]

    def drawPlot(
        self,
        plt,
        plotIndex=0,
        x=range(10),
        y=range(1, 11),
        fill=[],
        label="",
        b=None,
        rearrange=True,
        useCol=None,
        multipleLocator=5,
        alpha=globAlpha,
        title="",
        xlab="Retention time [minutes]",
        ylab="Intensity [counts]",
        plot=True,
        scatter=False,
        linestyle="-",
        gid=None,
        addDC=False,
    ):
        try:
            if b is None:
                b = predefinedColors

            if useCol is None:
                useCol = plotIndex

            if len(plt.twinxs) > 0 and not rearrange:
                plt.twinxs[0].get_ylim()
                plt.twinxs[0].get_xlim()

            if plotIndex == len(plt.twinxs):
                plt.twinxs.append(plt.axes.twinx())
            if plotIndex == 0:
                pass

            ax = plt.twinxs[plotIndex]

            if multipleLocator is not None:
                sf = ScalarFormatter(useOffset=True, useMathText=True)
                sf.set_scientific(True)
                sf.set_powerlimits((10, 0))
                ax.yaxis.set_major_formatter(sf)

            ax.set_title(title)
            ax.set_xlabel(xlab)
            ax.set_ylabel(ylab)

            plotCol = useCol
            if isinstance(useCol, int):
                plotCol = b[useCol % len(b)]

            if plot:
                ax.plot(x, y, color=plotCol, label=label, linestyle=linestyle, gid=gid)
                if addDC:
                    pass
                    # datacursor(line, formatter='{label}'.format)
                    # HighlightingDataCursor(line, formatter="".format)
            if scatter:
                ax.scatter(x, y, color=plotCol, gid=gid)
            if len(fill) > 0 and self.ui.plotMarkArea.checkState() == QtCore.Qt.Checked:
                ax.fill_between(
                    x[fill[0] : fill[1]],
                    y[fill[0] : fill[1]],
                    color=plotCol,
                    alpha=alpha,
                    gid=gid,
                )

        except Exception as ex:
            logging.warning("Exception caught", ex)

    def drawPoints(self, plt, x, y, plotIndex=0):
        ax = plt.twinxs[plotIndex]
        ax.scatter(x, y)

    def setLimts(self, plt, ylim, xlim):
        for ax in plt.twinxs:
            ax.set_ylim(ylim[0], ylim[1])
            ax.set_xlim(xlim[0], xlim[1])

    def drawCanvas(self, plt, ylim=None, xlim=None, showLegendOverwrite=True):
        # if ylim is None:
        #    ylim = plt.twinxs[0].get_ylim()
        # if xlim is None:
        #    xlim = plt.twinxs[0].get_xlim()

        for ax in plt.twinxs:
            if ylim is not None:
                if len(ylim) == 1:
                    ax.set_ylim(ylim[0])
                elif len(ylim) == 2:
                    ax.set_ylim(ylim[0], ylim[1])
            if xlim is not None:
                if len(xlim) == 1:
                    ax.set_xlim(xlim[0])
                else:
                    ax.set_xlim(xlim[0], xlim[1])
            if self.ui.showLegend.isChecked() and showLegendOverwrite:
                ax.legend()

        plt.canvas.draw()

    # </editor-fold>

    def multipleFilesPlotAdd(self):
        if self.ui.pl2A.pictureShown:
            self.clearPlot(self.ui.pl2A)
            self.ui.pl2A.pictureShown = False
            self.clearPlot(self.ui.pl2B)
            self.ui.pl2B.pictureShown = False

        maxInt = 0
        mzs = []
        peaks = []
        for item in self.ui.res_ExtractedData.selectedItems():
            if item.myType == "feature":
                if self.ui.pl2A.type is None or self.ui.pl2A.type == "feature":
                    self.ui.pl2A.type = "feature"
                else:
                    QtWidgets.QMessageBox.warning(
                        self,
                        "MetExtract",
                        "Selecting different result types in not supported. Please select only mz, mzbins or chromatographic peaks at a time",
                        QtWidgets.QMessageBox.Ok,
                    )
                    continue

                mzs.append(item.myData.id)
                peaks.append(item.myData.NPeakCenterMin / 60.0)
                xic = []
                times = []
                xics_df = self.currentOpenResultsFile.db_con.tables["XICs"].filter(pl.col("id") == item.myData.eicID)
                for row_dict in xics_df.to_dicts():
                    xic = [float(t) for t in row_dict["xic"].split(";")]
                    times = [float(t) / 60.0 for t in row_dict["times"].split(";")]
                    [float(t) for t in row_dict["xicL"].split(";")]
                self.ui.pl2A.xics.append(xic)
                self.ui.pl2A.times.append(times)
                self.ui.pl2A.peaks.append([item.myData])
            if item.myType == "MZs":
                if self.ui.pl2A.type is None or self.ui.pl2A.type == "MZs":
                    self.ui.pl2A.type = "MZs"
                else:
                    QtWidgets.QMessageBox.warning(
                        self,
                        "MetExtract",
                        "Selecting different result types is not supported. Please select only mz, mzbins or chromatographic peaks at a time",
                        QtWidgets.QMessageBox.Ok,
                    )
                    continue
                t = item

                x = []
                y = []
                for j in range(t.childCount()):
                    child = t.child(j)
                    assert child.myType == "mz"
                    x.append(child.myData[3] / 60.0)
                    y.append(child.myData[0])

                self.ui.pl2A.x_vals.append(x)
                self.ui.pl2A.y_vals.append(y)

        if self.ui.pl2A.type == "feature":
            for i in range(len(self.ui.pl2A.peaks)):
                maxInt = max(maxInt, max(self.ui.pl2A.xics[i]))
                maxindex, maxvalue = max(
                    enumerate(
                        xic[int(self.ui.pl2A.peaks[i][0].NPeakCenter - 3) : int(self.ui.pl2A.peaks[i][0].NPeakCenter + 3)],
                        start=int(self.ui.pl2A.peaks[i][0].NPeakCenter - 3),
                    ),
                    key=itemgetter(1),
                )
                if i == (len(self.ui.pl2A.peaks) - 1):
                    self.drawPlot(
                        self.ui.pl2A,
                        plotIndex=0,
                        x=self.ui.pl2A.times[i],
                        y=self.ui.pl2A.xics[i],
                        fill=[
                            int(self.ui.pl2A.peaks[i][0].NPeakCenter - self.ui.pl2A.peaks[i][0].NBorderLeft),
                            int(self.ui.pl2A.peaks[i][0].NPeakCenter + self.ui.pl2A.peaks[i][0].NBorderRight),
                        ],
                        useCol=i * 2,
                    )
                    if self.ui.plotAddLabels.checkState() == QtCore.Qt.Checked:
                        self.addAnnotation(
                            self.ui.pl2A,
                            "%s\n%.5f\n@ %.2f"
                            % (
                                self.ui.pl2A.peaks[i][0].assignedName,
                                self.ui.pl2A.peaks[i][0].mz,
                                self.ui.pl2A.peaks[i][0].NPeakCenterMin / 60.0,
                            ),
                            (
                                self.ui.pl2A.times[i][maxindex],
                                self.ui.pl2A.xics[i][maxindex],
                            ),
                            (
                                self.ui.pl2A.times[i][maxindex],
                                self.ui.pl2A.xics[i][maxindex],
                            ),
                            0,
                        )
            self.drawCanvas(self.ui.pl2A, ylim=[0, maxInt * 1.2])
        if self.ui.pl2A.type == "MZs":
            # matplotlib code taken from http://stackoverflow.com/questions/6508769/matplotlib-scatter-hist-with-stepfilled-histtype-in-histogram
            colour_LUT = [
                "#0000FF",
                "#FF0000",
                "#00FF00",
                "#FF00FF",
                "#00FFFF",
                "#FFFF00",
                "#FFFFFF",
            ]

            # the scatter plot:
            maxIntX = max([max(x) for x in self.ui.pl2A.x_vals])
            maxIntY = max([max(y) for y in self.ui.pl2A.y_vals])
            colors = []
            axScatter = self.ui.pl2A.axes

            for i in range(len(self.ui.pl2A.x_vals)):
                colour = colour_LUT[i % len(colour_LUT)]
                axScatter.scatter(self.ui.pl2A.x_vals[i], self.ui.pl2A.y_vals[i], color=colour)
                colors.append(colour)
            self.drawCanvas(self.ui.pl2A, ylim=[0, maxIntY], xlim=[0, maxIntX])

    def res_doubleClick(self, item):
        self.multipleFilesPlotAdd()

    def resChildNode(self, node):
        node.setHidden(False)
        for c in range(node.childCount()):
            self.resChildNode(node.child(c))

    def resetFilter(self):
        for i in range(self.ui.res_ExtractedData.topLevelItemCount()):
            d = self.ui.res_ExtractedData.topLevelItem(i)
            self.resChildNode(d)

    def filterHelp1_matchMZ(self, ref, akt, maxDiff, type="ppm"):
        ref = str(ref)
        if " " in ref:
            ref = ref[0 : ref.find(" ")]
        if "(" in ref:
            ref = ref[0 : ref.find("(")]
        ref = float(str(ref))
        akt = float(str(akt))
        maxDiff = float(str(maxDiff))
        if type == "ppm":
            if (abs(akt - ref) * 1000000.0 / akt) <= maxDiff:
                return True
        if type == "amu":
            if abs(akt - ref) <= maxDiff:
                return True
        return False

    def filterEdited(self, text):
        if self.currentOpenResultsFile is None:
            return

        text = str(text)
        self.resetFilter()

        if len(text) > 0:
            text = text.split(" ")
            textl = []
            for w in range(len(text)):
                textl.append(lambda x, _w=w: text[_w] in x)

            if "+-" in text[0]:
                h = text[0][0 : text[0].rfind("+")]
                textl[0] = lambda x: self.filterHelp1_matchMZ(x, h, 5.0)
                if "ppm" in text[0]:
                    ppm = text[0][(text[0].rfind("-") + 1) : text[0].find("p")]
                    textl[0] = lambda x: self.filterHelp1_matchMZ(x, h, ppm, "ppm")
                elif "a" in text[0]:
                    amu = text[0][(text[0].rfind("-") + 1) : text[0].find("a")]
                    textl[0] = lambda x: self.filterHelp1_matchMZ(x, h, amu, "amu")
                else:
                    amu = text[0][(text[0].rfind("-") + 1) :]
                    if len(amu) > 0:
                        textl[0] = lambda x: self.filterHelp1_matchMZ(x, h, amu, "amu")
                    else:
                        textl[0] = lambda x: self.filterHelp1_matchMZ(x, h, 1, "amu")

            # process first 3 result categories (mzs, bins, features)
            for i in range(3):
                d = self.ui.res_ExtractedData.topLevelItem(i)
                for c in range(d.childCount()):
                    dd = d.child(c)
                    show = True
                    for w in range(len(text)):
                        if len(text[w]) > 0 and not (textl[w](dd.text(w))):  # not(dd.text(w).contains(text[w])):
                            show = False
                    dd.setHidden(not show)

            # process feature groups (perform search on features within groups)
            d = self.ui.res_ExtractedData.topLevelItem(3)
            for v in range(d.childCount()):
                allShow = False
                fg = d.child(v)
                for c in range(fg.childCount()):
                    dd = fg.child(c)
                    show = True
                    for w in range(len(text)):
                        if len(text[w]) > 0 and not (textl[w](dd.text(w))):  # not(dd.text(w).contains(text[w])):
                            show = False
                    dd.setHidden(not show)
                    allShow = allShow or show
                fg.setHidden(not allShow)

    def _onExpFilterToggle(self, checked):
        self.ui.expFilterContent.setVisible(checked)
        self.ui.expFilterToggleBtn.setText("Hide filters \u25b4" if checked else "Show filters \u25be")
        if not checked:
            # Clear all filter fields when collapsing
            self.ui.expFilter_mz.clear()
            self.ui.expFilter_rt.clear()
            self.ui.expFilter_xn.clear()
            self.ui.expFilter_z.clear()
            self.ui.expFilter_polarity.setCurrentIndex(0)
            self.ui.expFilter_ms2.setCurrentIndex(0)

    def _onExpFilterReset(self):
        """Reset all filter fields to default values."""
        self.ui.expFilter_mz.clear()
        self.ui.expFilter_rt.clear()
        self.ui.expFilter_xn.clear()
        self.ui.expFilter_z.clear()
        self.ui.expFilter_polarity.setCurrentIndex(0)
        self.ui.expFilter_ms2.setCurrentIndex(0)
        # Show filter panel when resetting
        if not self.ui.expFilterToggleBtn.isChecked():
            self.ui.expFilterToggleBtn.setChecked(True)

    @staticmethod
    def _parse_filter_range(text):
        """Parse a filter text into a (value_contains, min_val, max_val) tuple.
        Handles flexible range syntax:
        - 'a-b', 'a - b', 'a- b', 'a -b' → (None, a, b)
        - '-b', '- b' → (None, -inf, b)  [up to b]
        - 'a-', 'a -' → (None, a, inf)  [from a onward]
        - Plain text → (text, None, None) for substring matching
        """
        text = text.strip()
        if not text:
            return None, None, None

        # Try to detect a range pattern (flexible spacing around dash)
        if "-" in text:
            # Split on the dash, handling cases like "a-b", "a - b", "a- b", etc.
            parts = text.split("-", 1)
            left = parts[0].strip()
            right = parts[1].strip() if len(parts) > 1 else ""

            left_val = None
            right_val = None

            # Try to parse left side
            if left:
                try:
                    left_val = float(left)
                except ValueError:
                    # Not a number, treat as plain substring match
                    return text, None, None

            # Try to parse right side
            if right:
                try:
                    right_val = float(right)
                except ValueError:
                    # Not a number, treat as plain substring match
                    return text, None, None

            # If we got here, at least one side parsed as a number
            if left_val is not None and right_val is not None:
                return None, left_val, right_val
            elif left_val is not None:
                # Only left (e.g., "100-") → from 100 to infinity
                return None, left_val, float("inf")
            elif right_val is not None:
                # Only right (e.g., "-100") → from -infinity to 100
                return None, float("-inf"), right_val

        # Plain text for substring matching
        return text, None, None

    def expFilterEdited(self, *args):
        tree = self.ui.resultsExperiment_TreeWidget

        # Gather filter values
        mz_text = self.ui.expFilter_mz.text().strip()
        rt_text = self.ui.expFilter_rt.text().strip()
        xn_text = self.ui.expFilter_xn.text().strip()
        z_text = self.ui.expFilter_z.text().strip()
        polarity_idx = self.ui.expFilter_polarity.currentIndex()  # 0=both, 1=pos, 2=neg
        ms2_idx = self.ui.expFilter_ms2.currentIndex()
        # 0=all, 1=without MS2, 2=with MS2, 3=with native MS2, 4=with labeled MS2

        # Handle MS2 filter change: if user selects non-"all" option, query MS2 forms from files
        prev_ms2_idx = getattr(self, "_prev_ms2_filter_idx", 0)
        if ms2_idx != 0 and ms2_idx != prev_ms2_idx:
            # User wants to filter by MS2 content; query all files
            msms_forms = self._build_msms_feature_forms()
        else:
            # Use cached data from updateMSMSList_exp or empty set
            msms_forms = getattr(self, "_exp_msms_feature_forms", {})
        self._prev_ms2_filter_idx = ms2_idx

        no_filters = not any([mz_text, rt_text, xn_text, z_text, polarity_idx != 0, ms2_idx != 0])

        mz_sub, mz_min, mz_max = self._parse_filter_range(mz_text)
        rt_sub, rt_min, rt_max = self._parse_filter_range(rt_text)
        xn_sub, xn_min, xn_max = self._parse_filter_range(xn_text)

        for i in range(tree.topLevelItemCount()):
            top = tree.topLevelItem(i)
            if no_filters:
                top.setHidden(False)
                for c in range(top.childCount()):
                    top.child(c).setHidden(False)
                continue

            any_child_shown = False
            for c in range(top.childCount()):
                child = top.child(c)
                bd = getattr(child, "bunchData", None)
                if bd is None:
                    child.setHidden(False)
                    any_child_shown = True
                    continue

                show = True

                # m/z filter
                if mz_text and show:
                    mz_val = getattr(bd, "mz", None)
                    if mz_val is not None:
                        mz_str = "%.4f" % mz_val
                        if mz_min is not None and mz_max is not None:
                            show = mz_min <= mz_val <= mz_max
                        elif mz_sub:
                            show = mz_sub in mz_str

                # RT filter
                if rt_text and show:
                    rt_val = getattr(bd, "rt", None)
                    if rt_val is not None:
                        rt_min_val = rt_val / 60.0
                        rt_str = "%.2f" % rt_min_val
                        if rt_min is not None and rt_max is not None:
                            show = rt_min <= rt_min_val <= rt_max
                        elif rt_sub:
                            show = rt_sub in rt_str

                # Xn filter
                if xn_text and show:
                    xn_val = getattr(bd, "xn", None)
                    if xn_val is not None:
                        xn_str = str(xn_val)
                        if xn_min is not None and xn_max is not None:
                            try:
                                show = xn_min <= float(xn_val) <= xn_max
                            except (TypeError, ValueError):
                                show = False
                        elif xn_sub:
                            show = xn_sub in xn_str

                # Z (charge state) filter
                if z_text and show:
                    z_val = getattr(bd, "charge", None)
                    if z_val is not None:
                        show = z_text in str(z_val)

                # Polarity filter
                if polarity_idx != 0 and show:
                    ion_mode = str(getattr(bd, "ionisationMode", "") or "")
                    if polarity_idx == 1:
                        show = "+" in ion_mode
                    elif polarity_idx == 2:
                        show = "-" in ion_mode

                # MS2 filter
                if ms2_idx != 0 and show:
                    feature_num = getattr(bd, "id", None)
                    forms = msms_forms.get(feature_num, set())
                    if ms2_idx == 1:
                        show = len(forms) == 0
                    elif ms2_idx == 2:
                        show = len(forms) > 0
                    elif ms2_idx == 3:
                        show = "native" in forms
                    elif ms2_idx == 4:
                        show = "labeled" in forms

                child.setHidden(not show)
                if show:
                    any_child_shown = True

            top.setHidden(not any_child_shown)

        # Sync feature map if it is open
        if self.ui.expFeatureMapContainer.isVisible():
            self._buildFeatureMap()

    def setChromPeakName(self):
        cpName = str(self.ui.chromPeakName.text())

        selectedItems = self.ui.res_ExtractedData.selectedItems()

        for item in selectedItems:
            if item.myType == "Features" or item.myType == "feature":
                self.currentOpenResultsFile.curs.execute("UPDATE chrompeaks SET assignedName='%s' WHERE id=%d" % (cpName, item.myID))
                self.currentOpenResultsFile.conn.commit()
                item.setText(0, cpName)
            if item.myType == "FeatureGroup" or item.myType == "featureGroup":
                self.currentOpenResultsFile.curs.execute("UPDATE featureGroups SET featureName='%s' WHERE id=%d" % (cpName, item.myID))
                self.currentOpenResultsFile.conn.commit()
                item.myData = cpName
                item.setText(0, cpName)

    def openRawFile(self):
        if self.currentOpenResultsFile.file is not None:
            os.startfile(self.currentOpenResultsFile.file)

    def removeChromPeaksFrom(self, item, idToRem):
        rem = 0
        todel = set()

        for j in range(item.childCount()):
            child = item.child(j)
            if child.myType == "feature" and child.myID == idToRem:
                todel.add(j)
        todel = list(todel)
        for j in sorted(todel, reverse=True):
            item.removeChild(item.child(j))
            rem += 1

        return rem

    def sortTreeChildren(self, node, sortColumn):
        children = []
        for ci in range(node.childCount() - 1, -1, -1):
            child = node.child(ci)
            children.append(child)
            node.removeChild(child)
        children = natSort(children, key=lambda x: str(x.text(sortColumn)))
        for child in children:
            node.addChild(child)

    def delNodeFromResults(self, item):
        chromPeaksRem = 0
        featureGroupsRem = 0
        if item.myType == "feature":
            cp = item.myData
            myID = item.myID
            self.currentOpenResultsFile.curs.execute("delete from ChromPeaks where id=%d" % cp.id)
            self.currentOpenResultsFile.curs.execute("delete from featureGroupFeatures where fID=%d" % cp.id)
            self.currentOpenResultsFile.curs.execute("delete from featureGroups where id not in (select distinct fGroupID from featureGroupFeatures)")
            self.ui.res_ExtractedData.setItemSelected(item, False)

            chromPeaksRem += self.removeChromPeaksFrom(self.ui.res_ExtractedData.topLevelItem(2), myID)

            todel = set()
            for i in range(self.ui.res_ExtractedData.topLevelItem(3).childCount()):
                featureGroup = self.ui.res_ExtractedData.topLevelItem(3).child(i)
                assert featureGroup.myType == "featureGroup"
                self.removeChromPeaksFrom(featureGroup, myID)
                featureGroup.setText(1, "%d" % featureGroup.childCount())
                if featureGroup.childCount() == 0:
                    todel.add(i)
            todel = list(todel)
            for i in sorted(todel, reverse=True):
                self.ui.res_ExtractedData.topLevelItem(3).removeChild(self.ui.res_ExtractedData.topLevelItem(3).child(i))
                featureGroupsRem += 1
        elif item.myType == "featureGroup":
            myID = item.myID
            partChromPeaks = []
            for j in range(item.childCount()):
                partChromPeaks.append(item.child(j))
            for child in partChromPeaks:
                z, u = self.delNodeFromResults(child)
                chromPeaksRem += z
                featureGroupsRem = featureGroupsRem + u

        self.ui.res_ExtractedData.topLevelItem(2).setText(1, "%d" % self.ui.res_ExtractedData.topLevelItem(2).childCount())
        self.ui.res_ExtractedData.topLevelItem(3).setText(1, "%d" % self.ui.res_ExtractedData.topLevelItem(3).childCount())

        return chromPeaksRem, featureGroupsRem

    def createGroupFrom(self, items, tracerID):
        fIDs = []
        for item in items:
            fIDs.append(item.myID)

        i = 0
        while True:
            i += 1
            found = 0
            fg_df = self.currentOpenResultsFile.db_con.tables["featureGroups"].filter(pl.col("featureName") == ("fg_%d" % i))
            for row_dict in fg_df.to_dicts():
                found += 1
            if found == 0:
                break

        SQLInsert(
            self.currentOpenResultsFile.curs,
            "featureGroups",
            featureName="fgU_%d" % i,
            tracer=tracerID,
        )
        found = 0
        newID = -1
        newName = ""
        tracerName = ""
        fg_df = self.currentOpenResultsFile.db_con.tables["featureGroups"].join(self.currentOpenResultsFile.db_con.tables["tracerConfiguration"], left_on="tracer", right_on="id", how="inner").filter(pl.col("featureName") == ("fgU_%d" % i))
        for row_dict in fg_df.to_dicts():
            found += 1
            newID = int(row_dict["id"])
            newName = str(row_dict["featureName"])
            tracerName = str(row_dict["name"])

        d = QtWidgets.QTreeWidgetItem([str(newName), str(len(items)), "", str(tracerName)])
        d.myType = "featureGroup"
        d.myID = newID
        d.myData = Bunch(fgID=newID, featureName=newName, tracerName=str(tracerName))
        self.ui.res_ExtractedData.topLevelItem(3).addChild(d)

        self.currentOpenResultsFile.curs.execute("update featureGroupFeatures set fGroupID=%d where fID in (%s)" % (newID, ", ".join([str(fId) for fId in fIDs])))
        sumRt = 0
        cpCount = 0
        for item in items:
            parent = item.parent()
            # if len(d.myData) == 3:
            #    d.myData.append(parent.myData[3])
            parent.removeChild(item)
            parent.setText(1, "%d" % parent.childCount())
            d.addChild(item)
            xp = item.myData
            sumRt = sumRt + xp.NPeakCenterMin
            cpCount += 1

        d.setText(1, "%d" % d.childCount())
        d.setText(2, "%.2f" % (sumRt / cpCount / 60.0))
        self.ui.res_ExtractedData.topLevelItem(3).setText(1, "%d" % self.ui.res_ExtractedData.topLevelItem(3).childCount())

        self.currentOpenResultsFile.curs.execute("delete from featureGroups where id not in (select distinct fGroupID from featureGroupFeatures)")

        self.currentOpenResultsFile.conn.commit()

        for ci in range(self.ui.res_ExtractedData.topLevelItem(3).childCount() - 1, -1, -1):
            child = self.ui.res_ExtractedData.topLevelItem(3).child(ci)
            if child.childCount() == 0:
                self.ui.res_ExtractedData.topLevelItem(3).removeChild(child)

        self.sortTreeChildren(self.ui.res_ExtractedData.topLevelItem(3), 2)

        return newName

    def _featureValuesFromSampleItem(self, item) -> dict:
        """Build the copy-value dict for a sample-results tree item (feature/featureGroup)."""
        values = {"ogroup": None, "num": None, "polarity": None, "rt": None, "mz": None, "lmz": None}
        myType = getattr(item, "myType", None)
        if myType == "feature":
            xp = getattr(item, "myData", None)
            values["num"] = getattr(item, "myID", None)
            if xp is not None:
                values["polarity"] = getattr(xp, "ionMode", None)
                mz = getattr(xp, "mz", None)
                if mz is not None:
                    values["mz"] = "%.4f" % float(mz)
                lmz = getattr(xp, "lmz", None)
                if lmz is not None:
                    values["lmz"] = "%.4f" % float(lmz)
                rt = getattr(xp, "NPeakCenterMin", None)
                if rt is not None:
                    values["rt"] = "%.3f" % (float(rt) / 60.0)
            parent = item.parent()
            if parent is not None and getattr(parent, "myType", None) == "featureGroup":
                values["ogroup"] = getattr(getattr(parent, "myData", None), "fgID", None)
        elif myType == "featureGroup":
            values["ogroup"] = getattr(getattr(item, "myData", None), "fgID", None)
            mzs, lmzs, rts, modes = [], [], [], set()
            for j in range(item.childCount()):
                xp = getattr(item.child(j), "myData", None)
                if xp is None:
                    continue
                if getattr(xp, "mz", None) is not None:
                    mzs.append(float(xp.mz))
                if getattr(xp, "lmz", None) is not None:
                    lmzs.append(float(xp.lmz))
                if getattr(xp, "NPeakCenterMin", None) is not None:
                    rts.append(float(xp.NPeakCenterMin) / 60.0)
                if getattr(xp, "ionMode", None) is not None:
                    modes.add(xp.ionMode)
            if mzs:
                values["mz"] = "%.4f" % (sum(mzs) / len(mzs))
            if lmzs:
                values["lmz"] = "%.4f" % (sum(lmzs) / len(lmzs))
            if rts:
                values["rt"] = "%.3f" % (sum(rts) / len(rts))
            if len(modes) == 1:
                values["polarity"] = next(iter(modes))
        return values

    def showPopup(self, position):
        selectedItems = self.ui.res_ExtractedData.selectedItems()

        types = set()
        parentTypes = set()
        parentItems = []

        for item in selectedItems:
            if hasattr(item, "myType"):
                types.add(item.myType)
                if item.parent() is not None:
                    parentItems.append(item.parent())
                    if hasattr(item, "myType"):
                        parentTypes.add(item.parent().myType)
            else:
                types.add("generic item")
        types = list(types)

        menu = QtWidgets.QMenu()
        actionAvailable = False
        tracerActions = []

        newGroupAction = -1
        if all([si == "feature" for si in types]) and all([pt == "featureGroup" for pt in parentTypes]):
            newGroupAction = menu.addMenu("Extract to new Group")
            actionAvailable = True
            tracer_df = self.currentOpenResultsFile.db_con.tables["tracerConfiguration"]
            for row_dict in tracer_df.to_dicts():
                tracerID = int(row_dict["id"])
                tracerName = str(row_dict["name"])
                tracerAction = newGroupAction.addAction(tracerName)
                tracerAction.tracerID = tracerID
                tracerActions.append(tracerAction)

        deleteAction = -1
        menu.addSeparator()

        if len(types) == 1 and (types[0] == "feature" or types[0] == "featureGroup"):
            deleteAction = menu.addAction("Delete")
            actionAvailable = True

        clipboardAction = -1
        menu.addSeparator()
        if len(types) == 1 and (types[0] == "feature" or types[0] == "featureGroup" or types[0] == "Features" or types[0] == "Feature Groups"):
            clipboardAction = menu.addAction("Copy")
            actionAvailable = True

        # Copy individual feature fields (OGroup, Feature-Num, polarity, mean RT, mean m/z)
        copyFieldActions = {}
        if len(selectedItems) == 1 and len(types) == 1 and types[0] in ("feature", "featureGroup"):
            menu.addSeparator()
            field_values = self._featureValuesFromSampleItem(selectedItems[0])
            copyFieldActions = self._addFeatureCopyMenu(menu, field_values)
            if copyFieldActions:
                actionAvailable = True

        if actionAvailable:
            action = menu.exec_(self.ui.res_ExtractedData.mapToGlobal(position))

            if action in copyFieldActions:
                pyperclip.copy(copyFieldActions[action])

            elif action == clipboardAction:
                clipboard = ["MZ\tRT\tXn\tZ\tIonMode\tFG\tTracer\tAdduct\tHeteroAtoms"]

                if len(selectedItems) == 1 and types[0] == "Features":
                    t = []
                    for j in range(selectedItems[0].childCount()):
                        child = item.child(j)
                        t.append(child)
                    selectedItems = t
                    types = ["feature"]

                if len(selectedItems) == 1 and types[0] == "Feature Groups":
                    t = []
                    for j in range(selectedItems[0].childCount()):
                        child = item.child(j)
                        t.append(child)
                    selectedItems = t
                    types = ["featureGroup"]

                if types[0] == "feature":
                    for item in selectedItems:
                        clipboard.append(
                            "%f\t%.2f\t%s\t%d\t%s\t%s\t%s\t%s\t%s"
                            % (
                                item.myData.mz,
                                item.myData.NPeakCenterMin / 60.0,
                                item.myData.xCount,
                                item.myData.loading,
                                item.myData.ionMode,
                                "-",
                                item.myData.tracer,
                                item.myData.adducts,
                                item.myData.heteroAtoms,
                            )
                        )
                if types[0] == "featureGroup":
                    for item in selectedItems:
                        for j in range(item.childCount()):
                            child = item.child(j)
                            clipboard.append(
                                "%f\t%.2f\t%d\t%d\t%s\t%s\t%s\t%s\t%s"
                                % (
                                    child.myData.mz,
                                    child.myData.NPeakCenterMin / 60.0,
                                    child.myData.xCount,
                                    child.myData.loading,
                                    child.myData.ionMode,
                                    item.myData.fgID,
                                    child.myData.tracer,
                                    child.myData.adducts,
                                    child.myData.heteroAtoms,
                                )
                            )

                pyperclip.copy("\n".join(clipboard))

            elif action == deleteAction:
                if (
                    QtWidgets.QMessageBox.question(
                        self,
                        "MetExtract",
                        "Are you sure you want to delete the selected result(s)?\nThis action cannot be undone",
                        QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    )
                    == QtWidgets.QMessageBox.Yes
                ):
                    pw = ProgressWrapper(1)
                    pw.getCallingFunction()("max")(0)
                    pw.getCallingFunction()("header")("working..")
                    pw.show()

                    featureGroupsRem = 0
                    chromPeaksRem = 0
                    for item in selectedItems:
                        z, u = self.delNodeFromResults(item)
                        chromPeaksRem += z
                        featureGroupsRem += u

                    self.currentOpenResultsFile.conn.commit()

                    pw.hide()

                    QtWidgets.QMessageBox.information(
                        self,
                        "MetExtract",
                        "%d features and %d feature groups have been deleted" % (chromPeaksRem, featureGroupsRem),
                        QtWidgets.QMessageBox.Ok,
                    )

            elif action in tracerActions:
                if (
                    QtWidgets.QMessageBox.question(
                        self,
                        "MetExtract",
                        "Are you sure you want to extract the selected features in a new group?\nThis action cannot be undone",
                        QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    )
                    == QtWidgets.QMessageBox.Yes
                ):
                    pw = ProgressWrapper(1)
                    pw.getCallingFunction()("max")(0)
                    pw.getCallingFunction()("header")("working..")
                    pw.show()

                    for item in selectedItems:
                        assert item.myType == "feature"
                    groupName = self.createGroupFrom(selectedItems, action.tracerID)

                    pw.hide()

                    QtWidgets.QMessageBox.information(
                        self,
                        "MetExtract",
                        "Created group %s with %d features" % (groupName, len(selectedItems)),
                        QtWidgets.QMessageBox.Ok,
                    )

    def loadAllSamples(self, selectedMZs=None, ppm=25.0):
        from .utilities import RunImapUnordered

        intensityThreshold = float(
            QtWidgets.QInputDialog.getDouble(
                self,
                "Intensity threshold",
                "Please enter the minimal threshold used for importing the data",
                10000,
                0,
                1000000,
                0,
            )[0]
        )

        definedGroups = self.getAllSampleGroups()
        self.loadedMZXMLs = {}

        filesToLoad = []
        for group in definedGroups:
            for i in range(len(group.files)):
                fi = str(group.files[i]).replace("\\", "/")
                filesToLoad.append(
                    {
                        "File": fi,
                        "Group": group.name,
                        "IntensityThreshold": intensityThreshold,
                        "selectedMZs": selectedMZs,
                        "ppm": ppm,
                    }
                )

        startTime = time.time()
        startMemory = 0
        try:
            startMemory = memory_usage_psutil()
        except Exception:
            pass
        res = RunImapUnordered.runImapUnordered(
            loadMZXMLFile,
            filesToLoad,
            processes=int(min(cpu_count() / 2, self.ui.cpuCores.value())),
            pw="createNewPW",
        )
        duration = time.time() - startTime
        usedMemory = -1
        try:
            usedMemory = memory_usage_psutil() - startMemory
        except Exception:
            pass

        for x in res:
            self.loadedMZXMLs[x["File"]] = x["mzXMLFile"]
            self.loadedMZXMLs[x["Group"]] = x["mzXMLFile"]

        logging.info(
            "Loading (%s, intensity threshold %d counts) took %.1f minutes and used %s of Memory"
            % (
                "entire chromatograms" if selectedMZs is None else "only detected feature pairs",
                intensityThreshold,
                duration / 60.0,
                "%.1f MB" % usedMemory if usedMemory < 1024 else "%.1f GB" % (usedMemory / 1024),
            )
        )

        # Update the per-feature MS2 spectra counts column now that raw data is loaded
        try:
            self._updateExperimentMSMSCountsColumn()
        except Exception:
            logging.exception("Could not update MS2 spectra counts column")

    def showCustomFeature(self):
        self.resultsExperimentChangedNew(askForFeature=True)

    def resultsExperimentChangedNew(self, askForFeature=False):
        if not askForFeature and len(self.ui.resultsExperiment_TreeWidget.selectedItems()) == 0:
            return

        if self.loadedMZXMLs is None:
            selectedMZs = None
            if (
                QtWidgets.QMessageBox.question(
                    self,
                    "MetExtract",
                    "Do you want to load the entire chromatograms (Yes) or just the m/z values of the detected results (No)\n"
                    + "If you don't have a lot of main memory (RAM) then the entire chromatograms should not be loaded.\n"
                    + "Please not that if you don't load the entire chromatograms, you won't be able to show a custom feature.",
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                )
                == QtWidgets.QMessageBox.No
            ):
                selectedMZs = []
                mzs = []
                for itemI in range(self.ui.resultsExperiment_TreeWidget.topLevelItemCount()):
                    item = self.ui.resultsExperiment_TreeWidget.topLevelItem(itemI)
                    if item.bunchData.type == "metaboliteGroup":
                        for i in range(item.childCount()):
                            child = item.child(i)
                            if child.bunchData.type == "featurePair":
                                mzs.append(child.bunchData.mz)
                                mzs.append(child.bunchData.lmz)
                    if item.bunchData.type == "featurePair":
                        mzs.append(item.bunchData.mz)
                        mzs.append(item.bunchData.lmz)
                for mz in mzs:
                    selectedMZs.append(mz)

            self.loadAllSamples(selectedMZs=selectedMZs, ppm=25.0)

        definedGroups = self.getAllSampleGroups()

        self.clearPlot(self.ui.resultsExperiment_plot)
        self.clearPlot(self.ui.resultsExperimentSeparatedPeaks_plot)
        self.clearPlot(self.ui.resultsExperimentMSScanPeaks_plot)

        plotItems = []

        if askForFeature:
            availableFilterLines = set()
            for grpInd, group in enumerate(definedGroups):
                for i in range(len(group.files)):
                    fi = str(group.files[i]).replace("\\", "/")
                    for fl in self.loadedMZXMLs[fi].getFilterLines(
                        includeMS1=True,
                        includeMS2=False,
                        includePosPolarity=True,
                        includeNegPolarity=True,
                    ):
                        availableFilterLines.add(fl)

            # --- Load saved custom features from JSON ---
            import json as _json

            _custom_features_path = os.path.join(local_folder, "custom_features.json")

            def _load_saved_features():
                try:
                    if os.path.exists(_custom_features_path):
                        with open(_custom_features_path, "r", encoding="utf-8") as _f:
                            return _json.load(_f)
                except Exception:
                    pass
                return []

            def _save_features(features):
                try:
                    with open(_custom_features_path, "w", encoding="utf-8") as _f:
                        _json.dump(features, _f, indent=2)
                except Exception as ex:
                    logging.warning(f"Could not save custom features: {ex}")

            saved_features = _load_saved_features()

            dlg = QtWidgets.QDialog(self)
            dlg.setWindowTitle("Custom feature")
            dlg.setMinimumWidth(450)
            form = QtWidgets.QFormLayout(dlg)

            # --- Saved compounds section ---
            saved_layout = QtWidgets.QHBoxLayout()
            saved_combo = QtWidgets.QComboBox()
            saved_combo.setMinimumWidth(200)
            saved_combo.addItem("-- select saved compound --")
            for sf in saved_features:
                saved_combo.addItem(sf.get("name", ""))
            saved_layout.addWidget(saved_combo)
            del_saved_btn = QtWidgets.QPushButton("Delete")
            saved_layout.addWidget(del_saved_btn)
            form.addRow("Saved compounds:", saved_layout)

            sep0 = QtWidgets.QFrame()
            sep0.setFrameShape(QtWidgets.QFrame.HLine)
            form.addRow(sep0)

            # --- Name field ---
            name_edit = QtWidgets.QLineEdit()
            name_edit.setPlaceholderText("Optional – set to save this feature")
            form.addRow("Name:", name_edit)

            # --- Sum formula helper ---
            formula_layout = QtWidgets.QHBoxLayout()
            native_formula_edit = QtWidgets.QLineEdit()
            native_formula_edit.setPlaceholderText("e.g. C15H20O6")
            labeled_formula_edit = QtWidgets.QLineEdit()
            labeled_formula_edit.setPlaceholderText("e.g. C15H20O6 (labeled)")
            formula_layout.addWidget(native_formula_edit)
            formula_layout.addWidget(QtWidgets.QLabel("→ labeled:"))
            formula_layout.addWidget(labeled_formula_edit)
            form.addRow("Sum formula (N / L):", formula_layout)

            adduct_combo = QtWidgets.QComboBox()
            _adduct_list = list(self.adducts) if hasattr(self, "adducts") and self.adducts else []
            adduct_combo.addItems([a.name for a in _adduct_list])
            form.addRow("Adduct:", adduct_combo)

            sep = QtWidgets.QFrame()
            sep.setFrameShape(QtWidgets.QFrame.HLine)
            form.addRow(sep)

            mz_spin = QtWidgets.QDoubleSpinBox()
            mz_spin.setDecimals(8)
            mz_spin.setRange(0, 100000)
            mz_spin.setValue(0)
            form.addRow("Native m/z:", mz_spin)

            lmz_spin = QtWidgets.QDoubleSpinBox()
            lmz_spin.setDecimals(8)
            lmz_spin.setRange(0, 100000)
            lmz_spin.setValue(0)
            form.addRow("Labeled m/z:", lmz_spin)

            def _calc_mz_from_formula():
                try:
                    fT = formulaTools()
                    idx = adduct_combo.currentIndex()
                    if idx < 0 or idx >= len(_adduct_list):
                        return
                    adduct = _adduct_list[idx]
                    nf = native_formula_edit.text().strip()
                    lf = labeled_formula_edit.text().strip()
                    if nf:
                        elems_n = fT.parseFormula(nf)
                        mass_n = fT.calcMolWeight(elems_n)
                        mz_n = (mass_n * adduct.mCount + adduct.mzoffset) / adduct.charge
                        mz_spin.setValue(mz_n)
                    if lf:
                        elems_l = fT.parseFormula(lf)
                        mass_l = fT.calcMolWeight(elems_l)
                        mz_l = (mass_l * adduct.mCount + adduct.mzoffset) / adduct.charge
                        lmz_spin.setValue(mz_l)
                except Exception as ex:
                    QtWidgets.QMessageBox.warning(dlg, "Formula error", str(ex))

            native_formula_edit.textEdited.connect(_calc_mz_from_formula)
            labeled_formula_edit.textEdited.connect(_calc_mz_from_formula)
            adduct_combo.currentIndexChanged.connect(_calc_mz_from_formula)

            rt_spin = QtWidgets.QDoubleSpinBox()
            rt_spin.setDecimals(3)
            rt_spin.setRange(0, 10000)
            rt_spin.setValue(0)
            rt_spin.setSuffix(" min")
            form.addRow("Retention time:", rt_spin)

            fl_combo = QtWidgets.QComboBox()
            fl_combo.addItems(sorted(availableFilterLines))
            form.addRow("Filter line:", fl_combo)

            def _populate_from_saved(index):
                """Fill form fields when a saved compound is selected."""
                if index <= 0 or index - 1 >= len(saved_features):
                    return
                sf = saved_features[index - 1]
                name_edit.setText(sf.get("name", ""))
                native_formula_edit.setText(sf.get("native_formula", ""))
                labeled_formula_edit.setText(sf.get("labeled_formula", ""))
                mz_spin.setValue(sf.get("mz", 0.0))
                lmz_spin.setValue(sf.get("lmz", 0.0))
                rt_spin.setValue(sf.get("rt", 0.0))
                fl_val = sf.get("filter_line", "")
                idx_fl = fl_combo.findText(fl_val)
                if idx_fl >= 0:
                    fl_combo.setCurrentIndex(idx_fl)
                adduct_name = sf.get("adduct", "")
                idx_ad = adduct_combo.findText(adduct_name)
                if idx_ad >= 0:
                    adduct_combo.setCurrentIndex(idx_ad)

            def _delete_saved():
                idx = saved_combo.currentIndex()
                if idx <= 0 or idx - 1 >= len(saved_features):
                    return
                saved_features.pop(idx - 1)
                _save_features(saved_features)
                saved_combo.removeItem(idx)

            saved_combo.currentIndexChanged.connect(_populate_from_saved)
            del_saved_btn.clicked.connect(_delete_saved)

            btn_box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
            btn_box.accepted.connect(dlg.accept)
            btn_box.rejected.connect(dlg.reject)
            form.addRow(btn_box)

            if dlg.exec() != QtWidgets.QDialog.Accepted:
                return

            mz = mz_spin.value()
            lmz = lmz_spin.value()
            rt = rt_spin.value()
            fl = fl_combo.currentText()

            # Save if a name was provided
            compound_name = name_edit.text().strip()
            if compound_name:
                # Update existing entry or append new one
                existing = next((sf for sf in saved_features if sf.get("name") == compound_name), None)
                entry = {
                    "name": compound_name,
                    "native_formula": native_formula_edit.text().strip(),
                    "labeled_formula": labeled_formula_edit.text().strip(),
                    "adduct": adduct_combo.currentText(),
                    "mz": mz,
                    "lmz": lmz,
                    "rt": rt,
                    "filter_line": fl,
                }
                if existing is not None:
                    existing.update(entry)
                else:
                    saved_features.append(entry)
                _save_features(saved_features)

            plotItems.append(Bunch(mz=mz, lmz=lmz, rt=rt * 60.0, scanEvent=fl))
        else:
            for item in self.ui.resultsExperiment_TreeWidget.selectedItems():
                if item.bunchData.type == "metaboliteGroup":
                    for i in range(item.childCount()):
                        child = item.child(i)
                        if child.bunchData.type == "featurePair":
                            plotItems.append(child.bunchData)
                if item.bunchData.type == "featurePair":
                    plotItems.append(item.bunchData)

        ppm = self.ui.doubleSpinBox_resultsExperiment_EICppm.value()
        borderOffset = self.ui.doubleSpinBox_resultsExperiment_PeakWidth.value()
        meanRT = []
        intlim = [0, 0]

        mostAbundantFile = None

        all = 0
        for grpInd, group in enumerate(definedGroups):
            for i in range(len(group.files)):
                all = all + 1

        pw = ProgressWrapper(1, parent=self, showIndProgress=False)
        pw.show()
        pw.getCallingFunction()("max")(len(plotItems) * all)
        pw.getCallingFunction()("value")(0)

        done = 0
        for h, pi in enumerate(plotItems):
            meanRT.append(pi.rt)

            rtBorderMin = pi.rt / 60.0 - borderOffset
            rtBorderMax = pi.rt / 60.0 + borderOffset

            singleOffset = 0
            offsetOrder = []
            for grpInd, group in enumerate(definedGroups):
                if self.ui.comboBox_separatePeaks.currentText() == "Group":
                    offsetOrder.append((group.name, group.color))
                for i in range(len(group.files)):
                    fi = str(group.files[i]).replace("\\", "/")
                    a = fi[fi.rfind("/") + 1 : fi.find(".mzXML")]

                    pw.getCallingFunction()("text")("MZ: %.5f\nFile: '%s'" % (pi.mz, a))
                    pw.getCallingFunction()("value")(done)
                    done = done + 1

                    if pi.scanEvent in self.loadedMZXMLs[fi].getFilterLines(
                        includeMS1=True,
                        includeMS2=False,
                        includePosPolarity=True,
                        includeNegPolarity=True,
                    ):
                        eic, times, scanIds, mzs = self.loadedMZXMLs[fi].getEIC(pi.mz, ppm=ppm, filterLine=pi.scanEvent)
                        eicL, times, scanIds, mzs = self.loadedMZXMLs[fi].getEIC(pi.lmz, ppm=ppm, filterLine=pi.scanEvent)

                        maxN = 1
                        maxL = 1

                        if self.ui.resultsExperimentNormaliseXICs_checkBox.isChecked() or self.ui.resultsExperimentNormaliseXICsSeparately_checkBox.isChecked():
                            m = 0
                            ml = 0
                            for j in range(len(eic)):
                                if abs(pi.rt / 60 - (times[j] / 60.0)) <= 0.2:
                                    m = max(m, eic[j])
                                    ml = max(ml, eicL[j])
                            if m != 0:
                                maxN = m
                            if ml != 0:
                                maxL = ml

                        if self.ui.resultsExperimentNormaliseXICs_checkBox.isChecked():
                            maxN = maxL

                        intlim[0] = min(
                            intlim[0],
                            min([1] + [-eicL[j] / maxL for j in range(len(eic)) if rtBorderMin <= times[j] / 60.0 <= rtBorderMax]),
                        )
                        intlim[1] = max(
                            intlim[1],
                            max([1] + [eic[j] / maxN for j in range(len(eic)) if rtBorderMin <= times[j] / 60.0 <= rtBorderMax]),
                        )

                        scan = self.loadedMZXMLs[fi].getClosestMS1Scan(pi.rt / 60.0, filterLine=pi.scanEvent)
                        peakID = scan.findMZ(pi.mz, ppm=ppm)
                        if peakID[0] != -1:
                            if mostAbundantFile is None or scan.intensity_list[peakID[0]] > mostAbundantFile[1]:
                                mostAbundantFile = (
                                    fi,
                                    scan.intensity_list[peakID[0]],
                                    scan,
                                    group.color,
                                )

                        mzs = scan.mz_list
                        ints = scan.intensity_list
                        use_inds = []
                        for ind_i in range(len(mzs)):
                            if mzs[ind_i] >= pi.mz - 5 and mzs[ind_i] <= pi.lmz + 5:
                                use_inds.append(ind_i)
                        if len(use_inds) > 0:
                            mzs = [mzs[i] for i in use_inds]
                            ints = [ints[i] for i in use_inds]

                            max_plotted_int = max(ints)
                            corrFact = 1.0 / max_plotted_int * 9
                            ints = [i * corrFact for i in ints]

                            self.ui.resultsExperimentMSScanPeaks_plot.axes.vlines(
                                x=mzs,
                                ymin=done * 10,
                                ymax=[done * 10 + i for i in ints],
                                color=group.color,
                                alpha=0.3,
                            )
                            for mz in [pi.mz + i * 1.00335484 for i in [0, 1, 2, 3]] + [pi.lmz + i * 1.00335484 for i in [0, 1, 2, 3]] + [pi.mz - i * 1.00335484 for i in [0, 1, 2, 3]] + [pi.lmz - i * 1.00335484 for i in [0, 1, 2, 3]]:
                                peakID = scan.findMZ(mz, ppm=ppm)
                                if peakID[0] != -1:
                                    self.ui.resultsExperimentMSScanPeaks_plot.axes.vlines(
                                        x=scan.mz_list[peakID[0]],
                                        ymin=done * 10,
                                        ymax=done * 10 + scan.intensity_list[peakID[0]] * corrFact,
                                        color=group.color,
                                    )
                            for mz in [pi.mz, pi.lmz]:
                                peakID = scan.findMZ(mz, ppm=ppm)
                                if peakID[0] != -1:
                                    self.ui.resultsExperimentMSScanPeaks_plot.axes.vlines(
                                        x=scan.mz_list[peakID[0]],
                                        ymin=done * 10,
                                        ymax=done * 10 + scan.intensity_list[peakID[0]] * corrFact,
                                        color=group.color,
                                        linewidth=2.0,
                                    )

                            self.ui.resultsExperimentMSScanPeaks_plot.axes.hlines(
                                y=done * 10,
                                xmin=pi.mz - 5,
                                xmax=pi.lmz + 5,
                                color=group.color,
                                alpha=0.33,
                            )
                            self.ui.resultsExperimentMSScanPeaks_plot.axes.text(
                                y=done * 10,
                                x=pi.lmz + 5.5,
                                color=group.color,
                                s="%s, max.int. %.3g" % (a, max_plotted_int),
                            )

                        self.ui.resultsExperiment_plot.axes.plot(
                            [t / 60.0 for t in times],
                            [e / maxN for e in eic],
                            color=group.color,
                            label="M: %s" % (a),
                        )
                        self.ui.resultsExperiment_plot.axes.plot(
                            [t / 60.0 for t in times],
                            [-e / maxL for e in eicL],
                            color=group.color,
                            label="M': %s" % (a),
                        )

                        offset = 0
                        if self.ui.comboBox_separatePeaks.currentText() == "Group":
                            offset = grpInd
                        else:
                            offset = singleOffset
                            offsetOrder.append((a, group.color))
                        self.ui.resultsExperimentSeparatedPeaks_plot.axes.plot(
                            [t / 60.0 + offset for t in times if rtBorderMin <= t / 60.0 <= rtBorderMax],
                            [eic[j] / maxN for j in range(len(eic)) if rtBorderMin <= times[j] / 60.0 <= rtBorderMax],
                            color=group.color,
                            label="M: %s" % (a),
                        )
                        self.ui.resultsExperimentSeparatedPeaks_plot.axes.plot(
                            [t / 60.0 + offset for t in times if rtBorderMin <= t / 60.0 <= rtBorderMax],
                            [-eicL[j] / maxL for j in range(len(eicL)) if rtBorderMin <= times[j] / 60.0 <= rtBorderMax],
                            color=group.color,
                            label="M': %s" % (a),
                        )

                        singleOffset += 1

            maxSigAbundance = 0
            if False and mostAbundantFile is not None:
                # self.ui.resultsExperiment_plot.axes.axvline(x=pi.rt/60., color=mostAbundantFile[3])

                self.ui.resultsExperimentMSScanPeaks_plot.axes.vlines(
                    x=mostAbundantFile[2].mz_list,
                    ymin=0,
                    ymax=mostAbundantFile[2].intensity_list,
                    color="lightgrey",
                )
                intLeft = 0
                mz = pi.mz
                peakID = mostAbundantFile[2].findMZ(mz, ppm=ppm)
                if peakID[0] != -1:
                    intLeft = mostAbundantFile[2].intensity_list[peakID[0]]

                intRight = 0
                mz = pi.lmz
                peakID = mostAbundantFile[2].findMZ(mz, ppm=ppm)
                if peakID[0] != -1:
                    intRight = mostAbundantFile[2].intensity_list[peakID[0]]

                for i in [0, 1, 2, 3, 4, 5, 6]:
                    mz = pi.mz + 1.00335484 * i
                    peakID = mostAbundantFile[2].findMZ(mz, ppm=ppm)
                    if peakID[0] != -1:
                        self.ui.resultsExperimentMSScanPeaks_plot.axes.vlines(
                            x=mz,
                            ymin=0,
                            ymax=mostAbundantFile[2].intensity_list[peakID[0]],
                            color=mostAbundantFile[3],
                            linewidth=2.0,
                        )
                    mz = pi.mz - 1.00335484 * i
                    peakID = mostAbundantFile[2].findMZ(mz, ppm=ppm)
                    if peakID[0] != -1:
                        self.ui.resultsExperimentMSScanPeaks_plot.axes.vlines(
                            x=mz,
                            ymin=0,
                            ymax=mostAbundantFile[2].intensity_list[peakID[0]],
                            color=mostAbundantFile[3],
                            linewidth=2.0,
                        )
                    mz = pi.lmz - 1.00335484 * i
                    peakID = mostAbundantFile[2].findMZ(mz, ppm=ppm)
                    if peakID[0] != -1:
                        self.ui.resultsExperimentMSScanPeaks_plot.axes.vlines(
                            x=mz,
                            ymin=0,
                            ymax=mostAbundantFile[2].intensity_list[peakID[0]],
                            color=mostAbundantFile[3],
                            linewidth=2.0,
                        )
                    mz = pi.lmz + 1.00335484 * i
                    peakID = mostAbundantFile[2].findMZ(mz, ppm=ppm)
                    if peakID[0] != -1:
                        self.ui.resultsExperimentMSScanPeaks_plot.axes.vlines(
                            x=mz,
                            ymin=0,
                            ymax=mostAbundantFile[2].intensity_list[peakID[0]],
                            color=mostAbundantFile[3],
                            linewidth=2.0,
                        )

                for i in [0, 1, 2, 3]:
                    annotationPPM = 5.0

                    self.ui.resultsExperimentMSScanPeaks_plot.axes.add_patch(
                        patches.Rectangle(
                            (
                                (pi.mz + (1.00335 * i)) * (1.0 - annotationPPM / 1000000.0),
                                intLeft * 0,
                            ),
                            (pi.mz + (1.00335 * i)) * (2 * annotationPPM / 1000000.0),
                            intLeft,
                            edgecolor="none",
                            facecolor="purple",
                            alpha=0.4,
                        )
                    )

                    self.ui.resultsExperimentMSScanPeaks_plot.axes.add_patch(
                        patches.Rectangle(
                            (
                                (pi.lmz + (1.00335 * i)) * (1.0 - annotationPPM / 1000000.0),
                                intRight * 0,
                            ),
                            (pi.lmz + (1.00335 * i)) * (2 * annotationPPM / 1000000.0),
                            intRight,
                            edgecolor="none",
                            facecolor="purple",
                            alpha=0.4,
                        )
                    )

                    rat = getNormRatio(0.9893, 15, i)
                    self.ui.resultsExperimentMSScanPeaks_plot.axes.add_patch(
                        patches.Rectangle(
                            (
                                (pi.mz + (1.00335 * i)) * (1.0 - 2 * annotationPPM / 1000000.0),
                                intLeft * rat * 0.98,
                            ),
                            (pi.mz + (1.00335 * i)) * (4 * annotationPPM / 1000000.0),
                            intLeft * rat * 0.04,
                            edgecolor="none",
                            facecolor="orange",
                            alpha=0.8,
                        )
                    )

                    rat = getNormRatio(0.993, 15, i)
                    self.ui.resultsExperimentMSScanPeaks_plot.axes.add_patch(
                        patches.Rectangle(
                            (
                                (pi.lmz + (1.00335 * i)) * (1.0 - 2 * annotationPPM / 1000000.0),
                                intRight * rat * 0.98,
                            ),
                            (pi.lmz + (1.00335 * i)) * (4 * annotationPPM / 1000000.0),
                            intRight * rat * 0.04,
                            edgecolor="none",
                            facecolor="orange",
                            alpha=0.8,
                        )
                    )

                peakID = mostAbundantFile[2].findMZ(pi.mz, ppm=ppm)
                peakID2 = mostAbundantFile[2].findMZ(pi.lmz, ppm=ppm)

                if peakID[0] != -1:
                    maxSigAbundance = max(maxSigAbundance, mostAbundantFile[2].intensity_list[peakID[0]])
                if peakID2[0] != -1:
                    maxSigAbundance = max(maxSigAbundance, mostAbundantFile[2].intensity_list[peakID2[0]])

        pw.hide()
        if done > 0:
            pi = plotItems[0]
            for oi, o in enumerate(offsetOrder):
                self.ui.resultsExperimentSeparatedPeaks_plot.axes.axvline(x=oi + pi.rt / 60.0, color=o[1])
                self.ui.resultsExperimentSeparatedPeaks_plot.axes.text(
                    x=oi - 0.05 + pi.rt / 60.0,
                    y=intlim[1] * 1.05,
                    s=o[0],
                    rotation=90,
                    horizontalalignment="left",
                    color=o[1],
                    backgroundcolor="white",
                    weight="bold",
                )
            intlim[1] = intlim[1] * 1.5

            self.ui.resultsExperiment_plot.axes.set_title("Overlaid EICs of selected feature pairs or groups")
            self.ui.resultsExperiment_plot.axes.set_xlabel("Retention time (min)")
            self.ui.resultsExperiment_plot.axes.set_ylabel("Intensity")
            self.ui.resultsExperimentSeparatedPeaks_plot.axes.set_title("Overlaid EICs of selected feature pairs or groups (separated artificially by respective %s)" % ("experimental group" if self.ui.comboBox_separatePeaks.currentText() == "Group" else "sample"))
            if len(plotItems) == 1:
                self.ui.resultsExperimentSeparatedPeaks_plot.axes.set_title(
                    "Overlaid EICs of %.5f (%.5f), %.2f min, %s\n(separated artificially by respective %s to improve comparison)\n"
                    % (
                        plotItems[0].mz,
                        plotItems[0].lmz,
                        plotItems[0].rt / 60.0,
                        plotItems[0].scanEvent,
                        "experimental group" if self.ui.comboBox_separatePeaks.currentText() == "Group" else "sample",
                    )
                )
            self.ui.resultsExperimentSeparatedPeaks_plot.axes.set_xlabel("Retention time (min) of feature + %s-index (0-based)" % ("experimental group" if self.ui.comboBox_separatePeaks.currentText() == "Group" else "sample"))
            self.ui.resultsExperimentSeparatedPeaks_plot.axes.set_ylabel("Intensity")
            self.ui.resultsExperimentMSScanPeaks_plot.axes.set_xlabel("M/Z")
            self.ui.resultsExperimentMSScanPeaks_plot.axes.set_ylabel("Normalized Intensity\nSeparated by sample")

            rtlim = [
                mean(meanRT) / 60.0 - borderOffset,
                mean(meanRT) / 60.0 + borderOffset,
            ]
            intlim = [intlim[0] * 1.1, intlim[1] * 1.1]
            self.drawCanvas(self.ui.resultsExperiment_plot, xlim=rtlim, ylim=intlim, showLegendOverwrite=False)
            self.drawCanvas(self.ui.resultsExperimentSeparatedPeaks_plot, showLegendOverwrite=self.ui.showLegend_experiment.isChecked())
            self.drawCanvas(
                self.ui.resultsExperimentMSScanPeaks_plot,
                xlim=[pi.mz - 5, pi.lmz + 10],
                ylim=[0, (done + 1) * 10],
            )

        # Update MSMS spectra list for selected features
        selectedItems = self.ui.resultsExperiment_TreeWidget.selectedItems()
        self.updateMSMSList_exp(selectedItems)

    def exportAsPDF(self, pdfFile=None):
        experimentalGroups = self.getAllSampleGroups()

        metabolitesToPlot = []
        for i in range(self.ui.resultsExperiment_TreeWidget.topLevelItemCount()):
            item = self.ui.resultsExperiment_TreeWidget.topLevelItem(i)
            features = []
            ogrp = ""
            for j in range(item.childCount()):
                child = item.child(j)
                features.append(
                    Bunch(
                        num=child.bunchData.id,
                        ogroup=item.bunchData.metaboliteGroupID,
                        mz=child.bunchData.mz,
                        lmz=child.bunchData.lmz,
                        xn=child.bunchData.xn,
                        rt=child.bunchData.rt / 60.0,
                        rtBorders=self.ui.doubleSpinBox_resultsExperiment_PeakWidth.value(),
                        filterLine=child.bunchData.scanEvent,
                        ionisationMode=child.bunchData.ionisationMode,
                        comment="",
                    )
                )
                ogrp = item.bunchData.metaboliteGroupID

            metabolitesToPlot.append(Bunch(name="unknown_" + str(ogrp), ogroup=ogrp, features=features))

        if pdfFile is None or not pdfFile:
            pdfFile = str(
                QtWidgets.QFileDialog.getSaveFileName(
                    caption="Select pdf file",
                    dir=self.lastOpenDir,
                    filter="PDF (*.pdf)",
                )
            )
            pdfFile = pdfFile.replace("\\", "/").replace(".PDF", ".pdf")
            if pdfFile == "":
                return

        if self.loadedMZXMLs is None:
            self.loadAllSamples()

        pw = ProgressWrapper()
        pw.show()

        from .resultsPostProcessing.GenericFeaturePDF.GenericFeaturePlotter import (
            generatePDF,
        )

        generatePDF(
            experimentalGroups,
            metabolitesToPlot,
            ppm=self.ui.doubleSpinBox_resultsExperiment_EICppm.value(),
            saveTo=pdfFile,
            mzXMLs=self.loadedMZXMLs,
            pw=pw,
        )

        pw.hide()

    def heteroAtomsConfiguration(self):
        t = heteroAtomEdit(heteroAtoms=deepcopy(self.heteroElements))
        if t.executeDialog() == QtWidgets.QDialog.Accepted:
            self.heteroElements = t.getHeteroAtoms()

            logging.info("Configured heteroatoms:")
            for ha in self.heteroElements:
                logging.info("  * " + ha.name)

    def relationShipConfiguration(self):
        t = adductsEdit(adds=deepcopy(self.adducts), nls=deepcopy(self.elementsForNL))
        if t.executeDialog() == QtWidgets.QDialog.Accepted:
            self.adducts = t.getAdducts()

            logging.info("Configured adducts")
            for add in self.adducts:
                logging.info("  *" + add.name)

            self.elementsForNL = t.getNeutralLosses()

            logging.info("Configured neutral loss elements")
            for elem in self.elementsForNL:
                logging.info("  *" + elem.name)

    def _openPeakPickingSettings(self):
        """Open the peak picking settings dialog."""
        sample_files = [(group.name, f) for group in self.getAllSampleGroups() for f in group.files]
        pos_items = [self.ui.positiveScanEvent.itemText(i) for i in range(self.ui.positiveScanEvent.count()) if self.ui.positiveScanEvent.itemText(i) not in ("None", "")]
        neg_items = [self.ui.negativeScanEvent.itemText(i) for i in range(self.ui.negativeScanEvent.count()) if self.ui.negativeScanEvent.itemText(i) not in ("None", "")]
        scan_events = {}
        if pos_items:
            scan_events["+"] = pos_items
        if neg_items:
            scan_events["-"] = neg_items

        dlg = PeakPickingSettingsDialog(
            parent=self,
            current_settings=self._peakPickingSettings,
            sample_files=sample_files,
            scan_events=scan_events,
            initial_eic_rows=getattr(self, "_peakPickingEicRows", []),
        )
        result = dlg.exec()
        # Persist EIC rows regardless of whether the user accepted or cancelled
        self._peakPickingEicRows = dlg.get_eic_rows()
        if result == QtWidgets.QDialog.Accepted:
            self._peakPickingSettings = dlg.get_settings()
            logging.info("Peak picking settings updated: algorithm=%s", self._peakPickingSettings.get("algorithm", "wavelettransform"))

    def updateCores(self):
        curMax = self.ui.cpuCores.maximum()
        curVal = self.ui.cpuCores.value()

        cpus = cpu_count()
        if self.ui.workingCore.isChecked():
            cpus -= 1
        self.ui.cpuCores.setMaximum(cpus)
        if curMax == curVal:
            self.ui.cpuCores.setValue(cpus)

    # <editor-fold desc="### Tracer Dialog">
    def setModuleUI(self, val):
        if val == TRACER:
            self.ui.label_50.setText("TracExtract experiment")

            self.ui.groupBox_ISOA.setVisible(False)
            self.ui.groupBox_ISOB.setVisible(False)
            self.ui.useCValidation.setVisible(True)

            self.ui.useRatio.setVisible(False)

            self.ui.setupTracers.setVisible(True)
            self.ui.tracerExperimentLabel.setVisible(True)
        elif val == METABOLOME:
            self.ui.label_50.setText("AllExtract experiment")

            self.ui.setupTracers.setVisible(False)
            self.ui.tracerExperimentLabel.setVisible(False)

            self.ui.groupBox_ISOA.setVisible(True)
            self.ui.groupBox_ISOB.setVisible(True)
            self.ui.useCValidation.setVisible(True)

            self.ui.useRatio.setVisible(True)
        else:
            raise Exception("undefined module provided")

    def showTracerEditor(self):
        tracerDialog = tracerEdit()
        print("Currently configured tracer")
        if self.configuredTracer is not None:
            print(self.configuredTracer.__dict__)
            tracerDialog.setTracer(deepcopy(self.configuredTracer))
        else:
            print("No tracer configured")
            # Initialize with default tracer if None
            self.configuredTracer = tracerDialog.getTracer()  # Get default tracer
            tracerDialog.setTracer(deepcopy(self.configuredTracer))
        if tracerDialog.executeDialog() == QtWidgets.QDialog.Accepted:
            self.configuredTracer = tracerDialog.getTracer()

            logging.info("Configured tracers:")
            tracer = self.configuredTracer
            logging.info(
                " * %s (%s/%s) isotopolog-ratio: %.2f, check ratio: %s %s"
                % (
                    tracer.name,
                    tracer.isotopeA,
                    tracer.isotopeB,
                    tracer.monoisotopicRatio,
                    tracer.checkRatio,
                    (
                        "(min. %.2f, max. %.2f)"
                        % (
                            tracer.monoisotopicRatio * tracer.maxRelNegBias,
                            tracer.monoisotopicRatio * tracer.maxRelPosBias,
                        )
                        if tracer.checkRatio
                        else ""
                    ),
                )
            )
        self.updateTracerInfo()

    def updateTracerInfo(self):
        trcs = []
        tracer = self.configuredTracer
        trcs.append(" * %s (%s/%s)" % (tracer.name, tracer.isotopeA, tracer.isotopeB))
        self.ui.tracerExperimentLabel.setText("\n\n".join(trcs))

    # </editor-fold>

    def isotopeATextChanged(self, text):
        text = str(text)
        self.isotopeAmass = getIsotopeMass(text)[0]
        if self.isotopeAmass == -1:
            self.ui.isotopeAMassLabel.setText("Unknown Isotope %s" % text)
        else:
            self.ui.isotopeAMassLabel.setText("%s mass is %.8f" % (text, self.isotopeAmass))

    def isotopeBTextChanged(self, text):
        text = str(text)
        self.isotopeBmass = getIsotopeMass(text)[0]
        if self.isotopeBmass == -1:
            self.ui.isotopeBMassLabel.setText("Unknown Isotope %s" % text)
        else:
            self.ui.isotopeBMassLabel.setText("%s mass is %.8f" % (text, self.isotopeBmass))

    def alignToggled(self, val):
        self.ui.polynomLabel.setEnabled(val)
        self.ui.polynomValue.setEnabled(val)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls:
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        if event.mimeData().hasUrls:
            event.setDropAction(QtCore.Qt.CopyAction)
            event.accept()

            links = []
            for url in event.mimeData().urls():
                links.append(str(url.toLocalFile()).replace("\\", "/"))

            links = natSort(links)

            if len(links) == 1 and (links[0].lower().endswith(".grp") or links[0].lower().endswith(".ini")):
                link = links[0]
                if (
                    link.lower().endswith(".grp")
                    and QtWidgets.QMessageBox.question(
                        self,
                        "MetExtract",
                        "Do you want to load the groups defined in this file?",
                        QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    )
                    == QtWidgets.QMessageBox.Yes
                ):
                    self.loadGroups(link, askLoadSettings=False)
                if (link.lower().endswith(".ini") or link.lower().endswith(".grp")) and QtWidgets.QMessageBox.question(
                    self,
                    "MetExtract",
                    "Do you want to load the settings saved in this file?",
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                ) == QtWidgets.QMessageBox.Yes:
                    self.loadSettingsFile(link)

            else:
                incorrectFiles = []

                for file in links:
                    if not (file.lower().endswith(".mzxml") or file.lower().endswith(".mzml")):
                        incorrectFiles.append(file)

                if len(incorrectFiles) > 0:
                    QtWidgets.QMessageBox.warning(
                        self,
                        "MetExtract",
                        "You are trying to load unknown file types that are not supported: \n\n  * " + "\n  * ".join(incorrectFiles) + "\n\nPlease only load mzXML or mzML files.",
                    )
                else:
                    for li in range(len(links)):
                        links[li] = links[li].replace("\\", "/")

                    if len(links) > 1:
                        pw = RegExTestDialog(strings=links)
                        pw.exec()

                        regEx = pw.getRegEx()
                        regExRes = pw.getRegExRes()

                        groups = {}
                        for link in links:
                            try:
                                res = re.match(regEx, link)
                                gName = regExRes.format(*res.groups())
                            except Exception:
                                continue
                            if gName not in groups:
                                groups[gName] = []
                            groups[gName].append(link)

                        if not groups:
                            return

                        gNames = {}
                        for gName, files in groups.items():
                            cleaned = gName.replace("\\", "/")
                            cleaned = cleaned[(cleaned.rfind("/") + 1) :]
                            gNames[cleaned] = len(files)

                        if (
                            QtWidgets.QMessageBox.question(
                                self,
                                "MetExtract",
                                "Do you want to automatically assign all files to separate groups? Everything before the last '_' character will be used as the group identifier and everything after the last '_' character will be used as the replicate identifier.\n\nThe following groups will be created: \n\n"
                                + "\n".join(["\u2022  " + gN + " (" + str(gNames[gN]) + " files)" for gN in natSort(gNames.keys())]),
                                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                            )
                            == QtWidgets.QMessageBox.Yes
                        ):
                            groupNames = natSort(list(groups.keys()))
                            for gName in groupNames:
                                files = groups[gName]
                                cleanedName = gName.replace("\\", "/")
                                cleanedName = cleanedName[(cleanedName.rfind("/") + 1) :]
                                color = predefinedColors[self.ui.groupsList.rowCount() % len(predefinedColors)]
                                self.addGroup(
                                    name=cleanedName,
                                    files=files,
                                    minGrpFound=1,
                                    omitFeatures=True,
                                    useForMetaboliteGrouping=True,
                                    removeAsFalsePositive=False,
                                    color=color,
                                    useAsMSMSTarget=False,
                                )
                            self.updateLCMSSampleSettings()
                            self.grpFileEdited = True
                    else:
                        self.showAddGroupDialog(initWithFiles=links)
        else:
            event.ignore()

    def loadAvailableSettingsFile(self, file):
        try:
            self.loadSettingsFile(file)
        except Exception:
            pass

    def addDB(self, events):
        dbFiles, _filter = QtWidgets.QFileDialog.getOpenFileNames(
            caption="Select database file(s)",
            dir=self.lastOpenDir,
            filter="Database (*.xlsx *.tsv);;Excel files (*.xlsx);;TSV files (*.tsv);;All files (*.*)",
        )

        for dbFile in dbFiles:
            dbFile = str(dbFile)
            if len(dbFile) > 0:
                self.lastOpenDir = dbFile.replace("\\", "/")
                self.lastOpenDir = self.lastOpenDir[: self.lastOpenDir.rfind("/")]

                dbFile = dbFile.replace("\\", "/")
                dbName = dbFile[dbFile.rfind("/") + 1 : dbFile.rfind(".")]

                item = QtGui.QStandardItem("%s (Database)" % dbName)
                item.setData(dbFile)
                self.ui.dbList_listView.model().appendRow(item)

    def addMZVaultRepository(self, events):
        dbFile = QtWidgets.QFileDialog.getOpenFileName(
            caption="Select database file",
            dir=self.lastOpenDir,
            filter="Database (*.db);;All files (*.*)",
        )
        dbFile = str(dbFile)

        if len(dbFile) > 0:
            self.lastOpenDir = str(dbFile).replace("\\", "/")
            self.lastOpenDir = self.lastOpenDir[: self.lastOpenDir.rfind("/")]

            dbFile = dbFile.replace("\\", "/")
            dbName = dbFile[dbFile.rfind("/") + 1 : dbFile.rfind(".")]

            item = QtGui.QStandardItem("%s (mzVault repository)" % dbName)
            item.setData(dbFile)
            self.ui.dbList_listView.model().appendRow(item)

    def removeDB(self, events):
        selectedRows = sorted(
            {index.row() for index in self.ui.dbList_listView.selectedIndexes()},
            reverse=True,
        )
        if not selectedRows:
            ind = self.ui.dbList_listView.currentIndex().row()
            if ind >= 0:
                selectedRows = [ind]
        for row in selectedRows:
            self.ui.dbList_listView.model().removeRows(row, 1)

    def generateDBTemplate(self, events):
        dbTemplateFile, _ = QtWidgets.QFileDialog.getSaveFileName(
            caption="Select database template",
            dir=self.lastOpenDir + "/DBTemplate.xlsx",
            filter="Database (*.xlsx);;All files (*.*)",
        )

        if len(dbTemplateFile) > 0:
            import polars as pl

            # Create template data DataFrame
            template_data = pl.DataFrame(
                {
                    "Num": [1, 2, 3, 4],
                    "Name": ["Phenylalanine", "Phenylalanine", "Phenylalanine", "Phenylalanine"],
                    "SumFormula": ["C9H11NO2", "C9H11NO2", "", "C9H11NO2"],
                    "Rt_min": [None, 5.5, 5.5, 5.5],
                    "MZ": [None, None, 166.085706, 166.085706],
                    "IonisationMode": ["", "", "+", "+"],
                    "AdditionalColumn1": ["", "", "", ""],
                    "AdditionalColumn2": ["", "", "", ""],
                }
            )

            # Create instructions DataFrame
            instructions_data = pl.DataFrame(
                {
                    "Instructions": [
                        "Any line starting with the hash-symbol (#) is a comment and will be skipped",
                        "The fields Num and Name are mandatory",
                        "Additionally, either the sum formula or the m/z value (and a polarity mode) have to be provided",
                        "If the sum formula is provided, the accurate mass will be calculated automatically",
                        "The retention time is optional and must be specified in minutes. If several retention times are possible, use different rows. Use the dot (.) as the decimal separator",
                        "Additional columns may be provided. These will be transferred to the results but not checked in any way. Do not include tab-stops in there.",
                        "Each column must have a unique name.",
                        "The first 6 columns are mandatory. The remaining columns can have any name and will not be considered for the annotation itself.",
                        "Each row must have a unique number.",
                        "The name and either a sum formula or a m/z value and ionisation mode are required.",
                        "If retention times are available, they must be a single number in minutes. If a compound has two or more retention times, generate a separate row for each retention time.",
                    ]
                }
            )

            # Write to Excel with multiple sheets using xlsxwriter
            from xlsxwriter import Workbook

            with pl.Config(tbl_width_chars=1000):
                with Workbook(dbTemplateFile) as wb:
                    template_data.write_excel(workbook=wb, worksheet="Template")
                    instructions_data.write_excel(workbook=wb, worksheet="Instructions")

        QtWidgets.QMessageBox.information(
            self,
            "MetExtract",
            "The database template has been generated successfully.\n\n"
            "Please consider the following advices:\n"
            "* Each column must have a unique name.\n"
            "* The first 6 columns are mandatory. The remaining columns can have any name and will not be considered for the annotaiton itself.\n\n"
            "* Each row must have a unique number.\n"
            "* The name and either a sum formula or a m/z value and ionisation mode are required.\n"
            "* If retention times are available, they must be a single number in minutes. If a compound has two or more retention times, generate a separate row for each retention time.",
            QtWidgets.QMessageBox.Ok,
        )

    def testDBs(self, events=None):
        dbFiles = []
        for entryInd in range(self.ui.dbList_listView.model().rowCount()):
            dbFile = str(self.ui.dbList_listView.model().item(entryInd, 0).data())
            dbFiles.append(dbFile)

        if not dbFiles:
            QtWidgets.QMessageBox.information(self, "MetExtract", "No database files added.", QtWidgets.QMessageBox.Ok)
            return

        # Show indeterminate progress dialog while importing
        progress = QtWidgets.QProgressDialog("Importing database files, please wait…", None, 0, 0, self)
        progress.setWindowTitle("Test DBs")
        progress.setWindowModality(QtCore.Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        progress.show()
        QtWidgets.QApplication.processEvents()

        worker = _DBTestWorker(dbFiles, parent=self)

        def _on_finished(results):
            progress.close()
            self._showDBTestResults(results)

        worker.finished.connect(_on_finished)
        worker.start()

    def _showDBTestResults(self, results):
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Database Import Test Results")
        dialog.resize(750, 450)
        layout = QtWidgets.QVBoxLayout(dialog)

        tree = QtWidgets.QTreeWidget(dialog)
        tree.setHeaderLabels(["Database / Message"])
        tree.setColumnCount(1)
        tree.header().setStretchLastSection(True)

        for result in results:
            has_errors = result["not_imported"] > 0 or len(result["errors"]) > 0
            status = "Issues found" if has_errors else "OK"
            top_item = QtWidgets.QTreeWidgetItem(
                tree,
                [f"{result['db_name']}  —  Imported: {result['imported']}, Errors: {result['not_imported']}  [{status}]"],
            )
            if not has_errors:
                QtWidgets.QTreeWidgetItem(top_item, [f"Successfully imported {result['imported']} entries"])
            for err in result["errors"]:
                QtWidgets.QTreeWidgetItem(top_item, [err])
            top_item.setExpanded(has_errors)

        layout.addWidget(tree)
        btn_close = QtWidgets.QPushButton("Close")
        btn_close.clicked.connect(dialog.accept)
        layout.addWidget(btn_close)
        dialog.exec_()

    # initialise main interface, triggers and command line parameters
    def __init__(self, module="TracExtract", parent=None, silent=False, disableR=False):
        super(Ui_MainWindow, self).__init__()
        QtWidgets.QMainWindow.__init__(self, parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setWindowTitle("MetExtract II - %s" % module)

        # Install relative-intensity bar delegate on column 0 of both result trees
        self._relative_bar_delegate_sample = _RelativeBarDelegate(self)
        self.ui.res_ExtractedData.setItemDelegateForColumn(0, self._relative_bar_delegate_sample)
        self._relative_bar_delegate_experiment = _RelativeBarDelegate(self)
        self.ui.resultsExperiment_TreeWidget.setItemDelegateForColumn(0, self._relative_bar_delegate_experiment)

        self.checkedLCMSFiles = {}

        self.silent = silent

        logging.info("MetExtract II started")

        logging.info(f"  Using sum formula generation cache at '{local_folder}/sfcache.pqts'")

        self.preConfigured_adducts = [
            ConfiguredAdduct(
                name="[M+H]+",
                mzoffset=1.007276,
                charge=1,
                polarity="+",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+NH4]+",
                mzoffset=18.033823,
                charge=1,
                polarity="+",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+Na]+",
                mzoffset=22.989218,
                charge=1,
                polarity="+",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+CH3OH+H]+",
                mzoffset=33.033489,
                charge=1,
                polarity="+",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+K]+",
                mzoffset=38.963158,
                charge=1,
                polarity="+",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+ACN+H]+",
                mzoffset=42.033823,
                charge=1,
                polarity="+",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+2Na-H]+",
                mzoffset=44.971160,
                charge=1,
                polarity="+",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+2K-H]+",
                mzoffset=76.919040,
                charge=1,
                polarity="+",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+CH3FeN]+",
                mzoffset=84.96094,
                charge=1,
                polarity="+",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+2H]++",
                mzoffset=1.007276,
                charge=2,
                polarity="+",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+H+NH4]++",
                mzoffset=9.520550,
                charge=2,
                polarity="+",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+H+Na]++",
                mzoffset=11.998247,
                charge=2,
                polarity="+",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+H+K]++",
                mzoffset=19.985217,
                charge=2,
                polarity="+",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+2Na]++",
                mzoffset=22.989218,
                charge=2,
                polarity="+",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[2M+H]+",
                mzoffset=1.007276,
                charge=1,
                polarity="+",
                mCount=2,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[2M+NH4]+",
                mzoffset=18.033823,
                charge=1,
                polarity="+",
                mCount=2,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[2M+Na]+",
                mzoffset=22.989218,
                charge=1,
                polarity="+",
                mCount=2,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[2M+K]+",
                mzoffset=38.963158,
                charge=1,
                polarity="+",
                mCount=2,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M-H2O-H]-",
                mzoffset=-19.01839,
                charge=1,
                polarity="-",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M-H]-",
                mzoffset=-1.007276,
                charge=1,
                polarity="-",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+Na-2H]-",
                mzoffset=20.974666,
                charge=1,
                polarity="-",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+Cl]-",
                mzoffset=34.969402,
                charge=1,
                polarity="-",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+K-2H]-",
                mzoffset=36.948606,
                charge=1,
                polarity="-",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[M+Br]-",
                mzoffset=78.918885,
                charge=1,
                polarity="-",
                mCount=1,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[2M-H]-",
                mzoffset=-1.007276,
                charge=1,
                polarity="-",
                mCount=2,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[2M+Fa-H]-",
                mzoffset=44.998201,
                charge=1,
                polarity="-",
                mCount=2,
                entryType="user",
            ),
            ConfiguredAdduct(
                name="[2M+Hac-H]-",
                mzoffset=59.013851,
                charge=1,
                polarity="-",
                mCount=2,
                entryType="user",
            ),
        ]
        self.adducts = self.preConfigured_adducts

        self.preConfigured_elementsForNL = [
            ConfiguredElement(name="C", weight=12.0, numberValenzElectrons=4, minCount=0, maxCount=3),
            ConfiguredElement(
                name="H",
                weight=1.00783,
                numberValenzElectrons=1,
                minCount=0,
                maxCount=30,
            ),
            ConfiguredElement(
                name="O",
                weight=15.99491,
                numberValenzElectrons=6,
                minCount=0,
                maxCount=20,
            ),
            ConfiguredElement(
                name="N",
                weight=14.00307,
                numberValenzElectrons=5,
                minCount=0,
                maxCount=2,
            ),
            ConfiguredElement(
                name="P",
                weight=30.97376,
                numberValenzElectrons=5,
                minCount=0,
                maxCount=2,
            ),
            ConfiguredElement(
                name="S",
                weight=31.97207,
                numberValenzElectrons=6,
                minCount=0,
                maxCount=2,
            ),
            ConfiguredElement(
                name="Cl",
                weight=34.968852,
                numberValenzElectrons=7,
                minCount=0,
                maxCount=1,
            ),
        ]
        self.elementsForNL = self.preConfigured_elementsForNL

        self.preConfigured_heteroElements = [
            ConfiguredHeteroAtom(
                name="34S",
                mzOffset=1.9957960000000021,
                relativeAbundance=0.044306461797516308,
                minCount=1,
                maxCount=4,
            ),
            ConfiguredHeteroAtom(
                name="37Cl",
                mzOffset=1.9970499999999944,
                relativeAbundance=0.31978355549689846,
                minCount=1,
                maxCount=2,
            ),
            ConfiguredHeteroAtom(
                name="41K",
                mzOffset=1.99812,
                relativeAbundance=0.07526881720430107526881720430108,
                minCount=1,
                maxCount=1,
            ),
        ]
        self.heteroElements = self.preConfigured_heteroElements

        self.lastOpenDir = "."
        if "USERPROFILE" in os.environ:
            self.lastOpenDir = os.getenv("USERPROFILE")
        elif "HOME" in os.environ:
            self.lastOpenDir = os.getenv("HOME")

        self.grpFile = None
        self.currentOpenResultsFile = None

        if module == "TracExtract":
            self.labellingExperiment = TRACER
            self.ui.xCountSearch.setText("12-15, 30, 45")
            logging.info("  Starting module TracExtract\n")
        elif module == "AllExtract":
            self.labellingExperiment = METABOLOME
            self.ui.xCountSearch.setText("3-60")
            logging.info("  Starting module AllExtract\n")
        else:
            logging.error("Error: invalid module '%s' selected.\nPlease specify either 'TracExtract' or 'AllExtract'\n" % module)
            QtWidgets.QMessageBox.warning(
                self,
                "MetExtract",
                "Error: invalid module '%s' selected.\nPlease specify either 'TracExtract' or 'AllExtract'\n" % module,
                QtWidgets.QMessageBox.Ok,
            )
            raise Exception("Error: invalid module '%s' selected.\nPlease specify either 'TracExtract' or 'AllExtract'")

        self.setModuleUI(self.labellingExperiment)

        # display version in main interface
        try:
            import platform

            module = "AllExtract" if self.labellingExperiment == METABOLOME else ("TracExtract" if self.labellingExperiment == TRACER else "MetExtract")
            self.ui.version.versionText = "%s II %s // Python %s (%s)" % (module, MetExtractVersion, platform.python_version(), platform.architecture()[0])
            self.ui.version.setText(self.ui.version.versionText)
        except Exception as exc:
            logging.info(f"Exception: {exc}")
            traceback.print_exc()
            module = "AllExtract" if self.labellingExperiment == METABOLOME else ("TracExtract" if self.labellingExperiment == TRACER else "MetExtract")
            self.ui.version.versionText = "%s II %s" % (module, MetExtractVersion)
            self.ui.version.setText(self.ui.version.versionText)

        # limit CPU usage to #cores-1 per default
        cpus = cpu_count()
        if cpus > 1:
            if self.ui.workingCore.isChecked():
                cpus -= 1
            self.ui.cpuCores.setMaximum(cpus)
            self.ui.cpuCores.setValue(cpus)
        else:
            self.ui.workingCore.setVisible(False)
            self.ui.cpuCores.setVisible(False)
            self.ui.label_43.setText("Calculations will be performed on one core only. System might be unresponsive")
            self.ui.cpuCores.setMaximum(1)
            self.ui.cpuCores.setValue(1)

        self.ui.workingCore.toggled.connect(self.updateCores)

        self.ui.tabWidget.setCurrentIndex(0)
        self.ui.tabWidget_2.setCurrentIndex(0)

        self.ui.isotopeAText.textChanged.connect(self.isotopeATextChanged)
        self.ui.isotopeBText.textChanged.connect(self.isotopeBTextChanged)

        self.ui.useRatio.setChecked(False)

        self.ui.groupBox_options1.hide()
        self.ui.groupBox_options2.hide()
        self.ui.groupBox_options3.hide()
        self.ui.pushButton_hideOptions1.hide()
        self.ui.pushButton_hideOptions2.hide()
        self.ui.pushButton_hideOptions3.hide()

        self.configuredTracer = ConfiguredTracer()
        self.updateTracerInfo()
        self.ui.setupTracers.clicked.connect(self.showTracerEditor)

        self.ui.processIndividualFiles.toggled.connect(self.procIndFilesChanges)
        self.ui.groupResults.toggled.connect(self.groupFilesChanges)
        self.ui.processMultipleFiles.toggled.connect(self.processMultipleFilesChanged)
        self.ui.annotateMetabolites_CheckBox.toggled.connect(self.annotateMetabolitesChanged)
        self.ui.generateMSMSInfo_CheckBox.toggled.connect(self.genrateMSMSListsChanged)

        self.ui.processIndividualFiles.setChecked(False)
        self.ui.groupResults.setChecked(False)
        self.ui.annotateMetabolites_CheckBox.setChecked(False)
        self.ui.generateMSMSInfo_CheckBox.setChecked(False)
        self.ui.generateMSMSInfo_CheckBox.setEnabled(False)
        self.ui.frame_bracketResults.setVisible(False)
        self.ui.frame_annotateMetabolites.setVisible(False)
        self.ui.frame_generateMSMSTargetLists.setVisible(False)

        self.ui.saveMZXML.toggled.connect(self.saveMZXMLChanged)
        self.updateIndividualFileProcessing = True

        self.ui.isotopologRatiosBox.setVisible(False)

        self.ui.smoothingPolynom_spinner.valueChanged.connect(self.smoothingWindowChanged)

        # RT alignment (PTW) has been removed – disable the checkbox
        self.ui.alignChromatograms.setChecked(False)
        self.ui.alignChromatograms.setEnabled(False)
        self.ui.alignChromatograms.setToolTip("RT alignment has been removed")
        self.ui.alignChromatograms.toggled.connect(self.alignToggled)

        self.ui.groupsSave.setText("./results.xlsx")
        self.ui.msms_fileLocation.setText("./MSMStargets.tsv")

        self.ui.addGroup.clicked.connect(self.showAddGroupDialogClicked)
        self.ui.saveGroups.clicked.connect(self.saveGroupsClicked)
        self.ui.loadGroups.clicked.connect(self.loadGroupsClicked)
        self.ui.removeGroup.clicked.connect(self.remGrp)
        self.ui.showFileStats.clicked.connect(self.showFileStatsPopup)

        self.ui.groupsList.doubleClicked.connect(self.editGroup)
        self.ui.groupsList.itemChanged.connect(self._onGroupTableItemChanged)
        self._configureExperimentalGroupsTableColumns()

        self.ui.startIdentification.clicked.connect(self.runProcess)

        self.ui.ppmRangeIdentification.valueChanged.connect(self.updateGroupPPM)

        self.ui.isotopeAText.setText("12C")
        self.ui.isotopeBText.setText("13C")

        self.ui.groupsSelectFile.clicked.connect(self.selectGroupsFile)

        self.ui.msms_selectFile.clicked.connect(self.selectMSMSTargetFile)

        self.ui.actionIsotopic_enrichment.triggered.connect(self.showCalcIsoEnrichmentDialog)
        self.ui.actionSet_working_directory.triggered.connect(self.setWorkingDirectoryDialog)
        self.ui.actionDownload_OBO_files.triggered.connect(self.openOboDownloadPage)

        self.ui.aboutMenue.triggered.connect(self.aboutMe)
        self.ui.actionShow_summary_of_previous_current_results.triggered.connect(self.showResultsSummary)
        self.ui.helpMenue.triggered.connect(self.helpMe)
        self.ui.openTempDir.triggered.connect(self.openTempDir)
        self.ui.actionLoad_Settings.triggered.connect(self.loadSettings)
        self.ui.actionSave_Settings.triggered.connect(self.saveSettings)
        self.ui.exitMenue.triggered.connect(self.exitMe)

        self.ui.processedFilesComboBox.currentIndexChanged.connect(self.selectedResultChanged)

        self.ui.res_ExtractedData.itemSelectionChanged.connect(self.selectedResChanged)

        self.ui.resultsExperiment_TreeWidget.itemSelectionChanged.connect(self.resultsExperimentChanged)
        self.ui.comboBox_separatePeaks.currentIndexChanged.connect(self._refreshExperimentEICs)
        self.ui.doubleSpinBox_separatePeaksShift.valueChanged.connect(self._refreshExperimentEICs)
        self.ui.resultsExperimentNormaliseXICs_checkBox.stateChanged.connect(self._refreshExperimentEICs)
        self.ui.resultsExperimentNormaliseXICsSeparately_checkBox.stateChanged.connect(self._refreshExperimentEICs)
        self.ui.showLegend_experiment.stateChanged.connect(self._refreshExperimentEICs)
        self.ui.doubleSpinBox_resultsExperiment_MSMSRTWindow.valueChanged.connect(self._refreshExperimentMSMS)
        self.ui.doubleSpinBox_resultsExperiment_MSMSPrecIntensPercent.valueChanged.connect(self._refreshExperimentMSMS)
        self.ui.doubleSpinBox_resultsExperiment_MSMSAbsIntensThreshold.valueChanged.connect(self._refreshExperimentMSMS)
        self.ui.lineEdit_msms_filter_regex.textChanged.connect(self._refreshExperimentMSMS)
        self.ui.pushButton_updateMSMSTree.clicked.connect(self._updateMSMSTreeClicked)
        self.ui.comboBox_abundancePlotType.currentIndexChanged.connect(self._refreshExperimentAbundancePlot)
        self.ui.comboBox_abundanceScale.currentIndexChanged.connect(self._refreshExperimentAbundancePlot)
        self.ui.comboBox_abundanceScalingMode.currentIndexChanged.connect(self._refreshExperimentAbundancePlot)
        self.ui.comboBox_abundanceValueType.currentIndexChanged.connect(self._refreshExperimentAbundancePlot)
        self.ui.pushButton_samplePeaksPrev.clicked.connect(self._samplePeaksPrevPage)
        self.ui.pushButton_samplePeaksNext.clicked.connect(self._samplePeaksNextPage)
        self.ui.spinBox_samplePeaksRows.valueChanged.connect(self._renderSamplePeaksPage)
        self.ui.spinBox_samplePeaksCols.valueChanged.connect(self._renderSamplePeaksPage)
        self.ui.comboBox_samplePeaksNorm.currentIndexChanged.connect(self._renderSamplePeaksPage)
        self.ui.checkBox_samplePeaksLabeledNegative.stateChanged.connect(self._renderSamplePeaksPage)
        self.ui.checkBox_samplePeaksLabeledNormSeparately.stateChanged.connect(self._renderSamplePeaksPage)

        self.ui.eicSmoothingWindow.currentIndexChanged.connect(self.smoothingWindowChanged)
        self.smoothingWindowChanged()

        self.ui.isoAbundance.stateChanged.connect(self.isoSearchChanged)
        self.isoSearchChanged()

        self.ui.autoZoomPlot.stateChanged.connect(self.selectedResChanged)
        self.ui.negEIC.stateChanged.connect(self.selectedResChanged)
        self.ui.plotAddLabels.stateChanged.connect(self.selectedResChanged)
        self.ui.plotMarkArea.stateChanged.connect(self.selectedResChanged)
        self.ui.scaleFeatures.stateChanged.connect(self.selectedResChanged)
        self.ui.scaleLabelledFeatures.stateChanged.connect(self.selectedResChanged)
        self.ui.showIsotopologues.stateChanged.connect(self.selectedResChanged)
        self.ui.showArtificialShoft_checkBox.stateChanged.connect(self.selectedResChanged)
        self.ui.showSmoothedEIC_checkBox.stateChanged.connect(self.selectedResChanged)
        self.ui.flattenXIC.stateChanged.connect(self.selectedResChanged)
        self.ui.showLegend.stateChanged.connect(self.selectedResChanged)
        self.ui.showDiagnostics.stateChanged.connect(self.selectedResChanged)
        self.ui.setPeakCentersToZero.stateChanged.connect(self.selectedResChanged)
        self.ui.MSLabels.stateChanged.connect(self.selectedResChanged)
        self.ui.MSIsos.stateChanged.connect(self.selectedResChanged)
        self.ui.drawFPIsotopologues.stateChanged.connect(self.selectedResChanged)
        self.ui.doubleSpinBox_isotopologAnnotationPPM.valueChanged.connect(self.selectedResChanged)

        self.ui.annotateMetabolites_CheckBox.stateChanged.connect(self.annotateMetabolitesChanged)

        self.dbListModel = QtGui.QStandardItemModel()
        self.ui.dbList_listView.setModel(self.dbListModel)
        self.ui.addDB_pushButton.clicked.connect(self.addDB)
        self.ui.addmzVaultRep_pushButton.clicked.connect(self.addMZVaultRepository)
        self.ui.removeDB_pushButton.clicked.connect(self.removeDB)
        self.ui.generateDBTemplate_pushButton.clicked.connect(self.generateDBTemplate)
        self.ui.testDBs_pushButton.clicked.connect(self.testDBs)
        self.ui.actionDownloadDBTemplate.triggered.connect(self.generateDBTemplate)

        # setup result plots
        # Setup first plot
        # http://eli.thegreenplace.net/2009/01/20/matplotlib-with-pyqt-guis/
        self.ui.pl_tic = QtCore.QObject()
        self.ui.pl_tic.dpi = 50
        self.ui.pl_tic.fig = Figure((5.0, 4.0), dpi=self.ui.pl_tic.dpi, facecolor="white")
        self.ui.pl_tic.fig.subplots_adjust(left=0.05, bottom=0.05, right=0.99, top=0.95)
        self.ui.pl_tic.canvas = FigureCanvas(self.ui.pl_tic.fig)
        self.ui.pl_tic.canvas.setParent(self.ui.visualizationWidget)
        self.ui.pl_tic.axes = self.ui.pl_tic.fig.add_subplot(111)
        simpleaxis(self.ui.pl_tic.axes)
        self.ui.pl_tic.twinxs = [self.ui.pl_tic.axes]
        self.ui.pl_tic.mpl_toolbar = NavigationToolbar(self.ui.pl_tic.canvas, self.ui.ticVisualisationWidget)

        vbox = QtWidgets.QVBoxLayout()
        vbox.addWidget(self.ui.pl_tic.canvas)
        vbox.addWidget(self.ui.pl_tic.mpl_toolbar)
        self.ui.ticVisualisationWidget.setLayout(vbox)

        # Setup first plot
        # http://eli.thegreenplace.net/2009/01/20/matplotlib-with-pyqt-guis/
        self.ui.pl1 = QtCore.QObject()
        self.ui.pl1.dpi = 50
        self.ui.pl1.fig = Figure((5.0, 4.0), dpi=self.ui.pl1.dpi, facecolor="white")
        self.ui.pl1.fig.subplots_adjust(left=0.05, bottom=0.1, right=0.99, top=0.95)
        self.ui.pl1.canvas = FigureCanvas(self.ui.pl1.fig)
        self.ui.pl1.canvas.setParent(self.ui.visualizationWidget)
        self.ui.pl1.axes = self.ui.pl1.fig.add_subplot(111)
        simpleaxis(self.ui.pl1.axes)
        self.ui.pl1.twinxs = [self.ui.pl1.axes]
        self.ui.pl1.mpl_toolbar = NavigationToolbar(self.ui.pl1.canvas, self.ui.visualizationWidget)

        vbox = QtWidgets.QVBoxLayout()
        vbox.addWidget(self.ui.pl1.mpl_toolbar)
        vbox.addWidget(self.ui.pl1.canvas)
        self.ui.visualizationWidget.setLayout(vbox)

        # setup second plot2
        self.ui.pl2A = QtCore.QObject()
        self.ui.pl2A.type = None
        self.ui.pl2A.dpi = 50
        self.ui.pl2A.fig = Figure((5.0, 4.0), dpi=self.ui.pl2A.dpi, facecolor="white")
        self.ui.pl2A.fig.subplots_adjust(left=0.15, bottom=0.05, right=0.99, top=0.85)
        self.ui.pl2A.canvas = FigureCanvas(self.ui.pl2A.fig)
        self.ui.pl2A.canvas.setParent(self.ui.pl2AWidget)
        self.ui.pl2A.axes = self.ui.pl2A.fig.add_subplot(111)
        # noaxis(self.ui.pl2A.axes)
        self.ui.pl2A.twinxs = [self.ui.pl2A.axes]
        self.ui.pl2A.mpl_toolbar = NavigationToolbar(self.ui.pl2A.canvas, self.ui.pl2AWidget)
        self.ui.pl2A.pictureShown = False

        vbox = QtWidgets.QVBoxLayout()
        vbox.addWidget(self.ui.pl2A.mpl_toolbar)
        vbox.addWidget(self.ui.pl2A.canvas)
        self.ui.pl2AWidget.setLayout(vbox)

        self.ui.pl2B = QtCore.QObject()
        self.ui.pl2B.type = None
        self.ui.pl2B.dpi = 50
        self.ui.pl2B.fig = Figure((5.0, 4.0), dpi=self.ui.pl2B.dpi, facecolor="white")
        self.ui.pl2B.fig.subplots_adjust(left=0.15, bottom=0.05, right=0.99, top=0.85)
        self.ui.pl2B.canvas = FigureCanvas(self.ui.pl2B.fig)
        self.ui.pl2B.canvas.setParent(self.ui.pl2BWidget)
        self.ui.pl2B.axes = self.ui.pl2B.fig.add_subplot(111)
        # noaxis(self.ui.pl2B.axes)
        self.ui.pl2B.twinxs = [self.ui.pl2B.axes]
        self.ui.pl2B.mpl_toolbar = NavigationToolbar(self.ui.pl2B.canvas, self.ui.pl2BWidget)
        self.ui.pl2B.pictureShown = False

        vbox = QtWidgets.QVBoxLayout()
        vbox.addWidget(self.ui.pl2B.mpl_toolbar)
        vbox.addWidget(self.ui.pl2B.canvas)
        self.ui.pl2BWidget.setLayout(vbox)

        # setup third plot
        self.ui.pl3 = QtCore.QObject()
        self.ui.pl3.type = None
        self.ui.pl3.dpi = 50
        self.ui.pl3.fig = Figure((5.0, 4.0), dpi=self.ui.pl3.dpi, facecolor="white")
        self.ui.pl3.fig.subplots_adjust(left=0.05, bottom=0.1, right=0.99, top=0.95)
        self.ui.pl3.canvas = FigureCanvas(self.ui.pl3.fig)
        self.ui.pl3.canvas.setParent(self.ui.visualizationWidget3)
        self.ui.pl3.axes = self.ui.pl3.fig.add_subplot(111)
        simpleaxis(self.ui.pl3.axes)
        self.ui.pl3.twinxs = [self.ui.pl3.axes]
        self.ui.pl3.mpl_toolbar = NavigationToolbar(self.ui.pl3.canvas, self.ui.visualizationWidget3)

        vbox = QtWidgets.QVBoxLayout()
        vbox.addWidget(self.ui.pl3.mpl_toolbar)
        vbox.addWidget(self.ui.pl3.canvas)
        self.ui.visualizationWidget3.setLayout(vbox)

        # Setup MSMS plot
        self.ui.plMSMS = QtCore.QObject()
        self.ui.plMSMS.dpi = 50
        self.ui.plMSMS.fig = Figure((5.0, 4.0), dpi=self.ui.plMSMS.dpi, facecolor="white")
        self.ui.plMSMS.fig.subplots_adjust(left=0.08, bottom=0.08, right=0.98, top=0.95, hspace=0.3)
        self.ui.plMSMS.canvas = FigureCanvas(self.ui.plMSMS.fig)
        self.ui.plMSMS.canvas.setParent(self.ui.plMSMSWidget)
        self.ui.plMSMS.axes = []  # Will hold multiple subplot axes
        self.ui.plMSMS.mpl_toolbar = NavigationToolbar(self.ui.plMSMS.canvas, self.ui.plMSMSWidget)

        vbox = QtWidgets.QVBoxLayout()
        vbox.addWidget(self.ui.plMSMS.mpl_toolbar)
        vbox.addWidget(self.ui.plMSMS.canvas)
        self.ui.plMSMSWidget.setLayout(vbox)

        # Setup experiment plot - overlaid EICs plot
        # http://eli.thegreenplace.net/2009/01/20/matplotlib-with-pyqt-guis/
        self.ui.resultsExperiment_plot = QtCore.QObject()
        self.ui.resultsExperiment_plot.dpi = 50
        self.ui.resultsExperiment_plot.fig = Figure((5.0, 4.0), dpi=self.ui.resultsExperiment_plot.dpi, facecolor="white")
        self.ui.resultsExperiment_plot.fig.subplots_adjust(left=0.05, bottom=0.05, right=0.99, top=0.95)
        self.ui.resultsExperiment_plot.canvas = FigureCanvas(self.ui.resultsExperiment_plot.fig)
        self.ui.resultsExperiment_plot.canvas.setParent(self.ui.visualizationWidget)
        self.ui.resultsExperiment_plot.axes = self.ui.resultsExperiment_plot.fig.add_subplot(111)
        simpleaxis(self.ui.resultsExperiment_plot.axes)
        self.ui.resultsExperiment_plot.twinxs = [self.ui.resultsExperiment_plot.axes]
        self.ui.resultsExperiment_plot.mpl_toolbar = NavigationToolbar(self.ui.resultsExperiment_plot.canvas, self.ui.visualizationWidget)

        vbox = QtWidgets.QVBoxLayout()
        vbox.addWidget(self.ui.resultsExperiment_plot.mpl_toolbar)
        vbox.addWidget(self.ui.resultsExperiment_plot.canvas)
        self.ui.resultsExperimentSeparatedPeaks_widget.setLayout(vbox)

        # Setup experiment plot - separate chrom. peaks plot
        # http://eli.thegreenplace.net/2009/01/20/matplotlib-with-pyqt-guis/
        self.ui.resultsExperimentSeparatedPeaks_plot = QtCore.QObject()
        self.ui.resultsExperimentSeparatedPeaks_plot.dpi = 50
        self.ui.resultsExperimentSeparatedPeaks_plot.fig = Figure(
            (5.0, 4.0),
            dpi=self.ui.resultsExperimentSeparatedPeaks_plot.dpi,
            facecolor="white",
        )
        self.ui.resultsExperimentSeparatedPeaks_plot.fig.subplots_adjust(left=0.05, bottom=0.05, right=0.99, top=0.95)
        self.ui.resultsExperimentSeparatedPeaks_plot.canvas = FigureCanvas(self.ui.resultsExperimentSeparatedPeaks_plot.fig)
        self.ui.resultsExperimentSeparatedPeaks_plot.canvas.setParent(self.ui.visualizationWidget)
        self.ui.resultsExperimentSeparatedPeaks_plot.axes = self.ui.resultsExperimentSeparatedPeaks_plot.fig.add_subplot(111)
        simpleaxis(self.ui.resultsExperimentSeparatedPeaks_plot.axes)
        self.ui.resultsExperimentSeparatedPeaks_plot.twinxs = [self.ui.resultsExperimentSeparatedPeaks_plot.axes]
        self.ui.resultsExperimentSeparatedPeaks_plot.mpl_toolbar = NavigationToolbar(
            self.ui.resultsExperimentSeparatedPeaks_plot.canvas,
            self.ui.visualizationWidget,
        )

        vbox = QtWidgets.QVBoxLayout()
        vbox.addWidget(self.ui.resultsExperimentSeparatedPeaks_plot.mpl_toolbar)
        vbox.addWidget(self.ui.resultsExperimentSeparatedPeaks_plot.canvas)
        self.ui.resultsExperiment_widget.setLayout(vbox)

        # Setup experiment plot - separate chrom. peaks plot
        # http://eli.thegreenplace.net/2009/01/20/matplotlib-with-pyqt-guis/
        self.ui.resultsExperimentMSScanPeaks_plot = QtCore.QObject()
        self.ui.resultsExperimentMSScanPeaks_plot.dpi = 50
        self.ui.resultsExperimentMSScanPeaks_plot.fig = Figure(
            (5.0, 4.0),
            dpi=self.ui.resultsExperimentMSScanPeaks_plot.dpi,
            facecolor="white",
        )
        self.ui.resultsExperimentMSScanPeaks_plot.fig.subplots_adjust(left=0.05, bottom=0.05, right=0.99, top=0.95)
        self.ui.resultsExperimentMSScanPeaks_plot.canvas = FigureCanvas(self.ui.resultsExperimentMSScanPeaks_plot.fig)
        self.ui.resultsExperimentMSScanPeaks_plot.canvas.setParent(self.ui.visualizationWidget)
        self.ui.resultsExperimentMSScanPeaks_plot.axes = self.ui.resultsExperimentMSScanPeaks_plot.fig.add_subplot(111)
        simpleaxis(self.ui.resultsExperimentMSScanPeaks_plot.axes)
        self.ui.resultsExperimentMSScanPeaks_plot.twinxs = [self.ui.resultsExperimentMSScanPeaks_plot.axes]
        self.ui.resultsExperimentMSScanPeaks_plot.mpl_toolbar = NavigationToolbar(
            self.ui.resultsExperimentMSScanPeaks_plot.canvas,
            self.ui.visualizationWidget,
        )

        vbox = QtWidgets.QVBoxLayout()
        vbox.addWidget(self.ui.resultsExperimentMSScanPeaks_plot.mpl_toolbar)
        vbox.addWidget(self.ui.resultsExperimentMSScanPeaks_plot.canvas)
        self.ui.resultsExperimentMSScan_widget.setLayout(vbox)

        # Setup experiment abundance plot
        self.ui.resultsExperimentAbundance_plot = QtCore.QObject()
        self.ui.resultsExperimentAbundance_plot.dpi = 50
        self.ui.resultsExperimentAbundance_plot.fig = Figure((5.0, 4.0), dpi=self.ui.resultsExperimentAbundance_plot.dpi, facecolor="white")
        self.ui.resultsExperimentAbundance_plot.fig.subplots_adjust(left=0.08, bottom=0.15, right=0.99, top=0.95)
        self.ui.resultsExperimentAbundance_plot.canvas = FigureCanvas(self.ui.resultsExperimentAbundance_plot.fig)
        self.ui.resultsExperimentAbundance_plot.canvas.setParent(self.ui.resultsExperimentAbundance_widget)
        self.ui.resultsExperimentAbundance_plot.axes = self.ui.resultsExperimentAbundance_plot.fig.add_subplot(111)
        simpleaxis(self.ui.resultsExperimentAbundance_plot.axes)
        self.ui.resultsExperimentAbundance_plot.twinxs = [self.ui.resultsExperimentAbundance_plot.axes]
        self.ui.resultsExperimentAbundance_plot.mpl_toolbar = NavigationToolbar(
            self.ui.resultsExperimentAbundance_plot.canvas,
            self.ui.resultsExperimentAbundance_widget,
        )

        vbox = QtWidgets.QVBoxLayout()
        vbox.addWidget(self.ui.resultsExperimentAbundance_plot.mpl_toolbar)
        vbox.addWidget(self.ui.resultsExperimentAbundance_plot.canvas)
        self.ui.resultsExperimentAbundance_widget.setLayout(vbox)

        # Setup sample peaks plot (dynamic per-sample subplot grid)
        self.ui.resultsExperimentSamplePeaks_plot = QtCore.QObject()
        self.ui.resultsExperimentSamplePeaks_plot.dpi = 72
        self.ui.resultsExperimentSamplePeaks_plot.fig = Figure(facecolor="white")
        self.ui.resultsExperimentSamplePeaks_plot.canvas = FigureCanvas(self.ui.resultsExperimentSamplePeaks_plot.fig)
        self.ui.resultsExperimentSamplePeaks_plot.canvas.setParent(self.ui.resultsExperimentSamplePeaks_widget)
        self.ui.resultsExperimentSamplePeaks_plot.mpl_toolbar = NavigationToolbar(
            self.ui.resultsExperimentSamplePeaks_plot.canvas,
            self.ui.resultsExperimentSamplePeaks_widget,
        )
        vbox_sp = QtWidgets.QVBoxLayout()
        vbox_sp.addWidget(self.ui.resultsExperimentSamplePeaks_plot.mpl_toolbar)
        vbox_sp.addWidget(self.ui.resultsExperimentSamplePeaks_plot.canvas)
        self.ui.resultsExperimentSamplePeaks_widget.setLayout(vbox_sp)

        # Setup feature map plot (RT vs MZ overview of all features)
        self.ui.expFeatureMap_plot = QtCore.QObject()
        self.ui.expFeatureMap_plot.dpi = 72
        self.ui.expFeatureMap_plot.fig = Figure(facecolor="white")
        self.ui.expFeatureMap_plot.fig.subplots_adjust(left=0.07, bottom=0.1, right=0.99, top=0.95)
        self.ui.expFeatureMap_plot.canvas = FigureCanvas(self.ui.expFeatureMap_plot.fig)
        self.ui.expFeatureMap_plot.canvas.setParent(self.ui.expFeatureMapContainer)
        self.ui.expFeatureMap_plot.axes = self.ui.expFeatureMap_plot.fig.add_subplot(111)
        self.ui.expFeatureMap_plot.mpl_toolbar = NavigationToolbar(
            self.ui.expFeatureMap_plot.canvas,
            self.ui.expFeatureMapContainer,
        )
        self.ui.expFeatureMap_plot.canvas.mpl_connect("motion_notify_event", self._onFeatureMapHover)
        self.ui.expFeatureMap_plot.canvas.mpl_connect("button_release_event", self._onFeatureMapClick)
        self._featureMapData = []
        self._featureMapAnnotation = None
        # Control row: dot-size metric selector
        self.ui.expFeatureMapSizeLabel = QtWidgets.QLabel("Dot size:")
        self.ui.expFeatureMapSizeCombo = QtWidgets.QComboBox()
        self.ui.expFeatureMapSizeCombo.addItems(["Average peak area", "Number of MSMS spectra", "Found in n samples"])
        self.ui.expFeatureMapSizeCombo.currentIndexChanged.connect(lambda _: self._buildFeatureMap() if self.ui.expFeatureMapContainer.isVisible() else None)
        ctrl_row = QtWidgets.QHBoxLayout()
        ctrl_row.addWidget(self.ui.expFeatureMapSizeLabel)
        ctrl_row.addWidget(self.ui.expFeatureMapSizeCombo)
        ctrl_row.addStretch(1)
        vbox_fm = QtWidgets.QVBoxLayout()
        vbox_fm.addWidget(self.ui.expFeatureMap_plot.mpl_toolbar)
        vbox_fm.addLayout(ctrl_row)
        vbox_fm.addWidget(self.ui.expFeatureMap_plot.canvas)
        self.ui.expFeatureMapContainer.setLayout(vbox_fm)

        # Setup experiment MSMS plot - multiple MS/MS spectra subplots
        self.ui.plMSMS_exp = QtCore.QObject()
        self.ui.plMSMS_exp.dpi = 50
        self.ui.plMSMS_exp.fig = Figure((5.0, 4.0), dpi=self.ui.plMSMS_exp.dpi, facecolor="white")
        self.ui.plMSMS_exp.canvas = FigureCanvas(self.ui.plMSMS_exp.fig)
        self.ui.plMSMS_exp.canvas.setParent(self.ui.plMSMSWidget_exp)
        self.ui.plMSMS_exp.mpl_toolbar = NavigationToolbar(self.ui.plMSMS_exp.canvas, self.ui.plMSMSWidget_exp)

        vbox = QtWidgets.QVBoxLayout()
        vbox.addWidget(self.ui.plMSMS_exp.mpl_toolbar)
        vbox.addWidget(self.ui.plMSMS_exp.canvas)
        self.ui.plMSMSWidget_exp.setLayout(vbox)

        # Install custom delegate on MSMS exp table to preserve row colors on selection
        self._msms_table_delegate = _MSMSTableDelegate(self.ui.msms_SpectraList_exp)
        self.ui.msms_SpectraList_exp.setItemDelegate(self._msms_table_delegate)

        self.ui.pl2A.xics = []
        self.ui.pl2A.times = []
        self.ui.pl2A.peaks = []
        self.ui.pl2A.x_vals = []
        self.ui.pl2A.y_vals = []

        font = {"size": 18}
        matplotlib.rc("font", **font)

        self.ui.dataFilter.textChanged.connect(self.filterEdited)
        self.ui.expDataFilter.textChanged.connect(self.expFilterEdited)
        # Connect collapsible filter toggle
        self.ui.expFilterToggleBtn.toggled.connect(self._onExpFilterToggle)
        self.ui.expFilterResetBtn.clicked.connect(self._onExpFilterReset)
        self.ui.expExportMGFBtn.clicked.connect(self._export_exp_msms_mgf)
        self.ui.expFeatureMapBtn.toggled.connect(self._toggleFeatureMap)
        self.ui.expMSMSPrecursorOverviewBtn.clicked.connect(self._showMSMSPrecursorOverview)
        self.ui.expGenInclusionListBtn.clicked.connect(self._generateInclusionLists)
        # Right-click context menu on the experiment-results tree (copy feature fields)
        self.ui.resultsExperiment_TreeWidget.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.ui.resultsExperiment_TreeWidget.customContextMenuRequested.connect(self._showExperimentTreeContextMenu)
        self.ui.expFilter_mz.textChanged.connect(self.expFilterEdited)
        self.ui.expFilter_rt.textChanged.connect(self.expFilterEdited)
        self.ui.expFilter_xn.textChanged.connect(self.expFilterEdited)
        self.ui.expFilter_z.textChanged.connect(self.expFilterEdited)
        self.ui.expFilter_polarity.currentIndexChanged.connect(self.expFilterEdited)
        self.ui.expFilter_ms2.currentIndexChanged.connect(self.expFilterEdited)
        self.ui.res_ExtractedData.itemDoubleClicked.connect(self.res_doubleClick)
        self.ui.msms_SpectraList.itemSelectionChanged.connect(self.plotSelectedMSMSSpectra)

        self.ui.openRAW.clicked.connect(self.openRawFile)
        self.ui.setChromPeakName.clicked.connect(self.setChromPeakName)

        self.ui.res_ExtractedData.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.ui.res_ExtractedData.customContextMenuRequested.connect(self.showPopup)

        self.ui.relationshipConfig.clicked.connect(self.relationShipConfiguration)

        self.ui.defineHeteroAtoms.clicked.connect(self.heteroAtomsConfiguration)

        # Hide the old "Chromatographic separation" groupBox_10 (min/max scale controls removed)
        self.ui.groupBox_10.setVisible(False)

        # Peak width / FWHM filter is now part of the peak picking settings dialog
        self.ui.groupBox_peakWidthFilter.setVisible(False)

        # Add "Peak Picking Settings" button into the MZ processing column, before Peak matching
        self._peakPickingSettings = PeakPickingSettingsDialog.get_default_settings()
        self.peakPickingSettingsButton = QtWidgets.QPushButton("Peak Picking Settings...")
        self.peakPickingSettingsButton.setToolTip("Configure chromatographic peak picking algorithm and parameters")
        # Insert into verticalLayout_10, right before groupBox_11 (Peak matching)
        _idx = self.ui.verticalLayout_10.indexOf(self.ui.groupBox_11)
        if _idx >= 0:
            self.ui.verticalLayout_10.insertWidget(_idx, self.peakPickingSettingsButton)
        else:
            self.ui.verticalLayout_10.addWidget(self.peakPickingSettingsButton)
        self.peakPickingSettingsButton.clicked.connect(self._openPeakPickingSettings)

        self.ui.visualConfig.setVisible(False)
        self.ui.scaleFeatures.setVisible(True)
        self.ui.correctcCount.setVisible(False)
        self.ui.snrThLabel.setVisible(False)
        self.ui.wavelet_SNRThreshold.setVisible(False)
        self.ui.scaleError.setVisible(False)
        self.ui.peak_scaleError.setVisible(False)

        # todo: implement results view with all brackated feature pairs
        self.ui.tabWidget.setTabEnabled(3, True)

        self.loadedMZXMLs = None
        # resultsExperimentChangedNew is used only by showCustomFeature (loads raw mzXML);
        # normal tree selection uses resultsExperimentChanged (reads pre-computed per-file DBs)
        self.ui.msms_SpectraList_exp.itemSelectionChanged.connect(self.plotSelectedMSMSSpectra_exp)
        self.ui.msms_SpectraList.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.ui.msms_SpectraList_exp.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.ui.msms_SpectraList.customContextMenuRequested.connect(lambda pos: self._show_msms_context_menu(self.ui.msms_SpectraList, pos))
        self.ui.msms_SpectraList_exp.customContextMenuRequested.connect(lambda pos: self._show_msms_context_menu(self.ui.msms_SpectraList_exp, pos))

        # Additional MS/MS controls in experiment-results tab
        self.ui.msms_controls_exp = QtWidgets.QHBoxLayout()
        self.ui.btn_msms_similarity_native = QtWidgets.QPushButton("Native similarity")
        self.ui.btn_msms_similarity_labeled = QtWidgets.QPushButton("Labeled similarity")
        self.ui.btn_msms_overview = QtWidgets.QPushButton("MS/MS overview")
        self.ui.btn_msms_export_mgf = QtWidgets.QPushButton("Export MGF")
        self.ui.btn_msms_filter_strings = QtWidgets.QPushButton("Show filter strings")
        self.ui.msms_controls_exp.addWidget(self.ui.btn_msms_similarity_native)
        self.ui.msms_controls_exp.addWidget(self.ui.btn_msms_similarity_labeled)
        self.ui.msms_controls_exp.addWidget(self.ui.btn_msms_overview)
        self.ui.msms_controls_exp.addWidget(self.ui.btn_msms_export_mgf)
        self.ui.msms_controls_exp.addWidget(self.ui.btn_msms_filter_strings)
        self.ui.msms_controls_exp.addStretch(1)
        self.ui.verticalLayout_msms_exp.insertLayout(1, self.ui.msms_controls_exp)
        self.ui.btn_msms_similarity_native.clicked.connect(lambda: self._show_msms_similarity_dialog("native"))
        self.ui.btn_msms_similarity_labeled.clicked.connect(lambda: self._show_msms_similarity_dialog("labeled"))
        self.ui.btn_msms_overview.clicked.connect(self._show_msms_overview)
        self.ui.btn_msms_export_mgf.clicked.connect(self._export_msms_mgf)
        self.ui.btn_msms_filter_strings.clicked.connect(self._show_msms_filter_strings_list)

        self.ui.showCustomFeature_pushButton.clicked.connect(self.showCustomFeature)

        self.ui.checkBox_expPeakArea.setVisible(False)
        self.ui.checkBox_expPeakApexIntensity.setVisible(False)
        self.ui.checkBox_expPeakSNR.setVisible(False)

        # Connect Statistics tab signals
        if hasattr(self.ui, "statisticsWidget"):
            self.ui.statisticsWidget.showFeatureInExperiment.connect(self._showFeatureInExperimentResults)

        p = self.ui.scrollAreaWidgetContents_5.palette()
        p.setColor(self.ui.scrollAreaWidgetContents_5.backgroundRole(), QtCore.Qt.white)
        self.ui.scrollAreaWidgetContents_5.setPalette(p)

        # fetch provided settings and show them as a menu
        try:
            try:
                self.loadSettingsFile(
                    get_main_dir() + "/Settings/defaultSettings.ini",
                    checkExperimentType=False,
                )
            except Exception as ex:
                print(str(ex))
                logging.warning("Warning: Default settings are invalid. Skipping..")
            if os.path.exists(get_main_dir() + "/Settings"):
                menus = {}
                menus[get_main_dir()] = "Predefined"
                for dirname, dirnames, filenames in os.walk(get_main_dir() + "/Settings/"):
                    for filename in filenames:
                        toMenu = self.ui.menuAvailable_Settings
                        cDir = get_main_dir()

                        for dir in filter(
                            None,
                            dirname[len(get_main_dir()) :].replace("\\", "/").split("/"),
                        ):
                            cDir = cDir + "/" + dir
                            if cDir not in menus:
                                menus[cDir] = QtWidgets.QMenu(toMenu)
                                menus[cDir].setTitle(dir)
                                toMenu.addMenu(menus[cDir])

                            toMenu = menus[cDir]

                        ac = QtGui.QAction(self.ui.menuAvailable_Settings)
                        ac.setText(filename.replace(".ini", ""))

                        ac.triggered.connect(
                            functools.partial(
                                self.loadAvailableSettingsFile,
                                os.path.join(dirname, filename).replace("\\", "/"),
                            )
                        )

                        toMenu.addAction(ac)
        except Exception as ex:
            traceback.print_exc()
            logging.error(str(traceback))

            logging.error("Error in %s: %s" % ("mainWindow", str(ex)))
            QtWidgets.QMessageBox.information(self, "MetExtract", "Cannot load settings", QtWidgets.QMessageBox.Ok)

        if disableR:
            self.ui.runTasksContainer.setVisible(False)
            QtWidgets.QMessageBox.warning(
                None,
                "MetExtract",
                "Error: R is not available. Processing of new samples is disabled, however\nalready processed results can be illustrated. ",
                QtWidgets.QMessageBox.Ok,
            )

    def closeEvent(self, event):
        self._contMemoryWatcher = False


def main():
    """Entry point for mexract command"""
    ## add freeze_support for multiprocessing
    freeze_support()

    # parse supplied options
    parser = OptionParser()
    parser.add_option(
        "-m",
        "--module",
        dest="module",
        default="TracExtract",
        metavar="MODULE",
        help="Select 'AllExtract' or 'TracExtract' module (default TracExtract)",
    )
    parser.add_option(
        "-e",
        "--exit",
        dest="exit",
        action="store_true",
        default=False,
        help="Exit after processing (for GUI debugging)",
    )

    (opts, args) = parser.parse_args()

    # create QAPP
    app = QtWidgets.QApplication(sys.argv)
    app.setApplicationName("MetExtract")
    app.setApplicationVersion(MetExtractVersion)
    _style_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "style.css")
    with open(_style_path, "r") as _f:
        app.setStyleSheet(_f.read())

    # setup and show main window
    mainWin = mainWindow(module=opts.module)
    mainWin._contMemoryWatcher = True  # Initialize memory watcher flag

    mainWin.setWindowTitle("MetExtract II (%s)" % MetExtractVersion)
    mainWin.show()

    if not opts.exit:
        QtWidgets.QMessageBox.information(
            None,
            "MetExtract",
            "When you start a new experiment, please <b>change the<br>working directory</b> to your experimental folder.<br>"
            + "You can set the working directory via the menu<br>('Tools'->'Set working directory').<br>"
            + "<br>"
            + "Please also consider <b>copying</b> any databases or<br>other resources to your working directory.<br>"
            + "<br>"
            + "To quickly change the values of drop-down and<br>integer/float spinner controls, <b>hold the CTRL-key<br>and use the mouse-wheel</b>.<br>"
            + "<br>"
            + f"If importing mzML files results in the error<br>of missing files, please find the correct version at<br><b>{OBO_DOWNLOAD_URL}</b>.<br>"
            + "Please download the corresponding obo-file and<br>save it to the folder in the error message.<br>"
            + "You can also open this page via the menu<br>(<b>'Tools'->'Download OBO files'</b>).<br>"
            + "<br>"
            + "In the experiment-results MS/MS tab you can <b>filter<br>spectra by their filter string</b> (cvParam MS:1000512)<br>"
            + "using a regular expression. Leave it empty to show all<br>spectra; a capturing group is shown in the first column.<br>"
            + "The 'Show filter strings' button lists all loaded filter<br>strings.<br>"
            + "<br>"
            + "To generate a template for a database, select<br><b>'Download Database Template'</b> from the 'Tools' menu.",
            QtWidgets.QMessageBox.Ok,
        )

    # status bar info thread
    def updateMemoryInfo():
        if mainWin._contMemoryWatcher:
            try:
                mainWin.ui.version.setText("%.2f MB memory // %s" % (memory_usage_psutil(), mainWin.ui.version.versionText))
            except Exception:
                mainWin.ui.version.setText("%s" % (mainWin.ui.version.versionText))
            if mainWin._contMemoryWatcher:
                threading.Timer(1, updateMemoryInfo).start()

    updateMemoryInfo()

    if opts.exit:
        QtWidgets.QApplication.exit()
        mainWin._contMemoryWatcher = False
    else:
        sys.exit(app.exec())


if __name__ in ["__main__", "src.MExtract"]:
    main()
