from __future__ import print_function, division, absolute_import

from os.path import basename

import string
import random

import time

from copy import deepcopy

import polars as pl

from .utils import Bunch

# functionality to store / read a table using polars DataFrames
# methods to insert / alter / delete data are provided (including bulk-update)


# class for storing metadata of one column
class ColDef:
    def __init__(self, name, format):
        self.name = name
        self.format = format

    def getName(self):
        return self.name

    def getFormat(self):
        return self.format

    def __str__(self):
        return "name: %s (format:%s)" % (self.name, self.format)


# class table provides generic methods to insert / alter / delete data stored in a polars DataFrame
# the use is abstracted as much as possible from polars syntax
class Table:
    def __initEmpty__(self, tableName=None):
        if tableName is None:
            tableName = "".join(random.choice(string.ascii_uppercase + string.ascii_lowercase) for x in range(6))
        self.tableName = tableName

        # Initialize with an empty DataFrame with an __internalID column
        self.df = pl.DataFrame({"__internalID": pl.Series([], dtype=pl.Int64)})
        self.comments = []
        self.saved = True

    def __init__(self, headers=None, rows=None, tableName=None, comments=None):
        if headers is None and rows is None:
            self.__initEmpty__(tableName)
            return

        assert headers is not None
        assert rows is not None

        for row in rows:
            if len(headers) < len(row):
                print(len(headers), len(row), row)
            assert len(headers) >= len(row), "Unequal data count"

        if tableName is None:
            tableName = "".join(random.choice(string.ascii_uppercase + string.ascii_lowercase) for x in range(6))
        self.tableName = tableName

        self.saved = True
        self.comments = comments if comments else []

        # Prepare data dictionary
        data_dict = {"__internalID": list(range(len(rows)))}

        # Process headers and rows
        processed_headers = []
        for h in headers:
            # Prepend underscore if header starts with a digit
            if str.isdigit(h[0]):
                h = "_" + h
            # Replace problematic characters
            h = h.replace("-", "__")
            h = h.replace(" ", "__")
            processed_headers.append(h)

        # Pad rows with empty strings if needed
        padded_rows = []
        for row in rows:
            padded_row = list(row) + [""] * (len(processed_headers) - len(row))
            padded_rows.append(padded_row)

        # Build column data
        for i, header in enumerate(processed_headers):
            col_data = [row[i] if i < len(row) else "" for row in padded_rows]
            data_dict[header] = col_data

        # Create DataFrame with automatic type inference
        self.df = pl.DataFrame(data_dict)

        # Try to cast columns to appropriate types
        for col in processed_headers:
            try:
                # Try integer first
                self.df = self.df.with_columns(pl.col(col).cast(pl.Int64, strict=False))
            except:
                try:
                    # Try float next
                    self.df = self.df.with_columns(pl.col(col).cast(pl.Float64, strict=False))
                except:
                    # Keep as string
                    self.df = self.df.with_columns(pl.col(col).cast(pl.Utf8, strict=False))

    def renameColumn(self, oldColName, newColName):
        if oldColName not in self.df.columns:
            raise Exception("Column %s not found" % oldColName)
        if newColName in self.df.columns:
            raise Exception("New column %s already exists" % newColName)

        self.df = self.df.rename({oldColName: newColName})
        self.saved = False

    def addPrefixToAllColumns(self, prefix):
        for col in self.getColumns():
            colName = col.name
            self.renameColumn(colName, prefix + colName)

    # get all available columns of the table
    def getColumns(self):
        cols = []
        for col_name in self.df.columns:
            if col_name != "__internalID":
                dtype = self.df[col_name].dtype
                # Map polars dtypes to SQLite-like format strings
                if dtype in [pl.Int8, pl.Int16, pl.Int32, pl.Int64, pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64]:
                    format_str = "integer"
                elif dtype in [pl.Float32, pl.Float64]:
                    format_str = "real"
                else:
                    format_str = "text"
                cols.append(ColDef(col_name, format_str))
        return cols

    # test if a specific column is present in the table
    def hasColumn(self, col):
        return col in self.df.columns

    # bulk-update: perform a function for each row present in the table
    def applyFunction(
        self,
        func,
        where="",
        pwMaxSet=None,
        pwValSet=None,
        showProgress=False,
        funcName="",
        printAfter=1,
    ):
        # Apply filtering if where clause is provided
        if len(where) > 0:
            # Convert simple where clause to polars expression
            # This is a simplified implementation - may need enhancement for complex queries
            filtered_df = self.df.filter(pl.SQLContext(tables={"df": self.df}).execute(f"SELECT * FROM df WHERE {where}"))
        else:
            filtered_df = self.df

        cols = self.getColumns()
        numOfRows = len(filtered_df)

        updates = []
        done = 0
        started = time.time()

        # Iterate through rows
        for row_idx in range(numOfRows):
            row = filtered_df.row(row_idx, named=True)
            iID = row["__internalID"]

            # Prepare data dict (excluding __internalID)
            data = {col.name: row[col.name] for col in cols}

            # Apply the function
            newData = func(deepcopy(data))

            # Find changed values
            changes = {}
            for col in cols:
                if col.name in newData.keys() and newData[col.name] != data[col.name]:
                    changes[col.name] = str(newData[col.name])

            if len(changes) > 0:
                updates.append((iID, changes))

            done += 1
            if showProgress:
                elapsed = time.time() - started
                if elapsed > printAfter:
                    doneP = 1.0 * done / numOfRows
                    s = (elapsed * (1 - doneP) / doneP) / 60.0
                    print(
                        "\rApplying (function: %s).. |%-30s| %.1f%% (approximately %.1f minutes remaining, running for %.1f minutes, total %.1f minutes)"
                        % (
                            funcName,
                            "*" * int(doneP * 30),
                            doneP * 100,
                            s,
                            elapsed / 60.0,
                            s + elapsed / 60.0,
                        ),
                    )

        if showProgress:
            s = (time.time() - started) / 60.0
            print("\rApplying (function: %s) took %.1f minutes                                                                                                                                                              " % (funcName, s))

        if pwMaxSet is not None:
            pwMaxSet(len(updates))
        if pwValSet is not None:
            pwValSet(0)

        doneS = 0
        started = time.time()

        # Apply updates
        for iID, changes in updates:
            for col_name, new_value in changes.items():
                mask = self.df["__internalID"] == iID
                self.df = self.df.with_columns(pl.when(mask).then(pl.lit(new_value)).otherwise(pl.col(col_name)).alias(col_name))

            done += 1
            if pwValSet is not None:
                pwValSet(done)

            doneS += 1
            if showProgress:
                elapsed = time.time() - started
                if elapsed > printAfter:
                    doneP = 1.0 * doneS / numOfRows
                    s = (elapsed * (1 - doneP) / doneP) / 60.0
                    print(
                        "\rUpdating.. |%-30s| %.1f%% (approximately %.1f minutes remaining)" % ("*" * int(doneP * 30), doneP * 100, s),
                    )

        if showProgress:
            s = (time.time() - started) / 60.0
            print("\rUpdating took %.1f minutes                                                                                                        " % s)

        self.saved = False

    # return the entire table
    def getData(self, cols=None, where=None, orderby=None, getResultsAsBunchObjects=False):
        ret = []

        df = self.df

        # Apply where filter if provided
        if where is not None and len(where) > 0:
            # Use SQL context for complex where clauses
            try:
                ctx = pl.SQLContext(tables={"tbl": df})
                df = ctx.execute(f"SELECT * FROM tbl WHERE {where}")
            except:
                # Fallback: assume it's a simple expression
                pass

        # Select columns
        if cols is None:
            cols = [x.getName() for x in self.getColumns()]

        # Handle columns that start with digits
        display_cols = []
        for col in cols:
            if col[0].isdigit():
                display_cols.append("_" + col)
            else:
                display_cols.append(col)

        # Select only the requested columns
        df = df.select(display_cols)

        # Apply ordering if provided
        if orderby is not None and len(orderby) > 0:
            # Parse orderby string (e.g., "col1, col2 DESC")
            order_parts = [part.strip() for part in orderby.split(",")]
            order_cols = []
            descending = []
            for part in order_parts:
                tokens = part.split()
                col_name = tokens[0]
                desc = len(tokens) > 1 and tokens[1].upper() == "DESC"
                order_cols.append(col_name)
                descending.append(desc)
            df = df.sort(order_cols, descending=descending)

        # Convert to return format
        for row_dict in df.iter_rows(named=True):
            if getResultsAsBunchObjects:
                ret.append(Bunch(_addFollowing=row_dict))
            else:
                values = list(row_dict.values())
                if len(values) == 1:
                    ret.append(values[0])
                else:
                    ret.append(tuple(values))

        return ret

    # set data specifically (similar to SQL update)
    def setData(self, cols, vals, where=None):
        assert len(cols) == len(vals)

        # Build update mask
        if where is not None:
            # Use SQL context for where clause
            try:
                ctx = pl.SQLContext(tables={"tbl": self.df})
                filtered_df = ctx.execute(f"SELECT __internalID FROM tbl WHERE {where}")
                filtered_ids = filtered_df["__internalID"].to_list()
                mask = pl.col("__internalID").is_in(filtered_ids)
            except:
                mask = pl.lit(True)
        else:
            mask = pl.lit(True)

        # Apply updates
        for col_name, val in zip(cols, vals):
            self.df = self.df.with_columns(pl.when(mask).then(pl.lit(val)).otherwise(pl.col(col_name)).alias(col_name))

        self.saved = False

    # add a new column to the table (in-memory only)
    def addColumn(self, col, colType, defaultValue=""):
        if self.hasColumn(col):
            raise Exception("Column %s already exists" % col)

        # Map SQLite types to polars types
        if colType.lower() == "integer":
            dtype = pl.Int64
        elif colType.lower() == "real":
            dtype = pl.Float64
        else:
            dtype = pl.Utf8

        # Add column with default value
        self.df = self.df.with_columns(pl.lit(defaultValue).cast(dtype).alias(col))
        self.saved = False

    # duplicate a new column to the table (in-memory only)
    def duplicateColumn(self, col, newColName):
        if self.hasColumn(newColName):
            raise Exception("Column %s already exists" % newColName)

        if not self.hasColumn(col):
            raise Exception("Column %s does not exists" % col)

        # Duplicate column
        self.df = self.df.with_columns(pl.col(col).alias(newColName))
        self.saved = False

    # append a new row at the end of the table
    def addRow(self, data):
        for dCol in data.keys():
            if not (self.hasColumn(dCol)):
                raise Exception("Column %s not found" % dCol)

        # Get next internal ID
        max_id = self.df["__internalID"].max()
        next_id = 0 if max_id is None else max_id + 1

        # Prepare row data with proper type casting
        row_data = {"__internalID": [next_id]}
        for col in self.df.columns:
            if col == "__internalID":
                continue
            value = data.get(col, "")

            # Try to match the column's existing dtype
            col_dtype = self.df[col].dtype
            if col_dtype in [pl.Int8, pl.Int16, pl.Int32, pl.Int64]:
                try:
                    value = int(value) if value != "" else None
                except:
                    value = None
            elif col_dtype in [pl.Float32, pl.Float64]:
                try:
                    value = float(value) if value != "" else None
                except:
                    value = None

            row_data[col] = [value]

        # Create new row DataFrame with matching schema
        new_row = pl.DataFrame(row_data, schema=self.df.schema)
        self.df = pl.concat([self.df, new_row], how="vertical")
        self.saved = False

    # reset entire column (affects all rows of the table)
    def resetColumn(self, col, value=""):
        if not (self.hasColumn(col)):
            raise Exception("Column %s not found" % col)

        self.df = self.df.with_columns(pl.lit(value).alias(col))
        self.saved = False

    # HELPER METHOD (kept for compatibility but no longer needed)
    def __updateTableName(self, sql):
        return sql.replace(":table:", self.tableName)

    # accepts a generic query (kept for compatibility)
    def updateQuery(self, sql):
        allCols = []
        for col in self.getColumns():
            allCols.append(col.name)
        allCols = ", ".join(allCols)

        sql = sql.replace(":table:", self.tableName)
        sql = sql.replace(":allCols:", allCols)

        return sql

    # tests, if the data has been edited
    def isSaved(self):
        return self.saved

    # sets, if the data has been edited
    def save(self):
        self.saved = True

    def getComments(self):
        return self.comments

    def deleteComments(self):
        self.comments = []

    def addComment(self, comment):
        self.comments.append(comment)

    # de-constructor: no longer needed for polars (no connection to close)
    def __del__(self):
        pass


# abstract class to import and save a table
class TableUtils:
    @staticmethod
    def readFile(file, fType="", commentLineStart="#", **args):
        fType = fType.lower()
        bn = basename(file).lower()

        if fType == "xls" or bn.endswith(".xls") or fType == "xlsx" or bn.endswith(".xlsx"):
            if not ("sheetName" in args):
                args["sheetName"] = ""
            return TableUtilsXLS.readFile(file, sheetName=args["sheetName"], commentLineStart=commentLineStart)
        elif fType == "csv" or bn.endswith(".csv"):
            if not ("delim" in args):
                args["delim"] = ";"
            return TableUtilsCSV.readFile(file, delim=args["delim"], commentLineStart=commentLineStart)
        elif fType == "tsv" or bn.endswith(".tsv") or bn.endswith(".txt") or bn.endswith(".csv"):
            if not ("delim" in args):
                args["delim"] = "\t"
            return TableUtilsCSV.readFile(file, delim=args["delim"], commentLineStart=commentLineStart)
        elif fType == "sqlite" or bn.endswith(".sqlite"):
            if not ("tableName" in args):
                args["tableName"] = ""
            return TableUtilsSQL.readFile(file, tableName=args["tableName"])
        else:
            raise Exception("Unknown file type. Please use xls, tsv, sqlite")

    @staticmethod
    def saveFile(table, file, fType="", writeComments=True, **args):
        fType = fType.lower()
        bn = basename(file).lower()

        if not ("where" in args):
            args["where"] = ""
        if not ("order" in args):
            args["order"] = "__internalID"
        if not ("select" in args):
            args["select"] = ""
        if not ("cols" in args):
            args["cols"] = ["*"]
        if "*" in args["cols"]:
            for col in table.getColumns():
                if col.name != "__internalID" and col.name not in args["cols"]:
                    args["cols"].insert(args["cols"].index("*"), col.name)
            args["cols"].remove("*")

        if fType == "xls" or bn.endswith(".xls") or fType == "xlsx" or bn.endswith(".xlsx"):
            if not ("sheetName" in args):
                args["sheetName"] = ""
            return TableUtilsXLS.saveFile(
                table,
                file,
                sheetName=args["sheetName"],
                where=args["where"],
                order=args["order"],
                select=args["select"],
                cols=args["cols"],
                writeComments=writeComments,
            )
        elif fType == "csv" or bn.endswith(".csv"):
            if not ("delim" in args):
                args["delim"] = ";"
            if not ("newLine" in args):
                args["newLine"] = ""
            return TableUtilsCSV.saveFile(
                table,
                file,
                delim=args["delim"],
                newLine=args["newLine"],
                where=args["where"],
                order=args["order"],
                select=args["select"],
                cols=args["cols"],
                writeComments=writeComments,
            )
        elif fType == "tsv" or bn.endswith(".tsv") or bn.endswith(".txt"):
            if not ("delim" in args):
                args["delim"] = "\t"
            if not ("newLine" in args):
                args["newLine"] = ""
            return TableUtilsCSV.saveFile(
                table,
                file,
                delim=args["delim"],
                newLine=args["newLine"],
                where=args["where"],
                order=args["order"],
                select=args["select"],
                cols=args["cols"],
                writeComments=writeComments,
            )
        elif fType == "sqlite" or bn.endswith(".sqlite"):
            if not ("tableName" in args):
                args["tableName"] = ""
            return TableUtilsSQL.saveFile(table, file, tableName=args["tableName"], writeComments=writeComments)
        else:
            raise Exception("Unknown file type. Please use xls, tsv, sqlite")


# read / write data table from / to a CSV file
class TableUtilsCSV:
    @staticmethod
    def readFile(file, delim="\t", commentLineStart="#"):
        # Read comments first
        comments = []
        with open(file, "r", encoding="utf-8") as fi:
            for line in fi:
                if line.startswith(commentLineStart):
                    comments.append(line.strip())

        # Read CSV using polars
        try:
            df = pl.read_csv(
                file,
                separator=delim,
                comment_prefix=commentLineStart,
                infer_schema_length=10000,
                ignore_errors=True,
            )
        except Exception as e:
            # Fallback to manual parsing if polars fails
            with open(file, "rb") as fi:
                delim_bytes = delim.encode("utf-8")
                commentLineStart_bytes = commentLineStart.encode("utf-8")

                headers = []
                rows = []
                i = 0
                for line in fi:
                    if len(line) > 0 and not line.startswith(commentLineStart_bytes):
                        cells = line.strip().split(delim_bytes)
                        if i == 0:
                            headers = [c.decode("utf-8").replace("-", "_") for c in cells]
                        else:
                            rows.append([c.decode("utf-8") for c in cells])
                        i += 1

                return Table(headers=headers, rows=rows, comments=comments)

        # Convert polars DataFrame to Table format
        headers = [col.replace("-", "_") for col in df.columns]
        rows = []
        for row_dict in df.iter_rows(named=True):
            rows.append([str(row_dict[col]) if row_dict[col] is not None else "" for col in df.columns])

        return Table(headers=headers, rows=rows, comments=comments)

    @staticmethod
    def saveFile(
        table,
        file,
        where="",
        order="",
        delim="\t",
        newLine="\n",
        select="",
        cols=None,
        writeComments=True,
    ):
        if cols is None:
            cols = []

        if len(delim) == 0:
            delim = "\t"
        if len(newLine) == 0:
            newLine = "\n"

        # Get the dataframe to write
        df = table.df

        # Apply where filter
        if len(where) > 0:
            try:
                ctx = pl.SQLContext(tables={"tbl": df})
                df = ctx.execute(f"SELECT * FROM tbl WHERE {where}")
            except:
                pass

        # Apply ordering (before selecting columns, so __internalID is available)
        if len(order) > 0:
            order_cols = [col.strip() for col in order.split(",")]
            # Only sort by columns that exist in the dataframe
            order_cols = [col for col in order_cols if col in df.columns]
            if order_cols:
                df = df.sort(order_cols)

        # Select columns (exclude __internalID unless explicitly requested)
        if len(cols) > 0:
            df = df.select(cols)
        else:
            # Exclude __internalID by default
            df = df.select([col for col in df.columns if col != "__internalID"])

        # Write CSV using polars
        df.write_csv(file, separator=delim)

        # Append comments if requested
        if writeComments and len(table.getComments()) > 0:
            with open(file, "a", encoding="utf-8") as fo:
                for comment in table.getComments():
                    fo.write(str(comment).replace(delim, "-DELIM-"))
                    fo.write(newLine)


from openpyxl import Workbook, load_workbook


# read / write data table from / to an excel file
class TableUtilsXLS:
    @staticmethod
    def readFile(file, sheetName="", tableName=None, useColumns=None, commentLineStart="#"):
        if sheetName == "":
            bn = basename(file)
            rf = bn.rfind(".")
            if rf > -1:
                sheetName = bn[:rf]
            else:
                sheetName = bn

        rb = load_workbook(file)

        ind = 0
        sheetInd = -1
        for s in rb.sheetnames:
            if s == sheetName:
                sheetInd = ind
            ind = ind + 1
        assert sheetInd != -1, "Sheet not found"
        sh = rb[sheetName]

        # Read data manually to handle comments and special cases
        i = 0
        remainingCols = True
        headers = []
        rows = []
        comments = []
        colIDs = []
        firstColFound = False
        curRow = 0

        while not firstColFound:
            val = sh.cell(row=curRow + 1, column=1).value
            if val is not None and str(val).startswith(commentLineStart):
                comments.append(str(val))
                curRow += 1
                continue
            else:
                firstColFound = True

            i = 0
            while remainingCols:
                col = i + 1
                row = curRow + 1
                try:
                    cell_val = sh.cell(row=row, column=col).value
                    if cell_val is None or str(cell_val) == "":
                        remainingCols = False
                        continue
                except:
                    remainingCols = False
                    continue

                header = str(cell_val)
                if useColumns is None or header in useColumns:
                    headers.append(header)
                    colIDs.append(i)
                i = i + 1

        r = 1 + curRow
        allEmpty = False
        while not allEmpty:
            val = sh.cell(row=r + 1, column=1).value
            if val is not None and str(val).startswith(commentLineStart):
                comments.append(str(val))
                r = r + 1
                continue

            rowValues = []
            allEmpty = True
            for j in colIDs:
                col = j + 1
                row = r + 1
                try:
                    cell_val = sh.cell(row=row, column=col).value
                    if cell_val is None or str(cell_val) == "":
                        rowValues.append("")
                        continue
                except:
                    rowValues.append("")
                    continue
                rowValues.append(str(cell_val))
                allEmpty = False
            if not allEmpty:
                rows.append(rowValues)

            r = r + 1

        ret = Table(headers=headers, rows=rows, tableName=tableName, comments=comments)
        return ret

    @staticmethod
    def saveFile(
        table,
        file,
        sheetName="",
        where="",
        order="",
        select="",
        cols=None,
        writeComments=True,
    ):
        if cols is None:
            cols = []

        if sheetName == "":
            bn = basename(file)
            rf = bn.rfind(".")
            if rf > -1:
                sheetName = bn[:rf]
            else:
                sheetName = bn

        # Get the dataframe to write
        df = table.df

        # Apply where filter
        if len(where) > 0:
            try:
                ctx = pl.SQLContext(tables={"tbl": df})
                df = ctx.execute(f"SELECT * FROM tbl WHERE {where}")
            except:
                pass

        # Apply ordering (before selecting columns, so __internalID is available)
        if len(order) > 0:
            order_cols = [col.strip() for col in order.split(",")]
            # Only sort by columns that exist in the dataframe
            order_cols = [col for col in order_cols if col in df.columns]
            if order_cols:
                df = df.sort(order_cols)

        # Select columns (exclude __internalID unless explicitly requested)
        if len(cols) > 0:
            df = df.select(cols)
        else:
            # Exclude __internalID by default
            df = df.select([col for col in df.columns if col != "__internalID"])

        rb = -1
        sheet = -1
        import os.path

        if os.path.isfile(file):
            rb = load_workbook(file)
            sheet = rb.create_sheet()
            sheet.title = sheetName
        else:
            rb = Workbook()
            sheet = rb.active
            sheet.title = sheetName

        # Write headers
        for i, col_name in enumerate(df.columns):
            sheet.cell(row=1, column=i + 1).value = col_name

        # Write data
        for row_idx, row_dict in enumerate(df.iter_rows(named=True)):
            for col_idx, col_name in enumerate(df.columns):
                sheet.cell(row=row_idx + 2, column=col_idx + 1).value = row_dict[col_name]

        # Write comments
        if writeComments:
            j = len(df) + 2
            for comment in table.getComments():
                sheet.cell(row=j, column=1).value = comment
                j += 1

        rb.save(file)


# read / write data table from / to an SQL table
class TableUtilsSQL:
    @staticmethod
    def readFile(file, tableName=""):
        raise Exception("Not implemented yet", "Not implemented yet")

    @staticmethod
    def saveFile(table, file, tableName, where="", order="", select="", writeComments=True):
        raise Exception("Not implemented yet", "Not implemented yet")
