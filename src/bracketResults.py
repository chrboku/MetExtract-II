from __future__ import print_function, division, absolute_import
import logging

import polars as pl
import xlsxwriter
from operator import itemgetter
import os
from math import floor
import ast
from collections import defaultdict

from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.graphics import renderPDF

from .XICAlignment import XICAlignment

from .utils import (
    ChromPeakPair,
    getSubGraphs,
    Bunch,
    natSort,
    get_main_dir,
    getSubGraphsFromDictDict,
    CallBackMethod,
)
from .runIdentification import ChromPeakPair
from .utils import getDBSuffix
from .PolarsDB import PolarsDB
from .MZHCA import HierarchicalClustering, cutTreeSized

import base64
from pickle import dumps, loads

from math import isnan

import time

from . import HCA_general
from .Chromatogram import Chromatogram
from .utils import mean, sd, corr


from . import exportAsFeatureML


# HELPER METHOD for writing first page of PDF (unused)
def _writeFirstPage(pdf, groupSizePPM, maxTimeDeviation, align, nPolynom):
    curHeight = 780

    pdf.drawString(60, curHeight, "Max. Group Size")
    pdf.drawString(200, curHeight, "%.1f" % groupSizePPM)
    curHeight -= 20

    pdf.drawString(60, curHeight, "Max. Time Deviation")
    pdf.drawString(200, curHeight, "%.2f seconds" % maxTimeDeviation)
    curHeight -= 20

    pdf.drawString(60, curHeight, "Align")
    pdf.drawString(200, curHeight, "%s" % ("yes" if align else "no"))
    curHeight -= 20

    if align:
        pdf.drawString(60, curHeight, "Polynom Order")
        pdf.drawString(200, curHeight, "%d" % nPolynom)
        curHeight -= 20

    pdf.showPage()


# store used configuration to DB file
def writeConfigToDB(
    db,
    align,
    file,
    groupSizePPM,
    maxLoading,
    maxTimeDeviation,
    xCounts,
    nPolynom,
    negativeScanEvent,
    positiveScanEvent,
    rVersion,
    meVersion,
):
    db.insert_row("config", {"key": "MEVersion", "value": str(meVersion)})
    db.insert_row("config", {"key": "RVersion", "value": str(rVersion)})
    db.insert_row("config", {"key": "FPBRACK_xCounts", "value": str(xCounts)})
    db.insert_row("config", {"key": "FPBRACK_groupSizePPM", "value": str(groupSizePPM)})
    db.insert_row("config", {"key": "FPBRACK_positiveScanEvent", "value": str(positiveScanEvent)})
    db.insert_row("config", {"key": "FPBRACK_negativeScanEvent", "value": str(negativeScanEvent)})
    db.insert_row("config", {"key": "FPBRACK_maxTimeDeviation", "value": str(maxTimeDeviation)})
    db.insert_row("config", {"key": "FPBRACK_maxLoading", "value": str(maxLoading)})
    db.insert_row("config", {"key": "FPBRACK_file", "value": str(file)})
    db.insert_row("config", {"key": "FPBRACK_align", "value": str(align)})
    db.insert_row("config", {"key": "FPBRACK_nPolynom", "value": str(nPolynom)})


# bracket results
def bracketResults(
    indGroups,
    xCounts,
    groupSizePPM,
    positiveScanEvent=None,
    negativeScanEvent=None,
    maxTimeDeviation=0.36 * 60,
    maxLoading=1,
    file="./results.tsv",
    align=True,
    nPolynom=1,
    expPeakArea=True,
    expApexIntensity=True,
    expPeakSNR=False,
    pwMaxSet=None,
    pwValSet=None,
    pwTextSet=None,
    rVersion="",
    meVersion="",
    generalProcessingParams=Bunch(),
    start=0,
):
    excel_file = file.replace(".tsv", ".xlsx")
    if os.path.exists(excel_file):
        os.remove(excel_file)

    # create results ParquetDB tables
    try:
        writePDF = False
        # used for debug purposes
        colos = [colors.red]

        cpf = get_main_dir() + "/src/XICAlignment.r"  # initialise chromatographic alignment script
        xicAlign = XICAlignment(cpf)

        results = []

        # generate results for each input file
        i = 1
        for key in natSort(indGroups.keys()):
            files = indGroups[key]

            for ident in files:
                conn = None
                curs = None
                if os.path.isfile(ident + getDBSuffix()):
                    conn = PolarsDB(ident + getDBSuffix())
                    curs = conn.cursor()
                fname = ident
                if ".mzxml" in ident.lower():
                    fname = ident[max(ident.rfind("/") + 1, ident.rfind("\\") + 1) : ident.lower().rfind(".mzxml")]
                if ".mzml" in ident.lower():
                    fname = ident[max(ident.rfind("/") + 1, ident.rfind("\\") + 1) : ident.lower().rfind(".mzml")]
                b = Bunch(
                    filePath=ident,
                    fileName=fname,
                    featurePairs=[],
                    conn=conn,
                    curs=curs,
                )

                results.append(b)

            i += 1

        # Initialize data collection for Excel output
        excel_data = {
            "Num": [],
            "Comment": [],
            "MZ": [],
            "L_MZ": [],
            "D_MZ": [],
            "MZ_Range": [],
            "RT": [],
            "RT_Range": [],
            "PeakScalesNL": [],
            "Xn": [],
            "Charge": [],
            "ScanEvent": [],
            "Ionisation_Mode": [],
            "Tracer": [],
            "OGroup": [],
            "Ion": [],
            "Loss": [],
            "M": [],
        }

        # Add file-specific columns
        for res in results:
            fname = res.fileName
            if ".mzxml" in fname.lower():
                fname = fname[max(fname.rfind("/") + 1, fname.rfind("\\") + 1) : fname.lower().rfind(".mzxml")]
            if ".mzml" in fname.lower():
                fname = fname[max(fname.rfind("/") + 1, fname.rfind("\\") + 1) : fname.lower().rfind(".mzml")]

            excel_data[fname + "_Area_N"] = []
            excel_data[fname + "_Area_L"] = []
            excel_data[fname + "_Abundance_N"] = []
            excel_data[fname + "_Abundance_L"] = []
            excel_data[fname + "_SNR_N"] = []
            excel_data[fname + "_SNR_L"] = []
            excel_data[fname + "_FID"] = []
            excel_data[fname + "_GroupID"] = []

        excel_data["doublePeak"] = []

        # get ionisation modes scan events
        ionModes = {}
        if positiveScanEvent != "None":
            ionModes["+"] = positiveScanEvent
        if negativeScanEvent != "None":
            ionModes["-"] = negativeScanEvent
        assert len(ionModes) > 0

        totalChromPeaks = 0
        totalChromPeaksProc = 0
        tracersDeltaMZ = {}

        ## TODO this used the last open file. May be invalid. improve
        isMetabolisationExperiment = False
        for res in results:
            if res.conn is not None:
                config_table = res.conn.get_table("config")
                if not config_table.is_empty():
                    config_rows = config_table.filter(pl.col("key") == "metabolisationExperiment")
                    if len(config_rows) > 0:
                        isMetabolisationExperiment = str(config_rows[0, "value"]).lower() == "true"
                        break

        ## TODO this used the last open file. May be invalid. improve
        if isMetabolisationExperiment:
            for res in results:
                if res.conn is not None:
                    tracer_table = res.conn.get_table("tracerConfiguration")
                    if not tracer_table.is_empty():
                        for row in tracer_table.iter_rows(named=True):
                            tracerName = str(row["name"])
                            dmz = float(row["deltaMZ"])

                            if tracerName not in tracersDeltaMZ:
                                tracersDeltaMZ[tracerName] = dmz

                            if abs(tracersDeltaMZ[tracerName] - dmz) < 0.00001:
                                logging.warning("Warning: Tracers have not been configured identical in all measurement files")
        else:
            ## TODO this used the last open file. May be invalid. improve
            for res in results:
                if res.conn is not None:
                    config_table = res.conn.get_table("config")
                    if not config_table.is_empty():
                        config_rows = config_table.filter(pl.col("key") == "xOffset")
                        if len(config_rows) > 0:
                            dmz = float(config_rows[0, "value"])

                            if "FLE" not in tracersDeltaMZ:
                                tracersDeltaMZ["FLE"] = dmz

                            if abs(tracersDeltaMZ["FLE"] - dmz) < 0.00001:
                                logging.warning("Warning: Tracers have not been configured identical in all measurement files")

        grp = 1
        if writePDF:
            pdf = canvas.Canvas(file + "_0.pdf")
        if writePDF:
            _writeFirstPage(pdf, groupSizePPM, maxTimeDeviation, align, nPolynom)
        xy = 1

        curNum = 1

        xCounts = set()
        for res in results:
            if res.conn is not None:
                chrom_peaks = res.conn.get_table("chromPeaks")
                if not chrom_peaks.is_empty():
                    unique_xcounts = chrom_peaks.select(pl.col("xcount")).unique()
                    for row in unique_xcounts.iter_rows(named=True):
                        xCounts.add(row["xcount"])
        xCounts = sorted(list(xCounts))

        # xCountshelp = []
        # a=xCounts.replace(" ", "").replace(";", ",").split(",")
        # for j in a:
        #     if "-" in j:
        #         xCountshelp.extend(range(int(j[0:j.find("-")]), int(j[j.find("-")+1:len(j)])+1))
        #     else:
        #         xCountshelp.append(int(j))
        # xCounts=sorted(list(set(xCountshelp)))

        # prefetch number of bracketing steps for the progress bar
        totalThingsToDo = 0
        for ionMode in ionModes:
            for tracer in tracersDeltaMZ:
                for xCount in xCounts:
                    for cLoading in range(maxLoading, 0, -1):
                        totalThingsToDo += 1

        doneSoFar = 0
        doneSoFarPercent = 0
        if pwMaxSet is not None:
            pwMaxSet.put(Bunch(mes="max", val=totalThingsToDo))

        # bracket data
        # process ionModes, tracers, xCount and loading separately
        for ionMode in ionModes:
            scanEvent = ionModes[ionMode]
            for xCount in xCounts:
                for cLoading in range(maxLoading, 0, -1):
                    # check each processed LC-HRMS file if the used data processing parameters match
                    for res in results:
                        res.featurePairs = []

                        if pwTextSet is not None:
                            # Log time used for bracketing
                            elapsed = (time.time() - start) / 60.0
                            hours = ""
                            if elapsed >= 60.0:
                                hours = "%d hours " % (elapsed // 60)
                            mins = "%.2f mins" % (elapsed % 60.0)
                            pwTextSet.put(
                                Bunch(
                                    mes="text",
                                    val="<p align='right' >%s%s elapsed</p>\n\n\n\nProcessing \n  (Ionmode: %s, XCount: %s, Charge: %d) \n  File: %s"
                                    % (
                                        hours,
                                        mins,
                                        ionMode,
                                        xCount,
                                        cLoading,
                                        res.fileName,
                                    ),
                                )
                            )
                        if res.conn is not None:
                            chrom_peaks = res.conn.get_table("chromPeaks")
                            feature_groups = res.conn.get_table("featureGroupFeatures")
                            tracer_config = res.conn.get_table("tracerConfiguration")

                            if not chrom_peaks.is_empty():
                                # Filter chromPeaks
                                filtered_peaks = chrom_peaks.filter((pl.col("ionMode") == ionMode) & (pl.col("xcount") == xCount) & (pl.col("Loading") == cLoading))

                                # Join with feature groups if available
                                if not feature_groups.is_empty():
                                    filtered_peaks = filtered_peaks.join(feature_groups.select(["fID", "fGroupID"]), left_on="id", right_on="fID", how="left")
                                else:
                                    filtered_peaks = filtered_peaks.with_columns(pl.lit(None).alias("fGroupID"))

                                # Join with tracer configuration if available
                                if not tracer_config.is_empty():
                                    filtered_peaks = filtered_peaks.join(tracer_config.select(["id", "name"]), left_on="tracer", right_on="id", how="left").rename({"name": "tracerName"})
                                else:
                                    filtered_peaks = filtered_peaks.with_columns(pl.lit(None).alias("tracerName"))

                                print(f"File {res.fileName}: Found {len(filtered_peaks)} chromPeaks for IonMode {ionMode}, XCount {xCount}, Loading {cLoading}")

                                for row in filtered_peaks.iter_rows(named=True):
                                    try:
                                        cp = ChromPeakPair(
                                            id=row["id"],
                                            tracer=row["tracer"],
                                            NPeakCenter=row["NPeakCenter"],
                                            NPeakCenterMin=row["NPeakCenterMin"],
                                            NPeakScale=row["NPeakScale"],
                                            NSNR=row["NSNR"],
                                            NPeakArea=row["NPeakArea"],
                                            mz=row["mz"],
                                            lmz=row["lmz"],
                                            tmz=row["tmz"],
                                            xCount=str(row["xcount"]),
                                            LPeakCenter=row["LPeakCenter"],
                                            LPeakCenterMin=row["LPeakCenterMin"],
                                            LPeakScale=row["LPeakScale"],
                                            LSNR=row["LSNR"],
                                            LPeakArea=row["LPeakArea"],
                                            loading=row["Loading"],
                                            fGroupID=row.get("fGroupID"),
                                            tracerName=row.get("tracerName"),
                                            ionMode=str(row["ionMode"]),
                                            NPeakAbundance=float(row["NPeakAbundance"]),
                                            LPeakAbundance=float(row["LPeakAbundance"]),
                                            NBorderLeft=float(row["NBorderLeft"]),
                                            NBorderRight=float(row["NBorderRight"]),
                                            LBorderLeft=float(row["LBorderLeft"]),
                                            LBorderRight=float(row["LBorderRight"]),
                                        )

                                        assert cp.ionMode in ionModes.keys()
                                        res.featurePairs.append(cp)
                                    except TypeError as err:
                                        print(
                                            "  TypeError in file %s, id %s, (Ionmode: %s, XCount: %s, Charge: %d)"
                                            % (
                                                str(res.fileName),
                                                str(row["id"]),
                                                ionMode,
                                                xCount,
                                                cLoading,
                                            ),
                                            str(err),
                                        )
                                    except:
                                        print(
                                            "  some general error in file %s, id %s, (Ionmode: %s, XCount: %s, Charge: %d)"
                                            % (
                                                str(res.fileName),
                                                str(row["id"]),
                                                ionMode,
                                                xCount,
                                                cLoading,
                                            )
                                        )

                        totalChromPeaks = totalChromPeaks + len(res.featurePairs)

                    if pwTextSet is not None:
                        # Log time used for bracketing
                        elapsed = (time.time() - start) / 60.0
                        hours = ""
                        if elapsed >= 60.0:
                            hours = "%d hours " % (elapsed // 60)
                        mins = "%.2f mins" % (elapsed % 60.0)
                        pwTextSet.put(
                            Bunch(
                                mes="text",
                                val="<p align='right' >%s%s elapsed</p>\n\n\n\nClustering \n  (Ionmode: %s, XCount: %s, Charge: %d)" % (hours, mins, ionMode, xCount, cLoading),
                            )
                        )

                    for tracer in tracersDeltaMZ:
                        ## TODO this does not work correctly. improve
                        if pwValSet is not None:
                            pwValSet.put(Bunch(mes="value", val=doneSoFar))
                        doneSoFar += 1

                        # get all results that match current bracketing criteria
                        chromPeaksAct = 0
                        allmz = []
                        for res in results:
                            chromPeaksAct = chromPeaksAct + len([chromPeak.mz for chromPeak in res.featurePairs])

                        allMZs = []
                        for res in results:
                            allMZs.extend([chromPeak.mz for chromPeak in res.featurePairs])

                        # cluster all current results with HCA
                        allMZs = sorted(allMZs)
                        lastMZ = None
                        u = 0
                        curAllMZs = []

                        ## pre-separate detected feature pairs to improve speed of HCA
                        while len(allMZs) > 0:
                            procCurSet = False
                            if u < len(allMZs) and (lastMZ is None or (allMZs[u] - lastMZ) < 3 * (groupSizePPM * allMZs[u] / 1e6)):
                                curAllMZs.append(allMZs[u])
                                lastMZ = allMZs[u]
                                u += 1
                            else:
                                lastMZ = None
                                allMZs = allMZs[u:]
                                u = 0
                                procCurSet = True

                            if procCurSet or len(allMZs) == u:
                                if len(curAllMZs) > 0:
                                    hc = HierarchicalClustering(
                                        curAllMZs,
                                        dist=lambda x, y: x.getValue() - y.getValue(),
                                        val=lambda x: x,
                                        mean=lambda x, y: x / y,
                                        add=lambda x, y: x + y,
                                    )

                                    # cut HCA tree in subclusters
                                    for n in cutTreeSized(hc.getTree(), groupSizePPM):
                                        try:
                                            lowMZ = min([kid.getValue() for kid in n.getKids()])
                                            minMZ = max([kid.getValue() for kid in n.getKids()])

                                            maxMZInGroup = lowMZ

                                            partChromPeaks = {}
                                            partXICs = {}
                                            toDel = {}
                                            allChromPeaks = []
                                            for res in results:
                                                chromPeaks = []
                                                toDel[res.filePath] = set()
                                                for i in range(len(res.featurePairs)):
                                                    chromPeak = res.featurePairs[i]
                                                    if chromPeak.mz <= minMZ:
                                                        toDel[res.filePath].add(i)
                                                        chromPeaks.append(chromPeak)
                                                        allChromPeaks.append(chromPeak)
                                                        maxMZInGroup = max(
                                                            maxMZInGroup,
                                                            chromPeak.mz,
                                                        )

                                                if align:
                                                    xics = []
                                                    # get EICs of current subCluster
                                                    for chromPeak in chromPeaks:
                                                        xics_table = res.conn.get_table("XICs")
                                                        chrom_peaks_table = res.conn.get_table("chromPeaks")

                                                        if not xics_table.is_empty() and not chrom_peaks_table.is_empty():
                                                            # Join XICs with chromPeaks
                                                            joined = chrom_peaks_table.filter(pl.col("id") == chromPeak.id).join(xics_table, left_on="eicID", right_on="id", how="inner")

                                                            if len(joined) > 0:
                                                                row = joined[0]
                                                                scanCount = float(row["scanCount"])
                                                                times = [float(r) for r in str(row["times"]).split(";")]
                                                                xicL = [float(r) for r in str(row["xicL"]).split(";")]
                                                                eicID = int(row["eicid"])
                                                                xics.append(
                                                                    [
                                                                        scanCount,
                                                                        times,
                                                                        xicL,
                                                                        eicID,
                                                                    ]
                                                                )
                                                partChromPeaks[res.filePath] = chromPeaks
                                                if align:
                                                    partXICs[res.filePath] = xics

                                            xy = xy + 1
                                            if (xy % 50) == 0:
                                                if writePDF:
                                                    pdf.save()
                                                    pdf = canvas.Canvas(file + "_%d.pdf" % xy)
                                                    _writeFirstPage(
                                                        pdf,
                                                        groupSizePPM,
                                                        maxTimeDeviation,
                                                        align,
                                                        nPolynom,
                                                    )
                                                    print(
                                                        "\r   new pdf %d metabolic features.." % xy,
                                                    )

                                            # debug purposes: create PDF for current subcluster
                                            if writePDF:
                                                drawing = Drawing(500, 350)

                                                lp = LinePlot()
                                                lp.x = 50
                                                lp.y = 30
                                                lp.height = 350
                                                lp.width = 500

                                                dd = []

                                                tr = 0
                                                pdf.drawString(
                                                    40,
                                                    810,
                                                    "Tracer: %s IonMode: %s" % (tracer, ionMode),
                                                )
                                                pdf.drawString(
                                                    40,
                                                    795,
                                                    "MZ: (min: %f - max: %f (tolerated: %f ppm: %.1f)) "
                                                    % (
                                                        lowMZ,
                                                        maxMZInGroup,
                                                        minMZ,
                                                        (maxMZInGroup - lowMZ) * 1000000 / lowMZ,
                                                    ),
                                                )
                                                pdf.drawString(
                                                    40,
                                                    780,
                                                    "xCount: %s Z: %d Feature pairs: %d"
                                                    % (
                                                        xCount,
                                                        cLoading,
                                                        len(partXICs),
                                                    ),
                                                )
                                                pdj = []
                                                pdjm = []
                                                for r in partXICs.keys():
                                                    xics = partXICs[r]
                                                    for xic in xics:
                                                        pdj.append(xic[2])
                                                        pdjm.append(xic[1])
                                                        dd.append(
                                                            [
                                                                (
                                                                    xic[1][i] / 60.0,
                                                                    xic[2][i],
                                                                )
                                                                for i in range(0, len(xic[1]))
                                                            ]
                                                        )
                                                        tr = tr + 1
                                                pdf.drawString(
                                                    40,
                                                    765,
                                                    "Aligning %d EICs" % (tr),
                                                )

                                                lp.data = dd

                                                i = 0
                                                for k in partXICs.keys():
                                                    lp.lines[i].strokeColor = colos[i % len(colos)]
                                                    lp.lines[i].strokeWidth = 0.01
                                                    i = i + 1

                                                lp.joinedLines = 1
                                                drawing.add(lp)
                                                renderPDF.draw(drawing, pdf, 10, 380)

                                                alignedEICs = xicAlign.getAligendXICs(
                                                    pdj,
                                                    pdjm,
                                                    align=align,
                                                    nPolynom=nPolynom,
                                                )[0]
                                                drawing = Drawing(500, 350)
                                                lp = LinePlot()
                                                lp.x = 50
                                                lp.y = 30
                                                lp.height = 350
                                                lp.width = 500

                                                dd = []

                                                tr = 0
                                                for xic in alignedEICs:
                                                    dd.append([(i / 60.0, xic[i]) for i in range(0, len(xic))])
                                                    tr = tr + 1

                                                lp.data = dd

                                                i = 0
                                                for k in range(len(alignedEICs)):
                                                    lp.lines[i].strokeColor = colos[i % len(colos)]
                                                    lp.lines[i].strokeWidth = 0.01
                                                    i = i + 1

                                                lp.joinedLines = 1
                                                drawing.add(lp)
                                                renderPDF.draw(drawing, pdf, 10, 0)

                                                pdf.showPage()

                                            eics = []
                                            peaks = []
                                            scantimes = []
                                            dd = []
                                            do = []
                                            for k in partChromPeaks.keys():
                                                for i in range(len(partChromPeaks[k])):
                                                    do.append(k)
                                                    dd.append(partChromPeaks[k][i].mz)
                                                    if align:
                                                        eics.append(partXICs[k][i][2])
                                                        scantimes.append(partXICs[k][i][1])
                                                    peaks.append([partChromPeaks[k][i]])

                                            # optional: align EICs; get bracketed chromatographic peaks
                                            aligned = xicAlign.alignXIC(
                                                eics,
                                                peaks,
                                                scantimes,
                                                align=align,
                                                maxTimeDiff=maxTimeDeviation,
                                                nPolynom=nPolynom,
                                            )
                                            aligned = [(x[0][0], int(x[0][1])) for x in aligned]

                                            if writePDF:
                                                uz = 800
                                                o = 0
                                                for peak in peaks:
                                                    for p in peak:
                                                        pdf.drawString(
                                                            40,
                                                            uz,
                                                            "%35s: %d %.2f -> %d %d"
                                                            % (
                                                                do[o][(1 + do[o].rfind("/")) :],
                                                                int(p.NPeakCenter),
                                                                p.NPeakCenterMin / 60.0,
                                                                int(aligned[o][0]),
                                                                aligned[o][1],
                                                            ),
                                                        )
                                                    uz -= 20
                                                    o += 1

                                            maxGroup = max(aligned, key=itemgetter(1))[1]
                                            minGroup = min(aligned, key=itemgetter(1))[1]

                                            groupedChromPeaks = []
                                            groupedChromPeaksAVGMz = []
                                            groupedChromPeaksAVGLMz = []
                                            groupedChromPeaksAVGTMz = []
                                            groupedChromPeaksAVGTimes = []
                                            groupedChromPeaksAVGNPeakScale = []
                                            groupedChromPeaksAVGLPeakScale = []

                                            for i in range(maxGroup + 1):
                                                groupedChromPeaks.append({})
                                                groupedChromPeaksAVGMz.append([])
                                                groupedChromPeaksAVGLMz.append([])
                                                groupedChromPeaksAVGTMz.append([])
                                                groupedChromPeaksAVGTimes.append([])
                                                groupedChromPeaksAVGNPeakScale.append([])
                                                groupedChromPeaksAVGLPeakScale.append([])

                                            # calculate average values (RT and mz) for feature pairs in the sub-subclusters
                                            j = 0
                                            for k in partChromPeaks.keys():
                                                for i in range(len(partChromPeaks[k])):
                                                    if not (k in groupedChromPeaks[aligned[j][1]]):
                                                        groupedChromPeaks[aligned[j][1]][k] = []
                                                    groupedChromPeaks[aligned[j][1]][k].append(
                                                        (
                                                            aligned[j],
                                                            partChromPeaks[k][i],
                                                        )
                                                    )
                                                    groupedChromPeaksAVGMz[aligned[j][1]].append(partChromPeaks[k][i].mz)
                                                    groupedChromPeaksAVGLMz[aligned[j][1]].append(partChromPeaks[k][i].lmz)
                                                    groupedChromPeaksAVGTMz[aligned[j][1]].append(partChromPeaks[k][i].tmz)
                                                    groupedChromPeaksAVGTimes[aligned[j][1]].append(partChromPeaks[k][i].NPeakCenterMin)
                                                    groupedChromPeaksAVGNPeakScale[aligned[j][1]].append(partChromPeaks[k][i].NPeakScale)
                                                    groupedChromPeaksAVGLPeakScale[aligned[j][1]].append(partChromPeaks[k][i].LPeakScale)

                                                    j = j + 1

                                            assert j == len(aligned)
                                            assert len(groupedChromPeaks) == len(groupedChromPeaksAVGMz) == len(groupedChromPeaksAVGTimes)

                                            # write results to data matrix and SQLite DB
                                            for i in range(minGroup, maxGroup + 1):
                                                if len(groupedChromPeaks[i]) > 0:
                                                    avgmz = sum(groupedChromPeaksAVGMz[i]) / len(groupedChromPeaksAVGMz[i])
                                                    avglmz = sum(groupedChromPeaksAVGLMz[i]) / len(groupedChromPeaksAVGLMz[i])
                                                    avgtmz = sum(groupedChromPeaksAVGTMz[i]) / len(groupedChromPeaksAVGTMz[i])
                                                    avgtime = sum(groupedChromPeaksAVGTimes[i]) / len(groupedChromPeaksAVGTimes[i])
                                                    avgNPeakScale = sum(groupedChromPeaksAVGNPeakScale[i]) / len(groupedChromPeaksAVGNPeakScale[i])
                                                    avgLPeakScale = sum(groupedChromPeaksAVGLPeakScale[i]) / len(groupedChromPeaksAVGLPeakScale[i])

                                                    # Append data to excel_data dictionary
                                                    excel_data["Num"].append(curNum)
                                                    excel_data["Comment"].append("")
                                                    excel_data["MZ"].append(avgmz)
                                                    excel_data["L_MZ"].append(avglmz)
                                                    excel_data["D_MZ"].append(avgtmz)
                                                    excel_data["MZ_Range"].append(
                                                        "%.6f - %.6f"
                                                        % (
                                                            min(groupedChromPeaksAVGMz[i]),
                                                            max(groupedChromPeaksAVGMz[i]),
                                                        )
                                                    )
                                                    excel_data["RT"].append(avgtime / 60.0)
                                                    excel_data["RT_Range"].append(
                                                        "%.3f - %.3f"
                                                        % (
                                                            min(groupedChromPeaksAVGTimes[i]) / 60.0,
                                                            max(groupedChromPeaksAVGTimes[i]) / 60.0,
                                                        )
                                                    )
                                                    excel_data["PeakScalesNL"].append(
                                                        "%.1f:%.1f"
                                                        % (
                                                            avgNPeakScale,
                                                            avgLPeakScale,
                                                        )
                                                    )
                                                    excel_data["Xn"].append(str(xCount))
                                                    excel_data["Charge"].append(cLoading)
                                                    excel_data["ScanEvent"].append(scanEvent)
                                                    excel_data["Ionisation_Mode"].append(ionMode)
                                                    excel_data["Tracer"].append(str(tracer))

                                                    excel_data["OGroup"].append("")
                                                    excel_data["Ion"].append("")
                                                    excel_data["Loss"].append("")
                                                    excel_data["M"].append("")

                                                    doublePeak = 0
                                                    for j in range(len(results)):
                                                        res = results[j].filePath
                                                        fname = results[j].fileName
                                                        if ".mzxml" in fname.lower():
                                                            fname = fname[max(fname.rfind("/") + 1, fname.rfind("\\") + 1) : fname.lower().rfind(".mzxml")]
                                                        if ".mzml" in fname.lower():
                                                            fname = fname[max(fname.rfind("/") + 1, fname.rfind("\\") + 1) : fname.lower().rfind(".mzml")]

                                                        if res in groupedChromPeaks[i] and len(groupedChromPeaks[i][res]) > 0:
                                                            doublePeak += 1 if len(groupedChromPeaks[i][res]) > 1 else 0

                                                            excel_data[fname + "_Area_N"].append(";".join([str(peak[1].NPeakArea) for peak in groupedChromPeaks[i][res]]))
                                                            excel_data[fname + "_Area_L"].append(";".join([str(peak[1].LPeakArea) for peak in groupedChromPeaks[i][res]]))
                                                            excel_data[fname + "_Abundance_N"].append(";".join([str(peak[1].NPeakAbundance) for peak in groupedChromPeaks[i][res]]))
                                                            excel_data[fname + "_Abundance_L"].append(";".join([str(peak[1].LPeakAbundance) for peak in groupedChromPeaks[i][res]]))
                                                            excel_data[fname + "_SNR_N"].append(";".join([str(peak[1].NSNR) for peak in groupedChromPeaks[i][res]]))
                                                            excel_data[fname + "_SNR_L"].append(";".join([str(peak[1].LSNR) for peak in groupedChromPeaks[i][res]]))
                                                            excel_data[fname + "_FID"].append(";".join([str(peak[1].id) for peak in groupedChromPeaks[i][res]]))
                                                            excel_data[fname + "_GroupID"].append(";".join([str(peak[1].fGroupID) for peak in groupedChromPeaks[i][res]]))

                                                        else:
                                                            excel_data[fname + "_Area_N"].append("")
                                                            excel_data[fname + "_Area_L"].append("")
                                                            excel_data[fname + "_Abundance_N"].append("")
                                                            excel_data[fname + "_Abundance_L"].append("")
                                                            excel_data[fname + "_SNR_N"].append("")
                                                            excel_data[fname + "_SNR_L"].append("")
                                                            excel_data[fname + "_FID"].append("")
                                                            excel_data[fname + "_GroupID"].append("")

                                                    excel_data["doublePeak"].append(doublePeak)

                                                    curNum = curNum + 1

                                            if writePDF:
                                                pdf.showPage()

                                        except Exception as e:
                                            import traceback

                                            traceback.print_exc()
                                            logging.error(str(traceback))

                                            logging.error("Error: %s" % (str(e)))
                                        grp = grp + 1

                                        for res in results:
                                            if res.filePath in toDel:
                                                toDel[res.filePath] = [a for a in toDel[res.filePath]]
                                                toDel[res.filePath].sort()
                                                toDel[res.filePath].reverse()
                                                for i in toDel[res.filePath]:
                                                    res.featurePairs.pop(i)
                                                    chromPeaksAct = chromPeaksAct - 1
                                                    totalChromPeaksProc = totalChromPeaksProc + 1

                                        if pwTextSet is not None:
                                            # Log time used for bracketing
                                            elapsed = (time.time() - start) / 60.0
                                            hours = ""
                                            if elapsed >= 60.0:
                                                hours = "%d hours " % (elapsed // 60)
                                            mins = "%.2f mins" % (elapsed % 60.0)
                                            pwTextSet.put(
                                                Bunch(
                                                    mes="text",
                                                    val="<p align='right' >%s%s elapsed</p>\n\n\n\nBracketing results\n%d feature pairs (%d individual pairs processed).. (Ionmode: %s, XCount: %s, Charge: %d)"
                                                    % (
                                                        hours,
                                                        mins,
                                                        curNum,
                                                        totalChromPeaksProc,
                                                        ionMode,
                                                        xCount,
                                                        cLoading,
                                                    ),
                                                )
                                            )
                                curAllMZs = []
            if writePDF:
                pdf.save()

            # Add metadata as a separate sheet to the Excel file
            import uuid
            import platform
            import datetime
            from openpyxl import load_workbook

            identifier = "%s_%s_%s" % (
                str(uuid.uuid1()),
                str(platform.node()),
                str(datetime.datetime.now()),
            )

            # Prepare metadata strings
            metadata_lines = []
            metadata_lines.append(
                "## MetExtract II %s"
                % (
                    Bunch(
                        MetExtractVersion=meVersion,
                        RVersion=rVersion,
                        UUID_ext=identifier,
                    )
                    .dumpAsJSon()
                    .replace('"', "'")
                )
            )
            metadata_lines.append("## Individual files processing parameters %s" % (generalProcessingParams.dumpAsJSon().replace('"', "'")))

            processingParams = Bunch()
            xCountsFmt = []
            try:
                for xn in sorted(xCounts):
                    if xn - 1 in xCounts and xn + 1 in xCounts:
                        xCountsFmt.append("-")
                    else:
                        if xn - 1 not in xCounts:
                            xCountsFmt.append(", ")
                        xCountsFmt.append(xn)
                xCountsFmt.pop(0)
                seen = set()
                seen_add = seen.add
                xCountsFmt = [x for x in xCountsFmt if not (x in seen or seen_add(x)) or x == ", "]
            except Exception:
                xCountsFmt = xCounts

            processingParams.FPBracketing_xCounts = "".join([str(t) for t in xCountsFmt])
            processingParams.FPBracketing_groupSizePPM = groupSizePPM
            processingParams.FPBracketing_positiveScanEvent = positiveScanEvent
            processingParams.FPBracketing_negativeScanEvent = negativeScanEvent
            processingParams.FPBracketing_maxTimeDeviation = maxTimeDeviation
            processingParams.FPBracketing_maxLoading = maxLoading
            processingParams.FPBracketing_resultsFile = file
            processingParams.FPBracketing_align = align
            if align:
                processingParams.FPBracketing_nPolynom = nPolynom
            metadata_lines.append("## Bracketing files processing parameters %s" % (processingParams.dumpAsJSon().replace('"', "'")))

            params_df = pl.DataFrame({"Parameters": metadata_lines})
            df = pl.DataFrame(excel_data).sort("RT")

            excel_file = file.replace(".tsv", ".xlsx")
            with xlsxwriter.Workbook(excel_file) as workbook:
                params_df.write_excel(workbook=workbook, worksheet="Parameters")
                df.write_excel(workbook=workbook, worksheet="1_afterBracketing")

        exportAsFeatureML.convertMEMatrixToFeatureMLSepPolarities(excel_file, postfix="_ab")

    except Exception as ex:
        import traceback

        traceback.print_exc()
        print(str(traceback))

        logging.error("Error during bracketing of files")

    finally:
        logging.info("Bracketing done...")


########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################


# store used configuration to DB file
def writeMetaboliteGroupingConfigToDB(db, minConnectionsInFiles, minConnectionRate, groups):
    db.insert_row("config", {"key": "MEGROUP_groups", "value": str(groups)})
    db.insert_row(
        "config",
        {"key": "MEGROUP_minConnectionsInFiles", "value": str(minConnectionsInFiles)},
    )
    db.insert_row("config", {"key": "MEGROUP_minConnectionRate", "value": str(minConnectionRate)})


def getBordersFor(db, fID, file):
    ffp_table = db.get_table("foundfeaturepairs")
    if not ffp_table.is_empty():
        rows = ffp_table.filter((pl.col("file") == file) & (pl.col("resID") == fID))
        if len(rows) > 0:
            row = rows[0]
            return (float(row["NBorderLeft"]), float(row["NBorderRight"]), float(row["LBorderLeft"]), float(row["LBorderRight"]))
    return None


def getPeak(times, rt, borderleft, borderright):
    ## find best matching rt
    ind, tim = min(enumerate(times), key=lambda x: abs(x[1] - rt))
    minind = int(max(0, ind - int(borderleft)))
    maxind = int(min(len(times) - 1, ind + borderright))

    return (minind, maxind)


def calculateMetaboliteGroups(
    file="./results.tsv",
    groups=[],
    eicPPM=10.0,
    maxAnnotationTimeWindow=0.05,
    minConnectionsInFiles=1,
    minConnectionRate=0.4,
    minPeakCorrelation=0.85,
    useRatio=False,
    runIdentificationInstance=None,
    pwMaxSet=None,
    pwValSet=None,
    pwTextSet=None,
    cpus=1,
    toFile=None,
):
    if toFile is None:
        toFile = file

    logging.info("Starting convoluting feature pairs..")

    resDB = Bunch(conn=None, curs=None)
    resDB.conn = PolarsDB(file + getDBSuffix())
    resDB.curs = resDB.conn.cursor()

    try:
        # Select only those groups that shall be used for metabolite group grouping
        ## load mzxml files for improved metabolite convolution
        useGroups = []
        useGroupsForConfig = []
        numFilesForConvolution = 0
        for group in groups:
            if group.useForMetaboliteGrouping:
                useGroups.append(group)
                useGroupsForConfig.append(str(group.name) + ":" + str(group.files))

                for fi in group.files:
                    numFilesForConvolution += 1

        if pwMaxSet is not None:
            pwMaxSet.put(Bunch(mes="max", val=numFilesForConvolution))
        if pwValSet is not None:
            pwValSet.put(Bunch(mes="value", val=0))

        # connect to results db
        writeMetaboliteGroupingConfigToDB(
            resDB.conn,
            minConnectionsInFiles,
            minConnectionRate,
            str(useGroupsForConfig).replace("'", "").replace('"', ""),
        )

        # read results table - determine file type from extension
        comments = []
        if file.lower().endswith(".tsv") or file.lower().endswith(".txt"):
            # Read comments first
            with open(file, "r", encoding="utf-8") as fi:
                for line in fi:
                    if line.startswith("#"):
                        comments.append(line.strip())

            # Read TSV file with polars
            table_df = pl.read_csv(
                file,
                separator="\t",
                comment_prefix="#",
                infer_schema_length=10000,
                ignore_errors=True,
            )
        elif file.lower().endswith(".csv"):
            with open(file, "r", encoding="utf-8") as fi:
                for line in fi:
                    if line.startswith("#"):
                        comments.append(line.strip())

            table_df = pl.read_csv(
                file,
                comment_prefix="#",
                infer_schema_length=10000,
                ignore_errors=True,
            )
        else:
            raise ValueError(f"Unsupported file extension for {file}")

        cols = table_df.columns

        # check if necessary columns are present
        ## overall columns
        assert "Num" in cols
        assert "OGroup" in cols
        assert "Ion" in cols
        assert "Loss" in cols
        assert "M" in cols
        assert "doublePeak" in cols

        doublePeaks = len(table_df.filter(pl.col("doublePeak") > 0))

        if doublePeaks > 0:
            logging.info("  found double peaks: %d" % doublePeaks)
            # Save double peaks to separate file
            double_peaks_df = table_df.filter(pl.col("doublePeak") > 0)
            double_peaks_file = toFile.replace(".tsv", ".doublePeaks.tsv")
            if double_peaks_file.lower().endswith(".tsv") or double_peaks_file.lower().endswith(".txt"):
                double_peaks_df.write_csv(double_peaks_file, separator="\t")
            elif double_peaks_file.lower().endswith(".csv"):
                double_peaks_df.write_csv(double_peaks_file)

        # Delete rows where doublePeak > 0
        table_df = table_df.filter(pl.col("doublePeak") <= 0)
        ## file specific columns

        for group in useGroups:
            for f in group.files:
                fShort = f[f.rfind("/") + 1 : f.rfind(".")]
                if fShort[0].isdigit():
                    fShort = "_" + fShort

                assert "%s_FID" % fShort in cols
                assert "%s_GroupID" % fShort in cols

        # fetch all feature pairs from the results file
        nodes = {}
        for row in table_df.select(
            [
                "Num",
                "MZ",
                "L_MZ",
                "RT",
                "Xn",
                "Charge",
                "ScanEvent",
                "Ionisation_Mode",
            ]
        ).iter_rows(named=True):
            nodes[row["Num"]] = Bunch(
                fpNum=row["Num"],
                mz=row["MZ"],
                lmz=row["L_MZ"],
                rt=row["RT"],
                xn=row["Xn"],
                charge=row["Charge"],
                scanEvent=row["ScanEvent"],
                ionMode=row["Ionisation_Mode"],
                oGroup=None,
                correlationToOtherFPs={},
            )

        ## generate all correlations and sil ratios
        fileCorrelations = {}
        fileSILRatios = {}

        borders = {}
        ffp_table = resDB.conn.get_table("foundfeaturepairs")
        if not ffp_table.is_empty():
            for row in ffp_table.iter_rows(named=True):
                fil = str(row["file"])
                resID = int(row["resID"])
                if fil not in borders.keys():
                    borders[fil] = {}
                borders[fil][resID] = (
                    float(row["NBorderLeft"]),
                    float(row["NBorderRight"]),
                    float(row["LBorderLeft"]),
                    float(row["LBorderRight"]),
                )

        ## Iterate all files
        processedFiles = 0
        procObjects = []
        for group in useGroups:
            for fi in group.files:
                if fi not in fileCorrelations.keys():
                    # if pwTextSet is not None: pwTextSet.put(Bunch(mes="text", val="Convoluting feature pairs in file %s" % (fi)))
                    # if pwValSet is not None: pwValSet.put(Bunch(mes="value", val=processedFiles))

                    s = ConvoluteFPsInFile(eicPPM, fi, maxAnnotationTimeWindow, nodes, borders)
                    procObjects.append(s)

                    # processedFiles += 1

        from multiprocessing import Pool, Manager

        p = Pool(processes=cpus, maxtasksperchild=1)  # only in python >=2.7; experimental
        manager = Manager()
        queue = manager.Queue()

        start = time.time()
        # start the multiprocessing pool
        res = p.imap_unordered(processConvoluteFPsInFile, procObjects)

        # wait until all subprocesses have finished re-integrating their respective LC-HRMS data file
        loop = True
        freeSlots = range(min(len(procObjects), cpus))
        assignedThreads = {}
        while loop:
            completed = res._index
            if completed == len(procObjects):
                loop = False
            else:
                if pwValSet is not None:
                    pwValSet.put(Bunch(mes="value", val=completed))

                elapsed = (time.time() - start) / 60.0
                hours = ""
                if elapsed >= 60.0:
                    hours = "%d hours " % (elapsed // 60)
                mins = "%.2f mins" % (elapsed % 60.0)

                if pwTextSet is not None:
                    pwTextSet.put(
                        Bunch(
                            mes="text",
                            val="<p align='right' >%s%s elapsed</p>\n\n%d / %d files done (%d parallel)" % (hours, mins, completed, len(procObjects), cpus),
                        )
                    )

                time.sleep(0.5)

        p.close()

        ## fetch multiprocessing results
        for rei, re in enumerate(res):
            fileCorrelations[procObjects[rei].fi] = re[0]
            fileSILRatios[procObjects[rei].fi] = re[1]

        logging.info("Testing SIL ratios")
        ## test all feature pair pairs for co-elution and similar SIL ratio
        connections = {}
        for fpNumA in nodes.keys():
            connections[fpNumA] = {}
        for fpNumA in nodes.keys():
            nodeA = nodes[fpNumA]

            for fpNumB in nodes.keys():
                nodeB = nodes[fpNumB]

                if nodeA.fpNum != nodeB.fpNum and nodeA.fpNum < nodeB.fpNum and abs(nodeB.rt - nodeA.rt) <= maxAnnotationTimeWindow:
                    ## test feature pair pair convolution in all samples
                    allCorrelations = []
                    allSILRatios = []
                    for group in useGroups:
                        for fi in group.files:
                            if fi in fileCorrelations.keys() and fpNumA in fileCorrelations[fi].keys() and fpNumB in fileCorrelations[fi][fpNumA].keys():
                                co = fileCorrelations[fi][fpNumA][fpNumB]
                                allCorrelations.append(co)
                            if fi in fileSILRatios.keys() and fpNumA in fileSILRatios[fi].keys() and fpNumB in fileSILRatios[fi][fpNumA].keys():
                                silRatio = fileSILRatios[fi][fpNumA][fpNumB]
                                allSILRatios.append(silRatio)

                    if len(allCorrelations) > 0 and len(allSILRatios) > 0:
                        if sum([co >= minPeakCorrelation for co in allCorrelations]) >= minConnectionRate * len(allCorrelations) and (not useRatio or sum([rat for rat in allSILRatios]) >= minConnectionRate * len(allSILRatios)):
                            nodes[nodeA.fpNum].correlationToOtherFPs[nodeB.fpNum] = True
                            nodes[nodeB.fpNum].correlationToOtherFPs[nodeA.fpNum] = True

                        m = mean(allCorrelations)
                        connections[fpNumA][fpNumB] = m
                        connections[fpNumB][fpNumA] = m
                    else:
                        connections[fpNumA][fpNumB] = 0
                        connections[fpNumB][fpNumA] = 0

        logging.info(" .. done")

        nodes2 = {}
        for fpNumA in nodes.keys():
            if fpNumA not in nodes2.keys():
                nodes2[fpNumA] = []
            for fpNumB in nodes[fpNumA].correlationToOtherFPs.keys():
                nodes2[fpNumA].append(fpNumB)

        for k in nodes2.keys():
            uniq = []
            for u in nodes2[k]:
                if u not in uniq:
                    uniq.append(u)
            nodes2[k] = uniq

        # get subgraphs from the feature pair graph. Each subgraph represents one convoluted
        # feature group
        tGroups = getSubGraphs(nodes2)

        def splitGroupWithHCA(tGroup, correlations):
            # if 1 or two feature pairs are in a group, automatically use them (no further splitting possible)
            if len(tGroup) <= 2:
                return [tGroup]

            # construct 2-dimensional data matrix
            data = []
            for tG1 in tGroup:
                t = []
                for tG2 in tGroup:
                    if tG1 in correlations.keys() and tG2 in correlations[tG1].keys():
                        t.append(correlations[tG1][tG2])
                    elif tG1 == tG2:
                        t.append(1.0)
                    else:
                        t.append(0.0)
                data.append(t)

            # calculate HCA tree using the correlations between all feature pairs
            hc = HCA_general.HCA_generic()
            tree = hc.generateTree(objs=data, ids=tGroup)

            # split the HCA tree in sub-clusters
            def checkSubCluster(tree, hca, corrThreshold, cutOffMinRatio):
                if isinstance(tree, HCA_general.HCALeaf):
                    return False
                elif isinstance(tree, HCA_general.HCAComposite):
                    corrs = hca.getLinkFor(tree)
                    inds = hca.getIndsFor(tree)

                    aboveThreshold = sum([corr > corrThreshold for i, corr in enumerate(corrs) if i in inds])

                    # hc.plotTree(tree)
                    # print(corrs)
                    # print([corrs[i] for i in inds])
                    # print(aboveThreshold)
                    # print(".....")
                    # print("")
                    return not (aboveThreshold * 1.0 / len(inds)) >= cutOffMinRatio

                else:
                    raise RuntimeError("Unknown if-branch")

            subClusts = hc.splitTreeWithCallback(
                tree,
                CallBackMethod(
                    _target=checkSubCluster,
                    corrThreshold=minPeakCorrelation,
                    cutOffMinRatio=minConnectionRate,
                ).getRunMethod(),
                recursive=True,
            )

            # convert the subclusters into arrays of feature pairs belonging to the same metabolite
            return [[leaf.getID() for leaf in subClust.getLeaves()] for subClust in subClusts]

        groups = []
        done = 0
        for tGroup in tGroups:
            sGroups = splitGroupWithHCA(tGroup, connections)
            groups.extend(sGroups)

            done = done + 1

        # Separate feature pair clusters; softer
        curGroup = 1
        for tGroup in groups:
            for i in range(len(tGroup)):
                # Update in DataFrame
                table_df = table_df.with_columns(pl.when(pl.col("Num") == tGroup[i]).then(pl.lit(curGroup)).otherwise(pl.col("OGroup")).alias("OGroup"))
                # Update in ParquetDB
                group_results = resDB.conn.get_table("GroupResults")
                if not group_results.is_empty():
                    resDB.conn.tables["GroupResults"] = group_results.with_columns(pl.when(pl.col("id") == tGroup[i]).then(pl.lit(curGroup)).otherwise(pl.col("OGroup")).alias("OGroup"))

            curGroup += 1

        logging.info("Annotating ions in feature groups")
        # Annotate groups with common adducts and in-source fragments
        if runIdentificationInstance is not None:
            groups = defaultdict(list)
            for row in table_df.select(
                [
                    "Num",
                    "OGroup",
                    "MZ",
                    "Ionisation_Mode",
                    "Charge",
                    "Ion",
                    "Xn",
                    "Loss",
                    "M",
                ]
            ).iter_rows(named=True):
                groups[row["OGroup"]].append(
                    ChromPeakPair(
                        id=row["Num"],
                        fGroupID=row["OGroup"],
                        mz=row["MZ"],
                        ionMode=row["Ionisation_Mode"],
                        loading=row["Charge"],
                        adducts=[],
                        heteroAtomsFeaturePairs=[],
                        xCount=row["Xn"],
                        fDesc=[],
                    )
                )

            for ogrp in groups.keys():
                chromPeaks = {}
                for fp in groups[ogrp]:
                    chromPeaks[fp.id] = fp

                runIdentificationInstance.annotateChromPeaks(chromPeaks.keys(), chromPeaks)

            for ogrp in groups.keys():
                for fp in groups[ogrp]:
                    # Update in DataFrame
                    table_df = table_df.with_columns(
                        pl.when(pl.col("Num") == fp.id).then(pl.lit(",".join([str(a) for a in fp.adducts]))).otherwise(pl.col("Ion")).alias("Ion"),
                        pl.when(pl.col("Num") == fp.id).then(pl.lit(",".join([str(a) for a in fp.fDesc]))).otherwise(pl.col("Loss")).alias("Loss"),
                        pl.when(pl.col("Num") == fp.id).then(pl.lit(",".join([str(a) for a in fp.Ms]))).otherwise(pl.col("M")).alias("M"),
                    )
                    # Update in ParquetDB
                    group_results = resDB.conn.get_table("GroupResults")
                    if not group_results.is_empty():
                        resDB.conn.tables["GroupResults"] = group_results.with_columns(
                            pl.when(pl.col("id") == fp.id).then(pl.lit(",".join([str(a) for a in fp.adducts]))).otherwise(pl.col("Ion")).alias("Ion"),
                            pl.when(pl.col("id") == fp.id).then(pl.lit(",".join([str(a) for a in fp.fDesc]))).otherwise(pl.col("LOSS")).alias("LOSS"),
                            pl.when(pl.col("id") == fp.id).then(pl.lit(",".join([str(a) for a in fp.Ms]))).otherwise(pl.col("M")).alias("M"),
                        )

        # Delete from GroupResults where id not in table nums
        valid_nums = table_df.select("Num").to_series().to_list()
        group_results = resDB.conn.get_table("GroupResults")
        if not group_results.is_empty():
            resDB.conn.tables["GroupResults"] = group_results.filter(pl.col("id").is_in(valid_nums))

        ## reassign feature group ids
        # Update OGroup column to prefix with 'X'
        group_results = resDB.conn.get_table("GroupResults")
        if not group_results.is_empty():
            resDB.conn.tables["GroupResults"] = group_results.with_columns(pl.concat_str([pl.lit("X"), pl.col("OGroup").cast(pl.Utf8)]).alias("OGroup"))
        # Ensure OGroup is string type before concatenation
        table_df = table_df.with_columns(pl.col("OGroup").cast(pl.Utf8))
        table_df = table_df.with_columns(pl.concat_str([pl.lit("X"), pl.col("OGroup")]).alias("OGroup"))
        oGrps = []
        curGrp = 1
        curFP = 1

        # Get unique OGroups ordered by average rt
        group_results = resDB.conn.get_table("GroupResults")
        if not group_results.is_empty():
            grouped = group_results.groupby("OGroup").agg(pl.col("rt").mean().alias("avg_rt")).sort("avg_rt")
            for row in grouped.iter_rows(named=True):
                oGrps.append(str(row["OGroup"]))

        for tgrp in oGrps:
            # Update in ParquetDB
            group_results = resDB.conn.get_table("GroupResults")
            if not group_results.is_empty():
                resDB.conn.tables["GroupResults"] = group_results.with_columns(pl.when(pl.col("OGroup") == tgrp).then(pl.lit(curGrp)).otherwise(pl.col("OGroup")).alias("OGroup"))
            # Update in DataFrame
            table_df = table_df.with_columns(pl.when(pl.col("OGroup") == tgrp).then(pl.lit(curGrp)).otherwise(pl.col("OGroup")).alias("OGroup"))
            curGrp = curGrp + 1

        processingParams = Bunch()
        processingParams.MEConvoluting_groups = str(useGroupsForConfig).replace("'", "").replace('"', "")
        processingParams.MEConvoluting_connThreshold = minConnectionRate
        processingParams.MEConvoluting_minPeakCorrelation = minPeakCorrelation
        comments.append("## Convolution FPs processing parameters %s" % (processingParams.dumpAsJSon().replace('"', "'")))

        # Filter out _CorrelatesTo columns
        writeCols = [col for col in table_df.columns if not col.endswith("_CorrelatesTo")]

        # Sort by OGroup, MZ, Xn
        table_df = table_df.sort(["OGroup", "MZ", "Xn"])

        # Write to file based on extension
        if toFile.lower().endswith(".tsv") or toFile.lower().endswith(".txt"):
            table_df.select(writeCols).write_csv(toFile, separator="\t")
        elif toFile.lower().endswith(".csv"):
            table_df.select(writeCols).write_csv(toFile)
        else:
            # Default to TSV
            table_df.select(writeCols).write_csv(toFile, separator="\t")

        # Append comments
        if len(comments) > 0:
            with open(toFile, "a", encoding="utf-8") as fo:
                for comment in comments:
                    fo.write(str(comment))
                    fo.write("\n")

        exportAsFeatureML.convertMEMatrixToFeatureMLSepPolarities(file, postfix="_ac")

    except Exception as ex:
        import traceback

        traceback.print_exc()

        raise ex

    finally:
        resDB.conn.commit()
        resDB.curs.close()
        resDB.conn.close()


from .utils import mapArrayToRefTimes


class ConvoluteFPsInFile:
    def __init__(self, eicPPM, fi, maxAnnotationTimeWindow, nodes, borders):
        self.eicPPM = eicPPM
        self.fi = fi
        self.maxAnnotationTimeWindow = maxAnnotationTimeWindow
        self.nodes = nodes
        self.borders = borders

        self.fileCorrelations = None
        self.fileSILRatios = None

    def getConvolutionInFile(self):
        startProc = time.time()

        eicPPM = self.eicPPM
        fi = self.fi
        maxAnnotationTimeWindow = self.maxAnnotationTimeWindow
        nodes = self.nodes
        borders = self.borders

        fileCorrelations = {}
        fileSILRatios = {}

        logging.info("  Convoluting feature pairs in file %s" % (fi))
        b = fi.replace("\\", "/")
        fiName = ""
        if ".mzXML" in b:
            fiName = b[(b.rfind("/") + 1) : b.rfind(".mzXML")]
        if ".mzML" in b:
            fiName = b[(b.rfind("/") + 1) : b.rfind(".mzML")]
        mzXML = Chromatogram()
        mzXML.parse_file(fi)

        filterLines = mzXML.getFilterLines(
            includeMS1=True,
            includeMS2=False,
            includePosPolarity=True,
            includeNegPolarity=True,
        )

        for fpNumA in nodes.keys():
            fileCorrelations[fpNumA] = {}
            fileSILRatios[fpNumA] = {}

        for fpNumA in nodes.keys():
            nodeA = nodes[fpNumA]

            for fpNumB in nodes.keys():
                nodeB = nodes[fpNumB]

                if nodeA.fpNum != nodeB.fpNum and nodeA.fpNum < nodeB.fpNum and abs(nodeB.rt - nodeA.rt) <= maxAnnotationTimeWindow:
                    ra = None
                    if fiName in borders.keys() and nodeA.fpNum in borders[fiName].keys():
                        ra = borders[fiName][nodeA.fpNum]
                    rb = None
                    if fiName in borders.keys() and nodeB.fpNum in borders[fiName].keys():
                        rb = borders[fiName][nodeB.fpNum]

                    if ra != None and rb != None:
                        nanbl, nanbr, nalbl, nalbr = ra
                        nbnbl, nbnbr, nblbl, nblbr = rb

                        if nodeA.scanEvent in filterLines and nodeB.scanEvent in filterLines:
                            meanRT = mean([nodeA.rt, nodeB.rt])

                            eicAL, timesA, scanIdsA, mzs = mzXML.getEIC(
                                nodeA.lmz,
                                ppm=eicPPM,
                                filterLine=nodeA.scanEvent,
                                removeSingles=True,
                                intThreshold=0,
                                useMS1=True,
                                useMS2=False,
                                startTime=meanRT * 60 - 120,
                                endTime=meanRT * 60 + 120,
                            )
                            eicBL, timesB, scanIdsB, mzs = mzXML.getEIC(
                                nodeB.lmz,
                                ppm=eicPPM,
                                filterLine=nodeB.scanEvent,
                                removeSingles=True,
                                intThreshold=0,
                                useMS1=True,
                                useMS2=False,
                                startTime=meanRT * 60 - 120,
                                endTime=meanRT * 60 + 120,
                            )

                            eicA, timesA, scanIdsA, mzs = mzXML.getEIC(
                                nodeA.mz,
                                ppm=eicPPM,
                                filterLine=nodeA.scanEvent,
                                removeSingles=True,
                                intThreshold=0,
                                useMS1=True,
                                useMS2=False,
                                startTime=meanRT * 60 - 120,
                                endTime=meanRT * 60 + 120,
                            )
                            eicB, timesB, scanIdsB, mzs = mzXML.getEIC(
                                nodeB.mz,
                                ppm=eicPPM,
                                filterLine=nodeB.scanEvent,
                                removeSingles=True,
                                intThreshold=0,
                                useMS1=True,
                                useMS2=False,
                                startTime=meanRT * 60 - 120,
                                endTime=meanRT * 60 + 120,
                            )

                            timesMin = [t for t in timesA]
                            timesMin.extend([t for t in timesB])
                            timesMin = sorted(timesMin)

                            eicAL = mapArrayToRefTimes(eicAL, timesA, timesMin)
                            eicBL = mapArrayToRefTimes(eicBL, timesB, timesMin)
                            eicA = mapArrayToRefTimes(eicA, timesA, timesMin)
                            eicB = mapArrayToRefTimes(eicB, timesB, timesMin)

                            timesMin = [t / 60.0 for t in timesMin]

                            ## A) Test correlation of different feature pairs
                            try:
                                lI, rI = getPeak(
                                    timesMin,
                                    meanRT,
                                    min(nanbl, nbnbl) * 2,
                                    min(nanbr, nbnbr) * 2,
                                )

                                eicAC = eicAL[lI:rI]
                                eicBC = eicBL[lI:rI]

                                co = corr(eicAC, eicBC)

                                if not (isnan(co)) and co != None:
                                    fileCorrelations[fpNumA][fpNumB] = co
                            except Exception as err:
                                logging.error("  Error during convolution of feature pairs (Peak-correlation, Nums: %s and %s, message: %s).." % (fpNumA, fpNumB, err.message))

                            ## B) Test similarity of native:labeled ratio
                            try:
                                lISIL, rISIL = getPeak(
                                    timesMin,
                                    nodeA.rt,
                                    min(nanbl, nanbr) * 0.8,
                                    min(nalbl, nalbr) * 0.8,
                                )

                                eicANCSIL = eicA[lISIL:rISIL]
                                eicALCSIL = eicAL[lISIL:rISIL]

                                folds = [eicANCSIL[i] / eicALCSIL[i] for i in range(len(eicANCSIL)) if eicALCSIL[i] > 0]
                                ma = mean(folds)
                                sa = sd(folds)

                                lISIL, rISIL = getPeak(
                                    timesMin,
                                    nodeB.rt,
                                    min(nanbl, nanbr) * 0.8,
                                    min(nalbl, nalbr) * 0.8,
                                )

                                eicBNCSIL = eicB[lISIL:rISIL]
                                eicBLCSIL = eicBL[lISIL:rISIL]

                                folds = [eicBNCSIL[i] / eicBLCSIL[i] for i in range(len(eicBNCSIL)) if eicBLCSIL[i] > 0]
                                mb = mean(folds)
                                sb = sd(folds)

                                if ma != None and mb != None and sa != None and sb != None:
                                    silRatioFold = 1
                                    try:
                                        silRatioFold = max([ma, mb]) / min([ma, mb])
                                    except RuntimeWarning as ex:
                                        pass
                                    # print(ma, mb, sa, sb, max([ma, mb]), min([ma, mb]))
                                    fileSILRatios[fpNumA][fpNumB] = silRatioFold <= 1 + max(0.5, 3 * mean([sa, sb]))

                                    # print(fiName, nodeA.mz, nodeB.mz, meanRT, silRatioFold, co)
                            except Exception as err:
                                logging.error("  Error during convolution of feature pairs (SIL-ratio, Nums: %s and %s, message: %s).." % (fpNumA, fpNumB, err.message))

        logging.info("  Finished convoluting feature pairs in file %s (%.1f minutes)" % (fi, (time.time() - startProc) / 60.0))
        return (fileCorrelations, fileSILRatios)


def processConvoluteFPsInFile(cif):
    se = cif.getConvolutionInFile()
    return se
