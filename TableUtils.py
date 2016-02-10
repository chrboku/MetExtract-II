from sqlite3 import *

from os.path import basename

import string
import random

import time


# functionality to store / read a table into a generic SQLite table
# methods to insert / alter / delete data are provided (including bulk-update)

# class for storing metadata of one column
class colDef:
    def __init__(self, name, format):
        self.name = name
        self.format = format

    def getName(self):
        return self.name

    def getFormat(self):
        return self.format

    def __str__(self):
        return "name: %s (%s)" % (self.name, self.format)

# class table provides generic methods to insert / alter / delete data stored in an in-memory SQLite database table
# the use is abstracted as much as possible from the SQLite syntax
class Table:
    def __initEmpty__(self, tableName=None):
        if tableName is None:
            tableName = ''.join(random.choice(string.ascii_uppercase + string.ascii_lowercase) for x in range(6))
        self.tableName = tableName

        self.conn = connect(":memory:")
        self.conn.text_factory = str
        self.curs = self.conn.cursor()
        self.saved = True

        self.curs.execute(self.__updateTableName("CREATE TABLE :table: (__internalID INTEGER PRIMARY KEY)"))
        self.curs.execute(self.__updateTableName("CREATE TABLE :table:_COMMENTS (__internalID INTEGER PRIMARY KEY, comment TEXT)"))

        self.conn.commit()

    def __init__(self, headers=None, rows=None, tableName=None, comments=None):
        if headers is None and rows is None:
            self.__initEmpty__(tableName)
            return

        assert headers is not None
        assert rows is not None

        for row in rows:
            assert len(headers) >= len(row), "Unequal data count"

        if tableName is None:
            tableName = ''.join(random.choice(string.ascii_uppercase + string.ascii_lowercase) for x in range(6))
        self.tableName = tableName

        self.conn = connect(":memory:")
        self.conn.text_factory = str
        self.curs = self.conn.cursor()
        self.saved = True

        cols = []
        for i, h in enumerate(headers):
            c = "integer"
            for row in rows:
                if len(row) > i and len(row[i]) > 0:
                    try:
                        row[i] = int(row[i])
                    except:
                        c = "real"
                    try:
                        row[i] = float(row[i])
                    except:
                        c = "text"
            if str.isdigit(h[0]):
                h = "_" + h
            h = h.replace("-", "__")
            h = h.replace(" ", "__")
            cols.append((h, c))

        self.curs.execute(self.__updateTableName("CREATE TABLE :table: (__internalID INTEGER PRIMARY KEY,  %s)" % (", ".join(["%s %s" % h for h in cols]))))
        self.curs.execute(self.__updateTableName("CREATE TABLE :table:_COMMENTS (__internalID INTEGER PRIMARY KEY, comment TEXT)"))

        for row in rows:
            while len(cols) > len(row):
                row.append("")
            self.curs.execute(self.__updateTableName("INSERT INTO :table: (%s) VALUES(%s)" % (
                ", ".join([h[0] for h in cols]), ", ".join(["?" for c in row]))), [str(c) for c in row])

        for comment in comments:
            self.curs.execute(self.__updateTableName("INSERT INTO :table:_COMMENTS(comment) VALUES(?)"), [comment])

        self.conn.commit()

    # get all available columns of the table
    def getColumns(self):
        cols = []
        for row in self.curs.execute(self.__updateTableName("PRAGMA table_info(:table:)")):
            if row[1] != "__internalID":
                cols.append(colDef(row[1], row[2]))
        return cols

    # test if a specific column is present in the table
    def hasColumn(self, col):
        cols = self.getColumns()
        colType = None
        for colD in cols:
            if colD.name == col:
                colType = colD.format

        return colType is not None

    # bulk-update: perform a function for each row present in the table
    def applyFunction(self, func, where="", pwMaxSet=None, pwValSet=None, showProgress=False):
        if len(where) > 0:
            where = "WHERE " + where

        updates = []
        cols = self.getColumns()
        r = range(1, 1 + len(cols))

        numOfCols=0
        for row in self.curs.execute(self.updateQuery("SELECT __internalID FROM :table:")):
            numOfCols+=1

        done=0
        started=time.time()
        for row in self.curs.execute(self.updateQuery("SELECT __internalID, :allCols: FROM :table: %s" % where)):
            iID = row[0]
            data = {}
            for i in r:
                data[cols[i - 1].name] = row[i]
            newData = func(data)
            f = {}
            for col in cols:
                if col.name in newData.keys():
                    f[col.name] = str(newData[col.name])

            if len(f)>0:
                updates.append((self.__updateTableName(
                    "UPDATE :table: SET %s WHERE __internalID=%d" % (", ".join(["%s=?" % x for x in f.keys()]), iID)),
                                [f[x] for x in f.keys()]))

            done+=1
            if showProgress:
                elapsed=time.time()-started
                doneP=(1.*done/numOfCols)
                s=(elapsed*(1-doneP)/doneP)/60.
                print "\rApplying.. |%-30s| %.1f%% (approximately %.1f minutes remaining)"%("*"*int(doneP*30), doneP*100, s),
        if showProgress:
            s=(time.time()-started)/60.
            print "(%.1f minutes)\n"%s

        if pwMaxSet!=None: pwMaxSet(len(updates))
        if pwValSet!=None: pwValSet(0)

        doneS=0
        started=time.time()
        for update in updates:
            self.curs.execute(self.updateQuery(update[0]), update[1])
            done+=1
            if pwValSet!=None: pwValSet(done)

            doneS+=1
            if showProgress:
                elapsed=time.time()-started
                doneP=(1.*doneS/numOfCols)
                s=(elapsed*(1-doneP)/doneP)/60.
                print "\rApplying.. |%-30s| %.1f%% (approximately %.1f minutes remaining)"%("*"*int(doneP*30), doneP*100, s),
        if showProgress:
            s=(time.time()-started)/60.
            print "(%.1f minutes)\n"%s

        self.conn.commit()

    # return the entire table
    def getData(self, cols=None, where=None, orderby=None):
        ret = []
        if where is None:
            where = ""
        else:
            where = " WHERE " + where

        if orderby is None:
            orderby = ""
        else:
            orderby = "ORDER BY " + orderby

        if cols is None:
            cols = [x.getName() for x in self.getColumns()]

        tCols=[]
        for col in cols:
            if col[0].isdigit():
                tCols.append("_"+col)
            else:
                tCols.append(col)
        cols=tCols

        for row in self.curs.execute(self.__updateTableName("SELECT %s FROM :table: %s %s" % (",".join([col for col in cols]), where, orderby))):
            if len(row) == 1:
                ret.append(row[0])
            else:
                ret.append(row)

        return ret

    # set data specifically (similar to SQL update)
    def setData(self, cols, vals, where=None):
        assert len(cols)==len(vals)

        sets=[]
        for coli, colVal in enumerate(cols):
            ca=""
            if isinstance(vals[coli], str):
                ca="'"
            sets.append("%s=%s%s%s"%(colVal, ca, vals[coli], ca))

        if where is not None:
            where="WHERE "+where

        self.curs.execute(self.__updateTableName("UPDATE :table: SET %s %s" % (",".join(sets), where)))

    # add a new column to the table (in-memory only)
    def addColumn(self, col, colType, defaultValue=""):
        if self.hasColumn(col):
            raise Exception("Column %s already exists" % col)

        self.curs.execute(self.__updateTableName("ALTER TABLE :table: ADD COLUMN %s %s DEFAULT '%s'" % (col, colType, defaultValue)))
        self.conn.commit()
        self.saved = False

    # append a new row at the end of the table
    def addRow(self, data):
        for dCol in data.keys():
            if not (self.hasColumn(dCol)):
                raise Exception("Column %s not found" % dCol)

        cols = ", ".join([str(col) for col in data.keys()])
        values = ", ".join("'%s'" % data[col] for col in data.keys())
        self.curs.execute(self.__updateTableName("INSERT INTO :table: (%s) VALUES(%s)" % (cols, values)))
        self.conn.commit()
        self.saved = False

    # reset entire column (affects all rows of the table)
    def resetColumn(self, col, value=""):
        if not (self.hasColumn(col)):
            raise Exception("Column %s not found" % col)

        self.executeSQL(self.updateQuery("UPDATE :table: SET %s='%s'" % (col, value)))
        self.conn.commit()
        self.saved = False

    # HELPER METHOD that updates a sql command with the temporary SQLite table name
    def __updateTableName(self, sql):
        return sql.replace(":table:", self.tableName)

    # accepts a generic SQLite command (without table name and columns) and adds the table name as well as all columns
    def updateQuery(self, sql):
        allCols = []
        for col in self.getColumns():
            allCols.append(col.name)
        allCols = ", ".join(allCols)

        sql = sql.replace(":table:", self.tableName)
        sql = sql.replace(":allCols:", allCols)

        return sql

    # executes an SQLite command on the table
    def executeSQL(self, sql, commit=True):
        self.curs.execute(self.updateQuery(sql))
        if commit: self.conn.commit()
        self.saved = False

    # tests, if the data has been edited
    def isSaved(self):
        return self.saved

    # sets, if the data has been edited
    def save(self):
        self.saved = True

    def getComments(self):
        ret=[]
        for row in self.curs.execute(self.__updateTableName("SELECT comment FROM :table:_COMMENTS")):
            if len(row) == 1:
                ret.append(row[0])
            else:
                ret.append(row)

        return ret

    def addComment(self, comment):
        self.curs.execute(self.__updateTableName("INSERT INTO :table:_COMMENTS (comment) VALUES(?)"), [comment])
        self.conn.commit()

    # de-constructor: close SQLite connection and cursor
    def __del__(self):
        self.curs.close()
        self.conn.close()


# abstract class to import and save a table
class TableUtils:
    @staticmethod
    def readFile(file, fType="", commentLineStart="#", **args):
        fType = fType.lower()
        bn = basename(file).lower()

        if fType == "xls" or bn.endswith(".xls") or fType == "xlsx" or bn.endswith(".xlsx"):
            if not (args.has_key("sheetName")):
                args["sheetName"] = ""
            return TableUtilsXLS.readFile(file, sheetName=args["sheetName"], commentLineStart=commentLineStart)
        elif fType == "tsv" or bn.endswith(".tsv") or bn.endswith(".txt") or bn.endswith(".csv"):
            if not (args.has_key("delim")):
                args["delim"] = "\t"
            return TableUtilsCSV.readFile(file, delim=args["delim"], commentLineStart=commentLineStart)
        elif fType == "sqlite" or bn.endswith(".sqlite"):
            if not (args.has_key("tableName")):
                args["tableName"] = ""
            return TableUtilsSQL.readFile(file, tableName=args["tableName"])
        else:
            raise Exception("Unknown file type. Please use xls, tsv, sqlite")

    @staticmethod
    def saveFile(table, file, fType="", writeComments=True, **args):
        fType = fType.lower()
        bn = basename(file).lower()

        if not (args.has_key("where")):
            args["where"] = ""
        if not (args.has_key("order")):
            args["order"] = "__internalID"
        if not (args.has_key("select")):
            args["select"] = ""
        if not (args.has_key("cols")):
            args["cols"] = []

        if fType == "xls" or bn.endswith(".xls") or fType == "xlsx" or bn.endswith(".xlsx"):
            if not (args.has_key("sheetName")):
                args["sheetName"] = ""
            return TableUtilsXLS.saveFile(table, file, sheetName=args["sheetName"], where=args["where"],
                                          order=args["order"], select=args["select"], cols=args["cols"], writeComments=writeComments)
        elif fType == "tsv" or bn.endswith(".tsv") or bn.endswith(".txt"):
            if not (args.has_key("delim")):
                args["delim"] = ""
            if not (args.has_key("newLine")):
                args["newLine"] = ""
            return TableUtilsCSV.saveFile(table, file, delim=args["delim"], newLine=args["newLine"],
                                          where=args["where"], order=args["order"], select=args["select"],
                                          cols=args["cols"], writeComments=writeComments)
        elif fType == "sqlite" or bn.endswith(".sqlite"):
            if not (args.has_key("tableName")):
                args["tableName"] = ""
            return TableUtilsSQL.saveFile(table, file, tableName=args["tableName"], writeComments=writeComments)
        else:
            raise Exception("Unknown file type. Please use xls, tsv, sqlite")


# read / write data table from / to a CSV file
class TableUtilsCSV:
    @staticmethod
    def readFile(file, delim="\t", commentLineStart="#"):
        ret = None
        with open(file, "rb") as fi:
            i = 0
            headers = []
            rows = []
            comments = []
            for line in fi:
                if len(line)>0 and not line.startswith(commentLineStart):
                    cells = line.strip().split(delim)
                    if i == 0:
                        headers = [c.replace("-","_") for c in cells]
                    else:
                        rows.append(cells)
                    i = i + 1
                if line.startswith(commentLineStart):
                    comments.append(line.strip())
            ret = Table(headers=headers, rows=rows, comments=comments)
        return ret

    @staticmethod
    def saveFile(table, file, where="", order="", delim="\t", newLine="\n", select="", cols=None, writeComments=True):
        if cols is None:
            cols = []

        if len(delim) == 0:
            delim = "\t"
        if len(newLine) == 0:
            newLine = "\n"
        if len(where) > 0:
            where = "where " + where
        if len(order) > 0:
            order = " order by " + order
        if len(select) == 0:
            if len(cols) == 0:
                select = "select :allCols: from :table: %s %s" % (where, order)
            else:
                select = "select %s from :table: %s %s" % (", ".join(cols), where, order)

        with open(file, "wb") as fo:
            headers = False
            rowsWritten=0
            for row in table.curs.execute(table.updateQuery(select)):
                if not headers:
                    fo.write(delim.join([str(d[0]) for d in table.curs.description]))
                    fo.write(newLine)
                    headers = True
                    rowsWritten+=1
                fo.write(delim.join([str(c) for c in row]))
                fo.write(newLine)
                rowsWritten+=1

            if rowsWritten==0:
                if not headers:
                    fo.write(delim.join([str(d[0]) for d in table.curs.description]))
                    fo.write(newLine)
                    headers = True
                    rowsWritten+=1


            if writeComments:
                for comment in table.getComments():
                    fo.write(str(comment))
                    fo.write(newLine)


from openpyxl import Workbook, load_workbook
from openpyxl.cell import get_column_letter

# read / write data table from / to an excel file
class TableUtilsXLS:
    @staticmethod
    def readFile(file, sheetName="", tableName=None, useColumns=None, commentLineStart="#"):
        if sheetName == "":
            bn = basename(file)
            rf = bn.rfind(".")
            if rf > -1:
                sheetName = bn[:rf]
            else:
                sheetName = bn

        ret = None
        rb = load_workbook(file)

        ind = 0
        sheetInd = -1
        for s in rb.get_sheet_names():
            if s == sheetName:
                sheetInd = ind
            ind = ind + 1
        assert sheetInd != -1, "Sheet not found"
        sh = rb.get_sheet_by_name(sheetName)

        i = 0
        remainingCols = True
        headers = []
        rows = []
        comments = []
        colIDs = []
        firstColFound=False
        curRow = 0
        while not firstColFound:
            val=sh.cell('%s%s' % (get_column_letter(0 + 1), curRow + 1)).value
            if str(val).startswith(commentLineStart):
                comments.append(val)
                curRow += 1
                continue
            else:
                firstColFound=True
            while remainingCols:
                col = get_column_letter(i + 1)
                row = curRow + 1
                try:
                    if sh.cell('%s%s' % (col, row)).value is None or str(sh.cell('%s%s' % (col, row)).value) == "":
                        remainingCols = False
                        continue
                except:
                    remainingCols = False
                    continue

                header = str(sh.cell('%s%s' % (col, row)).value)
                if useColumns is None or header in useColumns:
                    headers.append(header)
                    colIDs.append(i)
            i = i + 1

        r = 1 + curRow
        allEmpty = False
        while not allEmpty:
            val=sh.cell('%s%s' % (get_column_letter(0 + 1), r + 1)).value
            if str(val).startswith(commentLineStart):
                comments.append(val)
                r = r +1
                continue

            rowValues = []
            allEmpty = True
            for j in colIDs:
                col = get_column_letter(j + 1)
                row = r + 1
                try:
                    if sh.cell('%s%s' % (col, row)).value is None or str(sh.cell('%s%s' % (col, row)).value) == "":
                        rowValues.append("")
                        j = j + 1
                        continue
                except:
                    rowValues.append("")
                    j = j + 1
                    continue
                rowValues.append(str(sh.cell('%s%s' % (col, row)).value))
                allEmpty = False
            if not allEmpty:
                rows.append(rowValues)

            r = r + 1

        #assert all([cols.has_key(x) and cols[x]!=-1 for x in columns.values()]), "Could not find all specified columns in excel file"
        ret = Table(headers=headers, rows=rows, tableName=tableName, comments=comments)

        return ret

    @staticmethod
    def saveFile(table, file, sheetName="", where="", order="", select="", cols=None, writeComments=True):
        if cols is None:
            cols = []

        if sheetName == "":
            bn = basename(file)
            rf = bn.rfind(".")
            if rf > -1:
                sheetName = bn[:rf]
            else:
                sheetName = bn
        if len(where) > 0:
            where = "where " + where
        if len(order) > 0:
            order = " order by " + order
        if len(select) == 0:
            if len(cols) == 0:
                select = "select :allCols: from :table: %s %s" % (where, order)
            else:
                select = "select %s from :table: %s %s" % (", ".join(cols), where, order)

        rb = -1
        sheet = -1
        import os.path

        if os.path.isfile(file):
            rb = load_workbook(file)
            sheet = rb.create_sheet()
            sheet.title = sheetName
        else:
            rb = Workbook()
            sheet = rb.get_active_sheet()
            sheet.title = sheetName

        j = 0
        headers = False
        for rowValues in table.curs.execute(table.updateQuery(select)):
            i = 0
            if not headers:
                for d in table.curs.description:
                    col = get_column_letter(i + 1)
                    row = j + 1

                    sheet.cell('%s%s' % (col, row)).value = d[0]
                    i = i + 1
                headers = True
                i = 0
                j += 1

            for c in rowValues:
                col = get_column_letter(i + 1)
                row = j + 1
                sheet.cell('%s%s' % (col, row)).value = c
                i = i + 1
            j = j + 1

        if writeComments:
            for comment in table.getComments():
                col = get_column_letter(0 + 1)
                row = j + 1
                sheet.cell('%s%s' % (col, row)).value = comment

                j = j + 1

        rb.save(file)


# read / write data table from / to an SQL table
class TableUtilsSQL:
    @staticmethod
    def readFile(file, tableName=""):
        raise Exception("Not implemented yet", "Not implemented yet")

    @staticmethod
    def saveFile(table, file, tableName, where="", order="", select="", writeComments=True):
        raise Exception("Not implemented yet", "Not implemented yet")



